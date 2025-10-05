# import pandas as pd
# import openpyxl
# from openpyxl import load_workbook
# import tkinter as tk
# from tkinter import ttk, filedialog, messagebox
# import os

# class ObserFileGeneratorApp:
#     def __init__(self):
#         self.root = tk.Tk()
#         self.root.title("Obser Files Generator")
#         self.root.geometry("700x500")
#         self.root.minsize(700, 500)
        
#         # Center the window on screen
#         self.root.update_idletasks()
#         x = (self.root.winfo_screenwidth() // 2) - (700 // 2)
#         y = (self.root.winfo_screenheight() // 2) - (500 // 2)
#         self.root.geometry(f"700x500+{x}+{y}")
        
#         self.root.configure(bg="white")
        
#         self.workbook_path = None
#         self.nyuuryoku_params = {
#             'data_count': 8,
#             'prediction_years': 10,
#             'lambda_constant': 0.02,
#             'inspection_years': list(range(27, 43))
#         }
        
#         # Sheet mappings
#         self.sheet_mappings = {
#             'obser1.txt': '割算結果(補修考慮)',
#             'obser2.txt': '割算結果(補修無視)', 
#             'obser3.txt': '補修無視',
#             'obser4.txt': '補修考慮',
#             'obser5.txt': '新しい演算(補修無視)',
#             'obser6.txt': '新しい演算(補修考慮)',
#             'obser7.txt': '割算結果-新しい演算(補修無視)',
#             'obser8.txt': '割算結果-新しい演算(補修考慮)'
#         }
        
#         self.create_main_gui()
    
#     def create_main_gui(self):
#         """Create main GUI"""
#         main_frame = tk.Frame(self.root, bg="white", padx=40, pady=40)
#         main_frame.pack(fill="both", expand=True)
        
#         # Header section
#         header_frame = tk.Frame(main_frame, bg='white')
#         header_frame.pack(fill="x", pady=(0, 30))
        
#         # Title with icon effect
#         title_label = tk.Label(header_frame, text="📄 Obser Files Generator", 
#                               font=("Arial", 20, "bold"), fg="#2c3e50", bg='white')
#         title_label.pack()
        
#         # Subtitle
#         subtitle_label = tk.Label(header_frame, text="Advanced 8-File Generation System", 
#                                  font=("Arial", 12), fg="#7f8c8d", bg='white')
#         subtitle_label.pack(pady=(5, 0))
        
#         # Status section
#         status_frame = tk.LabelFrame(main_frame, text="📊 Status", font=("Arial", 12, "bold"), 
#                                    fg="#2c3e50", bg='white', bd=2, relief="solid")
#         status_frame.pack(fill="x", pady=(0, 20), ipady=10)
        
#         self.status_label = tk.Label(status_frame, text="💾 Ready to select Excel workbook...", 
#                                     font=("Arial", 11), fg="#3498db", bg='white')
#         self.status_label.pack(pady=5, padx=15, anchor="w")
        
#         # File info section
#         info_frame = tk.LabelFrame(main_frame, text="📄 Files to Generate", font=("Arial", 12, "bold"), 
#                                  fg="#2c3e50", bg='white', bd=2, relief="solid")
#         info_frame.pack(fill="x", pady=(0, 20), ipady=10)
        
#         files_text = ("• obser1.txt ← 割算結果(補修考慮) • obser2.txt ← 割算結果(補修無視)\n"
#                      "• obser3.txt ← 補修無視 • obser4.txt ← 補修考慮\n"
#                      "• obser5.txt ← 新しい演算(補修無視) • obser6.txt ← 新しい演算(補修考慮)\n"
#                      "• obser7.txt ← 割算結果-新しい演算(補修無視) • obser8.txt ← 割算結果-新しい演算(補修考慮)")
        
#         files_label = tk.Label(info_frame, text=files_text, 
#                               font=("Arial", 10), fg="#34495e", bg='white')
#         files_label.pack(pady=5, padx=15)
        
#         # Button section
#         button_frame = tk.Frame(main_frame, bg='white')
#         button_frame.pack(pady=(0, 20))
        
#         # Select Excel File button
#         select_btn = tk.Button(button_frame, text="📁 Select Excel Workbook", 
#                              command=self.select_workbook, 
#                              bg="#3498db", fg="white", 
#                              width=25, height=2, font=("Arial", 12, "bold"),
#                              relief="flat", cursor="hand2")
#         select_btn.pack(pady=10)
        
#         # Actions frame (initially hidden)
#         self.actions_frame = tk.Frame(main_frame, bg='white')
        
#         tk.Label(self.actions_frame, text="What would you like to do?", 
#                 font=("Arial", 14, "bold"), fg="#2c3e50", bg='white').pack(pady=(20, 15))
        
#         # Action buttons
#         btn_frame = tk.Frame(self.actions_frame, bg='white')
#         btn_frame.pack(pady=15)
        
#         tk.Button(btn_frame, text="⚙️ View/Edit Parameters", 
#                  command=self.show_parameter_editor,
#                  bg="#9b59b6", fg="white", width=22, height=2, 
#                  font=("Arial", 11, "bold"), relief="flat", cursor="hand2").pack(side="left", padx=10)
        
#         tk.Button(btn_frame, text="🚀 Generate Files", 
#                  command=self.generate_obser_files,
#                  bg="#e67e22", fg="white", width=22, height=2, 
#                  font=("Arial", 11, "bold"), relief="flat", cursor="hand2").pack(side="left", padx=10)
        
#         # Footer
#         footer_label = tk.Label(main_frame, text="Powered by Advanced File Generation Engine • 8 Obser Files in One Process", 
#                                font=("Arial", 9), fg="#95a5a6", bg='white')
#         footer_label.pack(side="bottom")

#     def select_workbook(self):
#         """Select and validate workbook"""
#         # Update status
#         self.status_label.config(text="🔍 Opening file browser...", fg="#e67e22")
#         self.root.update()
        
#         self.workbook_path = filedialog.askopenfilename(
#             title="Select Excel Workbook",
#             filetypes=[("Excel files", "*.xlsx *.xls")],
#             initialdir=os.path.expanduser("~")
#         )
        
#         if not self.workbook_path:
#             self.status_label.config(text="❌ No file selected", fg="#e74c3c")
#             return
        
#         # Update status
#         self.status_label.config(text="🔍 Validating Excel sheets...", fg="#e67e22")
#         self.root.update()
        
#         try:
#             wb = load_workbook(self.workbook_path)
#             required_sheets = list(self.sheet_mappings.values())
#             missing_sheets = [sheet for sheet in required_sheets 
#                             if sheet not in wb.sheetnames]
            
#             if missing_sheets:
#                 self.status_label.config(text=f"❌ Missing required sheets", fg="#e74c3c")
#                 messagebox.showerror("Missing Sheets", 
#                                    f"Required sheets not found:\n" + 
#                                    "\n".join(missing_sheets))
#                 return
            
#             # Load parameters from 入力値 sheet
#             self.status_label.config(text="📥 Loading parameters...", fg="#e67e22")
#             self.root.update()
            
#             self.load_nyuuryoku_parameters()
            
#             self.status_label.config(text=f"✅ File loaded: {os.path.basename(self.workbook_path)}", 
#                                    fg="#27ae60")
#             self.actions_frame.pack(fill="x", pady=20)
            
#         except Exception as e:
#             self.status_label.config(text="❌ Error loading file", fg="#e74c3c")
#             messagebox.showerror("Error", f"Error loading file:\n{str(e)}")

#     def load_nyuuryoku_parameters(self):
#         """Load parameters from 入力値 sheet"""
#         try:
#             # Load 入力値 sheet
#             nyuuryoku_df = pd.read_excel(self.workbook_path, sheet_name='入力値', header=None)
            
#             # Find parameters by matching column headers
#             if len(nyuuryoku_df) >= 2:
#                 headers = nyuuryoku_df.iloc[0]  # First row as headers
                
#                 # Find and extract parameters
#                 for i, header in enumerate(headers):
#                     if pd.notna(header):
#                         header_str = str(header)
#                         if 'データ個数' in header_str:
#                             try:
#                                 self.nyuuryoku_params['data_count'] = int(nyuuryoku_df.iloc[1, i])
#                             except (ValueError, TypeError):
#                                 pass
#                         elif '予測年数' in header_str:
#                             try:
#                                 self.nyuuryoku_params['prediction_years'] = int(nyuuryoku_df.iloc[1, i])
#                             except (ValueError, TypeError):
#                                 pass
#                         elif 'λ定数' in header_str:
#                             try:
#                                 self.nyuuryoku_params['lambda_constant'] = float(nyuuryoku_df.iloc[1, i])
#                             except (ValueError, TypeError):
#                                 pass
#                         elif '点検年度に対応した年' in header_str:
#                             # Extract years from this column
#                             years = []
#                             for row_idx in range(1, len(nyuuryoku_df)):
#                                 val = nyuuryoku_df.iloc[row_idx, i]
#                                 if pd.notna(val):
#                                     try:
#                                         year = int(val)
#                                         if 20 <= year <= 50:
#                                             years.append(year)
#                                     except (ValueError, TypeError):
#                                         break
                            
#                             if years:
#                                 self.nyuuryoku_params['inspection_years'] = years
            
#         except Exception as e:
#             print(f"Could not load 入力値 sheet: {e}")

#     def show_parameter_editor(self):
#         """Show parameter editor"""
#         editor = tk.Toplevel(self.root)
#         editor.title("Edit Parameters")
#         editor.geometry("650x600")
#         editor.minsize(650, 600)
        
#         # Center the window
#         editor.update_idletasks()
#         x = (editor.winfo_screenwidth() // 2) - (650 // 2)
#         y = (editor.winfo_screenheight() // 2) - (600 // 2)
#         editor.geometry(f"650x600+{x}+{y}")
        
#         editor.grab_set()
#         editor.configure(bg="white")
        
#         main_frame = tk.Frame(editor, bg="white", padx=30, pady=30)
#         main_frame.pack(fill="both", expand=True)
        
#         # Header
#         tk.Label(main_frame, text="⚙️ Edit Input Parameters", 
#                 font=("Arial", 16, "bold"), fg="#2c3e50", bg="white").pack(pady=(0, 25))
        
#         # Parameter fields
#         fields_frame = tk.LabelFrame(main_frame, text="📊 Parameters", font=("Arial", 12, "bold"), 
#                                    fg="#2c3e50", bg="white", bd=2, relief="solid")
#         fields_frame.pack(fill="x", pady=(0, 20), ipady=15)
        
#         params_grid = tk.Frame(fields_frame, bg="white")
#         params_grid.pack(pady=10, padx=15)
        
#         # データ個数
#         tk.Label(params_grid, text="データ個数:", font=("Arial", 11, "bold"), 
#                 bg="white", fg="#34495e").grid(row=0, column=0, sticky="w", pady=8, padx=(0, 15))
#         self.data_count_var = tk.StringVar(value=str(self.nyuuryoku_params['data_count']))
#         tk.Entry(params_grid, textvariable=self.data_count_var, width=20, 
#                 font=("Arial", 11)).grid(row=0, column=1, pady=8)
        
#         # 予測年数
#         tk.Label(params_grid, text="予測年数:", font=("Arial", 11, "bold"), 
#                 bg="white", fg="#34495e").grid(row=1, column=0, sticky="w", pady=8, padx=(0, 15))
#         self.prediction_years_var = tk.StringVar(value=str(self.nyuuryoku_params['prediction_years']))
#         tk.Entry(params_grid, textvariable=self.prediction_years_var, width=20, 
#                 font=("Arial", 11)).grid(row=1, column=1, pady=8)
        
#         # λ定数
#         tk.Label(params_grid, text="λ定数:", font=("Arial", 11, "bold"), 
#                 bg="white", fg="#34495e").grid(row=2, column=0, sticky="w", pady=8, padx=(0, 15))
#         self.lambda_var = tk.StringVar(value=str(self.nyuuryoku_params['lambda_constant']))
#         tk.Entry(params_grid, textvariable=self.lambda_var, width=20, 
#                 font=("Arial", 11)).grid(row=2, column=1, pady=8)
        
#         # Years section
#         years_frame = tk.LabelFrame(main_frame, text="📅 点検年度に対応した年", font=("Arial", 12, "bold"), 
#                                   fg="#2c3e50", bg="white", bd=2, relief="solid")
#         years_frame.pack(fill="both", expand=True, pady=(0, 20))
        
#         years_container = tk.Frame(years_frame, bg="white")
#         years_container.pack(fill="both", expand=True, pady=10, padx=15)
        
#         self.years_entry = tk.Text(years_container, height=6, width=60, font=("Arial", 11),
#                                   relief="solid", bd=1)
#         years_scrollbar = ttk.Scrollbar(years_container, orient="vertical", command=self.years_entry.yview)
#         self.years_entry.configure(yscrollcommand=years_scrollbar.set)
        
#         current_years = ' '.join(map(str, self.nyuuryoku_params['inspection_years']))
#         self.years_entry.insert("1.0", current_years)
        
#         self.years_entry.pack(side="left", fill="both", expand=True)
#         years_scrollbar.pack(side="right", fill="y")
        
#         # Info label
#         info_label = tk.Label(years_frame, text="💡 Enter years separated by spaces (e.g., 27 28 29 30...)", 
#                              font=("Arial", 9), fg="#7f8c8d", bg="white")
#         info_label.pack(pady=(0, 10))
        
#         # Buttons
#         button_frame = tk.Frame(main_frame, bg="white")
#         button_frame.pack(fill="x", pady=20)
        
#         def save_and_generate():
#             if self.validate_and_save_params():
#                 editor.destroy()
#                 self.generate_obser_files()
        
#         tk.Button(button_frame, text="💾 Save & Generate", command=save_and_generate,
#                  bg="#27ae60", fg="white", width=20, height=2, 
#                  font=("Arial", 11, "bold"), relief="flat", cursor="hand2").pack(side="left", padx=10)
        
#         tk.Button(button_frame, text="❌ Cancel", command=editor.destroy,
#                  bg="#e74c3c", fg="white", width=15, height=2, 
#                  font=("Arial", 11, "bold"), relief="flat", cursor="hand2").pack(side="left", padx=10)

#     def validate_and_save_params(self):
#         """Validate and save parameters"""
#         try:
#             # Validate inputs
#             data_count = int(self.data_count_var.get())
#             pred_years = int(self.prediction_years_var.get())
#             lambda_const = float(self.lambda_var.get())
            
#             if data_count <= 0 or pred_years <= 0 or lambda_const <= 0:
#                 messagebox.showerror("Error", "All values must be positive")
#                 return False
            
#             # Parse years
#             years_text = self.years_entry.get("1.0", tk.END).strip()
#             years = []
#             for year_str in years_text.split():
#                 try:
#                     year = int(year_str)
#                     if 20 <= year <= 50:
#                         years.append(year)
#                 except ValueError:
#                     pass
            
#             if not years:
#                 messagebox.showerror("Error", "Please enter valid years (20-50 range)")
#                 return False
            
#             # Update parameters
#             self.nyuuryoku_params.update({
#                 'data_count': data_count,
#                 'prediction_years': pred_years,
#                 'lambda_constant': lambda_const,
#                 'inspection_years': years
#             })
            
#             # Save to Excel
#             self.save_nyuuryoku_parameters()
#             return True
            
#         except ValueError:
#             messagebox.showerror("Error", "Please enter valid numeric values")
#             return False
#         except Exception as e:
#             messagebox.showerror("Error", f"Error saving parameters: {str(e)}")
#             return False

#     def save_nyuuryoku_parameters(self):
#         """Save parameters to 入力値 sheet"""
#         try:
#             wb = load_workbook(self.workbook_path)
            
#             if '入力値' in wb.sheetnames:
#                 wb.remove(wb['入力値'])
            
#             ws = wb.create_sheet('入力値')
            
#             # Headers in row 1
#             ws['A1'] = 'データ個数'
#             ws['B1'] = '予測年数' 
#             ws['C1'] = 'λ定数'
#             ws['D1'] = '点検年度に対応した年'
            
#             # Values in row 2
#             ws['A2'] = self.nyuuryoku_params['data_count']
#             ws['B2'] = self.nyuuryoku_params['prediction_years']
#             ws['C2'] = self.nyuuryoku_params['lambda_constant']
            
#             # Years in column D starting from row 2
#             for i, year in enumerate(self.nyuuryoku_params['inspection_years']):
#                 ws[f'D{i+2}'] = year
            
#             wb.save(self.workbook_path)
#             wb.close()
            
#         except Exception as e:
#             raise Exception(f"Error saving to Excel: {str(e)}")

#     def generate_obser_files(self):
#         """Generate all obser files with improved progress tracking"""
#         try:
#             output_directory = os.path.dirname(self.workbook_path)
#             generated_files = []
            
#             # Create professional progress window
#             progress_window = tk.Toplevel(self.root)
#             progress_window.title("Obser Files Generator - Processing")
#             progress_window.geometry("700x500")
#             progress_window.grab_set()
#             progress_window.configure(bg='white')
            
#             # Center the progress window
#             progress_window.update_idletasks()
#             x = (progress_window.winfo_screenwidth() // 2) - (700 // 2)
#             y = (progress_window.winfo_screenheight() // 2) - (500 // 2)
#             progress_window.geometry(f"700x500+{x}+{y}")
            
#             # Main container
#             main_frame = tk.Frame(progress_window, bg='white', padx=30, pady=20)
#             main_frame.pack(fill="both", expand=True)
            
#             # Header
#             title_label = tk.Label(main_frame, text="🚀 Generating Obser Files", 
#                                   font=("Arial", 18, "bold"), fg="#2c3e50", bg='white')
#             title_label.pack(pady=(0, 20))
            
#             # Status section
#             status_frame = tk.LabelFrame(main_frame, text="📊 Current Status", 
#                                        font=("Arial", 12, "bold"), fg="#2c3e50", bg='white', 
#                                        bd=2, relief="solid")
#             status_frame.pack(fill="x", pady=(0, 15), ipady=10)
            
#             status_label = tk.Label(status_frame, text="🔄 Initializing generation process...", 
#                                   font=("Arial", 11), fg="#3498db", bg='white')
#             status_label.pack(pady=5, padx=15, anchor="w")
            
#             # Progress section
#             progress_frame = tk.LabelFrame(main_frame, text="⏳ Progress", 
#                                          font=("Arial", 12, "bold"), fg="#2c3e50", bg='white', 
#                                          bd=2, relief="solid")
#             progress_frame.pack(fill="x", pady=(0, 20), ipady=15)
            
#             progress_label = tk.Label(progress_frame, text="⏸ Step 0/8: Starting generation...", 
#                                     font=("Arial", 11), fg="#e67e22", bg='white')
#             progress_label.pack(pady=(5, 10), padx=15, anchor="w")
            
#             progress_bar = ttk.Progressbar(progress_frame, mode='determinate', maximum=8, length=600)
#             progress_bar.pack(pady=(0, 10), padx=15)
            
#             # Files section
#             files_frame = tk.LabelFrame(main_frame, text="📄 Files to Generate", 
#                                       font=("Arial", 12, "bold"), fg="#2c3e50", bg='white', 
#                                       bd=2, relief="solid")
#             files_frame.pack(fill="both", expand=True, pady=(0, 15))
            
#             files_container = tk.Frame(files_frame, bg='white')
#             files_container.pack(fill="both", expand=True, pady=10, padx=15)
            
#             files_text = tk.Text(files_container, height=10, font=("Arial", 10), bg="#f8f9fa", 
#                                relief="solid", bd=1, wrap=tk.WORD)
#             files_scroll = ttk.Scrollbar(files_container, orient="vertical", command=files_text.yview)
#             files_text.configure(yscrollcommand=files_scroll.set)
            
#             # Add file list
#             for i, (obser_file, sheet_name) in enumerate(self.sheet_mappings.items(), 1):
#                 files_text.insert(tk.END, f"{i}. {obser_file} ← {sheet_name}\n")
            
#             files_text.config(state="disabled")
#             files_text.pack(side="left", fill="both", expand=True)
#             files_scroll.pack(side="right", fill="y")
            
#             progress_window.update()
            
#             # Generate files with improved progress tracking
#             total_files = len(self.sheet_mappings)
            
#             for i, (obser_file, sheet_name) in enumerate(self.sheet_mappings.items(), 1):
#                 # Update progress
#                 status_label.config(text=f"🔄 Generating {obser_file}...", fg="#e67e22")
#                 progress_label.config(text=f"▶️ Step {i}/{total_files}: Processing {sheet_name}...")
#                 progress_bar['value'] = i - 0.5
#                 progress_window.update()
                
#                 try:
#                     output_path = os.path.join(output_directory, obser_file)
#                     # Faster file creation without Excel updates
#                     self.create_obser_file_fast(sheet_name, output_path)
#                     generated_files.append(obser_file)
                    
#                     # Update completion status
#                     status_label.config(text=f"✅ Completed {obser_file}", fg="#27ae60")
#                     progress_bar['value'] = i
#                     progress_window.update()
                    
#                 except Exception as e:
#                     status_label.config(text=f"❌ Error with {obser_file}: {str(e)[:30]}...", fg="#e74c3c")
#                     print(f"Error generating {obser_file}: {e}")
#                     progress_window.update()
            
#             # Show completion
#             status_label.config(text="🎉 All files generated successfully!", fg="#27ae60")
#             progress_label.config(text=f"✅ Complete: Generated {len(generated_files)}/{total_files} files")
#             progress_bar['value'] = total_files
#             progress_window.update()
            
#             # Auto-close after 3 seconds with countdown
#             for countdown in range(3, 0, -1):
#                 status_label.config(text=f"🎉 Success! Auto-closing in {countdown} seconds...")
#                 progress_window.update()
#                 progress_window.after(1000)
            
#             progress_window.destroy()
            
#             # Show final completion message
#             completion_msg = f"✅ Successfully generated {len(generated_files)} obser files!\n\n"
#             completion_msg += "\n".join(f"• {file}" for file in generated_files)
#             completion_msg += f"\n\n📁 Location: {output_directory}"
            
#             messagebox.showinfo("Generation Complete", completion_msg)
            
#         except Exception as e:
#             if 'progress_window' in locals():
#                 progress_window.destroy()
#             messagebox.showerror("Error", f"Error generating files: {str(e)}")

#     def create_obser_file_fast(self, sheet_name, output_path):
#         """Create obser file without Excel updating for faster processing"""
#         try:
#             # Load sheet data (read-only, faster)
#             sheet_df = pd.read_excel(self.workbook_path, sheet_name=sheet_name)
            
#             # Sort by last column in descending order
#             if len(sheet_df) > 0 and len(sheet_df.columns) > 0:
#                 last_col = sheet_df.columns[-1]
#                 sheet_df = sheet_df.sort_values(by=last_col, ascending=False)
            
#             with open(output_path, 'w', encoding='utf-8') as f:
#                 # First line: parameters
#                 f.write(f"{self.nyuuryoku_params['data_count']} {self.nyuuryoku_params['prediction_years']} {self.nyuuryoku_params['lambda_constant']}\n")
                
#                 # Second line: years
#                 years_line = ' '.join(map(str, self.nyuuryoku_params['inspection_years']))
#                 f.write(f"{years_line}\n")
                
#                 # Third line: blank
#                 f.write("\n")
                
#                 # Find 構造物番号 column
#                 kozo_col_idx = None
#                 for i, col in enumerate(sheet_df.columns):
#                     if '構造物番号' in str(col):
#                         kozo_col_idx = i
#                         break
                
#                 if kozo_col_idx is None:
#                     raise Exception(f"構造物番号 column not found in {sheet_name}")
                
#                 # Get columns from 構造物番号 onwards
#                 columns_to_export = sheet_df.columns[kozo_col_idx:]
                
#                 # Write data rows
#                 for _, row in sheet_df.iterrows():
#                     row_data = []
#                     for col in columns_to_export:
#                         value = row[col]
                        
#                         if pd.isna(value) or value == '':
#                             row_data.append('')
#                         else:
#                             try:
#                                 numeric_val = float(value)
#                                 if numeric_val == 0:
#                                     row_data.append('0.1')
#                                 elif numeric_val == int(numeric_val):
#                                     row_data.append(str(int(numeric_val)))
#                                 else:
#                                     row_data.append(str(round(numeric_val, 3)))
#                             except (ValueError, TypeError):
#                                 if str(value) == '0':
#                                     row_data.append('0.1')
#                                 else:
#                                     row_data.append(str(value))
                    
#                     f.write('\t'.join(row_data) + '\n')
            
#         except Exception as e:
#             raise Exception(f"Error creating {output_path}: {str(e)}")

#     def create_obser_file(self, sheet_name, output_path):
#         """Create obser file with sorting and 0 value replacement"""
#         try:
#             # Load sheet data
#             sheet_df = pd.read_excel(self.workbook_path, sheet_name=sheet_name)
            
#             # Sort by last column in descending order (matching VBA logic)
#             if len(sheet_df) > 0 and len(sheet_df.columns) > 0:
#                 last_col = sheet_df.columns[-1]
#                 sheet_df = sheet_df.sort_values(by=last_col, ascending=False)
                
#                 # Save the sorted data back to Excel sheet
#                 with pd.ExcelWriter(self.workbook_path, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
#                     sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
#             with open(output_path, 'w', encoding='utf-8') as f:
#                 # First line: parameters separated by spaces
#                 f.write(f"{self.nyuuryoku_params['data_count']} {self.nyuuryoku_params['prediction_years']} {self.nyuuryoku_params['lambda_constant']}\n")
                
#                 # Second line: years separated by single spaces
#                 years_line = ' '.join(map(str, self.nyuuryoku_params['inspection_years']))
#                 f.write(f"{years_line}\n")
                
#                 # Third line: blank line
#                 f.write("\n")
                
#                 # Find 構造物番号 column and export from there onwards
#                 kozo_col_idx = None
#                 for i, col in enumerate(sheet_df.columns):
#                     if '構造物番号' in str(col):
#                         kozo_col_idx = i
#                         break
                
#                 if kozo_col_idx is None:
#                     raise Exception(f"構造物番号 column not found in {sheet_name}")
                
#                 # Get all columns from 構造物番号 onwards (including 構造物番号)
#                 columns_to_export = sheet_df.columns[kozo_col_idx:]
                
#                 # Write data rows (tab-separated)
#                 for _, row in sheet_df.iterrows():
#                     row_data = []
#                     for col in columns_to_export:
#                         value = row[col]
                        
#                         # Handle empty/NaN values
#                         if pd.isna(value) or value == '':
#                             row_data.append('')
#                         else:
#                             try:
#                                 numeric_val = float(value)
#                                 # Replace 0 with 0.1 (matching VBA logic)
#                                 if numeric_val == 0:
#                                     row_data.append('0.1')
#                                 elif numeric_val == int(numeric_val):
#                                     row_data.append(str(int(numeric_val)))
#                                 else:
#                                     row_data.append(str(round(numeric_val, 3)))
#                             except (ValueError, TypeError):
#                                 # Non-numeric values (like 構造物番号)
#                                 if str(value) == '0':
#                                     row_data.append('0.1')
#                                 else:
#                                     row_data.append(str(value))
                    
#                     f.write('\t'.join(row_data) + '\n')
            
#         except Exception as e:
#             raise Exception(f"Error creating {output_path}: {str(e)}")
        

#     def run(self):
#         """Run the application"""
#         self.root.mainloop()


# if __name__ == "__main__":
#     print("Obser Files Generator")
#     print("===================")
#     print("Sheet Mappings:")
#     sheet_mappings = {
#         'obser1.txt': '割算結果(補修考慮)',
#         'obser2.txt': '割算結果(補修無視)', 
#         'obser3.txt': '補修無視',
#         'obser4.txt': '補修考慮',
#         'obser5.txt': '新しい演算(補修無視)',
#         'obser6.txt': '新しい演算(補修考慮)',
#         'obser7.txt': '割算結果-新しい演算(補修無視)',
#         'obser8.txt': '割算結果-新しい演算(補修考慮)'
#     }
    
#     for obser_file, sheet_name in sheet_mappings.items():
#         print(f"• {obser_file} ← {sheet_name}")
#     print("===================")
    
#     app = ObserFileGeneratorApp()
#     app.run()


import pandas as pd
import openpyxl
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import time

class ObserFileGeneratorApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Obser Files Generator")
        self.root.geometry("700x580")  # Reduced height since removing Current Parameters
        self.root.minsize(700, 580)
        
        # Center the window on screen
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() // 2) - (700 // 2)
        y = (self.root.winfo_screenheight() // 2) - (580 // 2)
        self.root.geometry(f"700x580+{x}+{y}")
        
        self.root.configure(bg="white")
        
        self.workbook_path = None
        self.nyuuryoku_params = {
            'data_count': 8,
            'prediction_years': 10,
            'lambda_constant': 0.02,
            'inspection_years': list(range(27, 43))
        }
        
        # Sheet mappings
        self.sheet_mappings = {
            'obser1.txt': '割算結果(補修考慮)',
            'obser2.txt': '割算結果(補修無視)', 
            'obser3.txt': '補修無視',
            'obser4.txt': '補修考慮',
            'obser5.txt': '新しい演算(補修無視)',
            'obser6.txt': '新しい演算(補修考慮)',
            'obser7.txt': '割算結果-新しい演算(補修無視)',
            'obser8.txt': '割算結果-新しい演算(補修考慮)'
        }
        
        self.create_main_gui()
    
    def create_professional_button(self, parent, text, command, bg_color, hover_color, **kwargs):
        """Create a professional button with hover effects"""
        btn = tk.Button(parent, text=text, command=command, bg=bg_color, fg="white",
                       font=("Arial", 10, "bold"), relief="flat", cursor="hand2",
                       activebackground=hover_color, activeforeground="white",
                       bd=0, padx=20, pady=10, **kwargs)
        
        # Add hover effects
        def on_enter(e):
            if btn['state'] != 'disabled':
                btn.config(bg=hover_color)
        
        def on_leave(e):
            if btn['state'] != 'disabled':
                btn.config(bg=bg_color)
        
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)
        
        return btn
    
    def create_main_gui(self):
        """Create main GUI with organized layout"""
        # Main container
        main_container = tk.Frame(self.root, bg="white")
        main_container.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Header section
        header_frame = tk.Frame(main_container, bg='white')
        header_frame.pack(fill="x", pady=(0, 20))
        
        title_label = tk.Label(header_frame, text="📄 Obser Files Generator", 
                              font=("Arial", 18, "bold"), fg="#2c3e50", bg='white')
        title_label.pack()
        
        subtitle_label = tk.Label(header_frame, text="Advanced 8-File Generation System", 
                                 font=("Arial", 11), fg="#7f8c8d", bg='white')
        subtitle_label.pack(pady=(5, 0))
        
        # Status section
        status_frame = tk.LabelFrame(main_container, text="📊 Status", font=("Arial", 11, "bold"), 
                                   fg="#2c3e50", bg='white', bd=1, relief="solid")
        status_frame.pack(fill="x", pady=(0, 15), ipady=8)
        
        self.status_label = tk.Label(status_frame, text="💾 Ready to select Excel workbook...", 
                                    font=("Arial", 10), fg="#3498db", bg='white')
        self.status_label.pack(pady=5, padx=10, anchor="w")
        
        # File selection section
        file_frame = tk.Frame(main_container, bg='white')
        file_frame.pack(fill="x", pady=(0, 15))
        
        select_btn = self.create_professional_button(
            file_frame, "📁 Select Excel Workbook", self.select_workbook,
            "#3498db", "#2980b9", width=30, height=2
        )
        select_btn.pack()
        
        # Files info section - more compact
        info_frame = tk.LabelFrame(main_container, text="📄 Files to Generate (8 Files)", 
                                 font=("Arial", 11, "bold"), fg="#2c3e50", bg='white', 
                                 bd=1, relief="solid")
        info_frame.pack(fill="x", pady=(0, 15), ipady=8)
        
        # Create a scrollable text widget for files
        files_container = tk.Frame(info_frame, bg='white')
        files_container.pack(fill="x", padx=10, pady=5)
        
        files_text = tk.Text(files_container, height=5, font=("Arial", 9), 
                           bg='white', relief="flat", wrap=tk.WORD, cursor="arrow")
        files_scroll = ttk.Scrollbar(files_container, orient="vertical", command=files_text.yview)
        files_text.configure(yscrollcommand=files_scroll.set)
        
        # Add file mappings
        for i, (obser_file, sheet_name) in enumerate(self.sheet_mappings.items(), 1):
            files_text.insert(tk.END, f"{i}. {obser_file} ← {sheet_name}\n")
        
        files_text.config(state="disabled")
        files_text.pack(side="left", fill="both", expand=True)
        files_scroll.pack(side="right", fill="y")
        
        # Actions section - Always visible
        actions_frame = tk.LabelFrame(main_container, text="🚀 Actions", font=("Arial", 11, "bold"), 
                                    fg="#2c3e50", bg='white', bd=1, relief="solid")
        actions_frame.pack(fill="x", pady=(0, 15), ipady=15)
        
        # Action buttons in a grid
        btn_container = tk.Frame(actions_frame, bg='white')
        btn_container.pack(pady=15)
        
        # Professional buttons with proper colors and hover effects
        self.edit_btn = self.create_professional_button(
            btn_container, "⚙️ View/Edit Parameters", self.show_parameter_editor,
            "#27ae60", "#229954", width=25, height=2, state="disabled"
        )
        self.edit_btn.pack(side="left", padx=10)
        
        self.generate_btn = self.create_professional_button(
            btn_container, "🚀 Generate Files", self.generate_obser_files,
            "#e67e22", "#d35400", width=25, height=2, state="disabled"
        )
        self.generate_btn.pack(side="left", padx=10)
        
        # Footer
        footer_label = tk.Label(main_container, text="Powered by Advanced File Generation Engine", 
                               font=("Arial", 8), fg="#95a5a6", bg='white')
        footer_label.pack(side="bottom", pady=(20, 0))

    def select_workbook(self):
        """Select and validate workbook"""
        # Update status
        self.status_label.config(text="🔍 Opening file browser...", fg="#e67e22")
        self.root.update()
        
        self.workbook_path = filedialog.askopenfilename(
            title="Select Excel Workbook",
            filetypes=[("Excel files", "*.xlsx *.xls")],
            initialdir=os.path.expanduser("~")
        )
        
        if not self.workbook_path:
            self.status_label.config(text="❌ No file selected", fg="#e74c3c")
            return
        
        # Update status
        self.status_label.config(text="🔍 Validating Excel sheets...", fg="#e67e22")
        self.root.update()
        
        try:
            wb = load_workbook(self.workbook_path)
            required_sheets = list(self.sheet_mappings.values())
            missing_sheets = [sheet for sheet in required_sheets 
                            if sheet not in wb.sheetnames]
            
            if missing_sheets:
                self.status_label.config(text=f"❌ Missing required sheets", fg="#e74c3c")
                messagebox.showerror("Missing Sheets", 
                                   f"Required sheets not found:\n" + 
                                   "\n".join(missing_sheets))
                return
            
            # Load parameters from 入力値 sheet
            self.status_label.config(text="📥 Loading parameters...", fg="#e67e22")
            self.root.update()
            
            self.load_nyuuryoku_parameters()
            
            self.status_label.config(text=f"✅ File loaded: {os.path.basename(self.workbook_path)}", 
                                   fg="#27ae60")
            
            # Enable action buttons
            self.edit_btn.config(state="normal")
            self.generate_btn.config(state="normal")
            
        except Exception as e:
            self.status_label.config(text="❌ Error loading file", fg="#e74c3c")
            messagebox.showerror("Error", f"Error loading file:\n{str(e)}")

    def load_nyuuryoku_parameters(self):
        """Load parameters from 入力値 sheet"""
        try:
            # Load 入力値 sheet
            nyuuryoku_df = pd.read_excel(self.workbook_path, sheet_name='入力値', header=None)
            
            # Find parameters by matching column headers
            if len(nyuuryoku_df) >= 2:
                headers = nyuuryoku_df.iloc[0]  # First row as headers
                
                # Find and extract parameters
                for i, header in enumerate(headers):
                    if pd.notna(header):
                        header_str = str(header)
                        if 'データ個数' in header_str:
                            try:
                                self.nyuuryoku_params['data_count'] = int(nyuuryoku_df.iloc[1, i])
                            except (ValueError, TypeError):
                                pass
                        elif '予測年数' in header_str:
                            try:
                                self.nyuuryoku_params['prediction_years'] = int(nyuuryoku_df.iloc[1, i])
                            except (ValueError, TypeError):
                                pass
                        elif 'λ定数' in header_str:
                            try:
                                self.nyuuryoku_params['lambda_constant'] = float(nyuuryoku_df.iloc[1, i])
                            except (ValueError, TypeError):
                                pass
                        elif '点検年度に対応した年' in header_str:
                            # Extract years from this column
                            years = []
                            for row_idx in range(1, len(nyuuryoku_df)):
                                val = nyuuryoku_df.iloc[row_idx, i]
                                if pd.notna(val):
                                    try:
                                        year = int(val)
                                        if 20 <= year <= 50:
                                            years.append(year)
                                    except (ValueError, TypeError):
                                        break
                            
                            if years:
                                self.nyuuryoku_params['inspection_years'] = years
            
        except Exception as e:
            print(f"Could not load 入力値 sheet: {e}")

    def show_parameter_editor(self):
        """Show parameter editor"""
        editor = tk.Toplevel(self.root)
        editor.title("Edit Parameters")
        editor.geometry("650x600")
        editor.minsize(650, 600)
        
        # Center the window
        editor.update_idletasks()
        x = (editor.winfo_screenwidth() // 2) - (650 // 2)
        y = (editor.winfo_screenheight() // 2) - (600 // 2)
        editor.geometry(f"650x600+{x}+{y}")
        
        editor.grab_set()
        editor.configure(bg="white")
        
        main_frame = tk.Frame(editor, bg="white", padx=25, pady=25)
        main_frame.pack(fill="both", expand=True)
        
        # Header
        tk.Label(main_frame, text="⚙️ Edit Input Parameters", 
                font=("Arial", 16, "bold"), fg="#2c3e50", bg="white").pack(pady=(0, 20))
        
        # Parameter fields
        fields_frame = tk.LabelFrame(main_frame, text="📊 Parameters", font=("Arial", 11, "bold"), 
                                   fg="#2c3e50", bg="white", bd=1, relief="solid")
        fields_frame.pack(fill="x", pady=(0, 15), ipady=10)
        
        params_grid = tk.Frame(fields_frame, bg="white")
        params_grid.pack(pady=8, padx=15)
        
        # データ個数
        tk.Label(params_grid, text="データ個数:", font=("Arial", 10, "bold"), 
                bg="white", fg="#34495e").grid(row=0, column=0, sticky="w", pady=6, padx=(0, 15))
        self.data_count_var = tk.StringVar(value=str(self.nyuuryoku_params['data_count']))
        tk.Entry(params_grid, textvariable=self.data_count_var, width=18, 
                font=("Arial", 10)).grid(row=0, column=1, pady=6)
        
        # 予測年数
        tk.Label(params_grid, text="予測年数:", font=("Arial", 10, "bold"), 
                bg="white", fg="#34495e").grid(row=1, column=0, sticky="w", pady=6, padx=(0, 15))
        self.prediction_years_var = tk.StringVar(value=str(self.nyuuryoku_params['prediction_years']))
        tk.Entry(params_grid, textvariable=self.prediction_years_var, width=18, 
                font=("Arial", 10)).grid(row=1, column=1, pady=6)
        
        # λ定数
        tk.Label(params_grid, text="λ定数:", font=("Arial", 10, "bold"), 
                bg="white", fg="#34495e").grid(row=2, column=0, sticky="w", pady=6, padx=(0, 15))
        self.lambda_var = tk.StringVar(value=str(self.nyuuryoku_params['lambda_constant']))
        tk.Entry(params_grid, textvariable=self.lambda_var, width=18, 
                font=("Arial", 10)).grid(row=2, column=1, pady=6)
        
        # Years section
        years_frame = tk.LabelFrame(main_frame, text="📅 点検年度に対応した年", font=("Arial", 11, "bold"), 
                                  fg="#2c3e50", bg="white", bd=1, relief="solid")
        years_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        years_container = tk.Frame(years_frame, bg="white")
        years_container.pack(fill="both", expand=True, pady=8, padx=15)
        
        self.years_entry = tk.Text(years_container, height=6, width=55, font=("Arial", 10),
                                  relief="solid", bd=1)
        years_scrollbar = ttk.Scrollbar(years_container, orient="vertical", command=self.years_entry.yview)
        self.years_entry.configure(yscrollcommand=years_scrollbar.set)
        
        current_years = ' '.join(map(str, self.nyuuryoku_params['inspection_years']))
        self.years_entry.insert("1.0", current_years)
        
        self.years_entry.pack(side="left", fill="both", expand=True)
        years_scrollbar.pack(side="right", fill="y")
        
        # Info label
        info_label = tk.Label(years_frame, text="💡 Enter years separated by spaces (e.g., 27 28 29 30...)", 
                             font=("Arial", 8), fg="#7f8c8d", bg="white")
        info_label.pack(pady=(0, 8))
        
        # Buttons with professional styling
        button_frame = tk.Frame(main_frame, bg="white")
        button_frame.pack(fill="x", pady=15)
        
        save_btn = self.create_professional_button(
            button_frame, "💾 Save Parameters", lambda: self.save_parameters(editor),
            "#27ae60", "#229954", width=18, height=2
        )
        save_btn.pack(side="left", padx=8)
        
        cancel_btn = self.create_professional_button(
            button_frame, "❌ Cancel", editor.destroy,
            "#e74c3c", "#c0392b", width=15, height=2
        )
        cancel_btn.pack(side="left", padx=8)

    def save_parameters(self, editor_window):
        """Save parameters and close editor"""
        if self.validate_and_save_params():
            editor_window.destroy()
            messagebox.showinfo("Success", "Parameters saved successfully!")

    def validate_and_save_params(self):
        """Validate and save parameters"""
        try:
            # Validate inputs
            data_count = int(self.data_count_var.get())
            pred_years = int(self.prediction_years_var.get())
            lambda_const = float(self.lambda_var.get())
            
            if data_count <= 0 or pred_years <= 0 or lambda_const <= 0:
                messagebox.showerror("Error", "All values must be positive")
                return False
            
            # Parse years
            years_text = self.years_entry.get("1.0", tk.END).strip()
            years = []
            for year_str in years_text.split():
                try:
                    year = int(year_str)
                    if 20 <= year <= 50:
                        years.append(year)
                except ValueError:
                    pass
            
            if not years:
                messagebox.showerror("Error", "Please enter valid years (20-50 range)")
                return False
            
            # Update parameters
            self.nyuuryoku_params.update({
                'data_count': data_count,
                'prediction_years': pred_years,
                'lambda_constant': lambda_const,
                'inspection_years': years
            })
            
            # Save to Excel
            self.save_nyuuryoku_parameters()
            return True
            
        except ValueError:
            messagebox.showerror("Error", "Please enter valid numeric values")
            return False
        except Exception as e:
            messagebox.showerror("Error", f"Error saving parameters: {str(e)}")
            return False

    def save_nyuuryoku_parameters(self):
        """Save parameters to 入力値 sheet"""
        try:
            wb = load_workbook(self.workbook_path)
            
            if '入力値' in wb.sheetnames:
                wb.remove(wb['入力値'])
            
            ws = wb.create_sheet('入力値')
            
            # Headers in row 1
            ws['A1'] = 'データ個数'
            ws['B1'] = '予測年数' 
            ws['C1'] = 'λ定数'
            ws['D1'] = '点検年度に対応した年'
            
            # Values in row 2
            ws['A2'] = self.nyuuryoku_params['data_count']
            ws['B2'] = self.nyuuryoku_params['prediction_years']
            ws['C2'] = self.nyuuryoku_params['lambda_constant']
            
            # Years in column D starting from row 2
            for i, year in enumerate(self.nyuuryoku_params['inspection_years']):
                ws[f'D{i+2}'] = year
            
            wb.save(self.workbook_path)
            wb.close()
            
        except Exception as e:
            raise Exception(f"Error saving to Excel: {str(e)}")

    def generate_obser_files(self):
        """Generate all obser files with improved progress tracking"""
        try:
            output_directory = os.path.dirname(self.workbook_path)
            generated_files = []
            
            # Create professional progress window
            progress_window = tk.Toplevel(self.root)
            progress_window.title("Obser Files Generator - Processing")
            progress_window.geometry("650x500")
            progress_window.grab_set()
            progress_window.configure(bg='white')
            
            # Center the progress window
            progress_window.update_idletasks()
            x = (progress_window.winfo_screenwidth() // 2) - (650 // 2)
            y = (progress_window.winfo_screenheight() // 2) - (500 // 2)
            progress_window.geometry(f"650x500+{x}+{y}")
            
            # Main container
            main_frame = tk.Frame(progress_window, bg='white', padx=25, pady=20)
            main_frame.pack(fill="both", expand=True)
            
            # Header
            title_label = tk.Label(main_frame, text="🚀 Generating Obser Files", 
                                  font=("Arial", 16, "bold"), fg="#2c3e50", bg='white')
            title_label.pack(pady=(0, 15))
            
            # Status section
            status_frame = tk.LabelFrame(main_frame, text="📊 Current Status", 
                                       font=("Arial", 11, "bold"), fg="#2c3e50", bg='white', 
                                       bd=1, relief="solid")
            status_frame.pack(fill="x", pady=(0, 12), ipady=8)
            
            status_label = tk.Label(status_frame, text="🔄 Initializing generation process...", 
                                  font=("Arial", 10), fg="#3498db", bg='white')
            status_label.pack(pady=5, padx=10, anchor="w")
            
            # Progress section
            progress_frame = tk.LabelFrame(main_frame, text="⏳ Progress", 
                                         font=("Arial", 11, "bold"), fg="#2c3e50", bg='white', 
                                         bd=1, relief="solid")
            progress_frame.pack(fill="x", pady=(0, 15), ipady=10)
            
            progress_label = tk.Label(progress_frame, text="⏸ Step 0/8: Starting generation...", 
                                    font=("Arial", 10), fg="#e67e22", bg='white')
            progress_label.pack(pady=(5, 8), padx=10, anchor="w")
            
            progress_bar = ttk.Progressbar(progress_frame, mode='determinate', maximum=8, length=550)
            progress_bar.pack(pady=(0, 8), padx=10)
            
            # Files section
            files_frame = tk.LabelFrame(main_frame, text="📄 Generated Files", 
                                      font=("Arial", 11, "bold"), fg="#2c3e50", bg='white', 
                                      bd=1, relief="solid")
            files_frame.pack(fill="both", expand=True, pady=(0, 10))
            
            files_container = tk.Frame(files_frame, bg='white')
            files_container.pack(fill="both", expand=True, pady=8, padx=10)
            
            files_text = tk.Text(files_container, height=10, font=("Arial", 9), bg="#f8f9fa", 
                               relief="solid", bd=1, wrap=tk.WORD)
            files_scroll = ttk.Scrollbar(files_container, orient="vertical", command=files_text.yview)
            files_text.configure(yscrollcommand=files_scroll.set)
            
            # Add file list
            for i, (obser_file, sheet_name) in enumerate(self.sheet_mappings.items(), 1):
                files_text.insert(tk.END, f"{i}. {obser_file} ← {sheet_name}\n")
            
            files_text.config(state="disabled")
            files_text.pack(side="left", fill="both", expand=True)
            files_scroll.pack(side="right", fill="y")
            
            progress_window.update()
            
            # Generate files with improved progress tracking
            total_files = len(self.sheet_mappings)
            
            for i, (obser_file, sheet_name) in enumerate(self.sheet_mappings.items(), 1):
                # Update progress
                status_label.config(text=f"🔄 Generating {obser_file}...", fg="#e67e22")
                progress_label.config(text=f"▶️ Step {i}/{total_files}: Processing {sheet_name}...")
                progress_bar['value'] = i - 0.5
                progress_window.update()
                
                try:
                    output_path = os.path.join(output_directory, obser_file)
                    # Faster file creation without Excel updates
                    self.create_obser_file_fast(sheet_name, output_path)
                    generated_files.append(obser_file)
                    
                    # Update completion status
                    status_label.config(text=f"✅ Completed {obser_file}", fg="#27ae60")
                    progress_bar['value'] = i
                    progress_window.update()
                    
                except Exception as e:
                    status_label.config(text=f"❌ Error with {obser_file}: {str(e)[:30]}...", fg="#e74c3c")
                    print(f"Error generating {obser_file}: {e}")
                    progress_window.update()
            
            # Show completion
            status_label.config(text="🎉 All files generated successfully!", fg="#27ae60")
            progress_label.config(text=f"✅ Complete: Generated {len(generated_files)}/{total_files} files")
            progress_bar['value'] = total_files
            progress_window.update()
            
            
            # Auto-close after 3 seconds with countdown
            for countdown in range(3, 0, -1):
                status_label.config(text=f"🎉 Success! Auto-closing in {countdown} seconds...")
                progress_window.update()
                time.sleep(1)  # Actually wait 1 second
            
            progress_window.destroy()
            
            # Show final completion message
            completion_msg = f"✅ Successfully generated {len(generated_files)} obser files!\n\n"
            completion_msg += "\n".join(f"• {file}" for file in generated_files)
            completion_msg += f"\n\n📁 Location: {output_directory}"
            
            # messagebox.showinfo("Generation Complete", completion_msg)
            
            # # Auto-close main window after completion message
            # self.root.after(500, self.root.quit)  # Close main window after 0.5 seconds


            messagebox.showinfo("Generation Complete", completion_msg)
            self.root.quit()  # Close main window immediately
            
        except Exception as e:
            if 'progress_window' in locals():
                progress_window.destroy()
            messagebox.showerror("Error", f"Error generating files: {str(e)}")

    def create_obser_file_fast(self, sheet_name, output_path):
        """Create obser file without Excel updating for faster processing"""
        try:
            # Load sheet data (read-only, faster)
            sheet_df = pd.read_excel(self.workbook_path, sheet_name=sheet_name)
            
            # Sort by last column in descending order
            if len(sheet_df) > 0 and len(sheet_df.columns) > 0:
                last_col = sheet_df.columns[-1]
                sheet_df = sheet_df.sort_values(by=last_col, ascending=False)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                # First line: parameters
                f.write(f"{self.nyuuryoku_params['data_count']} {self.nyuuryoku_params['prediction_years']} {self.nyuuryoku_params['lambda_constant']}\n")
                
                # Second line: years
                years_line = ' '.join(map(str, self.nyuuryoku_params['inspection_years']))
                f.write(f"{years_line}\n")
                
                # Third line: blank
                f.write("\n")
                
                # Find 構造物番号 column
                kozo_col_idx = None
                for i, col in enumerate(sheet_df.columns):
                    if '構造物番号' in str(col):
                        kozo_col_idx = i
                        break
                
                if kozo_col_idx is None:
                    raise Exception(f"構造物番号 column not found in {sheet_name}")
                
                # Get columns from 構造物番号 onwards
                columns_to_export = sheet_df.columns[kozo_col_idx:]
                
                # Write data rows
                for _, row in sheet_df.iterrows():
                    row_data = []
                    for col in columns_to_export:
                        value = row[col]
                        
                        if pd.isna(value) or value == '':
                            row_data.append('')
                        else:
                            try:
                                numeric_val = float(value)
                                if numeric_val == 0:
                                    row_data.append('0.1')
                                elif numeric_val == int(numeric_val):
                                    row_data.append(str(int(numeric_val)))
                                else:
                                    row_data.append(str(round(numeric_val, 3)))
                            except (ValueError, TypeError):
                                if str(value) == '0':
                                    row_data.append('0.1')
                                else:
                                    row_data.append(str(value))
                    
                    f.write('\t'.join(row_data) + '\n')
            
        except Exception as e:
            raise Exception(f"Error creating {output_path}: {str(e)}")

    def run(self):
        """Run the application"""
        self.root.mainloop()


if __name__ == "__main__":
    app = ObserFileGeneratorApp()
    app.run()
                            