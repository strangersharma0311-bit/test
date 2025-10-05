import pandas as pd
import openpyxl
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import numpy as np
import re
import time

class ExcelProcessorApp:
    def __init__(self, root):
        self.root = root
        self.workbook = None
        self.selected_columns_for_weighting = []
        self.status_var = tk.StringVar()
        self.progress_var = tk.StringVar()
        self.create_gui()
        self.start_process()

    def create_gui(self):
        self.root.title("Excel Processor Pro")
        self.root.geometry("700x700")
        self.root.resizable(False, False)
        self.root.configure(bg='#f8f9fa')
        
        # Center the window
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() // 2) - (700 // 2)
        y = (self.root.winfo_screenheight() // 2) - (700 // 2)
        self.root.geometry(f"700x700+{x}+{y}")

        # Main frame
        main_frame = tk.Frame(self.root, bg='#f8f9fa', padx=40, pady=40)
        main_frame.pack(fill="both", expand=True)

        # Title
        title_frame = tk.Frame(main_frame, bg='#f8f9fa')
        title_frame.pack(pady=(0, 30))
        
        title_label = tk.Label(title_frame, text="⚡ Excel Processor Pro", 
                              font=("Segoe UI", 24, "bold"), fg="#2c3e50", bg="#f8f9fa")
        title_label.pack()
        
        subtitle_label = tk.Label(title_frame, text="Advanced Data Processing & Analysis", 
                                 font=("Segoe UI", 12), fg="#6c757d", bg="#f8f9fa")
        subtitle_label.pack(pady=(5, 0))

        # Status display
        status_frame = tk.Frame(main_frame, bg='#ffffff', relief='solid', bd=1)
        status_frame.pack(fill="x", pady=(0, 20))
        
        status_header = tk.Frame(status_frame, bg='#e9ecef')
        status_header.pack(fill="x", padx=20, pady=(15, 5))
        
        tk.Label(status_header, text="📊 Status", font=("Segoe UI", 14, "bold"), 
                fg="#2c3e50", bg="#e9ecef").pack(anchor="w")
        
        status_content = tk.Frame(status_frame, bg='#ffffff')
        status_content.pack(fill="x", padx=20, pady=(0, 15))
        
        self.status_label = tk.Label(status_content, textvariable=self.status_var, 
                                   font=("Segoe UI", 11), fg="#28a745", bg="#ffffff", 
                                   wraplength=600, justify="left")
        self.status_label.pack(anchor="w")

        # Progress display
        progress_frame = tk.Frame(main_frame, bg='#ffffff', relief='solid', bd=1)
        progress_frame.pack(fill="x", pady=(0, 20))
        
        progress_header = tk.Frame(progress_frame, bg='#e9ecef')
        progress_header.pack(fill="x", padx=20, pady=(15, 5))
        
        tk.Label(progress_header, text="⏳ Progress", font=("Segoe UI", 14, "bold"), 
                fg="#2c3e50", bg="#e9ecef").pack(anchor="w")
        
        progress_content = tk.Frame(progress_frame, bg='#ffffff')
        progress_content.pack(fill="x", padx=20, pady=(0, 15))
        
        self.progress_label = tk.Label(progress_content, textvariable=self.progress_var,
                             font=("Segoe UI", 11), fg="#007bff", bg="#ffffff", 
                             wraplength=600, justify="left", height=2)
        self.progress_label.pack(anchor="w")

        # Progress bar
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Modern.Horizontal.TProgressbar",
                       background='#28a745',
                       troughcolor='#e9ecef',
                       borderwidth=0)
        
        self.progress_bar = ttk.Progressbar(progress_content, mode='indeterminate', 
                                          style="Modern.Horizontal.TProgressbar", length=600)
        self.progress_bar.pack(fill="x", pady=(10, 0))

        # Add Select File button
        select_frame = tk.Frame(main_frame, bg='#f8f9fa')
        select_frame.pack(pady=(20, 10))

        select_btn = tk.Button(select_frame, text="📁 Select Excel File", command=self.select_workbook,
                            bg="#007bff", fg="white", width=20, height=2, font=("Segoe UI", 12, "bold"),
                            relief="flat", cursor="hand2")
        select_btn.pack()

        
        
        # Footer
        footer_frame = tk.Frame(main_frame, bg='#f8f9fa')
        footer_frame.pack(fill="x", pady=(10, 0))

        tk.Label(footer_frame, text="Powered by Advanced Analytics Engine", 
                font=("Segoe UI", 9), fg="#6c757d", bg="#f8f9fa").pack()

    def update_status(self, message):
        self.status_var.set(message)
        self.root.update()
        print(message)

    def update_progress(self, message):
        self.progress_var.set(message)
        self.root.update()

    def start_process(self):
        self.update_status("🔍 Click 'Select File' to choose your Excel workbook...")
        

    def select_workbook(self):
        self.workbook = filedialog.askopenfilename(
            title="Select Excel Workbook",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        
        if self.workbook:
            filename = self.workbook.split('/')[-1]
            self.update_status(f"✅ Selected: {filename}")
            self.progress_bar.start(10)

            
            self.root.after(500, self.run_automatic_process)
        else:
            self.update_status("❌ No file selected. Please restart the application.")
            self.root.after(3000, self.root.destroy)

    def run_automatic_process(self):
        """Run all processes automatically in sequence"""
        try:
            self.update_progress("🔄 Step 1/3: Extracting and merging data...")
            self.extract_and_merge_data()
            
            self.update_progress("🔄 Step 1/3: Creating extraction sheet...")
            self.create_chuushutsu_sheet()
            
            self.update_progress("🔄 Step 2/3: Processing weights...")
            self.apply_weights()
            
        except Exception as e:
            self.progress_bar.stop()
            self.update_status(f"❌ Error: {str(e)}")
            self.update_progress("💥 Process failed!")
            print(f"Error: {str(e)}")
            import traceback
            traceback.print_exc()

    def extract_and_merge_data(self):
        if not self.workbook:
            return

        start_time = time.time()

        wb = load_workbook(self.workbook)
        sheet_names = [sheet for sheet in wb.sheetnames if sheet.isnumeric()]
        
        chuushutsu_df = pd.read_excel(self.workbook, sheet_name="抽出列")
        ketsugou_df = pd.DataFrame()
        
        for year in sorted(sheet_names, reverse=True):
            year_col = None
            
            for col in chuushutsu_df.columns:
                if str(year) in str(col):
                    year_col = col
                    break
            
            if year_col is None:
                continue
            
            columns_to_extract = chuushutsu_df[year_col].dropna().tolist()
            year_df = pd.read_excel(self.workbook, sheet_name=year)
            available_columns = [col for col in columns_to_extract if col in year_df.columns]
            
            extracted_df = year_df[['調査番号'] + available_columns]
            extracted_df.columns = ['調査番号'] + [f"{year} {col}" for col in available_columns]
            
            if ketsugou_df.empty:
                ketsugou_df = extracted_df
            else:
                ketsugou_df = pd.merge(ketsugou_df, extracted_df, on='調査番号', how='outer', suffixes=('', '_duplicate'))
                ketsugou_df = ketsugou_df.loc[:, ~ketsugou_df.columns.str.endswith('_duplicate')]
        
        ketsugou_df = ketsugou_df.sort_values(by='調査番号')
        
        with pd.ExcelWriter(self.workbook, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            ketsugou_df.to_excel(writer, sheet_name='結合シート', index=False)
        
        end_time = time.time()
        self.update_status(f"✅ 結合シート created successfully in {end_time - start_time:.2f} seconds.")

    def create_chuushutsu_sheet(self):
        if not self.workbook:
            return

        wb = load_workbook(self.workbook)
        ketsugou_df = pd.read_excel(self.workbook, sheet_name='結合シート')
        tensuuka_df = pd.read_excel(self.workbook, sheet_name="点数化列")
        chuushutsu_df = pd.DataFrame()
        
        for col in tensuuka_df.columns:
            year = col
            columns_to_extract = tensuuka_df[year].dropna().tolist()
            for col_name in columns_to_extract:
                if f"{col_name}" in ketsugou_df.columns:
                    chuushutsu_df[f"{col_name}"] = ketsugou_df[f"{col_name}"]
        
        chuushutsu_df['路線名'] = ketsugou_df[[col for col in ketsugou_df.columns if '路線名' in col]].bfill(axis=1).iloc[:, 0]
        chuushutsu_df['構造物名称'] = ketsugou_df[[col for col in ketsugou_df.columns if '構造物名称' in col]].bfill(axis=1).iloc[:, 0]

        shubetsu_cols = [col for col in ketsugou_df.columns if '種別' in col]
        if shubetsu_cols:
            chuushutsu_df['種別'] = ketsugou_df[shubetsu_cols].bfill(axis=1).iloc[:, 0]

        tenken_cols = [col for col in ketsugou_df.columns if '点検区分1' in col]
        if tenken_cols:
            chuushutsu_df['点検区分1'] = ketsugou_df[tenken_cols].bfill(axis=1).iloc[:, 0]

        eki_hajimari_cols = [col for col in ketsugou_df.columns if '駅（始）' in col]
        if eki_hajimari_cols:
            chuushutsu_df['駅（始）'] = ketsugou_df[eki_hajimari_cols].bfill(axis=1).iloc[:, 0]

        eki_itaru_cols = [col for col in ketsugou_df.columns if '駅（至）' in col]
        if eki_itaru_cols:
            chuushutsu_df['駅（至）'] = ketsugou_df[eki_itaru_cols].bfill(axis=1).iloc[:, 0]

        base_cols = ['路線名', '構造物名称', '種別', '点検区分1', '駅（始）', '駅（至）']
        existing_base_cols = [col for col in base_cols if col in chuushutsu_df.columns]
        other_cols = [col for col in chuushutsu_df.columns if col not in base_cols]
        chuushutsu_df = chuushutsu_df[existing_base_cols + other_cols]

        with pd.ExcelWriter(self.workbook, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            chuushutsu_df.to_excel(writer, sheet_name='抽出シート', index=False)
        
        self.update_status("✅ 抽出シート created successfully.")

    def apply_weights(self):
        if not self.workbook:
            return

        try:
            wb = load_workbook(self.workbook)
            chuushutsu_df = pd.read_excel(self.workbook, sheet_name='抽出シート')
            tensuuka_df = pd.read_excel(self.workbook, sheet_name='点数化列')
            lookup_df = pd.read_excel(self.workbook, sheet_name='重みテーブル')
            
            lookup_df = self.expand_lookup_table_if_needed(lookup_df, tensuuka_df)
            lookup_dicts = self.create_lookup_dicts(lookup_df)
            not_found_values = [[] for _ in range(len(lookup_dicts))]

            for col_index, col in enumerate(tensuuka_df.columns):
                for i, column_name in enumerate(tensuuka_df[col].dropna()):
                    weight_col_name = f"{column_name} 重み"
                    
                    if column_name in chuushutsu_df.columns:
                        dict_index = i % len(lookup_dicts)
                        lookup_dict = lookup_dicts[dict_index]
                        
                        weight_values = []
                        for val in chuushutsu_df[column_name]:
                            weight = self.lookup_weight(lookup_dict, val, not_found_values[dict_index])
                            weight_values.append(weight)
                        
                        if weight_col_name in chuushutsu_df.columns:
                            chuushutsu_df[weight_col_name] = weight_values
                        else:
                            col_position = chuushutsu_df.columns.get_loc(column_name) + 1
                            chuushutsu_df.insert(col_position, weight_col_name, weight_values)

            missing_values_exist = any(len(values) > 0 for values in not_found_values)
            
            if missing_values_exist:
                self.progress_bar.stop()
                self.update_progress("⏸️ Step 2/3: Waiting for user input...")
                self.show_missing_values_choice(not_found_values, lookup_dicts, lookup_df, chuushutsu_df, tensuuka_df)
            else:
                self.write_to_excel(chuushutsu_df, tensuuka_df, lookup_df)
                self.continue_to_final_step()

        except Exception as e:
            self.progress_bar.stop()
            self.update_status(f"❌ Error in weight processing: {str(e)}")
            self.update_progress("💥 Process failed!")
            import traceback
            traceback.print_exc()

    def continue_to_final_step(self):
        """Continue to the final step after weights are applied"""
        self.update_status("✅ Weights applied successfully. Generating 演算結果 sheet...")
        self.update_progress("🔄 Step 3/3: Creating calculation results...")
        self.progress_bar.start(10)
        self.root.after(500, self.create_enzan_kekka_sheet)

    def expand_lookup_table_if_needed(self, lookup_df, tensuuka_df):
        """Expand lookup table if there are more columns in 点数化列 than lookup tables"""
        max_cols_per_year = 0
        for col in tensuuka_df.columns:
            col_count = len(tensuuka_df[col].dropna())
            max_cols_per_year = max(max_cols_per_year, col_count)
        
        current_table_pairs = len(lookup_df.columns) // 2
        
        if max_cols_per_year > current_table_pairs:
            additional_tables_needed = max_cols_per_year - current_table_pairs
            
            for i in range(additional_tables_needed):
                table_num = current_table_pairs + i + 1
                key_col_name = f"Table{table_num}_Key"
                value_col_name = f"Table{table_num}_Value"
                
                lookup_df[key_col_name] = None
                lookup_df[value_col_name] = None
        
        return lookup_df

    def create_enzan_kekka_sheet(self):
        """Create calculation results sheet based on operations"""
        if not self.workbook:
            return
            
        try:
            chuushutsu_df = pd.read_excel(self.workbook, sheet_name='抽出シート')
            tensuuka_df = pd.read_excel(self.workbook, sheet_name='点数化列')
            
            try:
                enzanshi_df = pd.read_excel(self.workbook, sheet_name='演算子')
            except:
                enzanshi_df = pd.DataFrame()
            
            result_df = pd.DataFrame()
            
            basic_columns = ['路線名', '構造物名称', '種別', '点検区分1', '駅（始）', '駅（至）']
            
            for col_name in basic_columns:
                if col_name in chuushutsu_df.columns:
                    result_df[col_name] = chuushutsu_df[col_name]
                else:
                    matching_cols = [col for col in chuushutsu_df.columns if col_name in col]
                    if matching_cols:
                        result_df[col_name] = chuushutsu_df[matching_cols].bfill(axis=1).iloc[:, 0]
                    else:
                        result_df[col_name] = ''
            
            for col_index, col in enumerate(tensuuka_df.columns):
                if pd.isna(col):
                    continue
                    
                year = str(col)
                
                if not enzanshi_df.empty and col_index < len(enzanshi_df.columns):
                    if len(enzanshi_df) > 1:
                        operation_formula = enzanshi_df.iloc[1, col_index]
                    else:
                        operation_formula = "A*B*C"
                else:
                    operation_formula = "A*B*C"
                
                year_columns = tensuuka_df[col].dropna().tolist()
                
                if len(year_columns) == 0:
                    continue
                
                weight_columns = []
                for year_col in year_columns:
                    weight_col_name = f"{year_col} 重み"
                    if weight_col_name in chuushutsu_df.columns:
                        weight_columns.append(weight_col_name)
                
                if len(weight_columns) == 0:
                    continue
                
                result_column_name = f"{year} 結果"
                result_values = []
                
                weight_data = chuushutsu_df[weight_columns].values
                
                for row_values in weight_data:
                    try:
                        valid_values = []
                        has_blank = False
                        
                        for val in row_values:
                            if pd.isna(val) or val == '':
                                has_blank = True
                                break
                            else:
                                valid_values.append(float(val))
                        
                        if has_blank or len(valid_values) == 0:
                            result_values.append('')
                        else:
                            if operation_formula and isinstance(operation_formula, str):
                                expression = operation_formula
                                for i, val in enumerate(valid_values):
                                    letter = chr(65 + i)
                                    expression = expression.replace(letter, str(val))
                                
                                try:
                                    result = eval(expression)
                                except:
                                    result = 0
                            else:
                                result = np.prod(valid_values)
                            
                            result_values.append(result)
                            
                    except Exception:
                        result_values.append('')
                
                result_df[result_column_name] = result_values
            
            with pd.ExcelWriter(self.workbook, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                chuushutsu_df.to_excel(writer, sheet_name='抽出シート', index=False)
                tensuuka_df.to_excel(writer, sheet_name='点数化列', index=False)
                
                lookup_df = pd.read_excel(self.workbook, sheet_name='重みテーブル')
                lookup_df.to_excel(writer, sheet_name='重みテーブル', index=False)
                
                if not enzanshi_df.empty:
                    enzanshi_df.to_excel(writer, sheet_name='演算子', index=False)
                
                result_df.to_excel(writer, sheet_name='演算結果', index=False)
            
            self.progress_bar.stop()
            self.update_status("🎉 All processing completed successfully!")
            self.update_progress("✅ 演算結果 sheet generated. Process complete.")
            
            self.root.after(3000, self.close_application)
            
        except Exception as e:
            self.progress_bar.stop()
            self.update_status(f"❌ Error creating 演算結果 sheet: {str(e)}")
            self.update_progress("💥 Process failed!")
            import traceback
            traceback.print_exc()

    def close_application(self):
        """Close the application"""
        self.root.destroy()

    def show_missing_values_choice(self, not_found_values, lookup_dicts, lookup_df, chuushutsu_df, tensuuka_df):
        """Show initial choice dialog for handling missing values"""
        choice_window = tk.Toplevel(self.root)
        choice_window.title("Missing Values Found")
        choice_window.geometry("700x700")
        choice_window.grab_set()
        choice_window.resizable(False, False)
        
        # Center the window
        choice_window.transient(self.root)
        choice_window.geometry("+%d+%d" % (self.root.winfo_rootx() + 100, self.root.winfo_rooty() + 50))

        # Main frame
        main_frame = tk.Frame(choice_window, padx=30, pady=20)
        main_frame.pack(fill="both", expand=True)

        # Title
        title_label = tk.Label(main_frame, text="Missing Weight Values Found", 
                              font=("Arial", 16, "bold"), fg="red")
        title_label.pack(pady=(0, 20))

        # Message
        msg_text = "Some values in your data don't have corresponding weights in the lookup table.\n\nHow would you like to proceed?"
        msg_label = tk.Label(main_frame, text=msg_text, justify="center", wraplength=500, font=("Arial", 12))
        msg_label.pack(pady=(0, 30))

        # Buttons frame
        button_frame = tk.Frame(main_frame)
        button_frame.pack(pady=20, expand=True)

        def assign_values():
            choice_window.destroy()
            self.show_assign_values_dialog(not_found_values, lookup_dicts, lookup_df, chuushutsu_df, tensuuka_df)

        def skip_assignment():
            choice_window.destroy()
            self.ask_default_value(not_found_values, lookup_dicts, lookup_df, chuushutsu_df, tensuuka_df)

        assign_btn = tk.Button(button_frame, text="Assign Values", command=assign_values,
                              bg="#4CAF50", fg="white", width=14, height=2, font=("Arial", 11))
        assign_btn.pack(side="left", padx=15, pady=10)

        skip_btn = tk.Button(button_frame, text="Skip", command=skip_assignment,
                            bg="#f44336", fg="white", width=14, height=2, font=("Arial", 11))
        skip_btn.pack(side="left", padx=15, pady=10)

    def ask_default_value(self, not_found_values, lookup_dicts, lookup_df, chuushutsu_df, tensuuka_df):
        """Ask for default value to assign to all missing values"""
        default_value = simpledialog.askfloat("Default Weight", 
                                            "Enter the default weight to assign to all missing values:",
                                            minvalue=0, maxvalue=20)
        if default_value is not None:
            for table_index, values in enumerate(not_found_values):
                for value in values:
                    if table_index < len(lookup_dicts):
                        lookup_dicts[table_index][value] = default_value
            
            self.recalculate_and_save(lookup_dicts, chuushutsu_df, tensuuka_df, lookup_df)

    def show_assign_values_dialog(self, not_found_values, lookup_dicts, lookup_df, chuushutsu_df, tensuuka_df):
        """Show dialog for assigning individual values"""
        assign_window = tk.Toplevel(self.root)
        assign_window.title("Assign Weight Values")
        assign_window.geometry("1000x750")
        assign_window.grab_set()
        assign_window.resizable(True, True)
        
        # Center the window
        assign_window.transient(self.root)

        # Main frame
        main_frame = tk.Frame(assign_window, padx=15, pady=15)
        main_frame.pack(fill="both", expand=True)

        # Title
        title_label = tk.Label(main_frame, text="Assign Weights to Missing Values", 
                              font=("Arial", 16, "bold"))
        title_label.pack(pady=(0, 15))

        # Scrollable frame for tables
        canvas = tk.Canvas(main_frame, height=550)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Pack canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Create entries for missing values
        entries = []
        not_found_values = [list(set(values)) for values in not_found_values]
        
        for table_index, values in enumerate(not_found_values):
            if values and table_index * 2 < len(lookup_df.columns):
                # Table frame
                table_frame = tk.LabelFrame(scrollable_frame, 
                                          text=f"Table {table_index + 1} - {lookup_df.columns[table_index * 2]}", 
                                          padx=15, pady=15, font=("Arial", 12, "bold"))
                table_frame.pack(fill="x", padx=15, pady=15)

                # Header
                header_frame = tk.Frame(table_frame)
                header_frame.pack(fill="x", pady=(0, 10))
                
                tk.Label(header_frame, text="Value", font=("Arial", 11, "bold"), width=35, anchor="w").pack(side="left")
                tk.Label(header_frame, text="Weight", font=("Arial", 11, "bold"), width=15, anchor="w").pack(side="left")

                # Values
                for value in values:
                    value_frame = tk.Frame(table_frame)
                    value_frame.pack(fill="x", pady=3)
                    
                    value_label = tk.Label(value_frame, text=str(value), width=35, anchor="w", 
                                         relief="solid", borderwidth=1, font=("Arial", 10))
                    value_label.pack(side="left", padx=(0, 10))
                    
                    entry = ttk.Combobox(value_frame, values=[i for i in range(21)], 
                                       state="normal", width=15, font=("Arial", 10))
                    entry.pack(side="left")
                    entries.append((value, entry, table_index))

        # Button frame
        button_frame = tk.Frame(main_frame)
        button_frame.pack(fill="x", pady=20, side="bottom")

        def submit_values():
            # Get values from entries before destroying window
            nonlocal lookup_df
            entry_data = []
            for value, entry, table_index in entries:
                weight_text = entry.get()
                entry_data.append((value, weight_text, table_index))
            
            # Group new values by table_index
            new_values_by_table = {}
            for value, weight_text, table_index in entry_data:
                if weight_text:
                    try:
                        weight = float(weight_text)
                        if table_index < len(lookup_dicts):
                            lookup_dicts[table_index][value] = weight
                            
                            if table_index not in new_values_by_table:
                                new_values_by_table[table_index] = []
                            new_values_by_table[table_index].append((value, weight))
                            
                    except ValueError:
                        continue
            
            # Add new values to lookup_df - Fixed: Proper dtype handling
            for table_index, new_values in new_values_by_table.items():
                if table_index * 2 + 1 < len(lookup_df.columns):
                    key_col = lookup_df.columns[table_index * 2]
                    value_col = lookup_df.columns[table_index * 2 + 1]
                    
                    # Find the last row with data in this table
                    last_row = lookup_df[key_col].last_valid_index()
                    if last_row is not None:
                        separator_row_index = last_row + 2
                    else:
                        separator_row_index = 0
                    
                    # Extend dataframe if needed
                    total_new_rows = len(new_values) + 1
                    while len(lookup_df) <= separator_row_index + total_new_rows:
                        new_row = pd.Series([None] * len(lookup_df.columns), index=lookup_df.columns)
                        lookup_df = pd.concat([lookup_df, new_row.to_frame().T], ignore_index=True)
                    
                    # Add separator row
                    lookup_df.at[separator_row_index, key_col] = None
                    lookup_df.at[separator_row_index, value_col] = None
                    
                    # Add all new values
                    for i, (value, weight) in enumerate(new_values):
                        new_row_index = separator_row_index + 1 + i
                        lookup_df.at[new_row_index, key_col] = value
                        lookup_df.at[new_row_index, value_col] = weight
            
            assign_window.destroy()
            
            # Check for unassigned values
            unassigned_values = [value for value, weight_text, table_index in entry_data if not weight_text]
            
            if unassigned_values:
                self.handle_unassigned_values(unassigned_values, lookup_dicts, lookup_df, chuushutsu_df, tensuuka_df, not_found_values)
            else:
                self.recalculate_and_save(lookup_dicts, chuushutsu_df, tensuuka_df, lookup_df)

        submit_btn = tk.Button(button_frame, text="Submit", command=submit_values,
                            bg="#4CAF50", fg="white", width=20, height=2, font=("Arial", 12))
        submit_btn.pack(side="right", padx=15)

    def handle_unassigned_values(self, unassigned_values, lookup_dicts, lookup_df, chuushutsu_df, tensuuka_df, not_found_values):
        """Handle unassigned values after submission"""
        unassigned_window = tk.Toplevel(self.root)
        unassigned_window.title("Unassigned Values")
        unassigned_window.geometry("500x350")
        unassigned_window.grab_set()
        unassigned_window.resizable(False, False)
        
        # Center the window
        unassigned_window.transient(self.root)
        unassigned_window.geometry("+%d+%d" % (self.root.winfo_rootx() + 150, self.root.winfo_rooty() + 100))

        main_frame = tk.Frame(unassigned_window, padx=25, pady=20)
        main_frame.pack(fill="both", expand=True)

        title_label = tk.Label(main_frame, text="Unassigned Values Found", 
                              font=("Arial", 14, "bold"), fg="orange")
        title_label.pack(pady=(0, 20))

        msg_text = f"Some values are still not assigned:\n{', '.join(unassigned_values[:5])}{'...' if len(unassigned_values) > 5 else ''}\n\nWhat would you like to do?"
        msg_label = tk.Label(main_frame, text=msg_text, justify="center", wraplength=450, font=("Arial", 11))
        msg_label.pack(pady=(0, 25))

        button_frame = tk.Frame(main_frame)
        button_frame.pack(pady=15, expand=True)

        def assign_default():
            unassigned_window.destroy()
            self.ask_default_value_for_remaining(unassigned_values, lookup_dicts, lookup_df, chuushutsu_df, tensuuka_df)

        def go_back():
            unassigned_window.destroy()
            self.show_assign_values_dialog(not_found_values, lookup_dicts, lookup_df, chuushutsu_df, tensuuka_df)

        default_btn = tk.Button(button_frame, text="Assign Default Value", command=assign_default,
                               bg="#2196F3", fg="white", width=18, height=2, font=("Arial", 11))
        default_btn.pack(side="left", padx=10, pady=10)

        back_btn = tk.Button(button_frame, text="Go Back", command=go_back,
                            bg="#FF9800", fg="white", width=18, height=2, font=("Arial", 11))
        back_btn.pack(side="left", padx=10, pady=10)

    def ask_default_value_for_remaining(self, unassigned_values, lookup_dicts, lookup_df, chuushutsu_df, tensuuka_df):
        """Ask for default value for remaining unassigned values"""
        default_window = tk.Toplevel(self.root)
        default_window.title("Default Value for Remaining")
        default_window.geometry("450x300")
        default_window.grab_set()
        default_window.resizable(False, False)
        
        # Center the window
        default_window.transient(self.root)
        default_window.geometry("+%d+%d" % (self.root.winfo_rootx() + 200, self.root.winfo_rooty() + 150))

        main_frame = tk.Frame(default_window, padx=25, pady=20)
        main_frame.pack(fill="both", expand=True)

        tk.Label(main_frame, text="Enter default weight for unassigned values:", 
                font=("Arial", 12)).pack(pady=20)

        entry_frame = tk.Frame(main_frame)
        entry_frame.pack(pady=20)

        tk.Label(entry_frame, text="Default Weight:", font=("Arial", 11)).pack(side="left")
        default_entry = tk.Entry(entry_frame, width=12, font=("Arial", 11))
        default_entry.pack(side="left", padx=15)
        default_entry.focus()

        button_frame = tk.Frame(main_frame)
        button_frame.pack(pady=30, expand=True)

        def submit_default():
            try:
                default_value = float(default_entry.get())
                # Assign default value to unassigned values
                for value in unassigned_values:
                    for table_index in range(len(lookup_dicts)):
                        lookup_dicts[table_index][value] = default_value
                
                default_window.destroy()
                self.recalculate_and_save(lookup_dicts, chuushutsu_df, tensuuka_df, lookup_df)
                
            except ValueError:
                messagebox.showerror("Error", "Please enter a valid number")

        submit_btn = tk.Button(button_frame, text="Submit", command=submit_default,
                            bg="#4CAF50", fg="white", width=18, height=2, font=("Arial", 12))
        submit_btn.pack()

        # Handle Enter key
        default_entry.bind('<Return>', lambda e: submit_default())

    def recalculate_and_save(self, lookup_dicts, chuushutsu_df, tensuuka_df, lookup_df):
        """Recalculate weights with updated lookup tables and save"""
        try:
            self.update_status("🔄 Recalculating weights with updated values...")
            self.progress_bar.start(10)
            
            # Recalculate all weight columns with updated lookup dictionaries
            for col_index, col in enumerate(tensuuka_df.columns):
                for i, column_name in enumerate(tensuuka_df[col].dropna()):
                    weight_col_name = f"{column_name} 重み"
                    
                    if column_name in chuushutsu_df.columns and weight_col_name in chuushutsu_df.columns:
                        dict_index = i % len(lookup_dicts)
                        lookup_dict = lookup_dicts[dict_index]
                        
                        # Recalculate weight values efficiently
                        weight_values = [self.lookup_weight(lookup_dict, val, None) for val in chuushutsu_df[column_name]]
                        
                        # Update the weight column
                        chuushutsu_df[weight_col_name] = weight_values

            # Write to Excel
            self.write_to_excel(chuushutsu_df, tensuuka_df, lookup_df)
            
        except Exception as e:
            self.progress_bar.stop()
            self.update_status(f"❌ Error in recalculation: {str(e)}")
            self.update_progress("💥 Process failed!")

    def write_to_excel(self, chuushutsu_df, tensuuka_df, lookup_df):
        """Write the updated DataFrames back to Excel"""
        self.update_status("💾 Saving updated data to Excel...")
        
        # Create a new Excel writer
        with pd.ExcelWriter(self.workbook, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            # Write the updated 抽出シート
            chuushutsu_df.to_excel(writer, sheet_name='抽出シート', index=False)
            
            # Also write the other sheets to preserve them
            tensuuka_df.to_excel(writer, sheet_name='点数化列', index=False)
            lookup_df.to_excel(writer, sheet_name='重みテーブル', index=False)
        
        self.continue_to_final_step()

    def create_lookup_dicts(self, lookup_df):
        lookup_dicts = []
        
        # Process pairs of columns (key, value)
        for i in range(0, len(lookup_df.columns), 2):
            if i + 1 < len(lookup_df.columns):
                key_col = lookup_df.columns[i]
                value_col = lookup_df.columns[i + 1]
                
                # Create dictionary with proper key-value pairs
                lookup_dict = {}
                for idx, row in lookup_df.iterrows():
                    key = row[key_col]
                    value = row[value_col]
                    
                    # Skip NaN values
                    if pd.isna(key) or pd.isna(value):
                        continue
                    
                    # Convert key to string and apply character conversion
                    key_str = str(key).strip()
                    key_converted = self.convert_to_hankaku(key_str)
                    
                    # Store the mapping
                    lookup_dict[key_converted] = int(value) if pd.notna(value) else 0
                
                lookup_dicts.append(lookup_dict)
        
        return lookup_dicts

    def extract_first_value(self, text):
        """Extract the first value before delimiters"""
        if pd.isna(text) or text == '':
            return ''
        
        text = str(text).strip()
        
        # Define delimiters
        delimiters = ['、', ',', ',', ' ', '　']
        
        # Find the position of the first delimiter
        min_pos = len(text)
        for delimiter in delimiters:
            pos = text.find(delimiter)
            if pos != -1 and pos < min_pos:
                min_pos = pos
        
        # Extract the first value before the delimiter
        if min_pos < len(text):
            result = text[:min_pos].strip()
        else:
            result = text.strip()
            
        return result

    def lookup_weight(self, lookup_dict, value, not_found_values):
        # Handle NaN or empty values
        if pd.isna(value) or value == 'nan' or value == '' or value is None:
            return ''
        
        # Convert to string and extract first value
        value_str = str(value).strip()
        first_value = self.extract_first_value(value_str)
        
        # Convert to hankaku
        converted_value = self.convert_to_hankaku(first_value)
        
        # Check if the converted value exists in lookup dictionary
        if converted_value in lookup_dict:
            return lookup_dict[converted_value]
        
        # If not found, try original value without conversion
        if first_value in lookup_dict:
            return lookup_dict[first_value]
        
        # Try exact match with original input
        if value_str in lookup_dict:
            return lookup_dict[value_str]
        
        # Add to not_found_values if not found and list is provided
        if not_found_values is not None and converted_value not in not_found_values:
            not_found_values.append(converted_value)
        return 1  # Default weight if not found

    def convert_to_hankaku(self, text):
        """Convert full-width characters to half-width"""
        if pd.isna(text) or text == '':
            return ''
        
        text = str(text)
        hankaku = ""
        
        for char in text:
            code = ord(char)
            # Full-width ASCII characters
            if 0xFF01 <= code <= 0xFF5E:
                hankaku += chr(code - 0xFF00 + 0x20)
            # Full-width digits
            elif 0xFF10 <= code <= 0xFF19:
                hankaku += chr(code - 0xFF10 + 0x30)
            # Full-width uppercase letters  
            elif 0xFF21 <= code <= 0xFF3A:
                hankaku += chr(code - 0xFF21 + 0x41)
            # Full-width lowercase letters
            elif 0xFF41 <= code <= 0xFF5A:
                hankaku += chr(code - 0xFF41 + 0x61)
            else:
                hankaku += char
        
        return hankaku

# Create the main window
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelProcessorApp(root)
    root.mainloop()