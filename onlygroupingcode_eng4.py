import pandas as pd
import openpyxl
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json
import os
from collections import defaultdict

class CleanDataGroupingApp:
    def __init__(self):
        self.processing = False
        self.root = tk.Tk()
        self.root.title("Data Grouping System")
        self.root.geometry("600x400")
        self.root.minsize(500, 300)
        self.root.configure(bg='#f5f5f5')
        
        # Center the window
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() // 2) - (600 // 2)
        y = (self.root.winfo_screenheight() // 2) - (400 // 2)
        self.root.geometry(f"600x400+{x}+{y}")
        
        self.workbook_path = None
        self.rules_file = None
        self.rules = []
        self.enzan_kekka_df = None
        self.structure_df = None
        
        self.create_main_gui()
    
    def abbreviate_sen_name(self, sen_name):
        """Convert route name to abbreviation"""
        if pd.isna(sen_name) or sen_name == '':
            return ''
        
        sen_name = str(sen_name).strip()
        
        abbreviation_map = {
            "東急多摩川線": "TM", "多摩川線": "TM", "東横線": "TY",
            "大井町線": "OM", "池上線": "IK", "田園都市線": "DT",
            "目黒線": "MG", "こどもの国線": "KD", "世田谷線": "SG"
        }
        
        return abbreviation_map.get(sen_name, sen_name)
    
    def lookup_structure_number(self, structure_df, rosen_name, kozo_name, ekikan):
        """Lookup 構造物番号 from structure sheet"""
        try:
            rosen_name = str(rosen_name).strip() if pd.notna(rosen_name) else ''
            
            if kozo_name and str(kozo_name).strip() not in ['', 'nan', 'NaN']:
                kozo_name = str(kozo_name).strip()
                matches = structure_df[
                    (structure_df['構造物名称'].astype(str).str.strip() == kozo_name) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    bangou = matches.iloc[0]['構造物番号']
                    if pd.notna(bangou) and str(bangou).strip() not in ['', 'nan']:
                        return str(bangou).strip()
            
            if ekikan and str(ekikan).strip() not in ['', 'nan', 'NaN']:
                ekikan = str(ekikan).strip()
                matches = structure_df[
                    (structure_df['駅間'].astype(str).str.strip() == ekikan) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    bangou = matches.iloc[0]['構造物番号']
                    if pd.notna(bangou) and str(bangou).strip() not in ['', 'nan']:
                        return str(bangou).strip()
            
            return ''
            
        except Exception as e:
            return ''
    
    def create_enhanced_grouping_key(self, shubetsu, tenken_kubun, structure_name, eki_start, eki_end, group_method):
        """Create grouping key with enhanced logic"""
        try:
            if group_method == "構造物名称":
                if tenken_kubun == "*":
                    key = f"{shubetsu}|{structure_name}"
                else:
                    key = f"{shubetsu}|{structure_name}|{tenken_kubun}"
            else:  # 駅間
                ekikan = f"{eki_start}→{eki_end}" if eki_start and eki_end else ""
                if tenken_kubun == "*":
                    key = f"{shubetsu}|{ekikan}"
                else:
                    key = f"{shubetsu}|{ekikan}|{tenken_kubun}"
            
            return key
            
        except Exception as e:
            return "UNKNOWN"
    
    def create_main_gui(self):
        """Create main GUI for file selection"""
        main_frame = tk.Frame(self.root, bg='#f5f5f5', padx=40, pady=40)
        main_frame.pack(fill="both", expand=True)
        
        # Title
        title_label = tk.Label(main_frame, text="Data Grouping System", 
                              font=("Arial", 18, "bold"), fg="#2c3e50", bg='#f5f5f5')
        title_label.pack(pady=(0, 30))
        
        # Status label for feedback
        self.status_label = tk.Label(main_frame, text="Ready to select file...", 
                                    font=("Arial", 10), fg="#666666", bg='#f5f5f5')
        self.status_label.pack(pady=(0, 20))
        
        # Select file button
        select_btn = tk.Button(main_frame, text="Select Excel File", 
                             command=self.select_workbook_with_feedback, 
                             bg="#4a90e2", fg="white", 
                             width=20, height=2, font=("Arial", 11, "bold"),
                             cursor="hand2", relief="flat", bd=1)
        select_btn.pack(pady=10)

    def select_workbook_with_feedback(self):
        """Select workbook with user feedback"""
        self.status_label.config(text="Opening file browser...", fg="#4a90e2")
        self.root.update()
        
        self.workbook_path = filedialog.askopenfilename(
            title="Select Excel Workbook",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        
        if not self.workbook_path:
            self.status_label.config(text="No file selected.", fg="#e74c3c")
            return
        
        self.status_label.config(text="Loading file...", fg="#4a90e2")
        self.root.update()
        
        self.root.after(100, self.validate_workbook)

    def validate_workbook(self):
        """Validate workbook with progress feedback"""
        try:
            if not os.path.exists(self.workbook_path):
                raise Exception("File not found")
            
            wb = load_workbook(self.workbook_path)
            required_sheets = ['演算結果']
            missing_sheets = [sheet for sheet in required_sheets if sheet not in wb.sheetnames]
            
            if missing_sheets:
                self.status_label.config(text="Required sheet not found!", fg="#e74c3c")
                messagebox.showerror("Error", f"Required sheet not found: {', '.join(missing_sheets)}")
                self.status_label.config(text="Ready to select file...", fg="#666666")
                return
            
            self.enzan_kekka_df = pd.read_excel(self.workbook_path, sheet_name='演算結果')
            
            if len(self.enzan_kekka_df) == 0:
                raise Exception("The calculation results sheet is empty")
            
            # Try to load structure data if it exists
            try:
                self.structure_df = pd.read_excel(self.workbook_path, sheet_name='構造物番号')
            except:
                self.structure_df = None
            
            self.rules_file = os.path.join(os.path.dirname(self.workbook_path), "grouping_rules.json")
            self.rules = self.load_rules()
            
            self.status_label.config(text="File loaded successfully!", fg="#27ae60")
            
            # Auto-proceed to grouping
            self.root.after(1000, self.start_auto_grouping)
            
        except Exception as e:
            self.status_label.config(text="Error loading file", fg="#e74c3c")
            messagebox.showerror("Error", f"Failed to load Excel file:\n{str(e)}")
            self.status_label.config(text="Ready to select file...", fg="#666666")

    def start_auto_grouping(self):
        """Start automatic grouping process"""
        self.status_label.config(text="Starting grouping process...", fg="#4a90e2")
        self.root.withdraw()  # Hide main window
        self.show_clean_grouping_manager()

    def load_rules(self):
        """Load existing rules from JSON file"""
        default_rules = [
            {"shubetsu": "停車場", "tenken_kubun": "*", "group_by": "構造物名称", "description": "Station grouped by structure name"},
            {"shubetsu": "擁壁・法面", "tenken_kubun": "*", "group_by": "駅間", "description": "Retaining walls grouped by station interval"},
            {"shubetsu": "線路設備", "tenken_kubun": "*", "group_by": "駅間", "description": "Track equipment grouped by station interval"},
            {"shubetsu": "トンネル", "tenken_kubun": "*", "group_by": "構造物名称", "description": "Tunnels grouped by structure name"},
            {"shubetsu": "高架橋", "tenken_kubun": "*", "group_by": "構造物名称", "description": "Elevated bridges grouped by structure name"}
        ]
        
        if os.path.exists(self.rules_file):
            try:
                with open(self.rules_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                return default_rules
        else:
            self.save_rules(default_rules)
            return default_rules
    
    def save_rules(self, rules=None):
        """Save rules to JSON file"""
        if rules is None:
            rules = self.rules
        with open(self.rules_file, 'w', encoding='utf-8') as f:
            json.dump(rules, f, ensure_ascii=False, indent=2)
    
    def show_clean_grouping_manager(self):
        """Show clean grouping rules management window - RESTORED ALL FUNCTIONALITY"""
        self.main_window = tk.Toplevel()
        self.main_window.title("Grouping Rules Management")
        self.main_window.geometry("1200x700")
        self.main_window.configure(bg='#f5f5f5')
        self.main_window.grab_set()
        self.main_window.resizable(True, True)
        
        def on_closing():
            if messagebox.askyesno("Close Application", "Are you sure?"):
                self.main_window.destroy()
                self.root.quit()
        
        self.main_window.protocol("WM_DELETE_WINDOW", on_closing)
        
        # Center window
        self.main_window.update_idletasks()
        x = (self.main_window.winfo_screenwidth() // 2) - (1200 // 2)
        y = (self.main_window.winfo_screenheight() // 2) - (700 // 2)
        self.main_window.geometry(f"1200x700+{x}+{y}")
        
                # Main frame
        main_frame = tk.Frame(self.main_window, bg='#f5f5f5', padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)
        
        # Title and file info
        title_frame = tk.Frame(main_frame, bg='#f5f5f5')
        title_frame.pack(fill="x", pady=(0, 20))
        
        title_label = tk.Label(title_frame, text="Data Grouping Rules Management", 
                              font=("Arial", 16, "bold"), fg="#2c3e50", bg='#f5f5f5')
        title_label.pack(anchor="w")
        
        file_label = tk.Label(title_frame, text=f"File: {os.path.basename(self.workbook_path)}", 
                             font=("Arial", 10), fg="#666666", bg='#f5f5f5')
        file_label.pack(anchor="w")
        
        data_info_label = tk.Label(title_frame, text=f"Data Count: {len(self.enzan_kekka_df):,} records", 
                                  font=("Arial", 10), fg="#4a90e2", bg='#f5f5f5')
        data_info_label.pack(anchor="w")
        
        # Enhancement status
        enhancement_text = "Enhanced with 構造物番号 & 路線名略称 columns" if self.structure_df is not None else "Basic version"
        enhancement_label = tk.Label(title_frame, text=enhancement_text, 
                                   font=("Arial", 9), fg="#27ae60" if self.structure_df is not None else "#e67e22", bg='#f5f5f5')
        enhancement_label.pack(anchor="w")
        
        # Rules display frame
        rules_frame = tk.LabelFrame(main_frame, text="Registered Rules", 
                                   font=("Arial", 12, "bold"), bg='#f5f5f5', fg="#2c3e50",
                                   relief="solid", bd=1, padx=15, pady=15)
        rules_frame.pack(fill="both", expand=True, pady=(0, 20))
        
        # Create treeview for rules display
        columns = ("No", "Type", "Inspection Category", "Grouping Method", "Description")
        self.rules_tree = ttk.Treeview(rules_frame, columns=columns, show="headings", height=12)
        
        # Define column headings and widths
        column_widths = {"No": 50, "Type": 150, "Inspection Category": 200, 
                        "Grouping Method": 150, "Description": 350}
        
        for col in columns:
            self.rules_tree.heading(col, text=col)
            self.rules_tree.column(col, width=column_widths.get(col, 100))
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(rules_frame, orient="vertical", command=self.rules_tree.yview)
        self.rules_tree.configure(yscrollcommand=scrollbar.set)
        
        # Pack treeview and scrollbar
        self.rules_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Populate rules
        self.refresh_rules_display()
        
        # Buttons frame - RESTORED ALL ORIGINAL BUTTONS
        button_frame = tk.Frame(main_frame, bg='#f5f5f5')
        button_frame.pack(fill="x", pady=20)
        
        # Rule management buttons
        edit_btn = tk.Button(button_frame, text="✏️ Edit Selected Rule", 
                           command=self.edit_selected_rule, bg="#4a90e2", fg="white", 
                           width=18, height=2, font=("Arial", 10, "bold"), cursor="hand2",
                           relief="solid", bd=1)
        edit_btn.pack(side="left", padx=5)
        
        add_btn = tk.Button(button_frame, text="➕ Add New Rule", 
                          command=self.add_new_rule, bg="#27ae60", fg="white", 
                          width=15, height=2, font=("Arial", 10, "bold"), cursor="hand2",
                          relief="solid", bd=1)
        add_btn.pack(side="left", padx=5)
        
        delete_btn = tk.Button(button_frame, text="🗑️ Delete Rule", 
                             command=self.delete_selected_rule, bg="#e74c3c", fg="white", 
                             width=15, height=2, font=("Arial", 10, "bold"), cursor="hand2",
                             relief="solid", bd=1)
        delete_btn.pack(side="left", padx=5)
        
        # Main action buttons
        action_frame = tk.Frame(main_frame, bg='#f5f5f5')
        action_frame.pack(fill="x", pady=(20, 0))
        
        self.continue_btn = tk.Button(action_frame, text="🚀 Start Grouping", 
                               command=self.start_clean_grouping_process, bg="#e67e22", fg="white", 
                               width=20, height=2, font=("Arial", 11, "bold"), cursor="hand2",
                               relief="solid", bd=1)
        self.continue_btn.pack(side="right", padx=15)
        
        back_btn = tk.Button(action_frame, text="⬅️ Back to File Selection", 
                           command=self.back_to_file_selection, bg="#95a5a6", fg="white", 
                           width=18, height=2, font=("Arial", 10), cursor="hand2",
                           relief="solid", bd=1)
        back_btn.pack(side="right", padx=15)
    
    def refresh_rules_display(self):
        """Refresh the rules display in treeview"""
        # Clear existing items
        for item in self.rules_tree.get_children():
            self.rules_tree.delete(item)
        
        # Add rules to treeview
        for i, rule in enumerate(self.rules, 1):
            tenken_display = "All" if rule["tenken_kubun"] == "*" else rule["tenken_kubun"]
            group_by_display = "Structure Name" if rule["group_by"] == "構造物名称" else "Station Interval"
            
            self.rules_tree.insert("", "end", values=(
                i, 
                rule["shubetsu"], 
                tenken_display,
                group_by_display,
                rule.get("description", "")
            ))
    
    def edit_selected_rule(self):
        """Edit the selected rule - RESTORED ORIGINAL FUNCTIONALITY"""
        selection = self.rules_tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a rule to edit.")
            return
        
        item = self.rules_tree.item(selection[0])
        rule_index = int(item['values'][0]) - 1
        self.show_rule_edit_dialog(rule_index)
    
    def add_new_rule(self):
        """Add a new rule - RESTORED ORIGINAL FUNCTIONALITY"""
        self.show_rule_edit_dialog(-1)
    
    def delete_selected_rule(self):
        """Delete the selected rule - RESTORED ORIGINAL FUNCTIONALITY"""
        selection = self.rules_tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a rule to delete.")
            return
        
        item = self.rules_tree.item(selection[0])
        rule_name = f"{item['values'][1]} - {item['values'][2]}"
        
        if messagebox.askyesno("Confirm Deletion", 
                              f"Delete this rule?\n\nRule: {rule_name}"):
            rule_index = int(item['values'][0]) - 1
            del self.rules[rule_index]
            self.save_rules()
            self.refresh_rules_display()
    
    def show_rule_edit_dialog(self, rule_index):
        """Show dialog for editing/adding rules - RESTORED ORIGINAL WITH CLEAN STYLING"""
        edit_window = tk.Toplevel(self.main_window)
        is_new = rule_index == -1
        title = "Add New Rule" if is_new else "Edit Rule"
        edit_window.title(title)
        edit_window.geometry("500x450")
        edit_window.configure(bg='#f5f5f5')
        edit_window.grab_set()
        edit_window.resizable(False, False)
        edit_window.transient(self.main_window)
        
        # Center the dialog
        edit_window.update_idletasks()
        x = (edit_window.winfo_screenwidth() // 2) - (500 // 2)
        y = (edit_window.winfo_screenheight() // 2) - (450 // 2)
        edit_window.geometry(f"500x450+{x}+{y}")
        
        main_frame = tk.Frame(edit_window, bg='#f5f5f5', padx=25, pady=25)
        main_frame.pack(fill="both", expand=True)
        
        # Title
        title_label = tk.Label(main_frame, text=title, font=("Arial", 14, "bold"), 
                              fg="#2c3e50", bg='#f5f5f5')
        title_label.pack(pady=(0, 20))
        
        # Get current rule data
        current_rule = self.rules[rule_index] if not is_new else {
            "shubetsu": "", "tenken_kubun": "*", "group_by": "構造物名称", "description": ""
        }
        
        # Form fields
        fields_frame = tk.Frame(main_frame, bg='#f5f5f5')
        fields_frame.pack(fill="x", pady=20)
        
        # Type field
        tk.Label(fields_frame, text="Structure Type:", font=("Arial", 10, "bold"), 
                bg='#f5f5f5').grid(row=0, column=0, sticky="w", pady=10)
        shubetsu_var = tk.StringVar(value=current_rule["shubetsu"])
        shubetsu_entry = ttk.Combobox(fields_frame, textvariable=shubetsu_var, width=30, font=("Arial", 10))
        
        if '種別' in self.enzan_kekka_df.columns:
            unique_shubetsu = sorted(self.enzan_kekka_df['種別'].dropna().unique().tolist())
            shubetsu_entry['values'] = unique_shubetsu
        
        shubetsu_entry.grid(row=0, column=1, sticky="w", padx=(10,0), pady=10)
        
        # Inspection Category field - with "All" option
        tk.Label(fields_frame, text="Inspection Category:", font=("Arial", 10, "bold"), 
                bg='#f5f5f5').grid(row=1, column=0, sticky="w", pady=10)
        tenken_var = tk.StringVar(value="All" if current_rule["tenken_kubun"] == "*" else current_rule["tenken_kubun"])
        tenken_entry = ttk.Combobox(fields_frame, textvariable=tenken_var, width=30, font=("Arial", 10))
        
        if '点検区分1' in self.enzan_kekka_df.columns:
            unique_tenken = sorted(self.enzan_kekka_df['点検区分1'].dropna().unique().tolist())
            unique_tenken.insert(0, "All")  # Use "All" instead of "*"
            tenken_entry['values'] = unique_tenken
        else:
            tenken_entry['values'] = ["All"]
        
        tenken_entry.grid(row=1, column=1, sticky="w", padx=(10,0), pady=10)
        
        # Help text
        help_label = tk.Label(fields_frame, text="※ \"All\" applies to all inspection categories", 
                             font=("Arial", 9), fg="#666666", bg='#f5f5f5')
        help_label.grid(row=2, column=0, columnspan=2, sticky="w", pady=(0,10))
        
        # Grouping Method field
        tk.Label(fields_frame, text="Grouping Method:", font=("Arial", 10, "bold"), 
                bg='#f5f5f5').grid(row=3, column=0, sticky="w", pady=10)
        group_by_var = tk.StringVar(value=current_rule["group_by"])
        
        radio_frame = tk.Frame(fields_frame, bg='#f5f5f5')
        radio_frame.grid(row=3, column=1, sticky="w", padx=(10,0), pady=10)
        
        structure_radio = tk.Radiobutton(radio_frame, text="Structure Name", 
                                       variable=group_by_var, value="構造物名称", 
                                       font=("Arial", 11), bg='#f5f5f5')
        structure_radio.pack(anchor="w")
        
        station_radio = tk.Radiobutton(radio_frame, text="Station Interval", 
                                     variable=group_by_var, value="駅間", 
                                     font=("Arial", 11), bg='#f5f5f5')
        station_radio.pack(anchor="w")
        
        # Description field
        tk.Label(fields_frame, text="Description:", font=("Arial", 10, "bold"), 
                bg='#f5f5f5').grid(row=4, column=0, sticky="w", pady=10)
        description_var = tk.StringVar(value=current_rule.get("description", ""))
        description_entry = tk.Entry(fields_frame, textvariable=description_var, width=40, 
                                   font=("Arial", 10), relief="solid", bd=1)
        description_entry.grid(row=4, column=1, sticky="w", padx=(10,0), pady=10)
        
        # Buttons
        button_frame = tk.Frame(main_frame, bg='#f5f5f5')
        button_frame.pack(fill="x", pady=(20, 0))
        
        def save_rule():
            shubetsu = shubetsu_var.get().strip()
            tenken = tenken_var.get().strip()
            group_by = group_by_var.get()
            description = description_var.get().strip()
            
            # Convert "All" back to "*" for internal storage
            if tenken == "All":
                tenken = "*"
                        
            if not shubetsu or not tenken:
                messagebox.showerror("Error", "Please enter both Structure Type and Inspection Category.")
                return
            
            new_rule = {
                "shubetsu": shubetsu,
                "tenken_kubun": tenken,
                "group_by": group_by,
                "description": description
            }
            
            if is_new:
                self.rules.append(new_rule)
            else:
                self.rules[rule_index] = new_rule
            
            self.save_rules()
            self.refresh_rules_display()
            
            edit_window.destroy()
        
        save_btn = tk.Button(button_frame, text="Save", command=save_rule,
                           bg="#27ae60", fg="white", width=12, height=2, 
                           font=("Arial", 10, "bold"), cursor="hand2", relief="solid", bd=1)
        save_btn.pack(side="right", padx=10)
        
        cancel_btn = tk.Button(button_frame, text="Cancel", command=edit_window.destroy,
                             bg="#95a5a6", fg="white", width=12, height=2, 
                             font=("Arial", 10), cursor="hand2", relief="solid", bd=1)
        cancel_btn.pack(side="right", padx=10)
    
    def back_to_file_selection(self):
        """Go back to file selection - RESTORED ORIGINAL FUNCTIONALITY"""
        self.main_window.destroy()
        self.root.deiconify()
    
    def start_clean_grouping_process(self):
        """Start the clean grouping process with progress indicator"""
        if self.processing:
            return  # Prevent multiple clicks
        
        self.processing = True
        
        # Disable the button and show loading
        self.continue_btn.config(text="⏳ Processing...", state="disabled", bg="#95a5a6")
        
        # Show progress window
        self.show_processing_dialog()
        
        # Start processing in background thread
        import threading
        threading.Thread(target=self.process_with_progress, daemon=True).start()
    
    def find_matching_rule(self, shubetsu, tenken_kubun):
        """Find matching rule for given shubetsu and tenken_kubun"""
        for rule in self.rules:
            if rule["shubetsu"] == shubetsu:
                if rule["tenken_kubun"] == "*" or rule["tenken_kubun"] == tenken_kubun:
                    return rule
        return None
    
    def show_processing_dialog(self):
        """Show processing dialog with progress indicator"""
        self.progress_dialog = tk.Toplevel(self.main_window)
        self.progress_dialog.title("Processing")
        self.progress_dialog.geometry("400x150")
        self.progress_dialog.configure(bg='#f5f5f5')
        self.progress_dialog.resizable(False, False)
        self.progress_dialog.transient(self.main_window)
        self.progress_dialog.grab_set()
        
        # Center dialog
        self.progress_dialog.geometry("+%d+%d" % (self.main_window.winfo_rootx() + 400, self.main_window.winfo_rooty() + 200))
        
        frame = tk.Frame(self.progress_dialog, bg='#f5f5f5', padx=30, pady=30)
        frame.pack(fill="both", expand=True)
        
        tk.Label(frame, text="🔄 Grouping in Progress...", 
                font=("Arial", 12, "bold"), fg="#4a90e2", bg='#f5f5f5').pack(pady=(0, 15))
        
        # Progress bar
        self.progress_bar = ttk.Progressbar(frame, mode='indeterminate', length=300)
        self.progress_bar.pack(pady=(0, 10))
        self.progress_bar.start(10)
        
        self.progress_status = tk.Label(frame, text="Analyzing data...", 
                                       font=("Arial", 10), fg="#666666", bg='#f5f5f5')
        self.progress_status.pack()

    def process_with_progress(self):
        """Process with progress updates"""
        try:
            # Update status
            self.main_window.after(0, lambda: self.progress_status.config(text="Checking rules..."))
            
            if '種別' not in self.enzan_kekka_df.columns or '点検区分1' not in self.enzan_kekka_df.columns:
                raise Exception("Required columns not found in data")
            
            unique_combinations = self.enzan_kekka_df[['種別', '点検区分1']].drop_duplicates()
            missing_rules = []
            
            # Check for missing rules
            for _, row in unique_combinations.iterrows():
                shubetsu = str(row['種別']) if pd.notna(row['種別']) else ""
                tenken = str(row['点検区分1']) if pd.notna(row['点検区分1']) else ""
                
                if not self.find_matching_rule(shubetsu, tenken):
                    missing_rules.append((shubetsu, tenken))
            
            if missing_rules:
                # Close progress dialog and show missing rules
                self.main_window.after(0, self.close_progress_dialog)
                self.main_window.after(100, lambda: self.show_missing_rules_dialog(missing_rules))
            else:
                # Update status and perform grouping
                self.main_window.after(0, lambda: self.progress_status.config(text="Processing grouping..."))
                self.perform_grouping_with_progress()
                
        except Exception as e:
            self.main_window.after(0, self.close_progress_dialog)
            self.main_window.after(100, lambda: messagebox.showerror("Error", f"Error: {str(e)}"))
            self.processing = False

    def perform_grouping_with_progress(self):
        """Perform grouping with progress updates"""
        try:
            self.main_window.after(0, lambda: self.progress_status.config(text="Creating groups..."))
            
            # Call the existing perform_grouping method
            self.perform_grouping()
            
            # Close progress and show completion
            self.main_window.after(0, self.close_progress_dialog)
            self.main_window.after(100, lambda: messagebox.showinfo("Complete", "グループ化点検履歴 sheet generated successfully!"))
            self.main_window.after(1500, self.auto_complete)
            
        except Exception as e:
            self.main_window.after(0, self.close_progress_dialog)
            self.main_window.after(100, lambda: messagebox.showerror("Error", f"Error during grouping: {str(e)}"))
            self.processing = False

    def close_progress_dialog(self):
        """Close progress dialog"""
        if hasattr(self, 'progress_dialog'):
            self.progress_dialog.destroy()
        self.processing = False

    def show_missing_rules_dialog(self, missing_rules):
        """Show clean dialog for missing rules - RESTORED ORIGINAL FUNCTIONALITY WITH BETTER STYLING"""
        missing_window = tk.Toplevel(self.main_window)
        missing_window.title("Configure Missing Rules")
        missing_window.geometry("1000x650")
        missing_window.configure(bg='#f5f5f5')
        missing_window.grab_set()
        missing_window.resizable(True, True)
        missing_window.transient(self.main_window)
        
        # Center window
        missing_window.update_idletasks()
        x = (missing_window.winfo_screenwidth() // 2) - (1000 // 2)
        y = (missing_window.winfo_screenheight() // 2) - (650 // 2)
        missing_window.geometry(f"1000x650+{x}+{y}")
        
        main_frame = tk.Frame(missing_window, bg='#f5f5f5', padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)
        
        # Title
        title_label = tk.Label(main_frame, text="Configure Missing Rules", 
                    font=("Arial", 18, "bold"), fg="#2c3e50", bg='#f5f5f5')
        title_label.pack(pady=(0, 10))
        
        subtitle_label = tk.Label(main_frame, 
                        text="Choose how to group each structure type", 
                        font=("Arial", 12), bg='#f5f5f5')
        subtitle_label.pack(pady=(0, 20))
        
        # Group missing rules by 種別
        shubetsu_groups = defaultdict(list)
        for shubetsu, tenken in missing_rules:
            shubetsu_groups[shubetsu].append(tenken)
        
                # Scrollable frame for 種別 groups
        canvas = tk.Canvas(main_frame, height=350, bg='#f5f5f5')
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#f5f5f5')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Store user choices
        self.shubetsu_choices = {}
        
        row_count = 0
        for shubetsu, tenken_list in shubetsu_groups.items():
            # Create frame for each 種別
            shubetsu_frame = tk.LabelFrame(scrollable_frame, 
                                        text=f"種別: {shubetsu}", 
                                        font=("Arial", 12, "bold"), 
                                        bg='#f5f5f5', fg="#2c3e50",
                                        relief="solid", bd=1,
                                        padx=15, pady=15)
            shubetsu_frame.grid(row=row_count, column=0, sticky="ew", padx=10, pady=10)
            scrollable_frame.grid_columnconfigure(0, weight=1)
            
            # Show what 点検区分1 values exist for this 種別
            tenken_info = tk.Label(shubetsu_frame, 
                                text=f"Found categories: {', '.join(tenken_list)}", 
                                font=("Arial", 10), fg="#4a90e2", bg='#f5f5f5', wraplength=400)
            tenken_info.pack(anchor="w", pady=(0, 10))
            
            # Choice variable for this 種別
            choice_var = tk.StringVar(value="universal")
            self.shubetsu_choices[shubetsu] = {
                'choice_var': choice_var,
                'tenken_list': tenken_list,
                'group_method_var': tk.StringVar(value="構造物名称"),
                'individual_methods': {}
            }
            
            # Option 1: Universal rule (use "All")
            universal_frame = tk.Frame(shubetsu_frame, bg='#f5f5f5')
            universal_frame.pack(fill="x", pady=5)
            
            universal_radio = tk.Radiobutton(universal_frame, 
                                        text="Create ONE rule for ALL inspection categories", 
                                        variable=choice_var, value="universal",
                                        font=("Arial", 11, "bold"), fg="#27ae60", bg='#f5f5f5')
            universal_radio.pack(anchor="w")
            
            # Grouping method for universal choice
            universal_method_frame = tk.Frame(universal_frame, bg='#f5f5f5')
            universal_method_frame.pack(fill="x", padx=20, pady=5)
            
            tk.Label(universal_method_frame, text="Group by:", font=("Arial", 10), bg='#f5f5f5').pack(side="left")
            
            struktur_radio = tk.Radiobutton(universal_method_frame, text="構造物名称", 
                                        variable=self.shubetsu_choices[shubetsu]['group_method_var'], 
                                        value="構造物名称", font=("Arial", 10), bg='#f5f5f5')
            struktur_radio.pack(side="left", padx=10)
            
            ekikan_radio = tk.Radiobutton(universal_method_frame, text="駅間", 
                                        variable=self.shubetsu_choices[shubetsu]['group_method_var'], 
                                        value="駅間", font=("Arial", 10), bg='#f5f5f5')
            ekikan_radio.pack(side="left", padx=10)
            
            # Option 2: Individual rules
            individual_frame = tk.Frame(shubetsu_frame, bg='#f5f5f5')
            individual_frame.pack(fill="x", pady=5)
            
            individual_radio = tk.Radiobutton(individual_frame, 
                                            text="Create SEPARATE rules for each inspection category", 
                                            variable=choice_var, value="individual",
                                            font=("Arial", 11, "bold"), fg="#e67e22", bg='#f5f5f5')
            individual_radio.pack(anchor="w")
            
            # Individual grouping methods for each 点検区分1
            individual_details_frame = tk.Frame(individual_frame, bg='#f5f5f5')
            individual_details_frame.pack(fill="x", padx=20, pady=5)
            
            for tenken in tenken_list:
                tenken_frame = tk.Frame(individual_details_frame, bg='#f5f5f5')
                tenken_frame.pack(fill="x", pady=2)
                
                tk.Label(tenken_frame, text=f"  {tenken}:", font=("Arial", 10), 
                        width=20, anchor="w", bg='#f5f5f5').pack(side="left")
                
                method_var = tk.StringVar(value="構造物名称")
                self.shubetsu_choices[shubetsu]['individual_methods'][tenken] = method_var
                
                struktur_radio2 = tk.Radiobutton(tenken_frame, text="構造物名称", 
                                            variable=method_var, value="構造物名称", 
                                            font=("Arial", 9), bg='#f5f5f5')
                struktur_radio2.pack(side="left", padx=5)
                
                ekikan_radio2 = tk.Radiobutton(tenken_frame, text="駅間", 
                                            variable=method_var, value="駅間", 
                                            font=("Arial", 9), bg='#f5f5f5')
                ekikan_radio2.pack(side="left", padx=5)
            
            row_count += 1
        
        # Apply smart rules function
        def apply_smart_rules():
            """Apply the smart rules based on user choices"""
            new_rules = []
            
            for shubetsu, choice_data in self.shubetsu_choices.items():
                choice = choice_data['choice_var'].get()
                
                if choice == "universal":
                    # Create one rule with "*" for all 点検区分1
                    group_method = choice_data['group_method_var'].get()
                    new_rule = {
                        "shubetsu": shubetsu,
                        "tenken_kubun": "*",
                        "group_by": group_method,
                        "description": f"Universal rule for {shubetsu}"
                    }
                    new_rules.append(new_rule)
                    
                elif choice == "individual":
                    # Create separate rules for each 点検区分1
                    for tenken in choice_data['tenken_list']:
                        group_method = choice_data['individual_methods'][tenken].get()
                        new_rule = {
                            "shubetsu": shubetsu,
                            "tenken_kubun": tenken,
                            "group_by": group_method,
                            "description": f"Individual rule for {shubetsu} - {tenken}"
                        }
                        new_rules.append(new_rule)
            
            # Add new rules to existing ones
            self.rules.extend(new_rules)
            self.save_rules()
            
            missing_window.destroy()
            self.show_processing_dialog()
            
            # Continue with grouping in background
            import threading
            threading.Thread(target=self.perform_grouping_with_progress, daemon=True).start()
        
        # Bulk assignment functions
        def set_all_universal_struktur():
            """Set all to universal with 構造物名称"""
            for choice_data in self.shubetsu_choices.values():
                choice_data['choice_var'].set("universal")
                choice_data['group_method_var'].set("構造物名称")
        
        def set_all_universal_ekikan():
            """Set all to universal with 駅間"""
            for choice_data in self.shubetsu_choices.values():
                choice_data['choice_var'].set("universal")
                choice_data['group_method_var'].set("駅間")
        
        def set_all_individual():
            """Set all to individual rules"""
            for choice_data in self.shubetsu_choices.values():
                choice_data['choice_var'].set("individual")
        
        # Bulk assignment buttons
        bulk_frame = tk.LabelFrame(main_frame, text="Quick Actions", 
                                  font=("Arial", 11, "bold"), bg='#f5f5f5', 
                                  fg="#2c3e50", relief="solid", bd=1, padx=15, pady=10)
        bulk_frame.pack(fill="x", pady=(0, 20))
        
        bulk_button_frame = tk.Frame(bulk_frame, bg='#f5f5f5')
        bulk_button_frame.pack(fill="x")
        
        tk.Button(bulk_button_frame, text="All Universal (構造物名称)", 
                command=set_all_universal_struktur, bg="#27ae60", fg="white", 
                width=20, font=("Arial", 10), cursor="hand2", relief="solid", bd=1).pack(side="left", padx=5)
        
        tk.Button(bulk_button_frame, text="All Universal (駅間)", 
              command=set_all_universal_ekikan, bg="#4a90e2", fg="white", 
              width=18, font=("Arial", 10), cursor="hand2", relief="solid", bd=1).pack(side="left", padx=5)
    
        tk.Button(bulk_button_frame, text="All Individual", 
                command=set_all_individual, bg="#e67e22", fg="white", 
                width=15, font=("Arial", 10), cursor="hand2", relief="solid", bd=1).pack(side="left", padx=5)
        
        # Main action buttons
        action_frame = tk.Frame(main_frame, bg='#f5f5f5')
        action_frame.pack(fill="x", pady=(20, 0))
        
        apply_btn = tk.Button(action_frame, text="Apply Rules & Start Grouping", 
                            command=apply_smart_rules, bg="#27ae60", fg="white", 
                            width=25, height=2, font=("Arial", 11, "bold"), 
                            cursor="hand2", relief="solid", bd=1)
        apply_btn.pack(side="right", padx=15)
        
        back_btn = tk.Button(action_frame, text="Back", 
                        command=missing_window.destroy, 
                        bg="#95a5a6", fg="white", width=12, height=2, 
                        font=("Arial", 10), cursor="hand2", relief="solid", bd=1)
        back_btn.pack(side="right", padx=15)

    def perform_grouping(self):
        """Perform the data grouping - CLEAN VERSION WITH AUTO-CLOSE"""
        try:
            # Create a copy of the data for processing
            df = self.enzan_kekka_df.copy()
            
            # Generate grouping keys
            grouping_keys = []
            grouping_methods = []
            
            for _, row in df.iterrows():
                shubetsu = str(row.get('種別', '')) if pd.notna(row.get('種別')) else ''
                tenken = str(row.get('点検区分1', '')) if pd.notna(row.get('点検区分1')) else ''
                
                # Find matching rule
                rule = self.find_matching_rule(shubetsu, tenken)
                
                if rule:
                    group_method = rule['group_by']
                    grouping_methods.append(group_method)
                    
                    if group_method == "構造物名称":
                        # Group by structure name
                        structure_name = str(row.get('構造物名称', '')) if pd.notna(row.get('構造物名称')) else ''
                        key = self.create_enhanced_grouping_key(shubetsu, rule['tenken_kubun'], structure_name, None, None, group_method)
                    else:  # 駅間
                        # Group by station interval
                        eki_start = str(row.get('駅（始）', '')) if pd.notna(row.get('駅（始）')) else ''
                        eki_end = str(row.get('駅（至）', '')) if pd.notna(row.get('駅（至）')) else ''
                        key = self.create_enhanced_grouping_key(shubetsu, rule['tenken_kubun'], None, eki_start, eki_end, group_method)
                    
                    grouping_keys.append(key)
                else:
                    # Fallback
                    grouping_keys.append(f"UNKNOWN|{shubetsu}|{tenken}")
                    grouping_methods.append("構造物名称")
            
            # Add grouping columns to dataframe
            df['Grouping Key'] = grouping_keys
            df['Grouping Method'] = grouping_methods
            
            # Get year columns (result columns)
            year_columns = [col for col in df.columns if col.endswith(' 結果') and any(year in col for year in ['2024', '2023', '2022', '2021', '2020', '2019', '2018'])]
            
            # Group data and aggregate
            grouped_data = []
            
            for group_key in df['Grouping Key'].unique():
                group_df = df[df['Grouping Key'] == group_key]
                
                # Get basic info from first row
                first_row = group_df.iloc[0]
                
                # Create result row
                result_row = {
                    'グループ化キー': group_key,
                    'グループ化方法': first_row['Grouping Method'],
                    'データ件数': len(group_df),
                    '路線名': first_row.get('路線名', ''),
                    '路線名略称': self.abbreviate_sen_name(first_row.get('路線名', '')),
                    '構造物番号': '',
                    '種別': first_row.get('種別', ''),
                    '構造物名称': first_row.get('構造物名称', '') if first_row['Grouping Method'] == "構造物名称" else '',
                    '駅（始）': first_row.get('駅（始）', '') if first_row['Grouping Method'] == "駅間" else '',
                    '駅（至）': first_row.get('駅（至）', '') if first_row['Grouping Method'] == "駅間" else '',
                    '点検区分1': first_row.get('点検区分1', '')
                }
                
                # Add 構造物番号 lookup if structure_df is available
                if self.structure_df is not None:
                    rosen_name = result_row['路線名']
                    kozo_name = result_row['構造物名称']
                    
                    # Create ekikan for lookup
                    ekikan = ''
                    if result_row['駅（始）'] and result_row['駅（至）']:
                        ekikan = f"{result_row['駅（始）']}→{result_row['駅（至）']}"
                    
                    # Lookup structure number
                    bangou = self.lookup_structure_number(self.structure_df, rosen_name, kozo_name, ekikan)
                    result_row['構造物番号'] = bangou
                
                # Aggregate year results
                for year_col in year_columns:
                    # Sum non-empty values
                    values = group_df[year_col].dropna()
                    values = values[values != '']
                    
                    if len(values) > 0:
                        try:
                            numeric_values = pd.to_numeric(values, errors='coerce').dropna()
                            if len(numeric_values) > 0:
                                result_row[year_col] = numeric_values.sum()
                            else:
                                result_row[year_col] = ''
                        except:
                            result_row[year_col] = ''
                    else:
                        result_row[year_col] = ''
                
                grouped_data.append(result_row)
            
            # Create grouped dataframe
            grouped_df = pd.DataFrame(grouped_data)
            
            # Sort by grouping key
            grouped_df = grouped_df.sort_values('グループ化キー')
            
            # Save to Excel
            self.save_grouped_data(grouped_df)
            
        except Exception as e:
            raise e

    def save_grouped_data(self, grouped_df):
        """Save grouped data to Excel sheet with proper column order"""
        # Enhanced column order: データ件数 → 路線名 → 路線名略称 → 構造物番号 → other columns → year columns
        base_columns = ['グループ化キー', 'グループ化方法', '種別','点検区分1', '構造物名称', '駅（始）', '駅（至）','データ件数', '路線名', '路線名略称', '構造物番号']
        
        # Get year columns dynamically and sort from oldest to newest
        year_columns = [col for col in grouped_df.columns if col.endswith(' 結果')]
        
        # Extract years and sort them to ensure correct chronological order
        def extract_year(col_name):
            """Extract year from column name like '2024 結果' """
            try:
                return int(col_name.split(' ')[0])
            except:
                return 0
        
        # Sort year columns by actual year value (oldest to newest)
        year_columns.sort(key=extract_year)
        
        final_columns = base_columns + year_columns
        existing_columns = [col for col in final_columns if col in grouped_df.columns]
        grouped_df = grouped_df[existing_columns]
        
        # Sort by the latest year column in descending order (highest values first)
        if year_columns:
            latest_year_col = year_columns[-1]  # Get the last (newest) year column
            # Convert to numeric for proper sorting, handle non-numeric values
            grouped_df[latest_year_col] = pd.to_numeric(grouped_df[latest_year_col], errors='coerce').fillna(0)
            grouped_df = grouped_df.sort_values(latest_year_col, ascending=False)
        
        # Save to Excel
        with pd.ExcelWriter(self.workbook_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            grouped_df.to_excel(writer, sheet_name='グループ化点検履歴', index=False)
            
            # Preserve other sheets
            try:
                original_wb = load_workbook(self.workbook_path)
                for sheet_name in original_wb.sheetnames:
                    if sheet_name != 'グループ化点検履歴':
                        try:
                            df_temp = pd.read_excel(self.workbook_path, sheet_name=sheet_name)
                            df_temp.to_excel(writer, sheet_name=sheet_name, index=False)
                        except Exception as e:
                            continue
            except Exception as e:
                pass

    def auto_complete(self):
        """Auto-complete and close the application"""
        self.main_window.destroy()
        self.root.quit()

    def run(self):
        """Run the application"""
        self.root.mainloop()


# Main execution
if __name__ == "__main__":
    app = CleanDataGroupingApp()
    app.run()