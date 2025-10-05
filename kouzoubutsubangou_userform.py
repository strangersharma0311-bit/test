import pandas as pd
import openpyxl
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import re
import threading
import time

class StructureDataEntryApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Structure Data Entry System")
        self.root.geometry("500x250")
        self.root.resizable(False, False)
        
        # Center the window on screen
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() // 2) - (500 // 2)
        y = (self.root.winfo_screenheight() // 2) - (250 // 2)
        self.root.geometry(f"500x250+{x}+{y}")
        
        self.workbook_path = None
        self.grouped_df = None
        self.structure_data_df = None
        
        self.create_main_gui()
    
    def create_main_gui(self):
        main_frame = tk.Frame(self.root, padx=30, pady=30)
        main_frame.pack(fill="both", expand=True)
        
        # Title
        title_label = tk.Label(main_frame, text="Structure Data Entry System", 
                              font=("Arial", 16, "bold"), fg="navy")
        title_label.pack(pady=(0, 30))
        
        # Select file button with hover effect
        self.select_btn = tk.Button(main_frame, text="Select Excel File", 
                                   command=self.select_and_load, 
                                   bg="#2196F3", fg="white", 
                                   width=20, height=2, font=("Arial", 12),
                                   cursor="hand2")
        self.select_btn.pack()
        
        # Add hover effects
        self.select_btn.bind("<Enter>", lambda e: e.widget.config(bg="#1976D2"))
        self.select_btn.bind("<Leave>", lambda e: e.widget.config(bg="#2196F3"))
        
        # Progress frame (initially hidden)
        self.progress_frame = tk.Frame(main_frame)
        
        self.status_label = tk.Label(self.progress_frame, text="", 
                                    font=("Arial", 12, "bold"), fg="blue")
        self.status_label.pack(pady=(20, 10))
        
        self.progress_bar = ttk.Progressbar(self.progress_frame, 
                                          length=350, mode='indeterminate')
        self.progress_bar.pack()

    # ... (keeping all the existing methods unchanged until create_structure_form) ...

    def create_structure_form(self, missing_entries):
        # Create form window with proper sizing
        self.form_window = tk.Toplevel()
        self.form_window.title("Structure Data Entry")
        self.form_window.geometry("1400x700")  # Increased width for better layout
        self.form_window.grab_set()
        self.form_window.resizable(True, True)
        
        # Center window
        self.form_window.update_idletasks()
        x = (self.form_window.winfo_screenwidth() // 2) - (1400 // 2)
        y = (self.form_window.winfo_screenheight() // 2) - (700 // 2)
        self.form_window.geometry(f"1400x700+{x}+{y}")
                
        # Main container
        main_container = tk.Frame(self.form_window)
        main_container.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Header section
        header_frame = tk.Frame(main_container)
        header_frame.pack(fill="x", pady=(0, 15))
        
        # Title
        title_label = tk.Label(header_frame, text="Enter Structure Data", 
                              font=("Arial", 16, "bold"), fg="navy")
        title_label.pack(pady=(0, 8))
        
        # Count info
        kozo_count = len([e for e in missing_entries if e['type'] == '構造物名称'])
        ekikan_count = len([e for e in missing_entries if e['type'] == '駅間'])
        info_text = f"Found {kozo_count} structure names + {ekikan_count} station intervals = {len(missing_entries)} total entries"
        info_label = tk.Label(header_frame, text=info_text, font=("Arial", 11), fg="blue")
        info_label.pack(pady=(0, 10))
        
        # Data table section (takes most of the space)
        table_section = tk.Frame(main_container)
        table_section.pack(fill="both", expand=True, pady=(0, 15))
        
        # Create scrollable frame with proper scrollbars
        canvas = tk.Canvas(table_section, bg="white")
        v_scrollbar = ttk.Scrollbar(table_section, orient="vertical", command=canvas.yview)
        h_scrollbar = ttk.Scrollbar(table_section, orient="horizontal", command=canvas.xview)
        
        canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Pack scrollbars and canvas
        v_scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x")
        canvas.pack(side="left", fill="both", expand=True)
        
        # Scrollable content frame
        scrollable_frame = tk.Frame(canvas, bg="white")
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        
        # Store all entry widgets
        self.entry_widgets = {}
        
        # Separate entries by type
        kozo_entries = [e for e in missing_entries if e['type'] == '構造物名称']
        ekikan_entries = [e for e in missing_entries if e['type'] == '駅間']
        
        current_row = 0
        
        # Configure column weights for proper sizing
        for col in range(10):
            scrollable_frame.grid_columnconfigure(col, weight=1, minsize=120)
        
        # Section 1: 構造物名称 entries
        if kozo_entries:
            # Section header
            section_label = tk.Label(scrollable_frame, text=f"構造物名称 Section ({len(kozo_entries)} entries)", 
                                   font=("Arial", 12, "bold"), fg="white", bg="navy",
                                   relief="solid", borderwidth=1, height=2)
            section_label.grid(row=current_row, column=0, columnspan=10, sticky="ew", padx=1, pady=2)
            current_row += 1
            
            # Column headers
            headers = ['路線名', '構造物名称', '構造物番号', '長さ(m)', '構造形式', 
                      '構造形式_重み', '角度', '角度_重み', '供用年数', '供用年数_重み']
            
            for col, header in enumerate(headers):
                header_label = tk.Label(scrollable_frame, text=header, 
                                      font=("Arial", 10, "bold"), bg="lightgray",
                                      relief="solid", borderwidth=1, wraplength=100)
                header_label.grid(row=current_row, column=col, sticky="ew", padx=1, pady=1)
            current_row += 1
            
            # Data rows for 構造物名称
            for entry in kozo_entries:
                item_key = f"kozo_{entry['value']}_{entry['rosen']}"
                self.entry_widgets[item_key] = {
                    'type': '構造物名称',
                    'rosen': entry['rosen'],
                    'main_value': entry['value'],
                    'widgets': {}
                }
                
                # 路線名 (display only)
                rosen_label = tk.Label(scrollable_frame, text=entry['rosen'], 
                                     font=("Arial", 9), bg="white", wraplength=100,
                                     relief="solid", borderwidth=1)
                rosen_label.grid(row=current_row, column=0, sticky="ew", padx=1, pady=1)
                
                # 構造物名称 (display only)
                kozo_label = tk.Label(scrollable_frame, text=entry['value'], 
                                    font=("Arial", 9), bg="white", wraplength=100,
                                    relief="solid", borderwidth=1)
                kozo_label.grid(row=current_row, column=1, sticky="ew", padx=1, pady=1)
                
                # Input fields
                input_fields = ['構造物番号', '長さ(m)', '構造形式', '構造形式_重み', 
                               '角度', '角度_重み', '供用年数', '供用年数_重み']
                
                for col, field in enumerate(input_fields, 2):
                    entry_widget = tk.Entry(scrollable_frame, font=("Arial", 9),
                                          relief="solid", borderwidth=1, justify="center")
                    entry_widget.grid(row=current_row, column=col, sticky="ew", padx=1, pady=1)
                    self.entry_widgets[item_key]['widgets'][field] = entry_widget
                
                current_row += 1
        
        # Section 2: 駅間 entries
        if ekikan_entries:
            # Empty row separator
            separator = tk.Label(scrollable_frame, text="", height=1, bg="white")
            separator.grid(row=current_row, column=0, columnspan=10)
            current_row += 1
            
            # Section header
            section_label = tk.Label(scrollable_frame, text=f"駅間 Section ({len(ekikan_entries)} entries)", 
                                   font=("Arial", 12, "bold"), fg="white", bg="darkgreen",
                                   relief="solid", borderwidth=1, height=2)
            section_label.grid(row=current_row, column=0, columnspan=10, sticky="ew", padx=1, pady=2)
            current_row += 1
            
            # Column headers
            headers = ['路線名', '駅間', '構造物番号', '長さ(m)', '構造形式', 
                      '構造形式_重み', '角度', '角度_重み', '供用年数', '供用年数_重み']
            
            for col, header in enumerate(headers):
                header_label = tk.Label(scrollable_frame, text=header, 
                                      font=("Arial", 10, "bold"), bg="lightgray",
                                      relief="solid", borderwidth=1, wraplength=100)
                header_label.grid(row=current_row, column=col, sticky="ew", padx=1, pady=1)
            current_row += 1
            
            # Data rows for 駅間
            for entry in ekikan_entries:
                item_key = f"ekikan_{entry['value']}_{entry['rosen']}"
                self.entry_widgets[item_key] = {
                    'type': '駅間',
                    'rosen': entry['rosen'],
                    'main_value': entry['value'],
                    'widgets': {}
                }
                
                # 路線名 (display only)
                rosen_label = tk.Label(scrollable_frame, text=entry['rosen'], 
                                     font=("Arial", 9), bg="white", wraplength=100,
                                     relief="solid", borderwidth=1)
                rosen_label.grid(row=current_row, column=0, sticky="ew", padx=1, pady=1)
                
                # 駅間 (display only)
                ekikan_label = tk.Label(scrollable_frame, text=entry['value'], 
                                      font=("Arial", 9), bg="white", wraplength=100,
                                      relief="solid", borderwidth=1)
                ekikan_label.grid(row=current_row, column=1, sticky="ew", padx=1, pady=1)
                
                # Input fields
                input_fields = ['構造物番号', '長さ(m)', '構造形式', '構造形式_重み', 
                               '角度', '角度_重み', '供用年数', '供用年数_重み']
                
                for col, field in enumerate(input_fields, 2):
                    entry_widget = tk.Entry(scrollable_frame, font=("Arial", 9),
                                          relief="solid", borderwidth=1, justify="center")
                    entry_widget.grid(row=current_row, column=col, sticky="ew", padx=1, pady=1)
                    self.entry_widgets[item_key]['widgets'][field] = entry_widget
                
                current_row += 1
        
        # FIXED: Bottom button section with proper positioning
        button_section = tk.Frame(main_container, bg="#f0f0f0", relief="raised", borderwidth=2)
        button_section.pack(fill="x", pady=(10, 0))
        
        # Create inner frame for button positioning
        button_inner_frame = tk.Frame(button_section, bg="#f0f0f0")
        button_inner_frame.pack(fill="x", padx=15, pady=12)
        
        # Left side buttons
        left_button_frame = tk.Frame(button_inner_frame, bg="#f0f0f0")
        left_button_frame.pack(side="left")
        
        # Auto-fill button
        auto_btn = tk.Button(left_button_frame, text="Auto-fill Remaining", 
                           command=lambda: self.show_smart_default_dialog(missing_entries), 
                           bg="#4CAF50", fg="white", width=18, height=1, 
                           font=("Arial", 11, "bold"), cursor="hand2",
                           relief="raised", borderwidth=2)
        auto_btn.pack(side="left", padx=(0, 10))
        
        # Add hover effects for auto button
        auto_btn.bind("<Enter>", lambda e: e.widget.config(bg="#45a049"))
        auto_btn.bind("<Leave>", lambda e: e.widget.config(bg="#4CAF50"))
        
        # Right side buttons
        right_button_frame = tk.Frame(button_inner_frame, bg="#f0f0f0")
        right_button_frame.pack(side="right")
        
        # Cancel button
        cancel_btn = tk.Button(right_button_frame, text="Cancel", 
                             command=self.close_all_windows, bg="#f44336", fg="white", 
                             width=10, height=1, font=("Arial", 11, "bold"), 
                             cursor="hand2", relief="raised", borderwidth=2)
        cancel_btn.pack(side="right", padx=(10, 0))
        
        # Save button  
        save_btn = tk.Button(right_button_frame, text="Save", 
                           command=self.save_and_process, 
                           bg="#2196F3", fg="white", width=12, height=1, 
                           font=("Arial", 11, "bold"), cursor="hand2",
                           relief="raised", borderwidth=2)
        save_btn.pack(side="right", padx=(10, 5))
        
        # Add hover effects for save and cancel buttons
        save_btn.bind("<Enter>", lambda e: e.widget.config(bg="#1976D2"))
        save_btn.bind("<Leave>", lambda e: e.widget.config(bg="#2196F3"))
        
        cancel_btn.bind("<Enter>", lambda e: e.widget.config(bg="#d32f2f"))
        cancel_btn.bind("<Leave>", lambda e: e.widget.config(bg="#f44336"))
        
        # Enable mouse wheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        def _bind_mousewheel(event):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        def _unbind_mousewheel(event):
            canvas.unbind_all("<MouseWheel>")
        
        canvas.bind('<Enter>', _bind_mousewheel)
        canvas.bind('<Leave>', _unbind_mousewheel)

    def show_smart_default_dialog(self, missing_entries):
        """SMART DEFAULT SYSTEM - Only shows fields that are empty"""
        # First, collect what's already filled
        filled_data = {}
        empty_fields = set()
        
        for item_key, entry_data in self.entry_widgets.items():
            for field_name, widget in entry_data['widgets'].items():
                value = widget.get().strip()
                if value:
                    if field_name not in filled_data:
                        filled_data[field_name] = set()
                    filled_data[field_name].add(value)
                else:
                    empty_fields.add(field_name)
        
        # Only show dialog for empty fields
        if not empty_fields:
            self.save_and_process()
            return
        
        # Create smart default dialog with improved sizing
        default_window = tk.Toplevel(self.form_window)
        default_window.title("Fill Remaining Empty Fields")
        default_window.geometry("650x750")  # Increased size
        default_window.grab_set()
        default_window.resizable(True, True)
        default_window.transient(self.form_window)
        
        # Center window
        default_window.update_idletasks()
        x = (default_window.winfo_screenwidth() // 2) - (650 // 2)
        y = (default_window.winfo_screenheight() // 2) - (750 // 2)
        default_window.geometry(f"650x750+{x}+{y}")
        
        main_frame = tk.Frame(default_window, padx=25, pady=20)
        main_frame.pack(fill="both", expand=True)
        
        # Title
        title_label = tk.Label(main_frame, text="Fill Remaining Empty Fields", 
                              font=("Arial", 16, "bold"), fg="navy")
        title_label.pack(pady=(0, 15))
        
        # Info about filled fields
        if filled_data:
            filled_info = "Already filled: " + ", ".join(filled_data.keys())
            info_label = tk.Label(main_frame, text=filled_info, font=("Arial", 10), 
                                fg="green", wraplength=600)
            info_label.pack(pady=(0, 10))
        
        # Empty fields info
        empty_info = f"Need defaults for: {len(empty_fields)} empty fields"
        empty_label = tk.Label(main_frame, text=empty_info, font=("Arial", 12), fg="blue")
        empty_label.pack(pady=(0, 20))
        
        # Create scrollable frame for fields
        fields_container = tk.Frame(main_frame)
        fields_container.pack(fill="both", expand=True, pady=(0, 20))
        
        canvas = tk.Canvas(fields_container, height=450, bg="white")
        scrollbar = ttk.Scrollbar(fields_container, orient="vertical", command=canvas.yview)
        scrollable_fields_frame = tk.Frame(canvas, bg="white")
        
        scrollable_fields_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_fields_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Create entry fields for empty fields only
        self.default_entries = {}
        
        default_values = {
            '構造物番号': '1', '長さ(m)': '100', '構造形式': 'RC', '構造形式_重み': '1',
            '角度': '0', '角度_重み': '1', '供用年数': '50', '供用年数_重み': '10'
        }
        
        row = 0
        for field in sorted(empty_fields):
            if field in default_values:
                # Create a frame for each field row
                field_frame = tk.Frame(scrollable_fields_frame, bg="white", pady=8)
                field_frame.pack(fill="x", padx=20, pady=5)
                
                # Label
                label = tk.Label(field_frame, text=f"{field}:", font=("Arial", 12), 
                               bg="white", width=20, anchor="w")
                label.pack(side="left", padx=(0, 15))
                
                # Entry
                entry = tk.Entry(field_frame, width=25, font=("Arial", 12),
                               relief="solid", borderwidth=1)
                entry.insert(0, default_values[field])
                entry.pack(side="left")
                
                self.default_entries[field] = entry
                row += 1
        
        # Progress frame for loading
        self.default_progress_frame = tk.Frame(main_frame)
        
        self.default_status_label = tk.Label(self.default_progress_frame, text="", 
                                           font=("Arial", 12, "bold"), fg="blue")
        self.default_status_label.pack(pady=(15, 8))
        
        self.default_progress_bar = ttk.Progressbar(self.default_progress_frame, 
                                                  length=400, mode='indeterminate')
        self.default_progress_bar.pack()
        
        # FIXED: Bottom button section with proper positioning
        button_frame = tk.Frame(main_frame, bg="#f0f0f0", relief="raised", borderwidth=2)
        button_frame.pack(fill="x", pady=(15, 0))
        
        button_inner = tk.Frame(button_frame, bg="#f0f0f0")
        button_inner.pack(fill="x", padx=15, pady=10)
        
        def apply_smart_defaults():
            # Hide buttons and show progress
            button_frame.pack_forget()
            self.default_progress_frame.pack(fill="x", pady=(15, 0))
            
            # Start processing in thread
            threading.Thread(target=self.apply_defaults_with_progress, 
                           args=(empty_fields, default_window), daemon=True).start()
        
        # Cancel button (left side)
        cancel_btn = tk.Button(button_inner, text="Cancel", 
                             command=default_window.destroy, bg="#f44336", fg="white", 
                             width=10, height=1, font=("Arial", 12, "bold"), 
                             cursor="hand2", relief="raised", borderwidth=2)
        cancel_btn.pack(side="left")
        
        # Apply button (right side)
        apply_btn = tk.Button(button_inner, text="Apply & Save", 
                            command=apply_smart_defaults, bg="#4CAF50", fg="white", 
                            width=15, height=1, font=("Arial", 12, "bold"), 
                            cursor="hand2", relief="raised", borderwidth=2)
        apply_btn.pack(side="right")
        
        # Add hover effects
        apply_btn.bind("<Enter>", lambda e: e.widget.config(bg="#45a049"))
        apply_btn.bind("<Leave>", lambda e: e.widget.config(bg="#4CAF50"))
        
        cancel_btn.bind("<Enter>", lambda e: e.widget.config(bg="#d32f2f"))
        cancel_btn.bind("<Leave>", lambda e: e.widget.config(bg="#f44336"))

    def apply_defaults_with_progress(self, empty_fields, default_window):
        try:
            # Update progress
            self.root.after(0, self.update_default_progress, "デフォルト値適用中...")
            time.sleep(0.5)
            
            # Get default values for empty fields
            defaults = {}
            for field in empty_fields:
                if field in self.default_entries:
                    defaults[field] = self.default_entries[field].get().strip()
            
            # Apply defaults to empty fields only
            self.root.after(0, self.update_default_progress, "空フィールドに値を設定中...")
            time.sleep(0.3)
            
            for item_key, entry_data in self.entry_widgets.items():
                for field_name, widget in entry_data['widgets'].items():
                    if not widget.get().strip() and field_name in defaults:
                        widget.insert(0, defaults[field_name])
            
            # Save all data
            self.root.after(0, self.update_default_progress, "データ保存中...")
            time.sleep(0.3)
            
            saved_count = 0
            for item_key, entry_data in self.entry_widgets.items():
                widgets = entry_data['widgets']
                entry_type = entry_data['type']
                rosen = entry_data['rosen']
                main_value = entry_data['main_value']
                
                # Check if any field has data
                has_data = any(widget.get().strip() for widget in widgets.values())
                
                if has_data:
                    new_row = {
                        '路線名': rosen,
                        '構造物名称': main_value if entry_type == '構造物名称' else '',
                        '駅間': main_value if entry_type == '駅間' else ''
                    }
                    
                    # Add field values
                    for field_name, widget in widgets.items():
                        new_row[field_name] = widget.get().strip()
                    
                    # Add to dataframe
                    self.structure_data_df = pd.concat([
                        self.structure_data_df, 
                        pd.DataFrame([new_row])
                    ], ignore_index=True)
                    
                    saved_count += 1
            
            # Save to Excel
            self.save_structure_data()
            
            # Complete
            self.root.after(0, self.update_default_progress, "完了")
            time.sleep(1)
            
            # Close windows
            self.root.after(0, lambda: self.close_all_windows_success(saved_count))
            
        except Exception as e:
            self.root.after(0, self.update_default_progress, f"エラー: {str(e)}")
            time.sleep(2)
            self.root.after(0, default_window.destroy)

    def update_default_progress(self, status):
        self.default_status_label.config(text=status)
        self.default_progress_bar.start()
        self.root.update()

    def save_and_process(self):
        # Show progress on main form
        self.create_save_progress()
        
        # Start saving in thread
        threading.Thread(target=self.save_with_progress, daemon=True).start()

    def create_save_progress(self):
        # Create progress overlay on form
        self.save_frame = tk.Frame(self.form_window, bg="white", relief="solid", borderwidth=3)
        self.save_frame.place(relx=0.5, rely=0.5, anchor="center")
        
        progress_frame = tk.Frame(self.save_frame, padx=40, pady=25)
        progress_frame.pack()
        
        self.save_status_label = tk.Label(progress_frame, text="保存中...", 
                                        font=("Arial", 14, "bold"), fg="blue")
        self.save_status_label.pack(pady=(0, 15))
        
        self.save_progress_bar = ttk.Progressbar(progress_frame, length=300, mode='indeterminate')
        self.save_progress_bar.pack()
        self.save_progress_bar.start()

    def save_with_progress(self):
        try:
            time.sleep(0.5)
            
            saved_count = 0
            for item_key, entry_data in self.entry_widgets.items():
                widgets = entry_data['widgets']
                entry_type = entry_data['type']
                rosen = entry_data['rosen']
                main_value = entry_data['main_value']
                
                # Check if any field has data
                has_data = any(widget.get().strip() for widget in widgets.values())
                
                if has_data:
                    new_row = {
                        '路線名': rosen,
                        '構造物名称': main_value if entry_type == '構造物名称' else '',
                        '駅間': main_value if entry_type == '駅間' else ''
                    }
                    
                    # Add field values
                    for field_name, widget in widgets.items():
                        new_row[field_name] = widget.get().strip()
                    
                    # Add to dataframe
                    self.structure_data_df = pd.concat([
                        self.structure_data_df, 
                        pd.DataFrame([new_row])
                    ], ignore_index=True)
                    
                    saved_count += 1
            
            # Save to Excel
            self.save_structure_data()
            
            # Complete
            self.root.after(0, lambda: self.close_all_windows_success(saved_count))
            
        except Exception as e:
            self.root.after(0, self.save_frame.destroy)
            messagebox.showerror("Error", f"Failed to save: {str(e)}")

    def close_all_windows_success(self, saved_count):
        # Stop all progress bars
        try:
            self.save_progress_bar.stop()
            self.default_progress_bar.stop()
        except:
            pass
        
        # Show success message
        messagebox.showinfo("Success", f"Successfully saved {saved_count} entries!")
        
        # Close all windows
        try:
            self.form_window.destroy()
        except:
            pass
        
        self.root.quit()
        self.root.destroy()

    def close_all_windows(self):
        try:
            self.form_window.destroy()
        except:
            pass
        
        self.root.quit()
        self.root.destroy()

    def save_structure_data(self):
        """Save structure data to Excel"""
        try:
            with pd.ExcelWriter(self.workbook_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                # Write structure data sheet
                self.structure_data_df.to_excel(writer, sheet_name='構造物番号', index=False)
                
                # Preserve other sheets
                try:
                    original_wb = load_workbook(self.workbook_path)
                    for sheet_name in original_wb.sheetnames:
                        if sheet_name != '構造物番号':
                            try:
                                df_temp = pd.read_excel(self.workbook_path, sheet_name=sheet_name)
                                df_temp.to_excel(writer, sheet_name=sheet_name, index=False)
                            except Exception as e:
                                continue
                except Exception as e:
                    pass
                    
        except Exception as e:
            raise Exception(f"Error saving structure data: {str(e)}")

    # Add the missing methods from the original code
    def load_structure_data(self):
        try:
            self.structure_data_df = pd.read_excel(self.workbook_path, sheet_name='構造物番号')
            required_columns = [
                '路線名', '構造物名称', '駅間', '構造物番号', '長さ(m)', 
                '構造形式', '構造形式_重み', '角度', '角度_重み', 
                '供用年数', '供用年数_重み'
            ]
            for col in required_columns:
                if col not in self.structure_data_df.columns:
                    self.structure_data_df[col] = ''
        except:
            self.structure_data_df = pd.DataFrame(columns=[
                '路線名', '構造物名称', '駅間', '構造物番号', '長さ(m)', 
                '構造形式', '構造形式_重み', '角度', '角度_重み', 
                '供用年数', '供用年数_重み'
            ])

    def get_missing_structure_entries(self):
        missing_entries = []
        
        if len(self.structure_data_df) == 0:
            self.structure_data_df = pd.DataFrame(columns=[
                '路線名', '構造物名称', '駅間', '構造物番号', '長さ(m)', 
                '構造形式', '構造形式_重み', '角度', '角度_重み', 
                '供用年数', '供用年数_重み'
            ])
        
        unique_kozo = set()
        unique_ekikan = set()
        
        for _, row in self.grouped_df.iterrows():
            rosen = str(row.get('路線名', '')).strip() if pd.notna(row.get('路線名', '')) else ''
            group_method = str(row.get('グループ化方法', '')).strip() if pd.notna(row.get('グループ化方法', '')) else ''
            
            if group_method == '構造物名称':
                kozo = str(row.get('構造物名称', '')).strip() if pd.notna(row.get('構造物名称', '')) else ''
                if kozo and kozo not in ['', 'nan', 'NaN']:
                    unique_kozo.add((rosen, kozo))
            
            elif group_method == '駅間':
                ekikan_start = str(row.get('駅（始）', '')).strip() if pd.notna(row.get('駅（始）', '')) else ''
                ekikan_end = str(row.get('駅（至）', '')).strip() if pd.notna(row.get('駅（至）', '')) else ''
                
                if ekikan_start and ekikan_end and ekikan_start not in ['', 'nan', 'NaN'] and ekikan_end not in ['', 'nan', 'NaN']:
                    ekikan = f"{ekikan_start}→{ekikan_end}"
                    unique_ekikan.add((rosen, ekikan))
        
        # Check missing entries
        for rosen, kozo in unique_kozo:
            exists = not self.structure_data_df[
                (self.structure_data_df['構造物名称'].astype(str).str.strip() == kozo) & 
                (self.structure_data_df['路線名'].astype(str).str.strip() == rosen)
            ].empty if len(self.structure_data_df) > 0 else False
            
            if not exists:
                missing_entries.append({
                    'type': '構造物名称',
                    'rosen': rosen,
                    'value': kozo,
                    'display_value': kozo
                })
        
        for rosen, ekikan in unique_ekikan:
            exists = not self.structure_data_df[
                (self.structure_data_df['駅間'].astype(str).str.strip() == ekikan) & 
                (self.structure_data_df['路線名'].astype(str).str.strip() == rosen)
            ].empty if len(self.structure_data_df) > 0 else False
            
            if not exists:
                    missing_entries.append({
                    'type': '駅間',
                    'rosen': rosen,
                    'value': ekikan,
                    'display_value': ekikan
                })
        
        missing_entries.sort(key=lambda x: (x['type'] == '駅間', x['rosen'], x['value']))
        return missing_entries

    def load_file(self):
        try:
            # Show loading
            self.root.after(0, self.update_progress, "ファイル読み込み中...")
            time.sleep(0.5)
            
            # Load data
            self.grouped_df = pd.read_excel(self.workbook_path, sheet_name='グループ化点検履歴')
            
            # Load structure data
            self.root.after(0, self.update_progress, "構造物データ読み込み中...")
            time.sleep(0.3)
            self.load_structure_data()
            
            # Complete loading and show form
            self.root.after(0, self.show_structure_form)
            
        except Exception as e:
            self.root.after(0, self.update_progress, f"エラー: {str(e)}")
            time.sleep(2)
            self.root.after(0, self.reset_to_main)

    def update_progress(self, status):
        self.status_label.config(text=status)
        self.progress_bar.start()
        self.root.update()

    def reset_to_main(self):
        self.progress_frame.pack_forget()
        self.select_btn.pack()
        self.progress_bar.stop()

    def show_structure_form(self):
        self.progress_bar.stop()
        missing_entries = self.get_missing_structure_entries()
        
        if not missing_entries:
            messagebox.showinfo("Info", "No missing entries found!")
            self.root.quit()
            return
        
        # Hide main window and show form
        self.root.withdraw()
        self.create_structure_form(missing_entries)

    def select_and_load(self):
        self.workbook_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        
        if self.workbook_path:
            self.select_btn.pack_forget()
            self.progress_frame.pack(fill="both", expand=True)
            
            # Start loading in separate thread
            threading.Thread(target=self.load_file, daemon=True).start()

    def run(self):
        """Run the application"""
        self.root.mainloop()


# Main execution
if __name__ == "__main__":
    app = StructureDataEntryApp()
    app.run()
                    