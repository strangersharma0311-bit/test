import pandas as pd
import openpyxl
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import re
import threading
import time

class ExcelProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Data Processor - Premium Edition")
        self.root.geometry("700x550")
        self.root.configure(bg='#2c3e50')
        self.root.resizable(True, True)
        
        # Workflow state management
        self.workflow_data = {}
        
        # Configure styles for premium look
        self.setup_styles()
        self.create_main_ui()

    def setup_styles(self):
        style = ttk.Style()
        
        # Configure modern theme
        style.theme_use('clam')
        
        # Title style
        style.configure('Title.TLabel', 
                    font=('Segoe UI', 18, 'bold'),
                    foreground='#ffffff',
                    background='#2c3e50')
        
        # Header style
        style.configure('Header.TLabel', 
                    font=('Segoe UI', 11, 'bold'),
                    foreground='#2c3e50',
                    background='#f8f9fa')
        
        # Description style
        style.configure('Desc.TLabel', 
                    font=('Segoe UI', 9),
                    foreground='#6c757d',
                    background='#f8f9fa')
        
        # Primary button style
        style.configure('Primary.TButton',
                    font=('Segoe UI', 10, 'bold'),
                    foreground='white',
                    background='#007bff',
                    borderwidth=0,
                    focuscolor='none',
                    padding=(20, 10))
        
        style.map('Primary.TButton',
                background=[('active', '#0056b3'),
                        ('pressed', '#004085')])
        
        # Success button style
        style.configure('Success.TButton',
                    font=('Segoe UI', 10, 'bold'),
                    foreground='white',
                    background='#28a745',
                    borderwidth=0,
                    focuscolor='none',
                    padding=(15, 8))
        
        style.map('Success.TButton',
                background=[('active', '#218838'),
                        ('pressed', '#1e7e34')])
        
        # Danger button style
        style.configure('Danger.TButton',
                    font=('Segoe UI', 10),
                    foreground='white',
                    background='#dc3545',
                    borderwidth=0,
                    focuscolor='none',
                    padding=(15, 8))
        
        style.map('Danger.TButton',
                background=[('active', '#c82333'),
                        ('pressed', '#bd2130')])
        
        # Secondary button style
        style.configure('Secondary.TButton',
                    font=('Segoe UI', 9),
                    foreground='#495057',
                    background='#e9ecef',
                    borderwidth=0,
                    focuscolor='none',
                    padding=(12, 6))
        
        style.map('Secondary.TButton',
                background=[('active', '#dee2e6'),
                        ('pressed', '#ced4da')])
        
        # Entry style
        style.configure('Modern.TEntry',
                    font=('Segoe UI', 10),
                    fieldbackground='white',
                    borderwidth=1,
                    relief='solid',
                    bordercolor='#ced4da',
                    focuscolor='#007bff',
                    padding=8)
        
        # LabelFrame style
        style.configure('Modern.TLabelframe',
                    background='#f8f9fa',
                    borderwidth=1,
                    relief='solid',
                    bordercolor='#dee2e6')
        
        style.configure('Modern.TLabelframe.Label',
                    font=('Segoe UI', 11, 'bold'),
                    foreground='#495057',
                    background='#f8f9fa')
        
        # Radiobutton style
        style.configure('Modern.TRadiobutton',
                    font=('Segoe UI', 10),
                    foreground='#495057',
                    background='#f8f9fa',
                    focuscolor='none')
        
        # Progressbar style
        style.configure('Modern.TProgressbar',
                    background='#007bff',
                    troughcolor='#e9ecef',
                    borderwidth=0,
                    lightcolor='#007bff',
                    darkcolor='#007bff')

    def create_main_ui(self):
        # Clear the window
        for widget in self.root.winfo_children():
            widget.destroy()
            
        # Main container
        main_container = tk.Frame(self.root, bg='#2c3e50')
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Title section
        title_frame = tk.Frame(main_container, bg='#2c3e50')
        title_frame.pack(fill=tk.X, pady=(0, 30))
        
        title_label = ttk.Label(title_frame, text="Excel Data Processor", style='Title.TLabel')
        title_label.pack()
        
        subtitle_label = ttk.Label(title_frame, text="Professional Data Processing Tool", 
                                  font=('Segoe UI', 10), foreground='#95a5a6', background='#2c3e50')
        subtitle_label.pack(pady=(5, 0))

        # Content frame - Changed background color
        content_frame = tk.Frame(main_container, bg='#f8f9fa', relief='solid', bd=1)
        content_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Padding frame - Changed background color
        padded_frame = tk.Frame(content_frame, bg='#f8f9fa')
        padded_frame.pack(fill=tk.BOTH, expand=True, padx=25, pady=25)

        # Workbook selection section
        workbook_frame = ttk.LabelFrame(padded_frame, text="Select Main Excel Workbook", 
                                       style='Modern.TLabelframe', padding=15)
        workbook_frame.pack(fill=tk.X, pady=(0, 20))
        
        workbook_input_frame = tk.Frame(workbook_frame, bg='#f8f9fa')
        workbook_input_frame.pack(fill=tk.X)
        
        self.workbook_entry = ttk.Entry(workbook_input_frame, width=70, style='Modern.TEntry')
        self.workbook_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        browse_workbook_button = ttk.Button(workbook_input_frame, text="Browse", 
                                           command=self.browse_workbook, style='Primary.TButton')
        browse_workbook_button.pack(side=tk.RIGHT)

        # New Data section
        new_data_frame = ttk.LabelFrame(padded_frame, text="Add New Data", 
                                       style='Modern.TLabelframe', padding=15)
        new_data_frame.pack(fill=tk.X, pady=(0, 20))

        # Radio buttons frame
        radio_frame = tk.Frame(new_data_frame, bg='#f8f9fa')
        radio_frame.pack(fill=tk.X, pady=(0, 15))
        
        question_label = ttk.Label(radio_frame, text="Do you want to add new data before processing?", 
                                  style='Desc.TLabel')
        question_label.pack(anchor=tk.W, pady=(0, 10))

        radio_options_frame = tk.Frame(radio_frame, bg='#f8f9fa')
        radio_options_frame.pack(anchor=tk.W)
        
        self.new_data_var = tk.BooleanVar()
        self.new_data_yes = ttk.Radiobutton(radio_options_frame, text="Yes", 
                                           variable=self.new_data_var, value=True, 
                                           command=self.toggle_new_data_options,
                                           style='Modern.TRadiobutton')
        self.new_data_yes.pack(side=tk.LEFT, padx=(0, 30))
        
        self.new_data_no = ttk.Radiobutton(radio_options_frame, text="No", 
                                          variable=self.new_data_var, value=False, 
                                          command=self.toggle_new_data_options,
                                          style='Modern.TRadiobutton')
        self.new_data_no.pack(side=tk.LEFT)

        # New data file selection frame (initially hidden)
        self.new_data_file_frame = tk.Frame(new_data_frame, bg='#f8f9fa')
        self.new_data_file_frame.pack(fill=tk.X, pady=(0, 15))
        
        new_file_label = ttk.Label(self.new_data_file_frame, text="Select New Data File:", 
                                  style='Desc.TLabel')
        new_file_label.pack(anchor=tk.W, pady=(0, 5))
        
        file_input_frame = tk.Frame(self.new_data_file_frame, bg='#f8f9fa')
        file_input_frame.pack(fill=tk.X)
        
        self.new_data_file_entry = ttk.Entry(file_input_frame, width=60, style='Modern.TEntry')
        self.new_data_file_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        browse_new_data_button = ttk.Button(file_input_frame, text="Browse", 
                                           command=self.browse_new_data_file, style='Primary.TButton')
        browse_new_data_button.pack(side=tk.RIGHT)

        # Year input frame
        self.year_frame = tk.Frame(new_data_frame, bg='#f8f9fa')
        self.year_frame.pack(fill=tk.X)
        
        self.year_label = ttk.Label(self.year_frame, text="Enter Year for New Data:", 
                                   style='Desc.TLabel')
        self.year_label.pack(anchor=tk.W, pady=(0, 5))

        self.year_entry = ttk.Entry(self.year_frame, width=20, style='Modern.TEntry')
        self.year_entry.pack(anchor=tk.W)
        
        # Initially hide new data options
        self.toggle_new_data_options()

        # Status frame (initially hidden)
        self.status_frame = tk.Frame(padded_frame, bg='#f8f9fa')
        
        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(self.status_frame, variable=self.progress_var, 
                                   maximum=100, length=400, mode='determinate')
        self.progress_bar.pack(pady=(0, 10))
        
        # Status label
        self.status_label = tk.Label(self.status_frame, text="", 
                                    font=('Segoe UI', 10), fg='#2c3e50', bg='#f8f9fa')
        self.status_label.pack()

        # Action buttons frame
        action_frame = tk.Frame(padded_frame, bg='#f8f9fa')
        action_frame.pack(fill=tk.X, pady=20)
        
        self.submit_button = ttk.Button(action_frame, text="Start Processing", 
                                       command=self.submit, style='Success.TButton')
        self.submit_button.pack(side=tk.RIGHT, padx=(10, 0))
        
        reset_button = ttk.Button(action_frame, text="Reset", 
                                 command=self.reset_form, style='Secondary.TButton')
        reset_button.pack(side=tk.RIGHT)

    def submit(self):
        workbook = self.workbook_entry.get().strip()
        if not workbook:
            messagebox.showerror("Error", "Please select a main workbook!")
            return

        add_new_data = self.new_data_var.get()
        
        if add_new_data:
            new_data_file = self.new_data_file_entry.get().strip()
            new_data_sheet = self.year_entry.get().strip()
            
            if not new_data_file:
                messagebox.showerror("Error", "Please select a new data file!")
                return
                
            if not new_data_sheet:
                messagebox.showerror("Error", "Please enter the year for the new data!")
                return
            
            # Store workflow data
            self.workflow_data = {
                'workbook': workbook,
                'new_data_file': new_data_file,
                'new_data_sheet': new_data_sheet,
                'add_new_data': add_new_data
            }
            
            # Start processing in background thread
            threading.Thread(target=self.process_workflow, daemon=True).start()
        else:
            # Process without new data
            threading.Thread(target=self.process_existing_data, args=(workbook,), daemon=True).start()

    def toggle_new_data_options(self):
        if self.new_data_var.get():
            self.new_data_file_frame.pack(fill=tk.X, pady=(0, 15))
            self.year_frame.pack(fill=tk.X)
        else:
            self.new_data_file_frame.pack_forget()
            self.year_frame.pack_forget()

    def browse_workbook(self):
        workbook = filedialog.askopenfilename(
            title="Select Main Excel Workbook",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if workbook:
            self.workbook_entry.delete(0, tk.END)
            self.workbook_entry.insert(0, workbook)

    def browse_new_data_file(self):
        new_data_file = filedialog.askopenfilename(
            title="Select New Data File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if new_data_file:
            self.new_data_file_entry.delete(0, tk.END)
            self.new_data_file_entry.insert(0, new_data_file)

    def reset_form(self):
        self.workbook_entry.delete(0, tk.END)
        self.new_data_file_entry.delete(0, tk.END)
        self.year_entry.delete(0, tk.END)
        self.new_data_var.set(False)
        self.toggle_new_data_options()
        if hasattr(self, 'status_frame'):
            self.status_frame.pack_forget()

    def show_progress(self):
        """Show progress frame"""
        self.status_frame.pack(fill=tk.X, pady=(0, 20))
        self.submit_button.config(state='disabled')

    def hide_progress(self):
        """Hide progress frame"""
        self.status_frame.pack_forget()
        self.submit_button.config(state='normal')

    def update_progress(self, value, message):
        """Update progress bar and status message - optimized"""
        self.progress_var.set(value)
        self.status_label.config(text=message)
        self.root.update()  # Use update() instead of update_idletasks() for faster response

    def process_workflow(self):
        """Process the complete workflow with faster progress updates"""
        try:
            self.show_progress()
            
            # Step 1: Validate new data file
            self.update_progress(10, "Validating new data file...")
            
            workbook = self.workflow_data['workbook']
            new_data_file = self.workflow_data['new_data_file']
            new_data_sheet = self.workflow_data['new_data_sheet']
            
            # Check if sheet exists (faster validation)
            try:
                xls = pd.ExcelFile(new_data_file)
                if new_data_sheet not in xls.sheet_names:
                    self.root.after(0, lambda: messagebox.showerror("Error", f"Sheet '{new_data_sheet}' not found in the selected file!"))
                    self.hide_progress()
                    return
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("Error", f"Error reading file: {str(e)}"))
                self.hide_progress()
                return
            
            # Step 2: Add new data to workbook
            self.update_progress(25, "Adding new data to workbook...")
            self.add_new_data_to_workbook(workbook, new_data_file, new_data_sheet)
            
            # Step 3: Get columns for user selection (faster loading)
            self.update_progress(40, "Loading column data...")
            new_data_df = pd.read_excel(workbook, sheet_name=new_data_sheet, nrows=1)  # Only read header for faster loading
            new_data_columns = list(new_data_df.columns)
            
            # Store original columns for back navigation
            self.workflow_data['original_columns'] = new_data_columns
            
            # Step 4: Continue with UI dialogs
            self.root.after(0, lambda: self.continue_workflow_ui(new_data_columns))
            
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("Error", f"An error occurred: {str(e)}"))
            self.hide_progress()

    def continue_workflow_ui(self, new_data_columns):
        """Continue workflow with UI dialogs"""
        try:
            new_data_sheet = self.workflow_data['new_data_sheet']
            workbook = self.workflow_data['workbook']
            
            self.update_progress(50, "Waiting for column selection...")
            
            # Step 1: Select columns for extraction
            selected_columns = self.show_fast_column_selection_dialog(
                new_data_columns, 
                f"Select Columns to Extract from {new_data_sheet}",
                "Please select columns for data extraction:",
                "#3498db",
                step=1
            )
            
            if not selected_columns:
                messagebox.showwarning("Warning", "Column selection is required for extraction!")
                self.hide_progress()
                return
            
            self.workflow_data['selected_columns'] = selected_columns
            
            self.update_progress(65, "Updating extraction sheet...")
            self.update_chuushutsu_sheet(workbook, selected_columns, new_data_sheet)
            
            # Step 2: Select columns for scoring with year prefix
            self.update_progress(75, "Waiting for scoring column selection...")
            
            # Create formatted column names for second popup
            formatted_columns = [f"{new_data_sheet} {col}" for col in selected_columns]
            
            scoring_columns = self.show_fast_column_selection_dialog(
                formatted_columns,
                f"Select Columns for Scoring (点数化)",
                "Please select columns for scoring calculation:",
                "#e67e22",
                step=2
            )
            
            if not scoring_columns:
                messagebox.showwarning("Warning", "Scoring column selection is required!")
                self.hide_progress()
                return
            
            # Remove year prefix for processing
            scoring_columns_clean = [col.replace(f"{new_data_sheet} ", "") for col in scoring_columns]
            self.workflow_data['scoring_columns'] = scoring_columns_clean
            
            self.update_progress(85, "Updating scoring sheet...")
            self.update_tensuka_sheet(workbook, scoring_columns_clean, new_data_sheet)
            
            # Step 3: Expression dialogs
            self.continue_expression_workflow()
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
            self.hide_progress()

    def continue_expression_workflow(self):
        """Continue with expression creation"""
        try:
            new_data_sheet = self.workflow_data['new_data_sheet']
            workbook = self.workflow_data['workbook']
            scoring_columns = self.workflow_data['scoring_columns']
            
            self.update_progress(90, "Creating expressions...")
            
            # Step 3: Create expression for 演算子
            expression = self.show_expression_dialog(scoring_columns, new_data_sheet, step=3)
            if expression:
                self.update_ensanshi_sheet(workbook, expression, new_data_sheet)
            else:
                messagebox.showwarning("Warning", "Expression is required for 演算子!")
                self.hide_progress()
                return
            
            # Step 4: Create expression for 演算子‐2
            expression2 = self.show_expression2_dialog(new_data_sheet, step=4)
            if expression2:
                self.update_ensanshi2_sheet(workbook, expression2, new_data_sheet)
            
            # Final processing
            self.update_progress(95, "Final data processing...")
            
            threading.Thread(target=self.finalize_processing, daemon=True).start()
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
            self.hide_progress()

    def finalize_processing(self):
        """Finalize the processing"""
        try:
            workbook = self.workflow_data['workbook']
            new_data_sheet = self.workflow_data['new_data_sheet']
            
            self.process_data(workbook, new_data_sheet)
            
            self.update_progress(100, "Processing completed successfully!")
            time.sleep(1)
            
            self.root.after(0, lambda: messagebox.showinfo("Success", "Data processed successfully!"))
            self.hide_progress()
            
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("Error", f"An error occurred during final processing: {str(e)}"))
            self.hide_progress()

    def show_fast_column_selection_dialog(self, columns, title, description, accent_color, step=1):
        """Fast loading column selection dialog with proper styling"""
        dialog = tk.Toplevel(self.root)
        dialog.title(title)
        dialog.geometry("600x700")
        dialog.configure(bg='#2c3e50')
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        
        # Center dialog
        dialog.geometry("+%d+%d" % (self.root.winfo_rootx() + 100, self.root.winfo_rooty() + 50))

        selected_columns = []
        checkboxes = []
        result = {'selected': None, 'action': None}
        
        # Header frame
        header_frame = tk.Frame(dialog, bg=accent_color, height=80)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        title_label = tk.Label(header_frame, text=title, 
                              font=('Segoe UI', 14, 'bold'), 
                              fg='white', bg=accent_color)
        title_label.pack(pady=20)
        
        # Content frame
        content_frame = tk.Frame(dialog, bg='#f8f9fa')
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Description
        desc_label = tk.Label(content_frame, text=description, 
                             font=('Segoe UI', 10), fg='#34495e', bg='#f8f9fa')
        desc_label.pack(pady=(0, 15))
        
        # Scrollable frame for checkboxes
        canvas = tk.Canvas(content_frame, bg='#f8f9fa', highlightthickness=0)
        scrollbar = ttk.Scrollbar(content_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#f8f9fa')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Create fast-loading checkboxes - single column for better performance
        for i, column in enumerate(columns):
            # Checkbox frame with better sizing
            cb_frame = tk.Frame(scrollable_frame, bg='white', relief='solid', bd=1)
            cb_frame.pack(fill=tk.X, padx=3, pady=1, ipady=8)  # Increased ipady for better height
            
            var = tk.BooleanVar()
            
            # Clean checkbox with proper padding
            chk = tk.Checkbutton(cb_frame, text=column, variable=var,
                            font=('Segoe UI', 10), fg='#2c3e50', bg='white',
                            activebackground='white', activeforeground='#2c3e50',
                            selectcolor='#ffffff',
                            relief='flat', borderwidth=0,
                            highlightthickness=0,
                            wraplength=500,  # Wrap long text
                            justify='left',
                            anchor='w',
                            onvalue=True, offvalue=False)
            chk.pack(anchor='w', padx=15, pady=3, fill=tk.X)  # Better padding
            
            checkboxes.append((column, var))
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Button frame
        button_frame = tk.Frame(dialog, bg='#2c3e50', height=70)
        button_frame.pack(fill=tk.X, side=tk.BOTTOM)
        button_frame.pack_propagate(False)
        
        button_container = tk.Frame(button_frame, bg='#2c3e50')
        button_container.pack(expand=True)
        
        def on_submit():
            for column, var in checkboxes:
                if var.get():
                    selected_columns.append(column)
            if not selected_columns:
                messagebox.showwarning("Warning", "Please select at least one column!")
                return
            result['selected'] = selected_columns
            result['action'] = 'submit'
            dialog.destroy()

        def on_back():
            result['action'] = 'back'
            dialog.destroy()
            
        def select_all():
            for _, var in checkboxes:
                var.set(True)
                
        def deselect_all():
            for _, var in checkboxes:
                var.set(False)

        # Buttons with proper back functionality
        submit_btn = tk.Button(button_container, text="Continue", command=on_submit,
                              font=('Segoe UI', 10, 'bold'), fg='white', bg='#27ae60',
                              relief='flat', padx=20, pady=8, cursor='hand2')
        submit_btn.pack(side=tk.RIGHT, padx=5)
        
        # Back button - functional for going to previous step
        if step > 1:
            back_btn = tk.Button(button_container, text="Back", command=on_back,
                                font=('Segoe UI', 10), fg='white', bg='#95a5a6',
                                relief='flat', padx=20, pady=8, cursor='hand2')
            back_btn.pack(side=tk.RIGHT, padx=5)
        
        select_all_btn = tk.Button(button_container, text="Select All", command=select_all,
                                  font=('Segoe UI', 10), fg='white', bg=accent_color,
                                  relief='flat', padx=15, pady=8, cursor='hand2')
        select_all_btn.pack(side=tk.LEFT, padx=5)
        
        deselect_all_btn = tk.Button(button_container, text="Clear All", command=deselect_all,
                                    font=('Segoe UI', 10), fg='white', bg='#e74c3c',
                                    relief='flat', padx=15, pady=8, cursor='hand2')
        deselect_all_btn.pack(side=tk.LEFT, padx=5)

        self.root.wait_window(dialog)
        
        # Handle back navigation
        if result['action'] == 'back':
            if step == 2:
                # Go back to step 1 - reload original columns
                new_data_df = pd.read_excel(self.workflow_data['workbook'], 
                                        sheet_name=self.workflow_data['new_data_sheet'], nrows=1)
                original_columns = list(new_data_df.columns)
                return self.show_fast_column_selection_dialog(
                    original_columns, 
                    f"Select Columns to Extract from {self.workflow_data['new_data_sheet']}",
                    "Please select columns for data extraction:",
                    "#3498db",
                    step=1
                )
            elif step == 3:
                # Go back to step 2 - reload formatted columns
                selected_columns = self.workflow_data.get('selected_columns', [])
                new_data_sheet = self.workflow_data['new_data_sheet']
                formatted_columns = [f"{new_data_sheet} {col}" for col in selected_columns]
                return self.show_fast_column_selection_dialog(
                    formatted_columns,
                    f"Select Columns for Scoring (点数化)",
                    "Please select columns for scoring calculation:",
                    "#e67e22",
                    step=2
                )
        
        return result['selected'] if result['selected'] else None

    def show_expression_dialog(self, scoring_columns, year, step=3):
        """Show expression dialog for 演算子"""
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Create Expression for 演算子 ({year})")
        dialog.geometry("700x500")
        dialog.configure(bg='#2c3e50')
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        
        # Center dialog
        dialog.geometry("+%d+%d" % (self.root.winfo_rootx() + 50, self.root.winfo_rooty() + 100))
        
        result = {'expression': None, 'action': None}
        
        # Header
        header_frame = tk.Frame(dialog, bg='#9b59b6', height=80)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        title_label = tk.Label(header_frame, text=f"Create Expression for 演算子 ({year})", 
                              font=('Segoe UI', 14, 'bold'), 
                              fg='white', bg='#9b59b6')
        title_label.pack(pady=20)
        
        # Content frame
        content_frame = tk.Frame(dialog, bg='#f8f9fa')
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Instructions frame
        instruction_frame = tk.Frame(content_frame, bg='#f8f9fa')
        instruction_frame.pack(fill=tk.X, pady=(0, 15))
        
        instruction_text = "Create mathematical expressions using the variables below:\n"
        instruction_text += "Variables represent: column_name_omomi (weighted columns)\n"
        instruction_text += "Supported operations: +, -, *, /, (), ^(power)"
        
        instruction_label = tk.Label(instruction_frame, text=instruction_text,
                                   font=('Segoe UI', 10), fg='#34495e', bg='#f8f9fa',
                                   justify=tk.LEFT)
        instruction_label.pack(anchor='w')
        
        # Variables display frame
        var_frame = tk.LabelFrame(content_frame, text="Available Variables", 
                                 font=('Segoe UI', 11, 'bold'), fg='#2c3e50', bg='#f8f9fa')
        var_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Create variable mapping
        variables = []
        for i, col in enumerate(scoring_columns):
            var_name = chr(65 + i)  # A, B, C, D...
            variables.append(f"{var_name} = {col}_omomi")
        
        var_text = "\n".join(variables)
        var_display = tk.Label(var_frame, text=var_text, font=('Segoe UI', 9), 
                              fg='#7f8c8d', bg='#f8f9fa', justify=tk.LEFT)
        var_display.pack(anchor='w', padx=10, pady=10)
        
        # Expression input frame
        expr_frame = tk.LabelFrame(content_frame, text="Enter Expression", 
                                  font=('Segoe UI', 11, 'bold'), fg='#2c3e50', bg='#f8f9fa')
        expr_frame.pack(fill=tk.X, pady=(0, 15))
        
        expression_entry = tk.Text(expr_frame, height=4, font=('Consolas', 11),
                                  bg='white', fg='#2c3e50', relief='solid', bd=1,
                                  wrap=tk.WORD)
        expression_entry.pack(fill=tk.X, padx=10, pady=10)
        
        # Button frame
        button_frame = tk.Frame(dialog, bg='#2c3e50', height=70)
        button_frame.pack(fill=tk.X, side=tk.BOTTOM)
        button_frame.pack_propagate(False)
        
        button_container = tk.Frame(button_frame, bg='#2c3e50')
        button_container.pack(expand=True)
        
        def validate_expression():
            expr = expression_entry.get(1.0, tk.END).strip()
            if not expr:
                messagebox.showerror("Error", "Please enter an expression!")
                return False
            return True
        
        def on_submit():
            if validate_expression():
                result['expression'] = expression_entry.get(1.0, tk.END).strip()
                result['action'] = 'submit'
                dialog.destroy()

        def on_back():
            result['action'] = 'back'
            dialog.destroy()

        # Main action buttons
        submit_btn = tk.Button(button_container, text="Create Expression", command=on_submit,
                              font=('Segoe UI', 10, 'bold'), fg='white', bg='#27ae60',
                              relief='flat', padx=20, pady=8, cursor='hand2')
        submit_btn.pack(side=tk.RIGHT, padx=5)
        
        back_btn = tk.Button(button_container, text="Back", command=on_back,
                            font=('Segoe UI', 10), fg='white', bg='#95a5a6',
                            relief='flat', padx=20, pady=8, cursor='hand2')
        back_btn.pack(side=tk.RIGHT, padx=5)

        self.root.wait_window(dialog)
        
        # Handle back navigation
        if result['action'] == 'back':
            # Go back to scoring column selection
            selected_columns = self.workflow_data.get('selected_columns', [])
            new_data_sheet = self.workflow_data['new_data_sheet']
            formatted_columns = [f"{new_data_sheet} {col}" for col in selected_columns]
            scoring_result = self.show_fast_column_selection_dialog(
                formatted_columns,
                f"Select Columns for Scoring (点数化)",
                "Please select columns for scoring calculation:",
                "#e67e22",
                step=2
            )
            if scoring_result:
                scoring_columns_clean = [col.replace(f"{new_data_sheet} ", "") for col in scoring_result]
                return self.show_expression_dialog(scoring_columns_clean, year, step)
            
        return result['expression']

    def show_expression2_dialog(self, year, step=4):
        """Show expression dialog for 演算子‐2"""
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Create Expression for 演算子‐2 ({year})")
        dialog.geometry("650x450")
        dialog.configure(bg='#2c3e50')
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        
        # Center dialog
        dialog.geometry("+%d+%d" % (self.root.winfo_rootx() + 75, self.root.winfo_rooty() + 125))
        
        result = {'expression': None, 'action': None}
        
        # Header
        header_frame = tk.Frame(dialog, bg='#e67e22', height=80)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        title_label = tk.Label(header_frame, text=f"Create Expression for 演算子‐2 ({year})", 
                              font=('Segoe UI', 14, 'bold'), 
                              fg='white', bg='#e67e22')
        title_label.pack(pady=20)
        
        # Content frame
        content_frame = tk.Frame(dialog, bg='#f8f9fa')
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Instructions
        instruction_text = "Create expressions using predefined variables for 演算子‐2:\n"
        instruction_text += "These variables are based on the kouzoubutsu bangou sheet"
        
        instruction_label = tk.Label(content_frame, text=instruction_text,
                                   font=('Segoe UI', 10), fg='#34495e', bg='#f8f9fa',
                                   justify=tk.LEFT)
        instruction_label.pack(anchor='w', pady=(0, 15))
        
        # Predefined variables
        var_frame = tk.LabelFrame(content_frame, text="Available Variables", 
                                 font=('Segoe UI', 11, 'bold'), fg='#2c3e50', bg='#f8f9fa')
        var_frame.pack(fill=tk.X, pady=(0, 15))
        
        predefined_vars = [
            "X = 合計重み",
            "A1 = 構造形式_重み", 
            "B1 = 角度_重み",
            "C1 = 供用年数_重み"
        ]
        
        var_text = "\n".join(predefined_vars)
        var_display = tk.Label(var_frame, text=var_text, font=('Segoe UI', 10), 
                              fg='#7f8c8d', bg='#f8f9fa', justify=tk.LEFT)
        var_display.pack(anchor='w', padx=10, pady=10)
        
        # Expression input
        expr_frame = tk.LabelFrame(content_frame, text="Enter Expression", 
                                  font=('Segoe UI', 11, 'bold'), fg='#2c3e50', bg='#f8f9fa')
        expr_frame.pack(fill=tk.X, pady=(0, 15))
        
        expression_entry = tk.Text(expr_frame, height=3, font=('Consolas', 11),
                                  bg='white', fg='#2c3e50', relief='solid', bd=1)
        expression_entry.pack(fill=tk.X, padx=10, pady=10)
        
        # Button frame
        button_frame = tk.Frame(dialog, bg='#2c3e50', height=70)
        button_frame.pack(fill=tk.X, side=tk.BOTTOM)
        button_frame.pack_propagate(False)
        
        button_container = tk.Frame(button_frame, bg='#2c3e50')
        button_container.pack(expand=True)
        
        def validate_expression2():
            expr = expression_entry.get(1.0, tk.END).strip()
            if not expr:
                messagebox.showerror("Error", "Please enter an expression!")
                return False
            return True
        
        def on_submit():
            if validate_expression2():
                result['expression'] = expression_entry.get(1.0, tk.END).strip()
                result['action'] = 'submit'
                dialog.destroy()

        def on_back():
            result['action'] = 'back'
            dialog.destroy()

        # Main action buttons
        submit_btn = tk.Button(button_container, text="Create Expression", command=on_submit,
                              font=('Segoe UI', 10, 'bold'), fg='white', bg='#27ae60',
                              relief='flat', padx=20, pady=8, cursor='hand2')
        submit_btn.pack(side=tk.RIGHT, padx=5)
        
        back_btn = tk.Button(button_container, text="Back", command=on_back,
                            font=('Segoe UI', 10), fg='white', bg='#95a5a6',
                            relief='flat', padx=20, pady=8, cursor='hand2')
        back_btn.pack(side=tk.RIGHT, padx=5)

        self.root.wait_window(dialog)
        
        # Handle back navigation
        if result['action'] == 'back':
            # Go back to expression 1 dialog
            scoring_columns = self.workflow_data.get('scoring_columns', [])
            expr_result = self.show_expression_dialog(scoring_columns, year, step=3)
            if expr_result:
                return self.show_expression2_dialog(year, step)
                
        return result['expression']

    def process_existing_data(self, workbook):
        """Process existing data without adding new data"""
        try:
            self.show_progress()
            self.update_progress(50, "Processing existing data...")
            time.sleep(1)
            
            self.process_data(workbook, None)
            
            self.update_progress(100, "Processing completed!")
            time.sleep(1)
            
            self.root.after(0, lambda: messagebox.showinfo("Success", "Data processed successfully!"))
            self.hide_progress()
            
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("Error", f"An error occurred: {str(e)}"))
            self.hide_progress()

    def add_new_data_to_workbook(self, workbook, new_data_file, new_data_sheet):
        """Add new data to workbook efficiently - O(m) time complexity"""
        # Load new data from the specified sheet
        new_data_df = pd.read_excel(new_data_file, sheet_name=new_data_sheet)

        # Use openpyxl for efficient sheet manipulation
        wb = load_workbook(workbook)
        
        # Remove existing sheet if present
        if new_data_sheet in wb.sheetnames:
            wb.remove(wb[new_data_sheet])
        
        # Create new sheet
        ws = wb.create_sheet(new_data_sheet)
        
        # Write data efficiently using bulk operations
        for col_num, column_title in enumerate(new_data_df.columns, 1):
            ws.cell(row=1, column=col_num, value=column_title)
        
        for row_num, row_data in enumerate(new_data_df.values, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=cell_value)
        
        # Move sheet to first position
        wb._sheets.remove(ws)
        wb._sheets.insert(0, ws)
        
        wb.save(workbook)

    def update_chuushutsu_sheet(self, workbook, columns_to_extract, new_data_sheet):
        """Update 抽出列 sheet - check for existing year first"""
        wb = load_workbook(workbook)
        
        # Create sheet if it doesn't exist
        if "抽出列" not in wb.sheetnames:
            ws = wb.create_sheet("抽出列")
            target_col = 1
        else:
            ws = wb["抽出列"]
            # Check if year already exists
            target_col = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == new_data_sheet:
                    target_col = col
                    break
            
            if target_col is None:
                # Year doesn't exist, insert at column 1 and shift existing data
                ws.insert_cols(1)
                target_col = 1

        # Add year header at target column
        ws.cell(row=1, column=target_col, value=new_data_sheet)
        
        # Clear existing data in this column first (if updating)
        if target_col != 1 or ws.max_row > 1:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=target_col, value=None)
        
        # Add selected columns starting from row 2
        for i, col in enumerate(columns_to_extract, start=2):
            ws.cell(row=i, column=target_col, value=col)

        wb.save(workbook)

    def update_tensuka_sheet(self, workbook, scoring_columns, new_data_sheet):
        """Update 点数化列 sheet with year prefix - check for existing year first"""
        wb = load_workbook(workbook)
        
        # Create sheet if it doesn't exist
        if "点数化列" not in wb.sheetnames:
            ws = wb.create_sheet("点数化列")
            target_col = 1
        else:
            ws = wb["点数化列"]
            # Check if year already exists
            target_col = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == new_data_sheet:
                    target_col = col
                    break
            
            if target_col is None:
                # Year doesn't exist, insert at column 1 and shift existing data
                ws.insert_cols(1)
                target_col = 1

        # Add year header at target column
        ws.cell(row=1, column=target_col, value=new_data_sheet)
        
        # Clear existing data in this column first (if updating)
        if target_col != 1 or ws.max_row > 1:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=target_col, value=None)
        
        # Add scoring columns with year prefix starting from row 2
        for i, col in enumerate(scoring_columns, start=2):
            formatted_column = f"{new_data_sheet} {col}"
            ws.cell(row=i, column=target_col, value=formatted_column)

        wb.save(workbook)

    def update_ensanshi_sheet(self, workbook, expression, new_data_sheet):
        """Update 演算子 sheet - check for existing year first"""
        wb = load_workbook(workbook)
        
        # Create sheet if it doesn't exist
        if "演算子" not in wb.sheetnames:
            ws = wb.create_sheet("演算子")
            target_col = 1
        else:
            ws = wb["演算子"]
            # Check if year already exists
            target_col = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == new_data_sheet:
                    target_col = col
                    break
            
            if target_col is None:
                # Year doesn't exist, insert at column 1 and shift existing data
                ws.insert_cols(1)
                target_col = 1

        # Add year header and expression at target column
        ws.cell(row=1, column=target_col, value=new_data_sheet)
        ws.cell(row=2, column=target_col, value=expression)

        wb.save(workbook)

    def update_ensanshi2_sheet(self, workbook, expression2, new_data_sheet):
        """Update 演算子‐2 sheet - check for existing year first"""
        wb = load_workbook(workbook)
        
        # Create sheet if it doesn't exist
        if "演算子‐2" not in wb.sheetnames:
            ws = wb.create_sheet("演算子‐2")
            target_col = 1
        else:
            ws = wb["演算子‐2"]
            # Check if year already exists
            target_col = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == new_data_sheet:
                    target_col = col
                    break
            
            if target_col is None:
                # Year doesn't exist, insert at column 1 and shift existing data
                ws.insert_cols(1)
                target_col = 1

        # Add year header and expression at target column
        ws.cell(row=1, column=target_col, value=new_data_sheet)
        ws.cell(row=2, column=target_col, value=expression2)

        wb.save(workbook)

    def process_data(self, workbook, new_data_sheet):
        """Process data efficiently - O(k*m) where k is sheets, m is rows per sheet"""
        xls = pd.ExcelFile(workbook)
        sheet_names = [sheet for sheet in xls.sheet_names if sheet.isnumeric()]

        if new_data_sheet:
            sheet_names = [new_data_sheet] + [sheet for sheet in sheet_names if sheet != new_data_sheet]

        # Get columns for extraction from 抽出列 sheet
        columns_to_extract = self.get_columns_from_chuushutsu_sheet(workbook, new_data_sheet)
        if not columns_to_extract:
            # If no specific columns found, use all columns as fallback
            if sheet_names:
                sample_df = pd.read_excel(workbook, sheet_name=sheet_names[0])
                columns_to_extract = [col for col in sample_df.columns if col != '調査番号']

        if columns_to_extract:
            ketsugou_df = self.create_ketsugou_sheet(workbook, sheet_names, columns_to_extract)
            
            # Sort by 調査番号 if it exists
            if '調査番号' in ketsugou_df.columns:
                ketsugou_df = ketsugou_df.sort_values(by='調査番号')

            # Save efficiently using ExcelWriter
            with pd.ExcelWriter(workbook, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                ketsugou_df.to_excel(writer, sheet_name='結合シート', index=False)

    def get_columns_from_chuushutsu_sheet(self, workbook, year_sheet):
        """Get columns from 抽出列 sheet efficiently - O(m) where m is rows in sheet"""
        try:
            wb = load_workbook(workbook)
            if "抽出列" not in wb.sheetnames:
                return []
            
            ws = wb["抽出列"]
            columns = []
            
            # Find the column with the year header
            year_col = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == year_sheet:
                    year_col = col
                    break
            
            if year_col:
                # Get all non-empty cells below the year header
                for row in range(2, ws.max_row + 1):
                    cell_value = ws.cell(row=row, column=year_col).value
                    if cell_value:
                        columns.append(str(cell_value))
                    else:
                        break  # Stop at first empty cell
            
            return columns
        except Exception as e:
            print(f"Error reading 抽出列 sheet: {e}")
            return []

    def create_ketsugou_sheet(self, workbook, sheet_names, columns_to_extract):
        """Create combined sheet efficiently - O(k*m*n) optimized with pandas operations"""
        ketsugou_df = pd.DataFrame()
        
        for sheet in sheet_names:
            try:
                # Read sheet data efficiently
                df = pd.read_excel(workbook, sheet_name=sheet)
                
                # Ensure 調査番号 exists for merging
                if '調査番号' not in df.columns and df.columns.size > 0:
                    df.insert(0, '調査番号', range(1, len(df) + 1))
                
                # Filter columns efficiently using list comprehension
                available_columns = [col for col in columns_to_extract if col in df.columns]
                base_columns = ['調査番号'] if '調査番号' in df.columns else []
                
                # Select only needed columns
                selected_cols = base_columns + available_columns
                df_selected = df[selected_cols].copy()
                
                # Rename columns with year prefix (except 調査番号)
                new_columns = base_columns + [f"{sheet} {col}" for col in available_columns]
                df_selected.columns = new_columns
                
                # Merge efficiently
                if ketsugou_df.empty:
                    ketsugou_df = df_selected
                else:
                    merge_on = base_columns if base_columns else [ketsugou_df.columns[0]]
                    ketsugou_df = pd.merge(ketsugou_df, df_selected, on=merge_on, how='outer')
                    
            except Exception as e:
                print(f"Error processing sheet {sheet}: {e}")
                continue
        
        return ketsugou_df

    def extract_and_merge_data(self, workbook, sheet_names):
        """Extract and merge data efficiently - O(k*m) where k is sheets, m is rows"""
        dfs = []  # Use list for efficient concatenation
        
        for sheet in sheet_names:
            try:
                df = pd.read_excel(workbook, sheet_name=sheet)
                df['Year'] = sheet
                dfs.append(df)
            except Exception as e:
                print(f"Error reading sheet {sheet}: {e}")
                continue
        
        # Concatenate all dataframes at once for better performance
        if dfs:
            master_df = pd.concat(dfs, ignore_index=True)
        else:
            master_df = pd.DataFrame()
            
        return master_df

# Create and run the application
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelProcessorApp(root)
    root.mainloop()