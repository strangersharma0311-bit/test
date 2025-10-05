import subprocess
import os
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import Workbook, load_workbook
import pandas as pd
import xlsxwriter
import tempfile
import time

class PostObserProcessorApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Post-Obser Files Processor")
        
        # Get screen dimensions
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # Set window size based on screen size (80% of screen or minimum required)
        min_width = 900
        min_height = 800
        
        window_width = max(min_width, int(screen_width * 0.8))
        window_height = max(min_height, int(screen_height * 0.8))
        
        # Ensure window doesn't exceed screen size
        window_width = min(window_width, screen_width - 100)
        window_height = min(window_height, screen_height - 100)
        
        self.root.geometry(f"{window_width}x{window_height}")
        self.root.minsize(min_width, min_height)
        
        # Center the window on screen - improved centering
        self.root.update_idletasks()
        x = max(0, (screen_width // 2) - (window_width // 2))
        y = max(0, (screen_height // 2) - (window_height // 2))
        self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        self.root.configure(bg="white")
        
        self.base_dir = None
        self.fortran_program_path = None
        self.year_ranges = None
        self.processing_settings = {
            'create_charts': True,  # Always True - charts are essential
            'backup_original': True,
            'detailed_logging': True
        }
        
        self.create_professional_gui()
    
    def create_professional_button(self, parent, text, command, bg_color, hover_color, **kwargs):
        """Create a professional button with hover effects"""
        btn = tk.Button(parent, text=text, command=command, bg=bg_color, fg="white",
                       font=("Arial", 11, "bold"), relief="flat", cursor="hand2",
                       activebackground=hover_color, activeforeground="white",
                       bd=0, padx=20, pady=10, **kwargs)
        
        # Add hover effects
        def on_enter(e):
            if btn['state'] != 'disabled':
                btn.config(bg=hover_color)
        
        def on_leave(e):
            if btn['state'] != 'disabled':
                btn.config(bg=bg_color)
        
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)
        
        return btn
    
    def create_professional_gui(self):
        """Create professional GUI with scrollable layout"""
        # Create main frame with scrollbars
        main_canvas = tk.Canvas(self.root, bg="white")
        v_scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=main_canvas.yview)
        h_scrollbar = ttk.Scrollbar(self.root, orient="horizontal", command=main_canvas.xview)
        
        # Configure scrollbars
        main_canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Create scrollable frame
        scrollable_frame = tk.Frame(main_canvas, bg="white")
        scrollable_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )
        
        # Create window in canvas
        canvas_frame = main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        
        # Pack scrollbars and canvas
        v_scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x")
        main_canvas.pack(side="left", fill="both", expand=True)
        
        # Configure canvas scrolling with mouse wheel and touchpad
        def _on_mousewheel(event):
            main_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        def _bind_to_mousewheel(event):
            main_canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        def _unbind_from_mousewheel(event):
            main_canvas.unbind_all("<MouseWheel>")
        
        main_canvas.bind('<Enter>', _bind_to_mousewheel)
        main_canvas.bind('<Leave>', _unbind_from_mousewheel)
        
        # Update scroll region when canvas size changes
        def configure_scroll_region(event):
            main_canvas.configure(scrollregion=main_canvas.bbox("all"))
            # Update the width of scrollable_frame to match canvas width
            canvas_width = event.width
            main_canvas.itemconfig(canvas_frame, width=canvas_width)
        
        main_canvas.bind('<Configure>', configure_scroll_region)
        
        # Main container inside scrollable frame
        main_container = tk.Frame(scrollable_frame, bg="white")
        main_container.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Header section - more compact
        header_frame = tk.Frame(main_container, bg='white')
        header_frame.pack(fill="x", pady=(0, 15))
        
        title_label = tk.Label(header_frame, text="📊 Post-Obser Files Processor", 
                              font=("Arial", 18, "bold"), fg="#2c3e50", bg='white')
        title_label.pack()
        
        subtitle_label = tk.Label(header_frame, text="Process obser files with Fortran program and create enhanced Excel outputs", 
                                 font=("Arial", 11), fg="#7f8c8d", bg='white')
        subtitle_label.pack(pady=(3, 0))
        
        # Requirements section - more compact
        req_frame = tk.LabelFrame(main_container, text="📋 Requirements", 
                                 font=("Arial", 11, "bold"), fg="#2c3e50", bg='white', 
                                 bd=1, relief="solid")
        req_frame.pack(fill="x", pady=(0, 15), ipady=10)
        
        req_container = tk.Frame(req_frame, bg='white')
        req_container.pack(fill="x", padx=12, pady=8)
        
        req_text = ("📁 Directory must contain:\n"
                   "  • obser1.txt to obser8.txt (generated by obser file generator)\n"
                   "  • 劣化予測プログラム .exe (Fortran program)\n"
                   "  • tamagawa-new method  - Copy .xlsx (structure data for charts)\n"
                   "  • All required txt files for Fortran program (pml.txt, logdensity.txt, ex1000.txt will be generated)\n\n"
                   "📊 Output will be created in 'output' subfolder:\n"
                   "  • Basic Excel files (出力1.xlsx - 出力8.xlsx)\n"
                   "  • Chart-enhanced files (作図付き出力1.xlsx - 作図付き出力8.xlsx)")
        
        tk.Label(req_container, text=req_text, font=("Arial", 9), 
                justify="left", bg='white', fg="#34495e").pack(anchor="w")
        
        # Status section - more compact
        status_frame = tk.LabelFrame(main_container, text="📊 Status", font=("Arial", 11, "bold"), 
                                   fg="#2c3e50", bg='white', bd=1, relief="solid")
        status_frame.pack(fill="x", pady=(0, 15), ipady=8)
        
        self.status_label = tk.Label(status_frame, text="💾 Ready to select processing directory...", 
                                    font=("Arial", 10), fg="#3498db", bg='white')
        self.status_label.pack(pady=5, padx=12, anchor="w")
        
        # Directory selection - more compact
        file_frame = tk.Frame(main_container, bg='white')
        file_frame.pack(fill="x", pady=(0, 15))
        
        select_btn = self.create_professional_button(
            file_frame, "📁 Select Processing Directory", self.select_directory,
            "#3498db", "#2980b9", width=30, height=1
        )
        select_btn.pack()
        
        # Processing settings section - more compact
        settings_frame = tk.LabelFrame(main_container, text="⚙️ Processing Settings", 
                                     font=("Arial", 11, "bold"), fg="#2c3e50", bg='white', 
                                     bd=1, relief="solid")
        settings_frame.pack(fill="x", pady=(0, 15), ipady=10)
        
        settings_container = tk.Frame(settings_frame, bg='white')
        settings_container.pack(fill="x", padx=12, pady=8)
        
        # Chart configuration - more compact
        chart_frame = tk.Frame(settings_container, bg='white')
        chart_frame.pack(fill="x", pady=(0, 8))
        
        tk.Label(chart_frame, text="Year range for logdensity charts:", 
                font=("Arial", 10, "bold"), bg='white', fg="#34495e").pack(anchor="w")
        
        year_controls = tk.Frame(chart_frame, bg='white')
        year_controls.pack(fill="x", pady=5)
        
        tk.Label(year_controls, text="Start:", font=("Arial", 9), bg='white').pack(side="left")
        self.start_year_var = tk.StringVar(value="1")
        tk.Entry(year_controls, textvariable=self.start_year_var, width=6, 
                font=("Arial", 9)).pack(side="left", padx=3)
        
        tk.Label(year_controls, text="End:", font=("Arial", 9), bg='white').pack(side="left", padx=(15, 0))
        self.end_year_var = tk.StringVar(value="5")
        tk.Entry(year_controls, textvariable=self.end_year_var, width=6, 
                font=("Arial", 9)).pack(side="left", padx=3)
        
        tk.Label(year_controls, text="(1-based indexing)", 
                font=("Arial", 8), fg="#7f8c8d", bg='white').pack(side="left", padx=(10, 0))
        
        # Processing options - REMOVED "Create charts" checkbox
        options_frame = tk.Frame(settings_container, bg='white')
        options_frame.pack(fill="x", pady=(8, 0))
        
        self.backup_var = tk.BooleanVar(value=True)
        tk.Checkbutton(options_frame, text="Backup original obser files during processing", 
                      variable=self.backup_var, bg='white', 
                      font=("Arial", 9)).pack(anchor="w", pady=1)
        
        self.detailed_log_var = tk.BooleanVar(value=True)
        tk.Checkbutton(options_frame, text="Show detailed processing logs", 
                      variable=self.detailed_log_var, bg='white', 
                      font=("Arial", 9)).pack(anchor="w", pady=1)
        
        # Action buttons section - more compact height
        actions_frame = tk.LabelFrame(main_container, text="🚀 Actions", font=("Arial", 11, "bold"), 
                                    fg="#2c3e50", bg='white', bd=1, relief="solid")
        actions_frame.pack(fill="x", pady=(0, 15), ipady=5)  # Reduced from ipady=12 to ipady=5
        
        button_frame = tk.Frame(actions_frame, bg='white')
        button_frame.pack(pady=5)  # Reduced from pady=10 to pady=5
        
        # Create buttons in a grid layout with smaller height
        button_grid = tk.Frame(button_frame, bg='white')
        button_grid.pack()
        
        self.start_btn = self.create_professional_button(
            button_grid, "🚀 Start Processing", self.start_processing,
            "#e67e22", "#d35400", width=18, height=1, state="disabled"  # Reduced from height=2 to height=1
        )
        self.start_btn.grid(row=0, column=0, padx=6, pady=3)  # Reduced padx from 8 to 6, pady from 5 to 3
        
        self.open_btn = self.create_professional_button(
            button_grid, "📁 Open Output Folder", self.open_output_folder,
            "#27ae60", "#229954", width=18, height=1, state="disabled"  # Reduced from height=2 to height=1
        )
        self.open_btn.grid(row=0, column=1, padx=6, pady=3)  # Reduced padding
        
        self.close_btn = self.create_professional_button(
            button_grid, "❌ Close Application", self.close_application,
            "#e74c3c", "#c0392b", width=18, height=1  # Reduced from height=2 to height=1
        )
        self.close_btn.grid(row=0, column=2, padx=6, pady=3)  # Reduced padding
        
        # Footer - more compact
        footer_label = tk.Label(main_container, text="Powered by Advanced File Processing Engine", 
                               font=("Arial", 8), fg="#95a5a6", bg='white')
        footer_label.pack(side="bottom", pady=(10, 0))
        
        # Add some bottom padding to ensure scrolling works properly
        bottom_spacer = tk.Frame(main_container, bg='white', height=20)
        bottom_spacer.pack(side="bottom", fill="x")

    def close_application(self):
        """Close the application"""
        self.root.quit()

    def select_directory(self):
        """Enhanced directory selection with better validation"""
        self.base_dir = filedialog.askdirectory(title="Select Directory with Obser Files")
        
        if not self.base_dir:
            return
        
        # Enhanced validation
        validation_results = self.validate_directory()
        
        if validation_results['valid']:
            self.fortran_program_path = os.path.join(self.base_dir, "劣化予測プログラム .exe")
            self.status_label.config(text=f"✅ Directory validated: {os.path.basename(self.base_dir)}", 
                                   fg="#27ae60")
            
            # Enable action buttons
            self.start_btn.config(state="normal")
            self.open_btn.config(state="normal")
        else:
            self.show_validation_error(validation_results)

    def validate_directory(self):
        """Enhanced directory validation"""
        results = {
            'valid': True,
            'missing_obser': [],
            'missing_programs': [],
            'missing_txt_files': [],
            'warnings': []
        }
        
        # Check obser files
        for i in range(1, 9):
            obser_file = f"obser{i}.txt"
            if not os.path.exists(os.path.join(self.base_dir, obser_file)):
                results['missing_obser'].append(obser_file)
        
        # Check required programs
        required_programs = ["劣化予測プログラム .exe"]
        for program in required_programs:
            if not os.path.exists(os.path.join(self.base_dir, program)):
                results['missing_programs'].append(program)
        
        # Check for any existing txt files that might be required by Fortran
        existing_txt_files = [f for f in os.listdir(self.base_dir) if f.endswith('.txt') and not f.startswith('obser')]
        if existing_txt_files:
            results['warnings'].append(f"Found additional txt files: {', '.join(existing_txt_files)}")
        
        # Check optional files
        if not os.path.exists(os.path.join(self.base_dir, "tamagawa-new method  - Copy .xlsx")):
            results['warnings'].append("tamagawa-new method  - Copy .xlsx not found - charts will not be generated")
        
        # Determine if valid
        if results['missing_obser'] or results['missing_programs']:
            results['valid'] = False
        
        return results

    def show_validation_error(self, results):
        """Show detailed validation error with responsive sizing"""
        error_window = tk.Toplevel(self.root)
        error_window.title("Directory Validation")
        
        # Responsive sizing based on screen
        screen_width = error_window.winfo_screenwidth()
        screen_height = error_window.winfo_screenheight()
        
        window_width = min(700, int(screen_width * 0.7))
        window_height = min(600, int(screen_height * 0.7))
        
        error_window.geometry(f"{window_width}x{window_height}")
        error_window.minsize(600, 500)
        error_window.grab_set()
        error_window.configure(bg="white")
        
        # Center the window
        x = (screen_width // 2) - (window_width // 2)
        y = (screen_height // 2) - (window_height // 2)
        error_window.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        main_frame = tk.Frame(error_window, bg="white", padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)
        
        if results['valid']:
            tk.Label(main_frame, text="✅ Directory Validation Passed", 
                    font=("Arial", 14, "bold"), fg="#27ae60", bg="white").pack(pady=(0, 15))
        else:
            tk.Label(main_frame, text="❌ Directory Validation Failed", 
                    font=("Arial", 14, "bold"), fg="#e74c3c", bg="white").pack(pady=(0, 15))
        
        # Create scrollable text area
        text_frame = tk.Frame(main_frame, bg="white")
        text_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        text_area = tk.Text(text_frame, wrap=tk.WORD, font=("Consolas", 9), bg="#f8f9fa")
        scrollbar = ttk.Scrollbar(text_frame, orient="vertical", command=text_area.yview)
        text_area.configure(yscrollcommand=scrollbar.set)
        
        text_area.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Build validation message
        validation_msg = f"Selected directory: {self.base_dir}\n\n"
        
        if results['missing_obser']:
            validation_msg += "❌ Missing Obser Files:\n"
            for file in results['missing_obser']:
                validation_msg += f"  • {file}\n"
            validation_msg += "\n"
        
        if results['missing_programs']:
            validation_msg += "❌ Missing Required Programs:\n"
            for program in results['missing_programs']:
                validation_msg += f"  • {program}\n"
            validation_msg += "\n"
        
        if results['warnings']:
            validation_msg += "⚠️ Information:\n"
            for warning in results['warnings']:
                validation_msg += f"  • {warning}\n"
            validation_msg += "\n"
        
        validation_msg += ("📋 Required Files:\n"
                         "  • obser1.txt to obser8.txt (from Obser File Generator)\n"
                         "  • 劣化予測プログラム .exe (Fortran program)\n"
                         "  • tamagawa-new method  - Copy .xlsx (optional, for charts)\n"
                         "  • Any additional txt files required by Fortran program\n\n"
                         "📊 Generated Output Files:\n"
                         "  • pml.txt, logdensity.txt, ex1000.txt (temporary files)\n"
                         "  • 8 Basic Excel files (出力1.xlsx - 出力8.xlsx)\n"
                         "  • 8 Chart-enhanced files (作図付き出力1.xlsx - 作図付き出力8.xlsx)\n\n"
                         "Please ensure all required files are in the selected directory.")
        
        text_area.insert("1.0", validation_msg)
        text_area.config(state="disabled")
        
        # Close button
        close_btn = self.create_professional_button(
            main_frame, "Close", error_window.destroy,
            "#3498db", "#2980b9", width=12, height=1
        )
        close_btn.pack()

    def start_processing(self):
        """Start processing with settings validation"""
        if not self.base_dir:
            messagebox.showerror("Error", "Please select a directory first")
            return
        
        # Validate settings
        try:
            start_year = int(self.start_year_var.get())
            end_year = int(self.end_year_var.get())
            if start_year < 1 or end_year < start_year:
                raise ValueError
            self.year_ranges = (start_year, end_year)
        except ValueError:
            messagebox.showerror("Error", "Invalid year range. Please enter valid positive integers.")
            return
        
        # Update processing settings
        self.processing_settings.update({
            'create_charts': True,  # Always True
            'backup_original': self.backup_var.get(),
            'detailed_logging': self.detailed_log_var.get()
        })
        
        # Show processing dialog
        self.show_processing_dialog()

    def show_processing_dialog(self):
        """Show processing dialog with responsive sizing"""
        self.progress_window = tk.Toplevel(self.root)
        self.progress_window.title("Processing Obser Files")
        
        # Responsive sizing
        screen_width = self.progress_window.winfo_screenwidth() 
        screen_height = self.progress_window.winfo_screenheight()
        
        window_width = min(750, int(screen_width * 0.8))
        window_height = min(650, int(screen_height * 0.8))
        
        self.progress_window.geometry(f"{window_width}x{window_height}")
        self.progress_window.minsize(700, 550)
        self.progress_window.grab_set()
        self.progress_window.configure(bg="white")
        
        # Center window
        x = (screen_width // 2) - (window_width // 2)
        y = (screen_height // 2) - (window_height // 2)
        self.progress_window.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        progress_frame = tk.Frame(self.progress_window, bg="white", padx=20, pady=15)
        progress_frame.pack(fill="both", expand=True)
        
        tk.Label(progress_frame, text="🔄 Processing Obser Files", 
                font=("Arial", 16, "bold"), fg="#2c3e50", bg="white").pack(pady=(0, 15))
        
        # Progress bars
        tk.Label(progress_frame, text="Overall Progress:", 
                font=("Arial", 11, "bold"), bg="white").pack(anchor="w")
        self.overall_progress = ttk.Progressbar(progress_frame, mode='determinate', maximum=10, length=500)
        self.overall_progress.pack(fill="x", pady=(5, 10))
        
        tk.Label(progress_frame, text="Current Step:", 
                font=("Arial", 11, "bold"), bg="white").pack(anchor="w")
        self.step_progress = ttk.Progressbar(progress_frame, mode='determinate', maximum=8, length=500)
        self.step_progress.pack(fill="x", pady=(5, 10))
        
        # Status
        self.current_status = tk.Label(progress_frame, text="Initializing...", 
                                      font=("Arial", 11, "bold"), fg="#e67e22", bg="white")
        self.current_status.pack(pady=(0, 10))
        
        # Log area
        log_frame = tk.Frame(progress_frame, bg="white")
        log_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        self.progress_text = tk.Text(log_frame, height=12, width=70, 
                                   font=("Consolas", 8), bg="#f8f9fa")
        log_scrollbar = ttk.Scrollbar(log_frame, orient="vertical", 
                                    command=self.progress_text.yview)
        self.progress_text.configure(yscrollcommand=log_scrollbar.set)
        
        self.progress_text.pack(side="left", fill="both", expand=True)
        log_scrollbar.pack(side="right", fill="y")
        
        # Start processing - NO CANCEL BUTTON
        self.root.after(100, self.execute_processing)

    def log_progress(self, message, level="INFO"):
        """Enhanced progress logging"""
        timestamp = pd.Timestamp.now().strftime("%H:%M:%S")
        
        if level == "ERROR":
            prefix = "❌"
        elif level == "WARNING":
            prefix = "⚠️"
        elif level == "SUCCESS":
            prefix = "✅"
        else:
            prefix = "ℹ️"
        
        log_message = f"[{timestamp}] {prefix} {message}\n"
        
        self.progress_text.insert(tk.END, log_message)
        self.progress_text.see(tk.END)
        
        if self.processing_settings['detailed_logging']:
            print(log_message.strip())
        
        self.progress_window.update()

    def update_status(self, status, step_progress=None, overall_progress=None):
        """Update status and progress bars"""
        self.current_status.config(text=status)
        
        if step_progress is not None:
            self.step_progress['value'] = step_progress
        
        if overall_progress is not None:
            self.overall_progress['value'] = overall_progress
        
        self.progress_window.update()

    def execute_processing(self):
        """Execute processing with better error handling"""
        try:
            self.update_status("🔧 Setting up directories...", 0, 1)
            
            output_dir = os.path.join(self.base_dir, "output")
            temp_dir = os.path.join(self.base_dir, "temp_obser")
            backup_dir = os.path.join(self.base_dir, "backup_obser") if self.processing_settings['backup_original'] else None
            
            # Setup directories
            self.setup_directories(output_dir, temp_dir, backup_dir)
            
            self.update_status("📝 Processing obser files...", 0, 2)
            files_to_write = ["pml.txt", "logdensity.txt", "ex1000.txt"]
            
            # Create backup if requested
            if backup_dir:
                self.create_backup(backup_dir)
            
            # Process first obser file
            self.update_status("🔄 Processing obser1.txt...", 1, 3)
            self.log_progress("Processing obser1.txt...")
            
            self.run_fortran_program_safe()
            self.write_to_excel_safe(self.base_dir, output_dir, files_to_write, "出力1.xlsx")
            
            # Store original obser1.txt
            original_obser1 = os.path.join(self.base_dir, "obser1.txt")
            temp_obser1 = os.path.join(temp_dir, "obser1_temp_1.txt")
            os.rename(original_obser1, temp_obser1)
            
            self.step_progress['value'] = 1
            
            # Process remaining obser files
            for i in range(2, 9):
                self.update_status(f"🔄 Processing obser{i}.txt...", i, 3)
                self.log_progress(f"Processing obser{i}.txt...")
                
                current_obser = os.path.join(self.base_dir, f"obser{i}.txt")
                if not os.path.exists(current_obser):
                    self.log_progress(f"Warning: obser{i}.txt not found, skipping...", "WARNING")
                    continue
                
                # Rename and process
                os.rename(current_obser, original_obser1)
                self.run_fortran_program_safe()
                
                workbook_name = f"出力{i}.xlsx"
                self.write_to_excel_safe(self.base_dir, output_dir, files_to_write, workbook_name)
                
                # Store processed file
                temp_obser_path = os.path.join(temp_dir, f"obser1_temp_{i}.txt")
                os.rename(original_obser1, temp_obser_path)
                
                self.step_progress['value'] = i
            
            # Restore original files
            self.update_status("🔄 Restoring original files...", 8, 4)
            self.restore_original_files(temp_dir)
            
            # Create charts (always enabled now)
            self.update_status("📊 Creating charts...", 0, 5)
            self.execute_chart_generation_safe(output_dir)
            
            # Cleanup
            self.cleanup_processing(temp_dir, backup_dir)
            
            self.update_status("✅ Processing completed successfully!", 8, 10)
            self.log_progress("All processing completed successfully!", "SUCCESS")
            
            # Auto-close and show completion
            self.root.after(3000, lambda: [self.progress_window.destroy(), 
                                          self.show_completion_dialog()])
                
        except Exception as e:
            self.log_progress(f"Critical error: {str(e)}", "ERROR")
            self.update_status("❌ Processing failed", 0, 0)
            messagebox.showerror("Processing Error", f"Processing failed:\n{str(e)}")
            self.progress_window.destroy()

    def setup_directories(self, output_dir, temp_dir, backup_dir):
        """Setup required directories"""
        self.log_progress("Setting up directories...")
        
        # Clear and create output directory
        if os.path.exists(output_dir):
            shutil.rmtree(output_dir)
        os.makedirs(output_dir)
        
        # Create temp directory
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        os.makedirs(temp_dir)
        
        # Create backup directory if needed
        if backup_dir:
            if os.path.exists(backup_dir):
                shutil.rmtree(backup_dir)
            os.makedirs(backup_dir)
        
        self.log_progress("Directories setup completed", "SUCCESS")

    def create_backup(self, backup_dir):
        """Create backup of original obser files"""
        self.log_progress("Creating backup of original obser files...")
        
        for i in range(1, 9):
            obser_file = f"obser{i}.txt"
            source_path = os.path.join(self.base_dir, obser_file)
            backup_path = os.path.join(backup_dir, obser_file)
            
            if os.path.exists(source_path):
                shutil.copy2(source_path, backup_path)
        
        self.log_progress("Backup created successfully", "SUCCESS")

    def run_fortran_program_safe(self):
        """Run Fortran program with error handling"""
        try:
            result = subprocess.run([self.fortran_program_path], 
                                cwd=self.base_dir, 
                                check=True, 
                                capture_output=True, 
                                text=True)
            
            if self.processing_settings['detailed_logging'] and result.stdout:
                self.log_progress(f"Fortran output: {result.stdout.strip()}")
                    
        except subprocess.CalledProcessError as e:
            error_msg = f"Fortran program failed with return code {e.returncode}"
            if e.stderr:
                error_msg += f"\nError output: {e.stderr}"
            raise Exception(error_msg)
        except FileNotFoundError:
            raise Exception(f"Fortran program not found: {self.fortran_program_path}")

    def write_to_excel_safe(self, source_dir, output_dir, files_to_write, workbook_name):
        """Write text files to Excel workbook"""
        try:
            workbook = Workbook()
            
            for file_name in files_to_write:
                sheet_name = file_name.split('.')[0]
                sheet = workbook.create_sheet(title=sheet_name)
                file_path = os.path.join(source_dir, file_name)
                
                if os.path.exists(file_path):
                    with open(file_path, 'r', encoding='utf-8') as file:
                        for row_idx, line in enumerate(file):
                            line = line.strip()
                            if line:
                                values = line.split()
                                for col_idx, value in enumerate(values):
                                    sheet.cell(row=row_idx + 1, column=col_idx + 1, value=value)
                else:
                    self.log_progress(f"Warning: {file_name} not found", "WARNING")
            
            # Remove the default sheet
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
            
            excel_path = os.path.join(output_dir, workbook_name)
            workbook.save(excel_path)
            self.log_progress(f"Created {workbook_name}", "SUCCESS")
            
        except Exception as e:
            raise Exception(f"Error writing to Excel {workbook_name}: {e}")

    def restore_original_files(self, temp_dir):
        """Restore original obser files"""
        self.log_progress("Restoring original obser files...")
        
        for i in range(1, 9):
            temp_obser_path = os.path.join(temp_dir, f"obser1_temp_{i}.txt")
            original_path = os.path.join(self.base_dir, f"obser{i}.txt")
            
            if os.path.exists(temp_obser_path):
                os.rename(temp_obser_path, original_path)
        
        self.log_progress("Original files restored", "SUCCESS")

    def execute_chart_generation_safe(self, output_folder):
        """Execute chart generation with error handling"""
        try:
            tamagawa_file = os.path.join(self.base_dir, 'tamagawa-new method  - Copy .xlsx')
            
            if not os.path.exists(tamagawa_file):
                self.log_progress("tamagawa-new method  - Copy .xlsx not found, skipping chart generation", "WARNING")
                return
            
            self.log_progress("Loading structure data...")
            
            # Load structure data with error handling
            try:
                wb = load_workbook(tamagawa_file, data_only=True)
                values_dict = {}
                
                for sheet in wb.worksheets:
                    values = {'構造物名称': [], '構造物番号': []}
                    try:
                        for row in sheet.iter_rows(min_row=2, values_only=True):
                            if row and len(row) >= 4:
                                values['構造物名称'].append(row[2] if row[2] else '')
                                values['構造物番号'].append(row[3] if row[3] else '')
                        values_dict[sheet.title] = values
                    except Exception as e:
                        self.log_progress(f"Error reading sheet {sheet.title}: {e}", "WARNING")
                
                wb.close()
                
            except Exception as e:
                self.log_progress(f"Error loading structure data: {e}", "WARNING")
                return
            
            # Process output files
            output_files = [f for f in os.listdir(output_folder) 
                          if f.endswith('.xlsx') and f.startswith('出力') 
                          and not f.startswith('作図付き')]
            
            for i, file in enumerate(output_files):
                self.log_progress(f"Creating charts for {file}...")
                file_path = os.path.join(output_folder, file)
                sheet_name = self.get_sheet_name(file)
                
                if sheet_name in values_dict:
                    try:
                        self.process_file_with_charts(file_path, values_dict[sheet_name], 
                                                    output_folder, self.year_ranges)
                        self.log_progress(f"Charts created for {file}", "SUCCESS")
                    except Exception as e:
                        self.log_progress(f"Error creating charts for {file}: {e}", "ERROR")
                else:
                    self.log_progress(f"No structure data found for {sheet_name}", "WARNING")
                
                # Update progress
                chart_progress = ((i + 1) / len(output_files)) * 8
                self.step_progress['value'] = chart_progress
                self.progress_window.update()
            
        except Exception as e:
            self.log_progress(f"Chart generation error: {e}", "ERROR")

    def cleanup_processing(self, temp_dir, backup_dir):
        """Cleanup temporary files"""
        self.log_progress("Cleaning up temporary files...")
        
        # Remove temp directory
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        
        # Keep backup directory if created
        if backup_dir and os.path.exists(backup_dir):
            self.log_progress(f"Backup files preserved in: {backup_dir}")
        
        self.log_progress("Cleanup completed", "SUCCESS")

    def open_output_folder(self):
        """Open output folder in file explorer"""
        if not self.base_dir:
            messagebox.showwarning("Warning", "Please select a directory first")
            return
        
        output_dir = os.path.join(self.base_dir, "output")
        
        if not os.path.exists(output_dir):
            messagebox.showinfo("Info", "Output folder does not exist yet")
            return
        
        try:
            os.startfile(output_dir)
        except:
            messagebox.showinfo("Info", f"Please open folder manually:\n{output_dir}")

    def show_completion_dialog(self):
        """Show completion dialog with responsive sizing"""
        completion_window = tk.Toplevel(self.root)
        completion_window.title("Processing Complete")
        
        # Responsive sizing
        screen_width = completion_window.winfo_screenwidth()
        screen_height = completion_window.winfo_screenheight()
        
        window_width = min(750, int(screen_width * 0.8))
        window_height = min(650, int(screen_height * 0.8))
        
        completion_window.geometry(f"{window_width}x{window_height}")
        completion_window.minsize(700, 550)
        completion_window.grab_set()
        completion_window.configure(bg="white")
        
        # Center window
        x = (screen_width // 2) - (window_width // 2)
        y = (screen_height // 2) - (window_height // 2)
        completion_window.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        main_frame = tk.Frame(completion_window, bg="white", padx=20, pady=15)
        main_frame.pack(fill="both", expand=True)
        
        # Success header
        header_frame = tk.Frame(main_frame, bg="white")
        header_frame.pack(fill="x", pady=(0, 20))
        
        tk.Label(header_frame, text="🎉", font=("Arial", 32), bg="white").pack(side="left")
        tk.Label(header_frame, text="Processing Complete!", 
                font=("Arial", 16, "bold"), fg="#27ae60", bg="white").pack(side="left", padx=(8, 0))
        
        # Summary statistics
        summary_frame = tk.LabelFrame(main_frame, text="Processing Summary", 
                                    font=("Arial", 11, "bold"), bg="white", padx=12, pady=8)
        summary_frame.pack(fill="x", pady=(0, 15))
        
        output_dir = os.path.join(self.base_dir, "output")
        basic_files = len([f for f in os.listdir(output_dir) if f.startswith('出力') and not f.startswith('作図付き')])
        chart_files = len([f for f in os.listdir(output_dir) if f.startswith('作図付き出力')])
        
        summary_text = (f"✅ Successfully processed {basic_files} obser files\n"
                       f"📊 Created {basic_files} basic Excel outputs\n"
                       f"📈 Created {chart_files} chart-enhanced outputs\n"
                       f"📁 All files saved to: output folder")
        
        tk.Label(summary_frame, text=summary_text, font=("Arial", 10), 
                justify="left", bg="white").pack(anchor="w")
        
        # File details
        details_frame = tk.LabelFrame(main_frame, text="Generated Files", 
                                    font=("Arial", 11, "bold"), bg="white", padx=12, pady=8)
        details_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        # Scrollable file list
        files_text = tk.Text(details_frame, height=8, font=("Consolas", 8), bg="#f8f9fa")
        files_scrollbar = ttk.Scrollbar(details_frame, orient="vertical", command=files_text.yview)
        files_text.configure(yscrollcommand=files_scrollbar.set)
        
        files_text.pack(side="left", fill="both", expand=True)
        files_scrollbar.pack(side="right", fill="y")
        
        # List all generated files
        files_text.insert("1.0", "Basic Output Files:\n")
        for i in range(1, 9):
            filename = f"出力{i}.xlsx"
            if os.path.exists(os.path.join(output_dir, filename)):
                files_text.insert(tk.END, f"  ✅ {filename}\n")
            else:
                files_text.insert(tk.END, f"  ❌ {filename} (missing)\n")
        
        files_text.insert(tk.END, "\nChart-Enhanced Files:\n")
        for i in range(1, 9):
            filename = f"作図付き出力{i}.xlsx"
            if os.path.exists(os.path.join(output_dir, filename)):
                files_text.insert(tk.END, f"  ✅ {filename}\n")
            else:
                files_text.insert(tk.END, f"  ❌ {filename} (missing)\n")
        
        files_text.config(state="disabled")
        
        # Action buttons - in a grid for better layout
        button_frame = tk.Frame(main_frame, bg="white")
        button_frame.pack(fill="x", pady=15)
        
        button_grid = tk.Frame(button_frame, bg="white")
        button_grid.pack()
        
        def open_output():
            try:
                os.startfile(output_dir)
                completion_window.after(1000, completion_window.destroy)
            except:
                messagebox.showinfo("Info", f"Please open folder manually:\n{output_dir}")
        
        def process_another():
            completion_window.destroy()
            self.reset_app()
        
        def exit_app():
            completion_window.destroy()
            self.root.quit()
        
        self.create_professional_button(
            button_grid, "📁 Open Output Folder", open_output,
            "#27ae60", "#229954", width=16, height=1
        ).grid(row=0, column=0, padx=8, pady=5)
        
        self.create_professional_button(
            button_grid, "🔄 Process Another", process_another,
            "#3498db", "#2980b9", width=16, height=1
        ).grid(row=0, column=1, padx=8, pady=5)
        
        self.create_professional_button(
            button_grid, "❌ Exit Application", exit_app,
            "#e74c3c", "#c0392b", width=16, height=1
        ).grid(row=0, column=2, padx=8, pady=5)

    def reset_app(self):
        """Reset application for processing another directory"""
        self.base_dir = None
        self.fortran_program_path = None
        self.year_ranges = None
        
        # Reset status and disable buttons
        self.status_label.config(text="💾 Ready to select processing directory...", fg="#3498db")
        self.start_btn.config(state="disabled")
        self.open_btn.config(state="disabled")
        
        # Reset settings to defaults
        self.start_year_var.set("1")
        self.end_year_var.set("5")
        self.backup_var.set(True)
        self.detailed_log_var.set(True)

    def get_sheet_name(self, output_filename):
        """Get sheet name mapping for output file"""
        sheet_mapping = {
            '出力1.xlsx': '割算結果(補修考慮)',
            '出力2.xlsx': '割算結果(補修無視)', 
            '出力3.xlsx': '補修無視',
            '出力4.xlsx': '補修考慮',
            '出力5.xlsx': '新しい演算(補修無視)',
            '出力6.xlsx': '新しい演算(補修考慮)',
            '出力7.xlsx': '割算結果-新しい演算(補修無視)',
            '出力8.xlsx': '割算結果-新しい演算(補修考慮)'
        }
        return sheet_mapping.get(output_filename, '割算結果(補修考慮)')

    def process_file_with_charts(self, file_path, values, output_folder_path, year_ranges):
        """Process file and add charts"""
        try:
            wb = load_workbook(file_path)
            
            # Get or create sheets
            ex_ws = wb['ex1000'] if 'ex1000' in wb.sheetnames else None
            log_ws = wb['logdensity'] if 'logdensity' in wb.sheetnames else wb.create_sheet('logdensity')
            pml_ws = wb['pml'] if 'pml' in wb.sheetnames else None

            # Format sheets
            if ex_ws:
                self.format_ex1000(ex_ws, values)
            self.format_logdensity(log_ws)
            if pml_ws:
                self.format_pml(pml_ws)

            # Create new filename
            base_name = os.path.basename(file_path)
            if '_' in base_name:
                modified_filename = f"作図付き出力{base_name.split('_')[1].split('.')[0]}.xlsx"
            else:
                modified_filename = f"作図付き出力{base_name.split('.')[0][-1]}.xlsx"

            new_file_path = os.path.join(output_folder_path, modified_filename)

            # Save to temporary file first
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                temp_path = tmp.name
            wb.save(temp_path)
            wb.close()

            # Read data and create charts with xlsxwriter
            self.create_charts_with_xlsxwriter(temp_path, new_file_path, year_ranges)
            
            # Clean up temp file
            os.remove(temp_path)
            
        except Exception as e:
            raise Exception(f"Error processing file {file_path}: {e}")

    def format_ex1000(self, ws, values):
        """Format ex1000 sheet"""
        if ws.cell(row=1, column=1).value != '順位':
            ws.insert_rows(1, amount=3)
            ws.cell(row=1, column=1, value='順位')
            
            for col_num in range(2, ws.max_column + 1):
                ws.cell(row=1, column=col_num, value=col_num - 1)
            
            # Get the actual data from the Excel sheet to find correct column values
            try:
                # Load the Excel file that was processed
                tamagawa_file = os.path.join(self.base_dir, 'tamagawa-new method  - Copy .xlsx')
                if os.path.exists(tamagawa_file):
                    # Find the correct sheet name for this output file
                    sheet_name = self.get_sheet_name(os.path.basename(ws.parent.path) if hasattr(ws.parent, 'path') else 'default')
                    
                    # Load the sheet data
                    df = pd.read_excel(tamagawa_file, sheet_name=sheet_name)
                    
                    # Find column indices
                    kouzou_bangou_col = None
                    kouzou_meisho_col = None
                    eki_hajime_col = None
                    eki_shuuryou_col = None
                    
                    for i, col_name in enumerate(df.columns):
                        col_str = str(col_name).strip()
                        if '構造物番号' in col_str:
                            kouzou_bangou_col = i
                        elif '構造物名称' in col_str:
                            kouzou_meisho_col = i
                        elif '駅（始）' in col_str or '駅(始)' in col_str:
                            eki_hajime_col = i
                        elif '駅（至）' in col_str or '駅(至)' in col_str:
                            eki_shuuryou_col = i
                    
                    # Row 2: 構造物番号
                    ws.cell(row=2, column=1, value='構造物番号')
                    if kouzou_bangou_col is not None:
                        for col_num in range(2, min(ws.max_column + 1, len(df) + 2)):
                            row_idx = col_num - 2
                            if row_idx < len(df):
                                bangou_value = df.iloc[row_idx, kouzou_bangou_col]
                                if pd.notna(bangou_value):
                                    ws.cell(row=2, column=col_num, value=f'({bangou_value})')
                                else:
                                    ws.cell(row=2, column=col_num, value='(-)')
                    else:
                        # Fallback to using values from tamagawa data
                        for col_num, number in enumerate(values['構造物番号'], start=2):
                            ws.cell(row=2, column=col_num, value=f'({number})')
                    
                    # Row 3: 構造物名称 or 駅（始）→駅（至）
                    ws.cell(row=3, column=1, value='構造物名称')
                    
                    # First try to use 構造物名称
                    if kouzou_meisho_col is not None:
                        for col_num in range(2, min(ws.max_column + 1, len(df) + 2)):
                            row_idx = col_num - 2
                            if row_idx < len(df):
                                meisho_value = df.iloc[row_idx, kouzou_meisho_col]
                                if pd.notna(meisho_value) and str(meisho_value).strip():
                                    ws.cell(row=3, column=col_num, value=str(meisho_value))
                                else:
                                    # If 構造物名称 is empty, try 駅（始）→駅（至）
                                    if eki_hajime_col is not None and eki_shuuryou_col is not None:
                                        hajime = df.iloc[row_idx, eki_hajime_col]
                                        shuuryou = df.iloc[row_idx, eki_shuuryou_col]
                                        if pd.notna(hajime) and pd.notna(shuuryou):
                                            combined_name = f"{hajime}→{shuuryou}"
                                            ws.cell(row=3, column=col_num, value=combined_name)
                                        else:
                                            ws.cell(row=3, column=col_num, value='-')
                                    else:
                                        ws.cell(row=3, column=col_num, value='-')
                    
                    # If no 構造物名称 column, use 駅（始）→駅（至）
                    elif eki_hajime_col is not None and eki_shuuryou_col is not None:
                        for col_num in range(2, min(ws.max_column + 1, len(df) + 2)):
                            row_idx = col_num - 2
                            if row_idx < len(df):
                                hajime = df.iloc[row_idx, eki_hajime_col]
                                shuuryou = df.iloc[row_idx, eki_shuuryou_col]
                                if pd.notna(hajime) and pd.notna(shuuryou):
                                    combined_name = f"{hajime}→{shuuryou}"
                                    ws.cell(row=3, column=col_num, value=combined_name)
                                else:
                                    ws.cell(row=3, column=col_num, value='-')
                    else:
                        # Fallback to original values
                        for col_num, name in enumerate(values['構造物名称'], start=2):
                            ws.cell(row=3, column=col_num, value=name)
                            
            except Exception as e:
                print(f"Error reading Excel data for formatting: {e}")
                # Fallback to original method
                ws.cell(row=2, column=1, value='構造物番号')
                for col_num, number in enumerate(values['構造物番号'], start=2):
                    ws.cell(row=2, column=col_num, value=f'({number})')
                
                ws.cell(row=3, column=1, value='構造物名称')
                for col_num, name in enumerate(values['構造物名称'], start=2):
                    ws.cell(row=3, column=col_num, value=name)
        
        # Format data
        for row in range(4, ws.max_row + 1):
            try:
                ws.cell(row=row, column=1).value = int(float(ws.cell(row=row, column=1).value))
            except:
                pass
            for col in range(2, ws.max_column + 1):
                try:
                    ws.cell(row=row, column=col).value = float(ws.cell(row=row, column=col).value)
                except:
                    ws.cell(row=row, column=col).value = None

    def format_logdensity(self, ws):
        """Format logdensity sheet"""
        ws.insert_rows(1)
        for col in range(2, ws.max_column + 1):
            ws.cell(row=1, column=col, value=f'経過{col - 1}年目')
        
        for row in range(2, ws.max_row + 1):
            try:
                ws.cell(row=row, column=1).value = int(float(ws.cell(row=row, column=1).value))
            except:
                pass
            for col in range(2, ws.max_column + 1):
                try:
                    ws.cell(row=row, column=col).value = float(ws.cell(row=row, column=col).value)
                except:
                    ws.cell(row=row, column=col).value = None

    def format_pml(self, ws):
        """Format pml sheet"""
        ws.insert_cols(1)
        ws.insert_rows(1)
        ws.cell(row=1, column=1, value='経過年数')
        ws.cell(row=1, column=2, value='年')
        ws.cell(row=1, column=3, value='NEL (0.5:0.5)')
        ws.cell(row=1, column=4, value='PML(0.9:0.1)')
        ws.cell(row=1, column=5, value='PML_0.95 (0.95:0.05)')
        
        for i in range(2, ws.max_row + 1):
            ws.cell(row=i, column=1, value=i - 1)
        
        for col in range(2, ws.max_column + 1):
            for row in range(2, ws.max_row + 1):
                try:
                    ws.cell(row=row, column=col).value = int(float(ws.cell(row=row, column=col).value))
                except:
                    pass

    def create_charts_with_xlsxwriter(self, temp_path, new_file_path, year_ranges):
        """Create charts using xlsxwriter"""
        try:
            # Read data from temp file
            df_ex = None
            df_log = None
            df_pml = None
            
            try:
                df_ex = pd.read_excel(temp_path, sheet_name='ex1000', header=None)
            except:
                pass
            
            try:
                df_log = pd.read_excel(temp_path, sheet_name='logdensity', header=None)
            except:
                pass
            
            try:
                df_pml = pd.read_excel(temp_path, sheet_name='pml', header=0)
            except:
                pass

            # Create new workbook with charts
            with pd.ExcelWriter(new_file_path, engine='xlsxwriter') as writer:
                workbook = writer.book

                # Process ex1000 sheet with chart
                if df_ex is not None:
                    df_ex = df_ex.where(pd.notnull(df_ex), None)
                    ws = workbook.add_worksheet('ex1000')
                    writer.sheets['ex1000'] = ws

                    # Formatting
                    fmt_bold = workbook.add_format({'bold': True})
                    fmt_bold = workbook.add_format({'bold': True})
                    fmt_int = workbook.add_format({'num_format': '0'})
                    fmt_float = workbook.add_format({'num_format': '0.00000'})

                    # Write data
                    for row_num in range(df_ex.shape[0]):
                        for col_num in range(df_ex.shape[1]):
                            val = df_ex.iat[row_num, col_num]
                            if pd.isna(val):
                                ws.write(row_num, col_num, None)
                            elif row_num < 3:
                                ws.write(row_num, col_num, val, fmt_bold)
                            elif col_num == 0:
                                ws.write(row_num, col_num, val, fmt_int)
                            else:
                                ws.write(row_num, col_num, val, fmt_float)

                    # Create chart for ex1000
                    last = df_ex.iloc[3:].dropna(how='all').index[-1] if not df_ex.iloc[3:].dropna(how='all').empty else 3
                    chart = workbook.add_chart({'type': 'line'})
                    
                    for i in range(1, df_ex.shape[1]):
                        chart.add_series({
                            'name': ['ex1000', 2, i],
                            'categories': ['ex1000', 3, 0, last, 0],
                            'values': ['ex1000', 3, i, last, i],
                        })
                    
                    chart.set_title({'name': '経過年 vs. しきい値の強度確率'})
                    chart.set_x_axis({'name': '経過年', 'position_axis': 'on_tick'})
                    chart.set_y_axis({'name': 'しきい値の強度確率', 'num_format': '0%'})
                    ws.insert_chart(f'A{last + 6}', chart)

                # Process logdensity sheet with chart
                if df_log is not None:
                    df_log = df_log.where(pd.notnull(df_log), None)
                    df_log.to_excel(writer, sheet_name='logdensity', index=False, header=False)
                    ws2 = writer.sheets['logdensity']

                    # Formatting
                    col_a_format = workbook.add_format({'num_format': '0'})
                    col_rest_format = workbook.add_format({'num_format': '0.00000'})
                    ws2.set_column('A:A', 8, col_a_format)
                    end_col_letter = chr(ord('A') + df_log.shape[1] - 1)
                    ws2.set_column(f'B:{end_col_letter}', 12, col_rest_format)

                    # Create chart for logdensity
                    start_year, end_year = year_ranges
                    
                    if (start_year is not None and end_year is not None and 
                        start_year >= 1 and end_year <= df_log.shape[1] - 1 and 
                        start_year <= end_year):
                        
                        last2 = df_log.iloc[1:].dropna(how='all').index[-1] if not df_log.iloc[1:].dropna(how='all').empty else 1
                        chart2 = workbook.add_chart({'type': 'line'})
                        
                        for i in range(start_year, end_year + 1):
                            chart2.add_series({
                                'name': ['logdensity', 0, i],
                                'categories': ['logdensity', 1, 0, last2, 0],
                                'values': ['logdensity', 1, i, last2, i],
                            })
                        
                        chart2.set_title({'name': '経過年 vs. しきい値の強度確率'})
                        chart2.set_x_axis({'name': '劣化点数', 'position_axis': 'on_tick'})
                        chart2.set_y_axis({'name': '確率密度関数', 'num_format': '0.00000'})

                        last_col = chr(ord('A') + df_log.shape[1] - 1)
                        insert_col = chr(ord(last_col) + 3)
                        ws2.insert_chart(f'{insert_col}2', chart2)

                # Process pml sheet with chart
                if df_pml is not None:
                    df_pml = df_pml.where(pd.notnull(df_pml), None)
                    df_pml.to_excel(writer, sheet_name='pml', index=False, header=True)
                    ws3 = writer.sheets['pml']

                    # Formatting
                    col_a_format = workbook.add_format({'num_format': '0'})
                    col_rest_format = workbook.add_format({'num_format': '0'})
                    ws3.set_column('A:A', 8, col_a_format)
                    end_col_letter = chr(ord('A') + df_pml.shape[1] - 1)
                    ws3.set_column(f'B:{end_col_letter}', 12, col_rest_format)

                    # Create chart for pml
                    last3 = df_pml.shape[0]
                    chart3 = workbook.add_chart({'type': 'line'})
                    
                    for i in range(2, df_pml.shape[1]):
                        chart3.add_series({
                            'name': ['pml', 0, i],
                            'categories': ['pml', 1, 0, last3, 0],
                            'values': ['pml', 1, i, last3, i],
                            'marker': {'type': 'circle'}
                        })
                    
                    chart3.set_title({'name': 'PML Data'})
                    chart3.set_x_axis({'name': '経過年数', 'position_axis': 'on_tick'})
                    chart3.set_y_axis({'name': '劣化点数'})

                    last_col_pml = chr(ord('A') + df_pml.shape[1] - 1)
                    insert_col_pml = chr(ord(last_col_pml) + 3)
                    ws3.insert_chart(f'{insert_col_pml}2', chart3)

        except Exception as e:
            raise Exception(f"Error creating charts: {e}")

    def run(self):
        """Run the application"""
        self.root.mainloop()


if __name__ == "__main__":
    print("Post-Obser Files Processor")
    print("=" * 50)
    print("🚀 Advanced Processing Tool")
    print("=" * 50)
    print("This tool processes obser files generated by the Obser File Generator")
    print("and creates enhanced Excel outputs with charts using Fortran programs.")
    print()
    print("📋 Requirements:")
    print("• obser1.txt to obser8.txt (from Obser File Generator)")
    print("• 劣化予測プログラム .exe (Fortran program)")
    print("• tamagawa-new method  - Copy .xlsx (structure data, for charts)")
    print("• Any additional txt files required by Fortran program")
    print()
    print("📊 Output:")
    print("• 8 basic Excel files (出力1.xlsx - 出力8.xlsx)")
    print("• 8 chart-enhanced files (作図付き出力1.xlsx - 作図付き出力8.xlsx)")
    print("=" * 50)
    print()
    
    app = PostObserProcessorApp()
    app.run()