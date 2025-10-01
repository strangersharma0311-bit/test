import pandas as pd
import openpyxl
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import re

class EnhancedKeijihenkaGeneratorApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Enhanced 経時変化 Sheets Generator")
        self.root.geometry("500x400")
        self.root.minsize(450, 350)
        
        # Center the window on screen
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() // 2) - (500 // 2)
        y = (self.root.winfo_screenheight() // 2) - (400 // 2)
        self.root.geometry(f"500x400+{x}+{y}")
        
        self.workbook_path = None
        self.structure_df = None
        self.operator_df = None
        
        self.create_main_gui()
    
    def abbreviate_sen_name(self, sen_name):
        """Convert route name to abbreviation"""
        if pd.isna(sen_name) or sen_name == '':
            return ''
        
        sen_name = str(sen_name).strip()
        
        abbreviation_map = {
            "東急多摩川線": "TM",
            "多摩川線": "TM", 
            "東横線": "TY",
            "大井町線": "OM",
            "池上線": "IK",
            "田園都市線": "DT",
            "目黒線": "MG",
            "こどもの国線": "KD",
            "世田谷線": "SG"
        }
        
        return abbreviation_map.get(sen_name, sen_name)
    
    def lookup_structure_number(self, structure_df, rosen_name, kozo_name, ekikan):
        """Lookup 構造物番号 from structure sheet"""
        try:
            if structure_df is None or len(structure_df) == 0:
                return ''
                
            rosen_name = str(rosen_name).strip() if pd.notna(rosen_name) else ''
            
            # First try to match by structure name
            if kozo_name and str(kozo_name).strip() not in ['', 'nan', 'NaN']:
                kozo_name = str(kozo_name).strip()
                matches = structure_df[
                    (structure_df['構造物名称'].astype(str).str.strip() == kozo_name) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    bangou = matches.iloc[0]['構造物番号']
                    if pd.notna(bangou) and str(bangou).strip() not in ['', 'nan']:
                        return str(bangou).strip()
            
            # If not found by structure name, try by station interval
            if ekikan and str(ekikan).strip() not in ['', 'nan', 'NaN']:
                ekikan = str(ekikan).strip()
                matches = structure_df[
                    (structure_df['駅間'].astype(str).str.strip() == ekikan) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    bangou = matches.iloc[0]['構造物番号']
                    if pd.notna(bangou) and str(bangou).strip() not in ['', 'nan']:
                        return str(bangou).strip()
            
            return ''
            
        except Exception as e:
            print(f"Error finding structure number: {e}")
            return ''
    
    def add_enhanced_columns(self, df, structure_df=None):
        """Add enhanced columns: 路線名略称 and 構造物番号"""
        enhanced_df = df.copy()
        
        # Add 路線名略称 column
        if '路線名' in enhanced_df.columns:
            enhanced_df['路線名略称'] = enhanced_df['路線名'].apply(self.abbreviate_sen_name)
        else:
            enhanced_df['路線名略称'] = ''
        
        # Add 構造物番号 column
        enhanced_df['構造物番号'] = ''
        
        if structure_df is not None:
            for index, row in enhanced_df.iterrows():
                rosen_name = row.get('路線名', '')
                kozo_name = row.get('構造物名称', '')
                
                # Create ekikan for lookup
                ekikan = ''
                if row.get('駅（始）', '') and row.get('駅（至）', ''):
                    ekikan = f"{row.get('駅（始）', '')}→{row.get('駅（至）', '')}"
                
                # Lookup structure number
                bangou = self.lookup_structure_number(structure_df, rosen_name, kozo_name, ekikan)
                enhanced_df.at[index, '構造物番号'] = bangou
        
        return enhanced_df
    
    def reorder_columns_enhanced(self, df):
        """Reorder columns: グループ化キー → グループ化方法 → 種別 → 構造物名称 → 駅（始） → 駅（至） → 点検区分1 → データ件数 → 路線名 → 路線名略称 → 構造物番号 → years"""
        
        # Define the correct enhanced column order
        priority_columns = [
            'グループ化キー',
            'グループ化方法', 
            '種別',
            '構造物名称',
            '駅（始）',
            '駅（至）',
            '点検区分1',
            'データ件数',
            '路線名',
            '路線名略称',
            '構造物番号'
        ]
        
        # Get year columns (various formats for 経時変化)
        year_columns = []
        
        for col in df.columns:
            if any(year in str(col) for year in ['2018', '2019', '2020', '2021', '2022', '2023', '2024']):
                year_columns.append(col)
        
        # Sort year columns chronologically
        def extract_year(col_name):
            try:
                for year in ['2018', '2019', '2020', '2021', '2022', '2023', '2024']:
                    if year in str(col_name):
                        return int(year)
                return 0
            except:
                return 0
        
        year_columns.sort(key=extract_year)
        
        # Create final column order
        final_columns = []
        
        # Add priority columns that exist
        for col in priority_columns:
            if col in df.columns:
                final_columns.append(col)
        
        # Add year columns
        final_columns.extend(year_columns)
        
        # Add any remaining columns that weren't in priority or year columns
        remaining_columns = [col for col in df.columns if col not in final_columns]
        final_columns.extend(remaining_columns)
        
        # Return reordered dataframe
        return df[final_columns]
    
    def create_main_gui(self):
        """Create main GUI for file selection"""
        main_frame = tk.Frame(self.root, padx=30, pady=30)
        main_frame.pack(fill="both", expand=True)
        
        # Title
        title_label = tk.Label(main_frame, text="Enhanced 経時変化 Sheets Generator", 
                              font=("Arial", 14, "bold"), fg="navy")
        title_label.pack(pady=(0, 15))
        
        # Instructions
        instruction_text = ("Enhanced time-series analysis sheets:\n"
                          "• 経時変化（橋長考慮） with enhanced columns\n"
                          "• 経時変化（橋長&形式考慮） with enhanced columns\n" 
                          "• 経時変化（橋長無視&形式考慮） with enhanced columns\n\n"
                          "Features: Dynamic weight application, 路線名略称, 構造物番号")
        instruction_label = tk.Label(main_frame, text=instruction_text, 
                                   font=("Arial", 10), justify="center")
        instruction_label.pack(pady=(0, 15))
        
        # Status label
        self.status_label = tk.Label(main_frame, text="Ready...", 
                                    font=("Arial", 9), fg="gray")
        self.status_label.pack(pady=(0, 10))
        
        # Select file button
        select_btn = tk.Button(main_frame, text="Browse & Select File", 
                             command=self.select_workbook, 
                             bg="#4CAF50", fg="white", 
                             width=20, height=1, font=("Arial", 10))
        select_btn.pack(pady=8)
        
        # Generate button (initially disabled)
        self.generate_btn = tk.Button(main_frame, text="Generate Enhanced 経時変化 Sheets", 
                                    command=self.generate_keijiheka_sheets, 
                                    bg="#FF9800", fg="white", 
                                    width=28, height=1, font=("Arial", 10),
                                    state="disabled")
        self.generate_btn.pack(pady=8)
        
        # Exit button
        exit_btn = tk.Button(main_frame, text="Exit", 
                           command=self.root.quit, bg="#f44336", fg="white", 
                           width=12, height=1, font=("Arial", 9))
        exit_btn.pack(pady=(15, 0))

    def select_workbook(self):
        """Select workbook with required sheets"""
        self.status_label.config(text="Opening browser...", fg="blue")
        self.root.update()
        
        # File selection
        self.workbook_path = filedialog.askopenfilename(
            title="Select Excel Workbook",
            filetypes=[("Excel files", "*.xlsx *.xls")],
            initialdir=os.path.expanduser("~")
        )
        
        if not self.workbook_path:
            self.status_label.config(text="No file selected", fg="orange")
            return

        self.status_label.config(text="Validating sheets...", fg="blue")
        self.root.update()
        
        # Validate required sheets
        try:
            wb = load_workbook(self.workbook_path)
            actual_sheets = [sheet.strip() for sheet in wb.sheetnames]
            print(f"Found sheets: {actual_sheets}")
            
            # Find the correct 演算子 sheet name with flexible matching
            operator_sheet_name = None
            for sheet_name in actual_sheets:
                if '演算子' in sheet_name and '2' in sheet_name:
                    operator_sheet_name = sheet_name
                    break
            
            if operator_sheet_name is None:
                self.status_label.config(text="演算子‐2 sheet not found!", fg="red")
                messagebox.showerror("Missing Sheet", "Could not find 演算子‐2 sheet (or similar)")
                return
            
            # Check other required sheets
            required_sheets = ['グループ化点検履歴', '構造物番号']
            missing_sheets = [sheet for sheet in required_sheets if sheet not in actual_sheets]
            
            if missing_sheets:
                self.status_label.config(text="Missing sheets!", fg="red")
                available_sheets_str = '\n'.join(actual_sheets)
                messagebox.showerror("Missing Sheets", 
                                f"Required sheets not found: {', '.join(missing_sheets)}\n\n"
                                f"Available sheets:\n{available_sheets_str}")
                return
            
            # Try to load required data for enhancements
            try:
                self.structure_df = pd.read_excel(self.workbook_path, sheet_name='構造物番号')
                # Use the dynamically found operator sheet name
                self.operator_df = pd.read_excel(self.workbook_path, sheet_name=operator_sheet_name)
                print(f"Successfully loaded operator sheet: '{operator_sheet_name}'")
            except Exception as e:
                self.status_label.config(text="Error loading data!", fg="red")
                messagebox.showerror("Error", f"Error loading required sheets:\n{str(e)}")
                return
            
            enhancement_status = "with enhancements" if self.structure_df is not None else "basic version"
            self.status_label.config(text="Ready to generate!", fg="green")
            self.generate_btn.config(state="normal")
            messagebox.showinfo("Success", f"All required sheets found ({enhancement_status})!\nReady to generate enhanced 経時変化 sheets.")
            
        except Exception as e:
            self.status_label.config(text="Error", fg="red")
            messagebox.showerror("Error", f"Error validating file:\n{str(e)}")

    def generate_keijiheka_sheets(self):
        """Generate enhanced 経時変化 sheets"""
        try:
            # Show progress dialog
            progress_window = tk.Toplevel(self.root)
            progress_window.title("Generating Enhanced 経時変化 Sheets")
            progress_window.geometry("400x120")
            progress_window.grab_set()
            progress_window.resizable(False, False)
            progress_window.transient(self.root)
            
            # Center the progress window
            progress_window.update_idletasks()
            x = (progress_window.winfo_screenwidth() // 2) - (400 // 2)
            y = (progress_window.winfo_screenheight() // 2) - (120 // 2)
            progress_window.geometry(f"400x120+{x}+{y}")
            
            progress_frame = tk.Frame(progress_window, padx=20, pady=20)
            progress_frame.pack(fill="both", expand=True)
            
            status_label = tk.Label(progress_frame, text="Processing enhanced 経時変化...", font=("Arial", 10))
            status_label.pack(pady=5)
            
            progress_bar = ttk.Progressbar(progress_frame, mode='indeterminate')
            progress_bar.pack(fill="x", pady=5)
            progress_bar.start()
            
            # Execute process
            self.root.after(100, lambda: self.execute_keijiheka_process(progress_window))
            
        except Exception as e:
            messagebox.showerror("Error", f"Error starting process:\n{str(e)}")

    def execute_keijiheka_process(self, progress_window):
        """Execute the enhanced 経時変化 process"""
        try:
            # Load required sheets
            grouped_df = pd.read_excel(self.workbook_path, sheet_name='グループ化点検履歴')
            structure_df = pd.read_excel(self.workbook_path, sheet_name='構造物番号')
            operator_df = pd.read_excel(self.workbook_path, sheet_name='演算子‐2')
            
            # Create enhanced 経時変化 results
            # Sheet 1: 経時変化（橋長考慮） - グループ化点検履歴 ÷ Length
            keiji_kyoucho_df = self.apply_enhanced_keiji_kyoucho_logic(grouped_df, structure_df)
            
            # Sheet 2: 経時変化（橋長&形式考慮） - Above × Structure weights
            keiji_both_df = self.apply_enhanced_keiji_both_logic(keiji_kyoucho_df, structure_df, operator_df)
            
            # Sheet 3: 経時変化（橋長無視&形式考慮） - グループ化点検履歴 × Structure weights
            keiji_mushi_df = self.apply_enhanced_keiji_mushi_logic(grouped_df, structure_df, operator_df)
            
            # Save to Excel
            self.save_enhanced_keijiheka_results(keiji_kyoucho_df, keiji_both_df, keiji_mushi_df)
            
            # Close progress window
            progress_window.destroy()
            
            # Show completion dialog
            self.show_enhanced_completion_dialog()
            
        except Exception as e:
            progress_window.destroy()
            messagebox.showerror("Error", f"Error during processing:\n{str(e)}")

    def apply_enhanced_keiji_kyoucho_logic(self, grouped_df, structure_df):
        """Apply enhanced 経時変化（橋長考慮） logic: グループ化点検履歴 ÷ Length"""
        result_df = grouped_df.copy()
        
        # Find year result columns
        year_columns = [col for col in result_df.columns if col.endswith('結果')]
        
        print(f"Processing enhanced 経時変化（橋長考慮） with {len(result_df)} rows")
        
        # Apply division by length for each row
        for index, row in result_df.iterrows():
            # Get length value from structure data
            length_value = self.get_length_value_for_keiji(structure_df, row)
            
            # Apply division to each year column
            for year_col in year_columns:
                original_value = row[year_col]
                
                if pd.notna(original_value) and str(original_value).strip() not in ['', 'nan']:
                    try:
                        numeric_value = float(original_value)
                        
                        # Divide by length if available
                        if length_value and length_value > 0:
                            divided_value = numeric_value / length_value
                            result_df.at[index, year_col] = round(divided_value, 3)
                        else:
                            # Handle missing length data - keep original value
                            result_df.at[index, year_col] = numeric_value
                        
                    except (ValueError, TypeError):
                        result_df.at[index, year_col] = original_value
                else:
                    result_df.at[index, year_col] = original_value
        
        # Add enhanced columns
        enhanced_df = self.add_enhanced_columns(result_df, self.structure_df)
        
        # Reorder columns
        final_df = self.reorder_columns_enhanced(enhanced_df)
        
        return final_df

    def apply_enhanced_keiji_both_logic(self, keiji_kyoucho_df, structure_df, operator_df):
        """Apply enhanced 経時変化（橋長&形式考慮） logic: Above × Structure weights"""
        result_df = keiji_kyoucho_df.copy()
        
        # Find year result columns
        year_columns = [col for col in result_df.columns if col.endswith('結果')]
        
        print(f"Processing enhanced 経時変化（橋長&形式考慮） with {len(result_df)} rows")
        
        # Apply structure weights multiplication for each row
        for index, row in result_df.iterrows():
            # Get structure weights using 演算子‐2 formulas
            weights = self.get_structure_weights_with_operator(structure_df, operator_df, row)
            
            # Apply multiplication to each year column
            for year_col in year_columns:
                original_value = row[year_col]
                
                if pd.notna(original_value) and str(original_value).strip() not in ['', 'nan']:
                    try:
                        numeric_value = float(original_value)
                        
                        # Multiply by structure weights
                        calculated_value = numeric_value * weights['total_weight']
                        result_df.at[index, year_col] = round(calculated_value, 3)
                        
                    except (ValueError, TypeError):
                        result_df.at[index, year_col] = original_value
                else:
                    result_df.at[index, year_col] = original_value
        
        # Add enhanced columns
        enhanced_df = self.add_enhanced_columns(result_df, self.structure_df)
        
        # Reorder columns
        final_df = self.reorder_columns_enhanced(enhanced_df)
        
        return final_df

    def apply_enhanced_keiji_mushi_logic(self, grouped_df, structure_df, operator_df):
        """Apply enhanced 経時変化（橋長無視&形式考慮） logic: グループ化点検履歴 × Structure weights"""
        result_df = grouped_df.copy()
        
        # Find year result columns
        year_columns = [col for col in result_df.columns if col.endswith('結果')]
        
        print(f"Processing enhanced 経時変化（橋長無視&形式考慮） with {len(result_df)} rows")
        
        # Apply structure weights multiplication for each row
        for index, row in result_df.iterrows():
            # Get structure weights using 演算子‐2 formulas
            weights = self.get_structure_weights_with_operator(structure_df, operator_df, row)
            
            # Apply multiplication to each year column
            for year_col in year_columns:
                original_value = row[year_col]
                
                if pd.notna(original_value) and str(original_value).strip() not in ['', 'nan']:
                    try:
                        numeric_value = float(original_value)
                        
                        # Multiply by structure weights
                        calculated_value = numeric_value * weights['total_weight']
                        result_df.at[index, year_col] = round(calculated_value, 3)
                        
                    except (ValueError, TypeError):
                        result_df.at[index, year_col] = original_value
                else:
                    result_df.at[index, year_col] = original_value
        
        # Add enhanced columns
        enhanced_df = self.add_enhanced_columns(result_df, self.structure_df)
        
        # Reorder columns
        final_df = self.reorder_columns_enhanced(enhanced_df)
        
        return final_df

    def get_length_value_for_keiji(self, structure_df, row):
        """Get length value from structure data for 経時変化"""
        try:
            rosen_name = str(row.get('路線名', '')).strip() if pd.notna(row.get('路線名', '')) else ''
            kozo_name = str(row.get('構造物名称', '')).strip() if pd.notna(row.get('構造物名称', '')) else ''
            
            # Try to construct ekikan from 駅（始） and 駅（至）
            eki_start = str(row.get('駅（始）', '')).strip() if pd.notna(row.get('駅（始）', '')) else ''
            eki_end = str(row.get('駅（至）', '')).strip() if pd.notna(row.get('駅（至）', '')) else ''
            
            ekikan = ''
            if eki_start and eki_end:
                ekikan = f"{eki_start}→{eki_end}"
            
            # First try to match by structure name
            if kozo_name:
                matches = structure_df[
                    (structure_df['構造物名称'].astype(str).str.strip() == kozo_name) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    length_val = matches.iloc[0]['長さ(m)']
                    if pd.notna(length_val) and str(length_val).strip() not in ['', 'nan']:
                        try:
                            return float(length_val)
                        except (ValueError, TypeError):
                            pass
            
            # If not found by structure name, try by station interval
            if ekikan:
                matches = structure_df[
                    (structure_df['駅間'].astype(str).str.strip() == ekikan) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    length_val = matches.iloc[0]['長さ(m)']
                    if pd.notna(length_val) and str(length_val).strip() not in ['', 'nan']:
                        try:
                            return float(length_val)
                        except (ValueError, TypeError):
                            pass
            
            # Return None if length not found
            print(f"Length not found for 経時変化: {kozo_name or ekikan} in {rosen_name}")
            return None
            
        except Exception as e:
            print(f"Error finding length value for 経時変化: {e}")
            return None

    def get_structure_weights_with_operator(self, structure_df, operator_df, row):
        """Get structure weights using 演算子‐2 formulas with A1, B1, C1... mapping"""
        try:
            rosen_name = str(row.get('路線名', '')).strip() if pd.notna(row.get('路線名', '')) else ''
            kozo_name = str(row.get('構造物名称', '')).strip() if pd.notna(row.get('構造物名称', '')) else ''
            
            # Try to construct ekikan from 駅（始） and 駅（至）
            eki_start = str(row.get('駅（始）', '')).strip() if pd.notna(row.get('駅（始）', '')) else ''
            eki_end = str(row.get('駅（至）', '')).strip() if pd.notna(row.get('駅（至）', '')) else ''
            
            ekikan = ''
            if eki_start and eki_end:
                ekikan = f"{eki_start}→{eki_end}"
            
            # Get values from structure data
            structure_values = {'A1': 1.0, 'B1': 1.0, 'C1': 1.0}  # Default values
            
            # First try to match by structure name
            if kozo_name:
                matches = structure_df[
                    (structure_df['構造物名称'].astype(str).str.strip() == kozo_name) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    match_row = matches.iloc[0]
                    
                    # Map structure data to A1, B1, C1 for 演算子‐2 formulas
                    if '構造形式_重み' in match_row and pd.notna(match_row['構造形式_重み']):
                        try:
                            structure_values['A1'] = float(match_row['構造形式_重み'])
                        except (ValueError, TypeError):
                            pass
                    
                    if '角度_重み' in match_row and pd.notna(match_row['角度_重み']):
                        try:
                            structure_values['B1'] = float(match_row['角度_重み'])
                        except (ValueError, TypeError):
                            pass
                    
                    if '供用年数_重み' in match_row and pd.notna(match_row['供用年数_重み']):
                        try:
                            structure_values['C1'] = float(match_row['供用年数_重み'])
                        except (ValueError, TypeError):
                            pass
            
            # If not found by structure name, try by station interval
            elif ekikan:
                matches = structure_df[
                    (structure_df['駅間'].astype(str).str.strip() == ekikan) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    match_row = matches.iloc[0]
                    
                    # Map structure data to A1, B1, C1 for 演算子‐2 formulas
                    if '構造形式_重み' in match_row and pd.notna(match_row['構造形式_重み']):
                        try:
                            structure_values['A1'] = float(match_row['構造形式_重み'])
                        except (ValueError, TypeError):
                            pass
                    
                    if '角度_重み' in match_row and pd.notna(match_row['角度_重み']):
                        try:
                            structure_values['B1'] = float(match_row['角度_重み'])
                        except (ValueError, TypeError):
                            pass
                    
                    if '供用年数_重み' in match_row and pd.notna(match_row['供用年数_重み']):
                        try:
                            structure_values['C1'] = float(match_row['供用年数_重み'])
                        except (ValueError, TypeError):
                            pass
            
            # Apply 演算子‐2 formulas with dynamic evaluation
            total_weight = self.evaluate_operator_formulas(operator_df, structure_values)
            
            return {
                'A1': structure_values['A1'],
                'B1': structure_values['B1'], 
                'C1': structure_values['C1'],
                'total_weight': total_weight
            }
            
        except Exception as e:
            print(f"Error getting structure weights with operator: {e}")
            # Return default values if error
            return {'A1': 1.0, 'B1': 1.0, 'C1': 1.0, 'total_weight': 1.0}

    def evaluate_operator_formulas(self, operator_df, structure_values):
        """Evaluate 演算子‐2 formulas with A1, B1, C1... mapping"""
        try:
            # Find the formula in 演算子‐2 sheet
            if len(operator_df) > 0:
                # Assume first row contains the formula
                formula_row = operator_df.iloc[0]
                
                # Look for formula column (usually contains formula like "A1*B1*C1")
                formula = None
                for col in operator_df.columns:
                    cell_value = str(formula_row[col]).strip()
                    if any(var in cell_value for var in ['A1', 'B1', 'C1']) and any(op in cell_value for op in ['*', '+', '-', '/']):
                        formula = cell_value
                        break
                
                if formula:
                    # Replace A1, B1, C1 with actual values
                    formula = formula.replace('A1', str(structure_values['A1']))
                    formula = formula.replace('B1', str(structure_values['B1']))
                    formula = formula.replace('C1', str(structure_values['C1']))
                    
                    # Safely evaluate the formula
                    try:
                        result = eval(formula)
                        return float(result)
                    except:
                        print(f"Error evaluating formula: {formula}")
                        return 1.0
            
            # Default calculation if no formula found
            return structure_values['A1'] * structure_values['B1'] * structure_values['C1']
            
        except Exception as e:
            print(f"Error evaluating operator formulas: {e}")
            return 1.0

    def save_enhanced_keijiheka_results(self, keiji_kyoucho_df, keiji_both_df, keiji_mushi_df):
        """Save enhanced 経時変化 results to Excel sheets"""
        try:
            with pd.ExcelWriter(self.workbook_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                # Write enhanced 経時変化 result sheets
                keiji_kyoucho_df.to_excel(writer, sheet_name='経時変化（橋長考慮）', index=False)
                keiji_both_df.to_excel(writer, sheet_name='経時変化（橋長&形式考慮）', index=False)
                keiji_mushi_df.to_excel(writer, sheet_name='経時変化（橋長無視&形式考慮）', index=False)
                
                # Preserve other sheets
                try:
                    original_wb = load_workbook(self.workbook_path)
                    sheet_names_to_preserve = [name for name in original_wb.sheetnames 
                                             if name not in ['経時変化（橋長考慮）', '経時変化（橋長&形式考慮）', 
                                                           '経時変化（橋長無視&形式考慮）']]
                    
                    for sheet_name in sheet_names_to_preserve:
                        try:
                            df_temp = pd.read_excel(self.workbook_path, sheet_name=sheet_name)
                            df_temp.to_excel(writer, sheet_name=sheet_name, index=False)
                        except Exception as e:
                            continue
                except Exception as e:
                    pass
                    
        except Exception as e:
            raise Exception(f"Error saving enhanced 経時変化 results: {str(e)}")

    def show_enhanced_completion_dialog(self):
        """Show enhanced completion dialog"""
        completion_window = tk.Toplevel(self.root)
        completion_window.title("Enhanced 経時変化 Complete")
        completion_window.geometry("450x400")
        completion_window.grab_set()
        completion_window.resizable(False, False)
        completion_window.transient(self.root)
        
        # Center window
        completion_window.update_idletasks()
        x = (completion_window.winfo_screenwidth() // 2) - (450 // 2)
        y = (completion_window.winfo_screenheight() // 2) - (400 // 2)
        completion_window.geometry(f"450x400+{x}+{y}")
        
        main_frame = tk.Frame(completion_window, padx=15, pady=15)
        main_frame.pack(fill="both", expand=True)
        
        # Title
        title_label = tk.Label(main_frame, text="Enhanced 経時変化 Complete!", 
                              font=("Arial", 12, "bold"), fg="green")
        title_label.pack(pady=(0, 10))
        
        # Enhanced features info
        features_text = ("✅ Enhanced Features Applied:\n\n"
                        "• 路線名略称 column added\n"
                        "• 構造物番号 column added\n"
                        "• Proper column ordering\n"
                        "• Dynamic weight application using 演算子‐2\n"
                        "• Length division and structure weight calculations")
        features_label = tk.Label(main_frame, text=features_text, font=("Arial", 10), 
                                 justify="left", fg="blue")
        features_label.pack(pady=(0, 10))
        
        # Processing info
        info_text = ("3 Enhanced sheets created:\n"
                    "• 経時変化（橋長考慮）\n"
                    "• 経時変化（橋長&形式考慮）\n"
                    "• 経時変化（橋長無視&形式考慮）")
        info_label = tk.Label(main_frame, text=info_text, font=("Arial", 10))
        info_label.pack(pady=(0, 10))
        
        # Calculation details
        calc_text = ("Calculations applied:\n"
                    "1. 橋長考慮: グループ化点検履歴 ÷ Length\n"
                    "2. 橋長&形式考慮: Above × Structure weights\n"
                    "3. 橋長無視&形式考慮: グループ化点検履歴 × Structure weights\n"
                    "Using 演算子‐2 formulas with A1, B1, C1 mapping")
        calc_label = tk.Label(main_frame, text=calc_text, font=("Arial", 9), fg="darkgreen")
        calc_label.pack(pady=(0, 15))
        
        # Enhancement status
        enhancement_status = "with 構造物番号 lookup" if self.structure_df is not None else "basic version"
        status_label = tk.Label(main_frame, text=f"Enhancement: {enhancement_status}", 
                               font=("Arial", 9), fg="green" if self.structure_df is not None else "orange")
        status_label.pack(pady=(0, 15))
        
        # Buttons
        button_frame = tk.Frame(main_frame)
        button_frame.pack()
        
        def open_excel():
            try:
                import os
                os.startfile(self.workbook_path)
                messagebox.showinfo("Excel Opened", 
                                  "✅ Enhanced Excel file opened!\n\n"
                                  "Check the 3 new enhanced 経時変化 sheets:\n"
                                  "• 経時変化（橋長考慮）\n"
                                  "• 経時変化（橋長&形式考慮）\n"
                                  "• 経時変化（橋長無視&形式考慮）\n\n"
                                  "With 路線名略称 and 構造物番号 columns!")
                completion_window.after(1000, completion_window.destroy)
                self.root.after(2000, self.root.quit)
            except:
                messagebox.showinfo("Info", f"Please open file manually:\n{self.workbook_path}")
                completion_window.destroy()

        def close_only():
            completion_window.destroy()
            messagebox.showinfo("Complete", 
                              "✅ Enhanced 経時変化 processing completed!\n\n"
                              "3 time-series analysis sheets created with:\n"
                              "• 路線名略称 columns\n"
                              "• 構造物番号 columns\n"
                              "• Proper column ordering\n"
                              "• Dynamic weight application\n"
                              "• 演算子‐2 formula evaluation")
            self.root.after(1000, self.root.quit)
        
        excel_btn = tk.Button(button_frame, text="Open Enhanced Excel", 
                            command=open_excel, bg="#4CAF50", fg="white", 
                            width=15, height=1, font=("Arial", 10))
        excel_btn.pack(side="left", padx=5)
        
        close_btn = tk.Button(button_frame, text="Complete", 
                            command=close_only, bg="#2196F3", fg="white", 
                            width=12, height=1, font=("Arial", 10))
        close_btn.pack(side="left", padx=5)

    def run(self):
        """Run the enhanced application"""
        self.root.mainloop()


# Main execution
if __name__ == "__main__":
    print("Enhanced 経時変化 Sheets Generator Starting...")
    print("=" * 60)
    print("🚀 Enhanced Features:")
    print("• 路線名略称 column (TM, TY, OM, IK, DT, MG, KD, SG)")
    print("• 構造物番号 column (auto-lookup from 構造物番号 sheet)")
    print("• Column order: グループ化キー → グループ化方法 → 種別 → 構造物名称 → 駅（始） → 駅（至） → 点検区分1 → データ件数 → 路線名 → 路線名略称 → 構造物番号 → years")
    print("• Dynamic weight application using 演算子‐2 formulas")
    print("• A1, B1, C1 mapping with structure weights")
    print("• Length division and structure weight calculations")
    print("=" * 60)
    print("Required input:")
    print("• グループ化点検履歴 sheet")
    print("• 構造物番号 sheet (for weights, length values and enhanced features)")
    print("• 演算子‐2 sheet (for dynamic formula evaluation)")
    print("=" * 60)
    print("Output:")
    print("• 経時変化（橋長考慮） sheet (グループ化点検履歴 ÷ Length)")
    print("• 経時変化（橋長&形式考慮） sheet (Above × Structure weights)")
    print("• 経時変化（橋長無視&形式考慮） sheet (グループ化点検履歴 × Structure weights)")
    print("=" * 60)
    
    app = EnhancedKeijihenkaGeneratorApp()
    app.run()