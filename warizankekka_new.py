import pandas as pd
import openpyxl
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, filedialog
import os
import re
import warnings

# Suppress pandas warnings
warnings.filterwarnings("ignore", category=FutureWarning)

class EnhancedDivisionSheetsApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Excel Processor Pro")
        self.root.geometry("600x500")
        self.root.minsize(600, 500)
        
        # Center the window on screen
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() // 2) - (600 // 2)
        y = (self.root.winfo_screenheight() // 2) - (500 // 2)
        self.root.geometry(f"600x500+{x}+{y}")
        
        # Set background color
        self.root.configure(bg='white')
        
        self.workbook_path = None
        self.structure_df = None
        
        self.create_professional_gui()
    
    def abbreviate_sen_name(self, sen_name):
        """Convert route name to abbreviation"""
        if pd.isna(sen_name) or sen_name == '':
            return ''
        
        sen_name = str(sen_name).strip()
        
        abbreviation_map = {
            "東急多摩川線": "TM",
            "多摩川線": "TM", 
            "東横線": "TY",
            "大井町線": "OM",
            "池上線": "IK",
            "田園都市線": "DT",
            "目黒線": "MG",
            "こどもの国線": "KD",
            "世田谷線": "SG"
        }
        
        return abbreviation_map.get(sen_name, sen_name)
    
    def lookup_structure_number(self, structure_df, rosen_name, kozo_name, ekikan):
        """Lookup 構造物番号 from structure sheet"""
        try:
            if structure_df is None or len(structure_df) == 0:
                return ''
                
            rosen_name = str(rosen_name).strip() if pd.notna(rosen_name) else ''
            
            # First try to match by structure name
            if kozo_name and str(kozo_name).strip() not in ['', 'nan', 'NaN']:
                kozo_name = str(kozo_name).strip()
                matches = structure_df[
                    (structure_df['構造物名称'].astype(str).str.strip() == kozo_name) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    bangou = matches.iloc[0]['構造物番号']
                    if pd.notna(bangou) and str(bangou).strip() not in ['', 'nan']:
                        return str(bangou).strip()
            
            # If not found by structure name, try by station interval
            if ekikan and str(ekikan).strip() not in ['', 'nan', 'NaN']:
                ekikan = str(ekikan).strip()
                matches = structure_df[
                    (structure_df['駅間'].astype(str).str.strip() == ekikan) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    bangou = matches.iloc[0]['構造物番号']
                    if pd.notna(bangou) and str(bangou).strip() not in ['', 'nan']:
                        return str(bangou).strip()
            
            return ''
            
        except Exception:
            return ''
    
    def add_enhanced_columns(self, df, structure_df=None):
        """Add enhanced columns: 路線名略称 and 構造物番号"""
        enhanced_df = df.copy()
        
        # Add 路線名略称 column
        if '路線名' in enhanced_df.columns:
            enhanced_df['路線名略称'] = enhanced_df['路線名'].apply(self.abbreviate_sen_name)
        else:
            enhanced_df['路線名略称'] = ''
        
        # Add 構造物番号 column
        enhanced_df['構造物番号'] = ''
        
        if structure_df is not None:
            for index, row in enhanced_df.iterrows():
                rosen_name = row.get('路線名', '')
                kozo_name = row.get('構造物名称', '')
                
                # Create ekikan for lookup
                ekikan = ''
                if row.get('駅（始）', '') and row.get('駅（至）', ''):
                    ekikan = f"{row.get('駅（始）', '')}→{row.get('駅（至）', '')}"
                
                # Lookup structure number
                bangou = self.lookup_structure_number(structure_df, rosen_name, kozo_name, ekikan)
                enhanced_df.at[index, '構造物番号'] = bangou
        
        return enhanced_df
    
    def reorder_columns_enhanced(self, df):
        """Reorder columns"""
        priority_columns = [
            'グループ化キー', 'グループ化方法', '種別', '構造物名称',
            '駅（始）', '駅（至）', '点検区分1', 'データ件数',
            '路線名', '路線名略称', '構造物番号'
        ]
        
        # Get year columns
        year_columns = []
        for col in df.columns:
            if '/長さ' in str(col) or str(col).endswith('結果') or any(year in str(col) for year in ['2018', '2019', '2020', '2021', '2022', '2023', '2024']):
                year_columns.append(col)
        
        # Sort year columns chronologically
        def extract_year(col_name):
            try:
                for year in ['2018', '2019', '2020', '2021', '2022', '2023', '2024']:
                    if year in str(col_name):
                        return int(year)
                return 0
            except:
                return 0
        
        year_columns.sort(key=extract_year)
        
        # Create final column order
        final_columns = []
        for col in priority_columns:
            if col in df.columns:
                final_columns.append(col)
        
        final_columns.extend(year_columns)
        
        # Add any remaining columns
        remaining_columns = [col for col in df.columns if col not in final_columns]
        final_columns.extend(remaining_columns)
        
        return df[final_columns]
    
    def create_professional_gui(self):
        """Create professional GUI similar to Excel Processor Pro"""
        # Main container
        main_frame = tk.Frame(self.root, bg='white', padx=40, pady=30)
        main_frame.pack(fill="both", expand=True)
        
        # Header section
        header_frame = tk.Frame(main_frame, bg='white')
        header_frame.pack(fill="x", pady=(0, 30))
        
        # Icon and title
        title_frame = tk.Frame(header_frame, bg='white')
        title_frame.pack()
        
        # Main title with icon effect
        title_label = tk.Label(title_frame, text="⚡ Excel Processor Pro", 
                              font=("Arial", 24, "bold"), fg="#2c3e50", bg='white')
        title_label.pack()
        
        # Subtitle
        subtitle_label = tk.Label(title_frame, text="Advanced Data Processing & Analysis", 
                                 font=("Arial", 12), fg="#7f8c8d", bg='white')
        subtitle_label.pack(pady=(5, 0))
        
        # Status section
        status_frame = tk.LabelFrame(main_frame, text="📊 Status", font=("Arial", 12, "bold"), 
                                   fg="#2c3e50", bg='white', bd=2, relief="solid")
        status_frame.pack(fill="x", pady=(0, 20), ipady=10)
        
        self.status_label = tk.Label(status_frame, text="💾 Ready to process Excel files...", 
                                    font=("Arial", 11), fg="#27ae60", bg='white')
        self.status_label.pack(pady=5, padx=15, anchor="w")
        
        # Progress section
        progress_frame = tk.LabelFrame(main_frame, text="⏳ Progress", font=("Arial", 12, "bold"), 
                                     fg="#2c3e50", bg='white', bd=2, relief="solid")
        progress_frame.pack(fill="x", pady=(0, 30), ipady=15)
        
        self.progress_label = tk.Label(progress_frame, text="⏸ Step 1/3: Waiting for user input...", 
                                      font=("Arial", 11), fg="#3498db", bg='white')
        self.progress_label.pack(pady=(5, 10), padx=15, anchor="w")
        
        # Progress bar
        self.progress_bar = ttk.Progressbar(progress_frame, mode='determinate', length=500)
        self.progress_bar.pack(pady=(0, 10), padx=15)
        
        # Button section
        button_frame = tk.Frame(main_frame, bg='white')
        button_frame.pack(pady=(0, 30))
        
        # Select Excel File button
        self.select_btn = tk.Button(button_frame, text="📁 Select Excel File", 
                                   command=self.select_and_process, 
                                   bg="#3498db", fg="white", 
                                   width=20, height=2, font=("Arial", 12, "bold"),
                                   relief="flat", cursor="hand2")
        self.select_btn.pack(pady=10)
        
        # Footer
        footer_label = tk.Label(main_frame, text="Powered by Advanced Analytics Engine", 
                               font=("Arial", 9), fg="#95a5a6", bg='white')
        footer_label.pack(side="bottom")

    def select_and_process(self):
        """Select workbook and automatically start processing"""
        # Update status
        self.status_label.config(text="🔍 Opening file browser...", fg="#e67e22")
        self.progress_label.config(text="⏸ Step 1/3: Selecting Excel file...")
        self.progress_bar.config(value=10)
        self.root.update()
        
        # File selection
        self.workbook_path = filedialog.askopenfilename(
            title="Select Excel Workbook",
            filetypes=[("Excel files", "*.xlsx *.xls")],
            initialdir=os.path.expanduser("~")
        )
        
        if not self.workbook_path:
            self.status_label.config(text="❌ No file selected", fg="#e74c3c")
            self.progress_label.config(text="⏸ Step 1/3: Waiting for user input...")
            self.progress_bar.config(value=0)
            return
        
                # Update status
        self.status_label.config(text="🔍 Validating Excel sheets...", fg="#e67e22")
        self.progress_label.config(text="▶️ Step 2/3: Validating workbook structure...")
        self.progress_bar.config(value=30)
        self.root.update()
        
        # Validate required sheets
        try:
            wb = load_workbook(self.workbook_path)
            required_sheets = ['補修無視', '補修考慮', '構造物番号']
            missing_sheets = [sheet for sheet in required_sheets if sheet not in wb.sheetnames]
            
            if missing_sheets:
                self.status_label.config(text=f"❌ Missing required sheets: {', '.join(missing_sheets)}", fg="#e74c3c")
                self.progress_label.config(text="❌ Validation failed - missing sheets")
                self.progress_bar.config(value=0)
                return
            
            # Load structure data for enhancements
            try:
                self.structure_df = pd.read_excel(self.workbook_path, sheet_name='構造物番号')
            except:
                self.structure_df = None
            
            # Update status and start processing
            self.status_label.config(text="✅ File validated successfully", fg="#27ae60")
            self.progress_label.config(text="▶️ Step 3/3: Processing division sheets...")
            self.progress_bar.config(value=50)
            self.root.update()
            
            # Disable button during processing
            self.select_btn.config(state="disabled", text="⏳ Processing...")
            
            # Start processing
            self.root.after(500, self.execute_division_process)
            
        except Exception as e:
            self.status_label.config(text=f"❌ Error: {str(e)[:60]}...", fg="#e74c3c")
            self.progress_label.config(text="❌ File validation failed")
            self.progress_bar.config(value=0)

    def execute_division_process(self):
        """Execute the enhanced division process"""
        try:
            # Update progress for step 1
            self.progress_label.config(text="▶️ Processing Warizan Kekka (Hoshuu Mushi) - 1/2")
            self.progress_bar.config(value=60)
            self.root.update()
            
            # Load required sheets
            max_df = pd.read_excel(self.workbook_path, sheet_name='補修無視')
            hoshuu_df = pd.read_excel(self.workbook_path, sheet_name='補修考慮')
            structure_df = pd.read_excel(self.workbook_path, sheet_name='構造物番号')
            
            # Create enhanced division results - First sheet
            max_division_df = self.apply_enhanced_division_logic(max_df, structure_df, "補修無視")
            
            # Update progress for step 2
            self.progress_label.config(text="▶️ Processing Warizan Kekka (Hoshuu Kouryo) - 2/2")
            self.progress_bar.config(value=80)
            self.root.update()
            
            hoshuu_division_df = self.apply_enhanced_division_logic(hoshuu_df, structure_df, "補修考慮")
            
            # Update progress for saving
            self.progress_label.config(text="💾 Saving enhanced division results...")
            self.progress_bar.config(value=90)
            self.root.update()
            
            # Save to Excel
            self.save_enhanced_division_results(max_division_df, hoshuu_division_df)
            
            # Complete processing
            self.status_label.config(text="🎉 Processing completed successfully!", fg="#27ae60")
            self.progress_label.config(text="✅ Both Warizan Kekka sheets created successfully")
            self.progress_bar.config(value=100)
            self.root.update()
            
            # Re-enable button
            self.select_btn.config(state="normal", text="📁 Select Another File")
            
            # Show completion message and auto-exit after 3 seconds
            self.root.after(3000, self.auto_exit)
            
        except Exception as e:
            self.status_label.config(text=f"❌ Processing failed: {str(e)[:50]}...", fg="#e74c3c")
            self.progress_label.config(text="❌ Error during processing")
            self.progress_bar.config(value=0)
            self.select_btn.config(state="normal", text="📁 Select Excel File")

    def auto_exit(self):
        """Auto exit after successful completion"""
        self.root.quit()

    def apply_enhanced_division_logic(self, source_df, structure_df, sheet_type):
        """Apply enhanced division logic to a dataframe"""
        result_df = source_df.copy()
        
        # Find year result columns
        year_columns = [col for col in result_df.columns if col.endswith('結果')]
        
        # Rename year columns to include "合計重み/長さ"
        column_mapping = {}
        for col in year_columns:
            # Extract year from column name (e.g., "2018 結果" -> "2018")
            year_match = re.search(r'(\d{4})', col)
            if year_match:
                year = year_match.group(1)
                new_col_name = f"{year} 合計重み/長さ"
                column_mapping[col] = new_col_name
        
        result_df = result_df.rename(columns=column_mapping)
        
        # Apply division for each row
        for index, row in result_df.iterrows():
            # Get structure name and station interval
            kozo_name = str(row.get('構造物名称', '')).strip() if pd.notna(row.get('構造物名称', '')) else ''
            
            # Try to construct ekikan from 駅（始） and 駅（至）
            eki_start = str(row.get('駅（始）', '')).strip() if pd.notna(row.get('駅（始）', '')) else ''
            eki_end = str(row.get('駅（至）', '')).strip() if pd.notna(row.get('駅（至）', '')) else ''
            
            ekikan = ''
            if eki_start and eki_end:
                ekikan = f"{eki_start}→{eki_end}"
            
            # Find corresponding length in structure data
            length_value = self.find_length_value(structure_df, kozo_name, ekikan, row.get('路線名', ''))
            
            # Divide year result columns by length
            for old_col, new_col in column_mapping.items():
                original_value = source_df.loc[index, old_col] if old_col in source_df.columns else None
                
                if pd.notna(original_value) and str(original_value).strip() not in ['', 'nan']:
                    try:
                        numeric_value = float(original_value)
                        if length_value and length_value > 0:
                            divided_value = numeric_value / length_value
                            # Fix FutureWarning by explicitly converting to float
                            result_df.loc[index, new_col] = float(round(divided_value, 3))
                        else:
                            result_df.loc[index, new_col] = float(numeric_value)
                    except (ValueError, TypeError):
                        result_df.loc[index, new_col] = original_value
                else:
                    result_df.loc[index, new_col] = original_value
        
        # Add enhanced columns
        enhanced_df = self.add_enhanced_columns(result_df, self.structure_df)
        
        # Reorder columns
        final_df = self.reorder_columns_enhanced(enhanced_df)
        
        return final_df

    def find_length_value(self, structure_df, kozo_name, ekikan, rosen_name):
        """Find length value from structure data"""
        try:
            rosen_name = str(rosen_name).strip() if pd.notna(rosen_name) else ''
            
            # First try to match by structure name
            if kozo_name:
                matches = structure_df[
                    (structure_df['構造物名称'].astype(str).str.strip() == kozo_name) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    length_val = matches.iloc[0]['長さ(m)']
                    if pd.notna(length_val) and str(length_val).strip() not in ['', 'nan']:
                        try:
                            return float(length_val)
                        except (ValueError, TypeError):
                            pass
            
            # If not found by structure name, try by station interval
            if ekikan:
                matches = structure_df[
                    (structure_df['駅間'].astype(str).str.strip() == ekikan) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    length_val = matches.iloc[0]['長さ(m)']
                    if pd.notna(length_val) and str(length_val).strip() not in ['', 'nan']:
                        try:
                            return float(length_val)
                        except (ValueError, TypeError):
                            pass
            
            # Default length if not found
            return 100.0
            
        except Exception:
            return 100.0

    def save_enhanced_division_results(self, max_division_df, hoshuu_division_df):
        """Save enhanced division results to Excel sheets"""
        try:
            with pd.ExcelWriter(self.workbook_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                # Write enhanced division result sheets
                max_division_df.to_excel(writer, sheet_name='割算結果(補修無視)', index=False)
                hoshuu_division_df.to_excel(writer, sheet_name='割算結果(補修考慮)', index=False)
                
        except Exception as e:
            raise Exception(f"Error saving enhanced division results: {str(e)}")

    def run(self):
        """Run the enhanced application"""
        self.root.mainloop()


# Main execution
if __name__ == "__main__":
    app = EnhancedDivisionSheetsApp()
    app.run()