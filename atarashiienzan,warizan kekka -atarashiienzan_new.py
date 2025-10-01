import pandas as pd
import openpyxl
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import re

class EnhancedNewCalculationSheetsApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Enhanced New Calculation Sheets Generator")
        self.root.geometry("500x400")
        self.root.minsize(450, 350)
        
        # Center the window on screen
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() // 2) - (500 // 2)
        y = (self.root.winfo_screenheight() // 2) - (400 // 2)
        self.root.geometry(f"500x400+{x}+{y}")
        
        self.workbook_path = None
        self.structure_df = None
        
        self.create_main_gui()
    
    def abbreviate_sen_name(self, sen_name):
        """Convert route name to abbreviation"""
        if pd.isna(sen_name) or sen_name == '':
            return ''
        
        sen_name = str(sen_name).strip()
        
        abbreviation_map = {
            "東急多摩川線": "TM",
            "多摩川線": "TM", 
            "東横線": "TY",
            "大井町線": "OM",
            "池上線": "IK",
            "田園都市線": "DT",
            "目黒線": "MG",
            "こどもの国線": "KD",
            "世田谷線": "SG"
        }
        
        return abbreviation_map.get(sen_name, sen_name)
    
    def lookup_structure_number(self, structure_df, rosen_name, kozo_name, ekikan):
        """Lookup 構造物番号 from structure sheet"""
        try:
            if structure_df is None or len(structure_df) == 0:
                return ''
                
            rosen_name = str(rosen_name).strip() if pd.notna(rosen_name) else ''
            
            # First try to match by structure name
            if kozo_name and str(kozo_name).strip() not in ['', 'nan', 'NaN']:
                kozo_name = str(kozo_name).strip()
                matches = structure_df[
                    (structure_df['構造物名称'].astype(str).str.strip() == kozo_name) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    bangou = matches.iloc[0]['構造物番号']
                    if pd.notna(bangou) and str(bangou).strip() not in ['', 'nan']:
                        return str(bangou).strip()
            
            # If not found by structure name, try by station interval
            if ekikan and str(ekikan).strip() not in ['', 'nan', 'NaN']:
                ekikan = str(ekikan).strip()
                matches = structure_df[
                    (structure_df['駅間'].astype(str).str.strip() == ekikan) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    bangou = matches.iloc[0]['構造物番号']
                    if pd.notna(bangou) and str(bangou).strip() not in ['', 'nan']:
                        return str(bangou).strip()
            
            return ''
            
        except Exception as e:
            print(f"Error finding structure number: {e}")
            return ''
    
    def add_enhanced_columns(self, df, structure_df=None):
        """Add enhanced columns: 路線名略称 and 構造物番号"""
        enhanced_df = df.copy()
        
        # Add 路線名略称 column
        if '路線名' in enhanced_df.columns:
            enhanced_df['路線名略称'] = enhanced_df['路線名'].apply(self.abbreviate_sen_name)
        else:
            enhanced_df['路線名略称'] = ''
        
        # Add 構造物番号 column
        enhanced_df['構造物番号'] = ''
        
        if structure_df is not None:
            for index, row in enhanced_df.iterrows():
                rosen_name = row.get('路線名', '')
                kozo_name = row.get('構造物名称', '')
                
                # Create ekikan for lookup
                ekikan = ''
                if row.get('駅（始）', '') and row.get('駅（至）', ''):
                    ekikan = f"{row.get('駅（始）', '')}→{row.get('駅（至）', '')}"
                
                # Lookup structure number
                bangou = self.lookup_structure_number(structure_df, rosen_name, kozo_name, ekikan)
                enhanced_df.at[index, '構造物番号'] = bangou
        
        return enhanced_df
    
    def reorder_columns_enhanced(self, df):
        """Reorder columns: グループ化キー → グループ化方法 → 種別 → 構造物名称 → 駅（始） → 駅（至） → 点検区分1 → データ件数 → 路線名 → 路線名略称 → 構造物番号 → years"""
        
        # Define the correct enhanced column order
        priority_columns = [
            'グループ化キー',
            'グループ化方法', 
            '種別',
            '構造物名称',
            '駅（始）',
            '駅（至）',
            '点検区分1',
            'データ件数',
            '路線名',
            '路線名略称',
            '構造物番号'
        ]
        
        # Get year columns (various formats)
        year_columns = []
        
        for col in df.columns:
            if any(year in str(col) for year in ['2018', '2019', '2020', '2021', '2022', '2023', '2024']):
                year_columns.append(col)
        
        # Sort year columns chronologically
        def extract_year(col_name):
            try:
                for year in ['2018', '2019', '2020', '2021', '2022', '2023', '2024']:
                    if year in str(col_name):
                        return int(year)
                return 0
            except:
                return 0
        
        year_columns.sort(key=extract_year)
        
        # Create final column order
        final_columns = []
        
        # Add priority columns that exist
        for col in priority_columns:
            if col in df.columns:
                final_columns.append(col)
        
        # Add year columns
        final_columns.extend(year_columns)
        
        # Add any remaining columns that weren't in priority or year columns
        remaining_columns = [col for col in df.columns if col not in final_columns]
        final_columns.extend(remaining_columns)
        
        # Return reordered dataframe
        return df[final_columns]
    
    def create_main_gui(self):
        """Create main GUI for file selection"""
        main_frame = tk.Frame(self.root, padx=30, pady=30)
        main_frame.pack(fill="both", expand=True)
        
        # Title
        title_label = tk.Label(main_frame, text="Enhanced New Calculation Sheets Generator", 
                              font=("Arial", 14, "bold"), fg="navy")
        title_label.pack(pady=(0, 15))
        
        # Instructions
        instruction_text = ("Enhanced new calculation sheets:\n"
                          "• 新しい演算(補修無視) with enhanced columns\n"
                          "• 新しい演算(補修考慮) with enhanced columns\n" 
                          "• 割算結果-新しい演算(補修無視) with enhanced columns\n"
                          "• 割算結果-新しい演算(補修考慮) with enhanced columns\n\n"
                          "Features: X*A*B*C calculations, 路線名略称, 構造物番号")
        instruction_label = tk.Label(main_frame, text=instruction_text, 
                                   font=("Arial", 10), justify="center")
        instruction_label.pack(pady=(0, 15))
        
        # Status label
        self.status_label = tk.Label(main_frame, text="Ready...", 
                                    font=("Arial", 9), fg="gray")
        self.status_label.pack(pady=(0, 10))
        
        # Select file button
        select_btn = tk.Button(main_frame, text="Browse & Select File", 
                             command=self.select_workbook, 
                             bg="#4CAF50", fg="white", 
                             width=20, height=1, font=("Arial", 10))
        select_btn.pack(pady=8)
        
        # Generate button (initially disabled)
        self.generate_btn = tk.Button(main_frame, text="Generate Enhanced Calculation Sheets", 
                                    command=self.generate_calculation_sheets, 
                                    bg="#FF9800", fg="white", 
                                    width=28, height=1, font=("Arial", 10),
                                    state="disabled")
        self.generate_btn.pack(pady=8)
        
        # Exit button
        exit_btn = tk.Button(main_frame, text="Exit", 
                           command=self.root.quit, bg="#f44336", fg="white", 
                           width=12, height=1, font=("Arial", 9))
        exit_btn.pack(pady=(15, 0))

    def select_workbook(self):
        """Select workbook with required sheets"""
        self.status_label.config(text="Opening browser...", fg="blue")
        self.root.update()
        
        # File selection
        self.workbook_path = filedialog.askopenfilename(
            title="Select Excel Workbook",
            filetypes=[("Excel files", "*.xlsx *.xls")],
            initialdir=os.path.expanduser("~")
        )
        
        if not self.workbook_path:
            self.status_label.config(text="No file selected", fg="orange")
            return
        
        self.status_label.config(text="Validating sheets...", fg="blue")
        self.root.update()
        
        # Validate required sheets
        try:
            wb = load_workbook(self.workbook_path)
            required_sheets = ['補修無視', '補修考慮', '構造物番号']
            missing_sheets = [sheet for sheet in required_sheets if sheet not in wb.sheetnames]
            
            if missing_sheets:
                self.status_label.config(text="Missing sheets!", fg="red")
                messagebox.showerror("Missing Sheets", 
                                   f"Required sheets not found:\n" + "\n".join(missing_sheets))
                return
            
            # Try to load structure data for enhancements
            try:
                self.structure_df = pd.read_excel(self.workbook_path, sheet_name='構造物番号')
                print("Found 構造物番号 sheet - enhanced features enabled")
            except:
                self.structure_df = None
                print("No 構造物番号 sheet found - basic features only")
            
            enhancement_status = "with enhancements" if self.structure_df is not None else "basic version"
            self.status_label.config(text="Ready to generate!", fg="green")
            self.generate_btn.config(state="normal")
            messagebox.showinfo("Success", f"All required sheets found ({enhancement_status})!\nReady to generate enhanced calculation sheets.")
            
        except Exception as e:
            self.status_label.config(text="Error", fg="red")
            messagebox.showerror("Error", f"Error validating file:\n{str(e)}")

    def generate_calculation_sheets(self):
        """Generate enhanced calculation sheets"""
        try:
            # Show progress dialog
            progress_window = tk.Toplevel(self.root)
            progress_window.title("Generating Enhanced Calculation Sheets")
            progress_window.geometry("400x120")
            progress_window.grab_set()
            progress_window.resizable(False, False)
            progress_window.transient(self.root)
            
            # Center the progress window
            progress_window.update_idletasks()
            x = (progress_window.winfo_screenwidth() // 2) - (400 // 2)
            y = (progress_window.winfo_screenheight() // 2) - (120 // 2)
            progress_window.geometry(f"400x120+{x}+{y}")
            
            progress_frame = tk.Frame(progress_window, padx=20, pady=20)
            progress_frame.pack(fill="both", expand=True)
            
            status_label = tk.Label(progress_frame, text="Processing enhanced calculations...", font=("Arial", 10))
            status_label.pack(pady=5)
            
            progress_bar = ttk.Progressbar(progress_frame, mode='indeterminate')
            progress_bar.pack(fill="x", pady=5)
            progress_bar.start()
            
            # Execute process
            self.root.after(100, lambda: self.execute_calculation_process(progress_window))
            
        except Exception as e:
            messagebox.showerror("Error", f"Error starting process:\n{str(e)}")

    def execute_calculation_process(self, progress_window):
        """Execute the enhanced calculation process"""
        try:
            # Load required sheets
            max_df = pd.read_excel(self.workbook_path, sheet_name='補修無視')
            hoshuu_df = pd.read_excel(self.workbook_path, sheet_name='補修考慮')
            structure_df = pd.read_excel(self.workbook_path, sheet_name='構造物番号')
            
            # Create enhanced calculation results
            # Sheet 1: 新しい演算(補修無視) - X*A*B*C
            new_calc_max_df = self.apply_enhanced_new_calculation_logic(max_df, structure_df, "補修無視")
            
            # Sheet 2: 新しい演算(補修考慮) - X*A*B*C  
            new_calc_hoshuu_df = self.apply_enhanced_new_calculation_logic(hoshuu_df, structure_df, "補修考慮")
            
            # Sheet 3: 割算結果-新しい演算(補修無視) - X*A*B*C ÷ Length
            division_calc_max_df = self.apply_enhanced_division_calculation_logic(max_df, structure_df, "補修無視")
            
            # Sheet 4: 割算結果-新しい演算(補修考慮) - X*A*B*C ÷ Length
            division_calc_hoshuu_df = self.apply_enhanced_division_calculation_logic(hoshuu_df, structure_df, "補修考慮")
            
            # Save to Excel
            self.save_enhanced_calculation_results(new_calc_max_df, new_calc_hoshuu_df, 
                                                 division_calc_max_df, division_calc_hoshuu_df)
            
            # Close progress window
            progress_window.destroy()
            
            # Show completion dialog
            self.show_enhanced_completion_dialog()
            
        except Exception as e:
            progress_window.destroy()
            messagebox.showerror("Error", f"Error during processing:\n{str(e)}")

    def apply_enhanced_new_calculation_logic(self, source_df, structure_df, sheet_type):
        """Apply enhanced new calculation logic: X*A*B*C"""
        result_df = source_df.copy()
        
        # Find year result columns
        year_columns = [col for col in result_df.columns if col.endswith('結果')]
        
        print(f"Processing enhanced new calculation {sheet_type} with {len(result_df)} rows")
        
        # Apply X*A*B*C calculation for each row
        for index, row in result_df.iterrows():
            # Get weight values from structure data
            weights = self.get_structure_weights(structure_df, row)
            
            # Apply calculation to each year column
            for year_col in year_columns:
                original_value = row[year_col]  # X value
                
                if pd.notna(original_value) and str(original_value).strip() not in ['', 'nan']:
                    try:
                        x_value = float(original_value)
                        
                        # X * A * B * C calculation
                        calculated_value = x_value * weights['A'] * weights['B'] * weights['C']
                        result_df.at[index, year_col] = round(calculated_value, 3)
                        
                    except (ValueError, TypeError):
                        result_df.at[index, year_col] = original_value
                else:
                    result_df.at[index, year_col] = original_value
        
        # Add enhanced columns
        enhanced_df = self.add_enhanced_columns(result_df, self.structure_df)
        
        # Reorder columns
        final_df = self.reorder_columns_enhanced(enhanced_df)
        
        return final_df

    def apply_enhanced_division_calculation_logic(self, source_df, structure_df, sheet_type):
        """Apply enhanced division calculation logic: X*A*B*C ÷ Length"""
        result_df = source_df.copy()
        
        # Find year result columns and rename them
        year_columns = [col for col in result_df.columns if col.endswith('結果')]
        
        # Rename year columns to include division indicator
        column_mapping = {}
        for col in year_columns:
            year_match = re.search(r'(\d{4})', col)
            if year_match:
                year = year_match.group(1)
                new_col_name = f"{year} 新演算/長さ"
                column_mapping[col] = new_col_name
        
        result_df = result_df.rename(columns=column_mapping)
        
        print(f"Processing enhanced division calculation {sheet_type} with {len(result_df)} rows")
        
        # Apply X*A*B*C ÷ Length calculation for each row
        for index, row in result_df.iterrows():
            # Get weight values and length from structure data
            weights = self.get_structure_weights(structure_df, row)
            length_value = self.get_length_value(structure_df, row)
            
            # Apply calculation to each year column
            for old_col, new_col in column_mapping.items():
                original_value = source_df.loc[index, old_col] if old_col in source_df.columns else None
                
                if pd.notna(original_value) and str(original_value).strip() not in ['', 'nan']:
                    try:
                        x_value = float(original_value)
                        
                        # X * A * B * C calculation
                        calculated_value = x_value * weights['A'] * weights['B'] * weights['C']
                        
                        # Divide by length if available
                        if length_value and length_value > 0:
                            final_value = calculated_value / length_value
                            result_df.at[index, new_col] = round(final_value, 3)
                        else:
                            # Handle missing length data - keep calculated value without division
                            result_df.at[index, new_col] = round(calculated_value, 3)
                        
                    except (ValueError, TypeError):
                        result_df.at[index, new_col] = original_value
                else:
                    result_df.at[index, new_col] = original_value
        
        # Add enhanced columns
        enhanced_df = self.add_enhanced_columns(result_df, self.structure_df)
        
        # Reorder columns
        final_df = self.reorder_columns_enhanced(enhanced_df)
        
        return final_df

    def get_structure_weights(self, structure_df, row):
        """Get structure weights (A, B, C) from structure data"""
        try:
            rosen_name = str(row.get('路線名', '')).strip() if pd.notna(row.get('路線名', '')) else ''
            kozo_name = str(row.get('構造物名称', '')).strip() if pd.notna(row.get('構造物名称', '')) else ''
            
            # Try to construct ekikan from 駅（始） and 駅（至）
            eki_start = str(row.get('駅（始）', '')).strip() if pd.notna(row.get('駅（始）', '')) else ''
            eki_end = str(row.get('駅（至）', '')).strip() if pd.notna(row.get('駅（至）', '')) else ''
            
            ekikan = ''
            if eki_start and eki_end:
                ekikan = f"{eki_start}→{eki_end}"
            
            # Default weights
            weights = {'A': 1.0, 'B': 1.0, 'C': 1.0}
            
            # First try to match by structure name
            if kozo_name:
                matches = structure_df[
                    (structure_df['構造物名称'].astype(str).str.strip() == kozo_name) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    match_row = matches.iloc[0]
                    
                    # Get weight values (A = 構造形式_重み, B = 角度_重み, C = 供用年数_重み)
                    if '構造形式_重み' in match_row and pd.notna(match_row['構造形式_重み']):
                        try:
                            weights['A'] = float(match_row['構造形式_重み'])
                        except (ValueError, TypeError):
                            pass
                    
                    if '角度_重み' in match_row and pd.notna(match_row['角度_重み']):
                        try:
                            weights['B'] = float(match_row['角度_重み'])
                        except (ValueError, TypeError):
                            pass
                    
                    if '供用年数_重み' in match_row and pd.notna(match_row['供用年数_重み']):
                        try:
                            weights['C'] = float(match_row['供用年数_重み'])
                        except (ValueError, TypeError):
                            pass
                    
                    return weights
            
            # If not found by structure name, try by station interval
            if ekikan:
                matches = structure_df[
                    (structure_df['駅間'].astype(str).str.strip() == ekikan) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    match_row = matches.iloc[0]
                    
                    # Get weight values
                    if '構造形式_重み' in match_row and pd.notna(match_row['構造形式_重み']):
                        try:
                            weights['A'] = float(match_row['構造形式_重み'])
                        except (ValueError, TypeError):
                            pass
                    
                    if '角度_重み' in match_row and pd.notna(match_row['角度_重み']):
                        try:
                            weights['B'] = float(match_row['角度_重み'])
                        except (ValueError, TypeError):
                            pass
                    
                    if '供用年数_重み' in match_row and pd.notna(match_row['供用年数_重み']):
                        try:
                            weights['C'] = float(match_row['供用年数_重み'])
                        except (ValueError, TypeError):
                            pass
            
            return weights
            
        except Exception as e:
            print(f"Error getting structure weights: {e}")
            # Return default weights if error
            return {'A': 1.0, 'B': 1.0, 'C': 1.0}

    def get_length_value(self, structure_df, row):
        """Get length value from structure data"""
        try:
            rosen_name = str(row.get('路線名', '')).strip() if pd.notna(row.get('路線名', '')) else ''
            kozo_name = str(row.get('構造物名称', '')).strip() if pd.notna(row.get('構造物名称', '')) else ''
            
            # Try to construct ekikan from 駅（始） and 駅（至）
            eki_start = str(row.get('駅（始）', '')).strip() if pd.notna(row.get('駅（始）', '')) else ''
            eki_end = str(row.get('駅（至）', '')).strip() if pd.notna(row.get('駅（至）', '')) else ''
            
            ekikan = ''
            if eki_start and eki_end:
                ekikan = f"{eki_start}→{eki_end}"
            
            # First try to match by structure name
            if kozo_name:
                matches = structure_df[
                    (structure_df['構造物名称'].astype(str).str.strip() == kozo_name) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    length_val = matches.iloc[0]['長さ(m)']
                    if pd.notna(length_val) and str(length_val).strip() not in ['', 'nan']:
                        try:
                            return float(length_val)
                        except (ValueError, TypeError):
                            pass
            
            # If not found by structure name, try by station interval
            if ekikan:
                matches = structure_df[
                    (structure_df['駅間'].astype(str).str.strip() == ekikan) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    length_val = matches.iloc[0]['長さ(m)']
                    if pd.notna(length_val) and str(length_val).strip() not in ['', 'nan']:
                        try:
                            return float(length_val)
                        except (ValueError, TypeError):
                            pass
            
            # Return None if length not found (no default value)
            print(f"Length not found for: {kozo_name or ekikan} in {rosen_name}")
            return None
            
        except Exception as e:
            print(f"Error finding length value: {e}")
            return None

    def save_enhanced_calculation_results(self, new_calc_max_df, new_calc_hoshuu_df, 
                                        division_calc_max_df, division_calc_hoshuu_df):
        """Save enhanced calculation results to Excel sheets"""
        try:
            with pd.ExcelWriter(self.workbook_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                # Write enhanced calculation result sheets
                new_calc_max_df.to_excel(writer, sheet_name='新しい演算(補修無視)', index=False)
                new_calc_hoshuu_df.to_excel(writer, sheet_name='新しい演算(補修考慮)', index=False)
                division_calc_max_df.to_excel(writer, sheet_name='割算結果-新しい演算(補修無視)', index=False)
                division_calc_hoshuu_df.to_excel(writer, sheet_name='割算結果-新しい演算(補修考慮)', index=False)
                
                # Preserve other sheets
                try:
                    original_wb = load_workbook(self.workbook_path)
                    sheet_names_to_preserve = [name for name in original_wb.sheetnames 
                                             if name not in ['新しい演算(補修無視)', '新しい演算(補修考慮)', 
                                                           '割算結果-新しい演算(補修無視)', '割算結果-新しい演算(補修考慮)']]
                    
                    for sheet_name in sheet_names_to_preserve:
                        try:
                            df_temp = pd.read_excel(self.workbook_path, sheet_name=sheet_name)
                            df_temp.to_excel(writer, sheet_name=sheet_name, index=False)
                        except Exception as e:
                            continue
                except Exception as e:
                    pass
                    
        except Exception as e:
            raise Exception(f"Error saving enhanced calculation results: {str(e)}")

    def show_enhanced_completion_dialog(self):
        """Show enhanced completion dialog"""
        completion_window = tk.Toplevel(self.root)
        completion_window.title("Enhanced Calculation Complete")
        completion_window.geometry("450x350")
        completion_window.grab_set()
        completion_window.resizable(False, False)
        completion_window.transient(self.root)
        
        # Center window
        completion_window.update_idletasks()
        x = (completion_window.winfo_screenwidth() // 2) - (450 // 2)
        y = (completion_window.winfo_screenheight() // 2) - (350 // 2)
        completion_window.geometry(f"450x350+{x}+{y}")
        
        main_frame = tk.Frame(completion_window, padx=15, pady=15)
        main_frame.pack(fill="both", expand=True)
        
        # Title
        title_label = tk.Label(main_frame, text="Enhanced Calculation Complete!", 
                              font=("Arial", 12, "bold"), fg="green")
        title_label.pack(pady=(0, 10))
        
        # Enhanced features info
        features_text = ("✅ Enhanced Features Applied:\n\n"
                        "• 路線名略称 column added\n"
                        "• 構造物番号 column added\n"
                        "• Proper column ordering\n"
                        "• X*A*B*C calculations\n"
                        "• X*A*B*C ÷ Length calculations")
        features_label = tk.Label(main_frame, text=features_text, font=("Arial", 10), 
                                 justify="left", fg="blue")
        features_label.pack(pady=(0, 10))
        
        # Processing info
        info_text = ("4 Enhanced sheets created:\n"
                    "• 新しい演算(補修無視)\n"
                    "• 新しい演算(補修考慮)\n"
                    "• 割算結果-新しい演算(補修無視)\n"
                    "• 割算結果-新しい演算(補修考慮)")
        info_label = tk.Label(main_frame, text=info_text, font=("Arial", 10))
        info_label.pack(pady=(0, 10))
        
        # Calculation details
        calc_text = ("Calculations applied:\n"
                    "X = Original year result value\n"
                    "A = 構造形式_重み, B = 角度_重み, C = 供用年数_重み")
        calc_label = tk.Label(main_frame, text=calc_text, font=("Arial", 9), fg="darkgreen")
        calc_label.pack(pady=(0, 15))
        
        # Enhancement status
        enhancement_status = "with 構造物番号 lookup" if self.structure_df is not None else "basic version"
        status_label = tk.Label(main_frame, text=f"Enhancement: {enhancement_status}", 
                               font=("Arial", 9), fg="green" if self.structure_df is not None else "orange")
        status_label.pack(pady=(0, 15))
        
        # Buttons
        button_frame = tk.Frame(main_frame)
        button_frame.pack()
        
        def open_excel():
            try:
                import os
                os.startfile(self.workbook_path)
                messagebox.showinfo("Excel Opened", 
                                  "✅ Enhanced Excel file opened!\n\n"
                                  "Check the 4 new enhanced calculation sheets:\n"
                                  "• 新しい演算(補修無視)\n"
                                  "• 新しい演算(補修考慮)\n"
                                  "• 割算結果-新しい演算(補修無視)\n"
                                  "• 割算結果-新しい演算(補修考慮)\n\n"
                                  "With 路線名略称 and 構造物番号 columns!")
                completion_window.after(1000, completion_window.destroy)
                self.root.after(2000, self.root.quit)
            except:
                messagebox.showinfo("Info", f"Please open file manually:\n{self.workbook_path}")
                completion_window.destroy()

        def close_only():
            completion_window.destroy()
            messagebox.showinfo("Complete", 
                              "✅ Enhanced calculation processing completed!\n\n"
                              "4 calculation sheets created with:\n"
                              "• 路線名略称 columns\n"
                              "• 構造物番号 columns\n"
                              "• Proper column ordering\n"
                              "• X*A*B*C calculations\n"
                              "• Division by length values")
            self.root.after(1000, self.root.quit)
        
        excel_btn = tk.Button(button_frame, text="Open Enhanced Excel", 
                            command=open_excel, bg="#4CAF50", fg="white", 
                            width=15, height=1, font=("Arial", 10))
        excel_btn.pack(side="left", padx=5)
        
        close_btn = tk.Button(button_frame, text="Complete", 
                            command=close_only, bg="#2196F3", fg="white", 
                            width=12, height=1, font=("Arial", 10))
        close_btn.pack(side="left", padx=5)

    def run(self):
        """Run the enhanced application"""
        self.root.mainloop()


# Main execution
if __name__ == "__main__":
    print("Enhanced New Calculation Sheets Generator Starting...")
    print("=" * 60)
    print("🚀 Enhanced Features:")
    print("• 路線名略称 column (TM, TY, OM, IK, DT, MG, KD, SG)")
    print("• 構造物番号 column (auto-lookup from 構造物番号 sheet)")
    print("• Column order: グループ化キー → グループ化方法 → 種別 → 構造物名称 → 駅（始） → 駅（至） → 点検区分1 → データ件数 → 路線名 → 路線名略称 → 構造物番号 → years")
    print("• X*A*B*C calculations with weight values from 構造物番号 sheet")
    print("• X*A*B*C ÷ Length calculations")
    print("• A = 構造形式_重み, B = 角度_重み, C = 供用年数_重み")
    print("=" * 60)
    print("Required input:")
    print("• 補修無視 sheet")
    print("• 補修考慮 sheet") 
    print("• 構造物番号 sheet (for weights, length values and enhanced features)")
    print("=" * 60)
    print("Output:")
    print("• 新しい演算(補修無視) sheet (X*A*B*C calculation)")
    print("• 新しい演算(補修考慮) sheet (X*A*B*C calculation)")
    print("• 割算結果-新しい演算(補修無視) sheet (X*A*B*C ÷ Length)")
    print("• 割算結果-新しい演算(補修考慮) sheet (X*A*B*C ÷ Length)")
    print("=" * 60)
    
    app = EnhancedNewCalculationSheetsApp()
    app.run()