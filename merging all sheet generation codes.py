import pandas as pd
import openpyxl
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, filedialog
import os
import re
import warnings

# Suppress pandas warnings for better performance
warnings.filterwarnings("ignore", category=FutureWarning)

class UnifiedExcelProcessorApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Excel Processor Pro - Complete Suite")
        self.root.geometry("1000x700")
        self.root.minsize(1000, 700)
        
        # Center the window on screen
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() // 2) - (1000 // 2)
        y = (self.root.winfo_screenheight() // 2) - (700 // 2)
        self.root.geometry(f"1000x700+{x}+{y}")
        
        # Set background color
        self.root.configure(bg='white')
        
        self.workbook_path = None
        self.structure_df = None
        self.operator_df = None
        
        self.create_professional_gui()
    
    def abbreviate_sen_name(self, sen_name):
        """Convert route name to abbreviation"""
        if pd.isna(sen_name) or sen_name == '':
            return ''
        
        sen_name = str(sen_name).strip()
        
        abbreviation_map = {
            "東急多摩川線": "TM", "多摩川線": "TM", "東横線": "TY",
            "大井町線": "OM", "池上線": "IK", "田園都市線": "DT",
            "目黒線": "MG", "こどもの国線": "KD", "世田谷線": "SG"
        }
        
        return abbreviation_map.get(sen_name, sen_name)
    
    def lookup_structure_number(self, structure_df, rosen_name, kozo_name, ekikan):
        """Lookup 構造物番号 from structure sheet"""
        try:
            if structure_df is None or len(structure_df) == 0:
                return ''
                
            rosen_name = str(rosen_name).strip() if pd.notna(rosen_name) else ''
            
            # First try to match by structure name
            if kozo_name and str(kozo_name).strip() not in ['', 'nan', 'NaN']:
                kozo_name = str(kozo_name).strip()
                matches = structure_df[
                    (structure_df['構造物名称'].astype(str).str.strip() == kozo_name) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    bangou = matches.iloc[0]['構造物番号']
                    if pd.notna(bangou) and str(bangou).strip() not in ['', 'nan']:
                        return str(bangou).strip()
            
            # If not found by structure name, try by station interval
            if ekikan and str(ekikan).strip() not in ['', 'nan', 'NaN']:
                ekikan = str(ekikan).strip()
                matches = structure_df[
                    (structure_df['駅間'].astype(str).str.strip() == ekikan) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    bangou = matches.iloc[0]['構造物番号']
                    if pd.notna(bangou) and str(bangou).strip() not in ['', 'nan']:
                        return str(bangou).strip()
            
            return ''
            
        except Exception:
            return ''
    
    def add_enhanced_columns(self, df, structure_df=None):
        """Add enhanced columns: 路線名略称 and 構造物番号"""
        enhanced_df = df.copy()
        
        # Add 路線名略称 column
        if '路線名' in enhanced_df.columns:
            enhanced_df['路線名略称'] = enhanced_df['路線名'].apply(self.abbreviate_sen_name)
        else:
            enhanced_df['路線名略称'] = ''
        
        # Add 構造物番号 column
        enhanced_df['構造物番号'] = ''
        
        if structure_df is not None:
            for index, row in enhanced_df.iterrows():
                rosen_name = row.get('路線名', '')
                kozo_name = row.get('構造物名称', '')
                
                # Create ekikan for lookup
                ekikan = ''
                if row.get('駅（始）', '') and row.get('駅（至）', ''):
                    ekikan = f"{row.get('駅（始）', '')}→{row.get('駅（至）', '')}"
                
                # Lookup structure number
                bangou = self.lookup_structure_number(structure_df, rosen_name, kozo_name, ekikan)
                enhanced_df.at[index, '構造物番号'] = bangou
        
        return enhanced_df
    
    def reorder_columns_enhanced(self, df):
        """Reorder columns"""
        priority_columns = [
            'グループ化キー', 'グループ化方法', '種別', '構造物名称',
            '駅（始）', '駅（至）', '点検区分1', 'データ件数',
            '路線名', '路線名略称', '構造物番号'
        ]
        
        # Get year columns
        year_columns = []
        for col in df.columns:
            if any(year in str(col) for year in ['2018', '2019', '2020', '2021', '2022', '2023', '2024']):
                year_columns.append(col)
        
        # Sort year columns chronologically
        def extract_year(col_name):
            try:
                for year in ['2018', '2019', '2020', '2021', '2022', '2023', '2024']:
                    if year in str(col_name):
                        return int(year)
                return 0
            except:
                return 0
        
        year_columns.sort(key=extract_year)
        
        # Create final column order
        final_columns = []
        for col in priority_columns:
            if col in df.columns:
                final_columns.append(col)
        
        final_columns.extend(year_columns)
        
        # Add any remaining columns
        remaining_columns = [col for col in df.columns if col not in final_columns]
        final_columns.extend(remaining_columns)
        
        return df[final_columns]
    
    def create_professional_gui(self):
        """Create professional GUI"""
        # Main container
        main_frame = tk.Frame(self.root, bg='white', padx=40, pady=30)
        main_frame.pack(fill="both", expand=True)
        
        # Header section
        header_frame = tk.Frame(main_frame, bg='white')
        header_frame.pack(fill="x", pady=(0, 30))
        
        # Title with icon effect
        title_label = tk.Label(header_frame, text="⚡ Excel Processor Pro - Complete Suite", 
                              font=("Arial", 22, "bold"), fg="#2c3e50", bg='white')
        title_label.pack()
        
        # Subtitle
        subtitle_label = tk.Label(header_frame, text="Advanced 9-Sheet Generation System", 
                                 font=("Arial", 12), fg="#7f8c8d", bg='white')
        subtitle_label.pack(pady=(5, 0))
        
        # Status section
        status_frame = tk.LabelFrame(main_frame, text="📊 Status", font=("Arial", 12, "bold"), 
                                   fg="#2c3e50", bg='white', bd=2, relief="solid")
        status_frame.pack(fill="x", pady=(0, 20), ipady=10)
        
        self.status_label = tk.Label(status_frame, text="💾 Ready to process Excel workbook for 9 sheets...", 
                                    font=("Arial", 11), fg="#27ae60", bg='white')
        self.status_label.pack(pady=5, padx=15, anchor="w")
        
        # Progress section
        progress_frame = tk.LabelFrame(main_frame, text="⏳ Progress", font=("Arial", 12, "bold"), 
                                     fg="#2c3e50", bg='white', bd=2, relief="solid")
        progress_frame.pack(fill="x", pady=(0, 30), ipady=15)
        
        self.progress_label = tk.Label(progress_frame, text="⏸ Step 1/3: Waiting for workbook selection...", 
                                      font=("Arial", 11), fg="#3498db", bg='white')
        self.progress_label.pack(pady=(5, 10), padx=15, anchor="w")
        
        # Progress bar
        self.progress_bar = ttk.Progressbar(progress_frame, mode='determinate', length=500)
        self.progress_bar.pack(pady=(0, 10), padx=15)
        
        # Sheets info section
        sheets_frame = tk.LabelFrame(main_frame, text="📄 Sheets to Generate", font=("Arial", 12, "bold"), 
                                   fg="#2c3e50", bg='white', bd=2, relief="solid")
        sheets_frame.pack(fill="x", pady=(0, 20), ipady=10)
        
        sheets_text = ("• 割算結果(補修無視) • 割算結果(補修考慮)\n"
                      "• 新しい演算(補修無視) • 新しい演算(補修考慮)\n"
                      "• 割算結果-新しい演算(補修無視) • 割算結果-新しい演算(補修考慮)\n"
                      "• 経時変化（橋長考慮） • 経時変化（橋長&形式考慮） • 経時変化（橋長無視&形式考慮）")
        
        sheets_label = tk.Label(sheets_frame, text=sheets_text, 
                               font=("Arial", 10), fg="#34495e", bg='white')
        sheets_label.pack(pady=5, padx=15)
        
        # Button section
        button_frame = tk.Frame(main_frame, bg='white')
        button_frame.pack(pady=(0, 30))
        
        # Select Excel File button
        self.select_btn = tk.Button(button_frame, text="📁 Select Excel Workbook", 
                                   command=self.select_and_process, 
                                   bg="#3498db", fg="white", 
                                   width=25, height=2, font=("Arial", 12, "bold"),
                                   relief="flat", cursor="hand2")
        self.select_btn.pack(pady=10)
        
        # Footer
        footer_label = tk.Label(main_frame, text="Powered by Advanced Analytics Engine • 9 Sheets in One Process", 
                               font=("Arial", 9), fg="#95a5a6", bg='white')
        footer_label.pack(side="bottom")

    def select_and_process(self):
        """Select workbook and automatically start processing"""
        # Update status
        self.status_label.config(text="🔍 Opening file browser...", fg="#e67e22")
        self.progress_label.config(text="⏸ Step 1/3: Selecting Excel workbook...")
        self.progress_bar.config(value=5)
        self.root.update()
        
        # File selection
        self.workbook_path = filedialog.askopenfilename(
            title="Select Excel Workbook",
            filetypes=[("Excel files", "*.xlsx *.xls")],
            initialdir=os.path.expanduser("~")
        )
        
        if not self.workbook_path:
            self.status_label.config(text="❌ No file selected", fg="#e74c3c")
            self.progress_label.config(text="⏸ Step 1/3: Waiting for workbook selection...")
            self.progress_bar.config(value=0)
            return
        
        # Update status
        self.status_label.config(text="🔍 Validating Excel sheets...", fg="#e67e22")
        self.progress_label.config(text="▶️ Step 2/3: Validating workbook structure...")
        self.progress_bar.config(value=15)
        self.root.update()
        
        # Validate required sheets
        try:
            wb = load_workbook(self.workbook_path)
            actual_sheets = [sheet.strip() for sheet in wb.sheetnames]
            
            # Find operator sheet with flexible matching
            operator_sheet_name = None
            for sheet_name in actual_sheets:
                if '演算子' in sheet_name and '2' in sheet_name:
                    operator_sheet_name = sheet_name
                    break
            
            # Check required sheets
            required_sheets = ['補修無視', '補修考慮', '構造物番号', 'グループ化点検履歴']
            missing_sheets = [sheet for sheet in required_sheets if sheet not in actual_sheets]
            
            if missing_sheets:
                self.status_label.config(text=f"❌ Missing required sheets: {', '.join(missing_sheets)}", fg="#e74c3c")
                self.progress_label.config(text="❌ Validation failed - missing sheets")
                self.progress_bar.config(value=0)
                return
            
            if operator_sheet_name is None:
                self.status_label.config(text="❌ 演算子‐2 sheet not found!", fg="#e74c3c")
                self.progress_label.config(text="❌ Validation failed - missing operator sheet")
                self.progress_bar.config(value=0)
                return
            
            # Load structure and operator data
            try:
                self.structure_df = pd.read_excel(self.workbook_path, sheet_name='構造物番号')
                self.operator_df = pd.read_excel(self.workbook_path, sheet_name=operator_sheet_name)
            except Exception as e:
                self.status_label.config(text=f"❌ Error loading data: {str(e)[:50]}...", fg="#e74c3c")
                self.progress_label.config(text="❌ Data loading failed")
                self.progress_bar.config(value=0)
                return
            
            # Update status and start processing
            self.status_label.config(text="✅ All sheets validated successfully", fg="#27ae60")
            self.progress_label.config(text="▶️ Step 3/3: Processing all 9 sheets...")
            self.progress_bar.config(value=25)
            self.root.update()
            
            # Disable button during processing
            self.select_btn.config(state="disabled", text="⏳ Processing 9 Sheets...")
            
            # Start processing
            self.root.after(500, self.execute_complete_process)
            
        except Exception as e:
            self.status_label.config(text=f"❌ Error: {str(e)[:60]}...", fg="#e74c3c")
            self.progress_label.config(text="❌ File validation failed")
            self.progress_bar.config(value=0)

    def execute_complete_process(self):
        """Execute the complete 9-sheet generation process"""
        try:
            # Load all required sheets
            max_df = pd.read_excel(self.workbook_path, sheet_name='補修無視')
            hoshuu_df = pd.read_excel(self.workbook_path, sheet_name='補修考慮')
            structure_df = pd.read_excel(self.workbook_path, sheet_name='構造物番号')
            grouped_df = pd.read_excel(self.workbook_path, sheet_name='グループ化点検履歴')
            
            # Sheet 1: 割算結果(補修無視)
            self.progress_label.config(text="▶️ Processing 割算結果(補修無視) - 1/9")
            self.progress_bar.config(value=30)
            self.root.update()
            sheet1 = self.apply_division_logic(max_df, structure_df, "補修無視")
            
            # Sheet 2: 割算結果(補修考慮)
            self.progress_label.config(text="▶️ Processing 割算結果(補修考慮) - 2/9")
            self.progress_bar.config(value=35)
            self.root.update()
            sheet2 = self.apply_division_logic(hoshuu_df, structure_df, "補修考慮")
            
            # Sheet 3: 新しい演算(補修無視)
            self.progress_label.config(text="▶️ Processing 新しい演算(補修無視) - 3/9")
            self.progress_bar.config(value=40)
            self.root.update()
            sheet3 = self.apply_new_calculation_logic(max_df, structure_df, "補修無視")
            
            # Sheet 4: 新しい演算(補修考慮)
            self.progress_label.config(text="▶️ Processing 新しい演算(補修考慮) - 4/9")
            self.progress_bar.config(value=45)
            self.root.update()
            sheet4 = self.apply_new_calculation_logic(hoshuu_df, structure_df, "補修考慮")
            
            # Sheet 5: 割算結果-新しい演算(補修無視)
            self.progress_label.config(text="▶️ Processing 割算結果-新しい演算(補修無視) - 5/9")
            self.progress_bar.config(value=55)
            self.root.update()
            sheet5 = self.apply_division_calculation_logic(max_df, structure_df, "補修無視")
            
            # Sheet 6: 割算結果-新しい演算(補修考慮)
            self.progress_label.config(text="▶️ Processing 割算結果-新しい演算(補修考慮) - 6/9")
            self.progress_bar.config(value=65)
            self.root.update()
            sheet6 = self.apply_division_calculation_logic(hoshuu_df, structure_df, "補修考慮")
            
            # Sheet 7: 経時変化（橋長考慮）
            self.progress_label.config(text="▶️ Processing 経時変化（橋長考慮） - 7/9")
            self.progress_bar.config(value=75)
            self.root.update()
            sheet7 = self.apply_keiji_kyoucho_logic(grouped_df, structure_df)
            
            # Sheet 8: 経時変化（橋長&形式考慮）
            self.progress_label.config(text="▶️ Processing 経時変化（橋長&形式考慮） - 8/9")
            self.progress_bar.config(value=85)
            self.root.update()
            sheet8 = self.apply_keiji_both_logic(sheet7, structure_df, self.operator_df)
            
            # Sheet 9: 経時変化（橋長無視&形式考慮）
            self.progress_label.config(text="▶️ Processing 経時変化（橋長無視&形式考慮） - 9/9")
            self.progress_bar.config(value=90)
            self.root.update()
            sheet9 = self.apply_keiji_mushi_logic(grouped_df, structure_df, self.operator_df)
            
            # Save all sheets
            self.progress_label.config(text="💾 Saving all 9 enhanced sheets...")
            self.progress_bar.config(value=95)
            self.root.update()
            
            self.save_all_results(sheet1, sheet2, sheet3, sheet4, sheet5, sheet6, sheet7, sheet8, sheet9)
            
            # Complete processing
            self.status_label.config(text="🎉 All 9 sheets generated successfully!", fg="#27ae60")
            self.progress_label.config(text="✅ Complete: All 9 enhanced sheets created successfully")
            self.progress_bar.config(value=100)
            self.root.update()
            
            # Re-enable button
            self.select_btn.config(state="normal", text="📁 Process Another Workbook")
            
            # Auto-exit after 3 seconds
            self.root.after(3000, self.auto_exit)
            
        except Exception as e:
            self.status_label.config(text=f"❌ Processing failed: {str(e)[:50]}...", fg="#e74c3c")
            self.progress_label.config(text="❌ Error during 9-sheet processing")
            self.progress_bar.config(value=0)
            self.select_btn.config(state="normal", text="📁 Select Excel Workbook")

    def apply_division_logic(self, source_df, structure_df, sheet_type):
        """Apply division logic: Original ÷ Length"""
        result_df = source_df.copy()
        year_columns = [col for col in result_df.columns if col.endswith('結果')]
        
        # Rename year columns to include "合計重み/長さ"
        column_mapping = {}
        for col in year_columns:
            year_match = re.search(r'(\d{4})', col)
            if year_match:
                year = year_match.group(1)
                new_col_name = f"{year} 合計重み/長さ"
                column_mapping[col] = new_col_name
        
        result_df = result_df.rename(columns=column_mapping)
        
        # Apply division for each row
        for index, row in result_df.iterrows():
            kozo_name = str(row.get('構造物名称', '')).strip() if pd.notna(row.get('構造物名称', '')) else ''
            eki_start = str(row.get('駅（始）', '')).strip() if pd.notna(row.get('駅（始）', '')) else ''
            eki_end = str(row.get('駅（至）', '')).strip() if pd.notna(row.get('駅（至）', '')) else ''
            
            ekikan = f"{eki_start}→{eki_end}" if eki_start and eki_end else ''
            length_value = self.find_length_value(structure_df, kozo_name, ekikan, row.get('路線名', ''))
            
            # Divide year result columns by length
            for old_col, new_col in column_mapping.items():
                original_value = source_df.loc[index, old_col] if old_col in source_df.columns else None
                
                if pd.notna(original_value) and str(original_value).strip() not in ['', 'nan']:
                    try:
                        numeric_value = float(original_value)
                        if length_value and length_value > 0:
                            divided_value = numeric_value / length_value
                            result_df.loc[index, new_col] = float(round(divided_value, 3))
                        else:
                            result_df.loc[index, new_col] = float(numeric_value)
                    except (ValueError, TypeError):
                        result_df.loc[index, new_col] = original_value
                else:
                    result_df.loc[index, new_col] = original_value
        
        enhanced_df = self.add_enhanced_columns(result_df, self.structure_df)
        return self.reorder_columns_enhanced(enhanced_df)

    def apply_new_calculation_logic(self, source_df, structure_df, sheet_type):
        """Apply new calculation logic: X*A*B*C"""
        result_df = source_df.copy()
        year_columns = [col for col in result_df.columns if col.endswith('結果')]
        
        # Apply X*A*B*C calculation for each row
        for index, row in result_df.iterrows():
            weights = self.get_structure_weights(structure_df, row)
            
            for year_col in year_columns:
                original_value = row[year_col]
                
                if pd.notna(original_value) and str(original_value).strip() not in ['', 'nan']:
                    try:
                        x_value = float(original_value)
                        calculated_value = x_value * weights['A'] * weights['B'] * weights['C']
                        result_df.loc[index, year_col] = float(round(calculated_value, 3))
                    except (ValueError, TypeError):
                        result_df.loc[index, year_col] = original_value
                else:
                    result_df.loc[index, year_col] = original_value
        
        enhanced_df = self.add_enhanced_columns(result_df, self.structure_df)
        return self.reorder_columns_enhanced(enhanced_df)

    def apply_division_calculation_logic(self, source_df, structure_df, sheet_type):
        """Apply division calculation logic: X*A*B*C ÷ Length"""
        result_df = source_df.copy()
        year_columns = [col for col in result_df.columns if col.endswith('結果')]
        
        # Rename year columns to include division indicator
        column_mapping = {}
        for col in year_columns:
            year_match = re.search(r'(\d{4})', col)
            if year_match:
                year = year_match.group(1)
                new_col_name = f"{year} 新演算/長さ"
                column_mapping[col] = new_col_name
        
        result_df = result_df.rename(columns=column_mapping)
        
        # Apply X*A*B*C ÷ Length calculation for each row
        for index, row in result_df.iterrows():
            weights = self.get_structure_weights(structure_df, row)
            length_value = self.get_length_value_direct(structure_df, row)
            
            for old_col, new_col in column_mapping.items():
                original_value = source_df.loc[index, old_col] if old_col in source_df.columns else None
                
                if pd.notna(original_value) and str(original_value).strip() not in ['', 'nan']:
                    try:
                        x_value = float(original_value)
                        calculated_value = x_value * weights['A'] * weights['B'] * weights['C']
                        
                        if length_value and length_value > 0:
                            final_value = calculated_value / length_value
                            result_df.loc[index, new_col] = float(round(final_value, 3))
                        else:
                            result_df.loc[index, new_col] = float(round(calculated_value, 3))
                    except (ValueError, TypeError):
                        result_df.loc[index, new_col] = original_value
                else:
                    result_df.loc[index, new_col] = original_value
        
        enhanced_df = self.add_enhanced_columns(result_df, self.structure_df)
        return self.reorder_columns_enhanced(enhanced_df)

    def apply_keiji_kyoucho_logic(self, grouped_df, structure_df):
        """Apply 経時変化（橋長考慮） logic: グループ化点検履歴 ÷ Length"""
        result_df = grouped_df.copy()
        year_columns = [col for col in result_df.columns if col.endswith('結果')]
        
        # Apply division by length for each row
        for index, row in result_df.iterrows():
            length_value = self.get_length_value_direct(structure_df, row)
            
            for year_col in year_columns:
                original_value = row[year_col]
                
                if pd.notna(original_value) and str(original_value).strip() not in ['', 'nan']:
                    try:
                        numeric_value = float(original_value)
                        if length_value and length_value > 0:
                            divided_value = numeric_value / length_value
                            result_df.loc[index, year_col] = float(round(divided_value, 3))
                        else:
                            result_df.loc[index, year_col] = float(numeric_value)
                    except (ValueError, TypeError):
                        result_df.loc[index, year_col] = original_value
                else:
                    result_df.loc[index, year_col] = original_value
        
        enhanced_df = self.add_enhanced_columns(result_df, self.structure_df)
        return self.reorder_columns_enhanced(enhanced_df)

    def apply_keiji_both_logic(self, keiji_kyoucho_df, structure_df, operator_df):
        """Apply 経時変化（橋長&形式考慮） logic: Above × Structure weights"""
        result_df = keiji_kyoucho_df.copy()
        year_columns = [col for col in result_df.columns if col.endswith('結果')]
        
        # Apply structure weights multiplication for each row
        for index, row in result_df.iterrows():
            weights = self.get_structure_weights_with_operator(structure_df, operator_df, row)
            
            for year_col in year_columns:
                original_value = row[year_col]
                
                if pd.notna(original_value) and str(original_value).strip() not in ['', 'nan']:
                    try:
                        numeric_value = float(original_value)
                        calculated_value = numeric_value * weights['total_weight']
                        result_df.loc[index, year_col] = float(round(calculated_value, 3))
                    except (ValueError, TypeError):
                        result_df.loc[index, year_col] = original_value
                else:
                    result_df.loc[index, year_col] = original_value
        
        enhanced_df = self.add_enhanced_columns(result_df, self.structure_df)
        return self.reorder_columns_enhanced(enhanced_df)

    def apply_keiji_mushi_logic(self, grouped_df, structure_df, operator_df):
        """Apply 経時変化（橋長無視&形式考慮） logic: グループ化点検履歴 × Structure weights"""
        result_df = grouped_df.copy()
        year_columns = [col for col in result_df.columns if col.endswith('結果')]
        
        # Apply structure weights multiplication for each row
        for index, row in result_df.iterrows():
            weights = self.get_structure_weights_with_operator(structure_df, operator_df, row)
            
            for year_col in year_columns:
                original_value = row[year_col]
                
                if pd.notna(original_value) and str(original_value).strip() not in ['', 'nan']:
                    try:
                        numeric_value = float(original_value)
                        calculated_value = numeric_value * weights['total_weight']
                        result_df.loc[index, year_col] = float(round(calculated_value, 3))
                    except (ValueError, TypeError):
                        result_df.loc[index, year_col] = original_value
                else:
                    result_df.loc[index, year_col] = original_value
        
        enhanced_df = self.add_enhanced_columns(result_df, self.structure_df)
        return self.reorder_columns_enhanced(enhanced_df)

    def find_length_value(self, structure_df, kozo_name, ekikan, rosen_name):
        """Find length value from structure data"""
        try:
            rosen_name = str(rosen_name).strip() if pd.notna(rosen_name) else ''
            
            # First try to match by structure name
            if kozo_name:
                matches = structure_df[
                    (structure_df['構造物名称'].astype(str).str.strip() == kozo_name) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    length_val = matches.iloc[0]['長さ(m)']
                    if pd.notna(length_val) and str(length_val).strip() not in ['', 'nan']:
                        try:
                            return float(length_val)
                        except (ValueError, TypeError):
                            pass
            
            # If not found by structure name, try by station interval
            if ekikan:
                matches = structure_df[
                    (structure_df['駅間'].astype(str).str.strip() == ekikan) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    length_val = matches.iloc[0]['長さ(m)']
                    if pd.notna(length_val) and str(length_val).strip() not in ['', 'nan']:
                        try:
                            return float(length_val)
                        except (ValueError, TypeError):
                            pass
            
            # Default length if not found
            return 100.0
            
        except Exception:
            return 100.0

    def get_structure_weights(self, structure_df, row):
        """Get structure weights (A, B, C) from structure data"""
        try:
            rosen_name = str(row.get('路線名', '')).strip() if pd.notna(row.get('路線名', '')) else ''
            kozo_name = str(row.get('構造物名称', '')).strip() if pd.notna(row.get('構造物名称', '')) else ''
            
            eki_start = str(row.get('駅（始）', '')).strip() if pd.notna(row.get('駅（始）', '')) else ''
            eki_end = str(row.get('駅（至）', '')).strip() if pd.notna(row.get('駅（至）', '')) else ''
            
            ekikan = f"{eki_start}→{eki_end}" if eki_start and eki_end else ''
            
            # Default weights
            weights = {'A': 1.0, 'B': 1.0, 'C': 1.0}
            
            # First try to match by structure name
            if kozo_name:
                matches = structure_df[
                    (structure_df['構造物名称'].astype(str).str.strip() == kozo_name) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    match_row = matches.iloc[0]
                    
                    if '構造形式_重み' in match_row and pd.notna(match_row['構造形式_重み']):
                        try:
                            weights['A'] = float(match_row['構造形式_重み'])
                        except (ValueError, TypeError):
                            pass
                    
                    if '角度_重み' in match_row and pd.notna(match_row['角度_重み']):
                        try:
                            weights['B'] = float(match_row['角度_重み'])
                        except (ValueError, TypeError):
                            pass
                    
                    if '供用年数_重み' in match_row and pd.notna(match_row['供用年数_重み']):
                        try:
                            weights['C'] = float(match_row['供用年数_重み'])
                        except (ValueError, TypeError):
                            pass
                    
                    return weights
            
            # If not found by structure name, try by station interval
            if ekikan:
                matches = structure_df[
                    (structure_df['駅間'].astype(str).str.strip() == ekikan) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    match_row = matches.iloc[0]
                    
                    if '構造形式_重み' in match_row and pd.notna(match_row['構造形式_重み']):
                        try:
                            weights['A'] = float(match_row['構造形式_重み'])
                        except (ValueError, TypeError):
                            pass
                    
                    if '角度_重み' in match_row and pd.notna(match_row['角度_重み']):
                        try:
                            weights['B'] = float(match_row['角度_重み'])
                        except (ValueError, TypeError):
                            pass
                    
                    if '供用年数_重み' in match_row and pd.notna(match_row['供用年数_重み']):
                        try:
                            weights['C'] = float(match_row['供用年数_重み'])
                        except (ValueError, TypeError):
                            pass
            
            return weights
            
        except Exception:
            return {'A': 1.0, 'B': 1.0, 'C': 1.0}

    def get_length_value_direct(self, structure_df, row):
        """Get length value directly from structure data"""
        try:
            rosen_name = str(row.get('路線名', '')).strip() if pd.notna(row.get('路線名', '')) else ''
            kozo_name = str(row.get('構造物名称', '')).strip() if pd.notna(row.get('構造物名称', '')) else ''
            
            eki_start = str(row.get('駅（始）', '')).strip() if pd.notna(row.get('駅（始）', '')) else ''
            eki_end = str(row.get('駅（至）', '')).strip() if pd.notna(row.get('駅（至）', '')) else ''
            
            ekikan = f"{eki_start}→{eki_end}" if eki_start and eki_end else ''
            
            # First try to match by structure name
            if kozo_name:
                matches = structure_df[
                    (structure_df['構造物名称'].astype(str).str.strip() == kozo_name) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    length_val = matches.iloc[0]['長さ(m)']
                    if pd.notna(length_val) and str(length_val).strip() not in ['', 'nan']:
                        try:
                            return float(length_val)
                        except (ValueError, TypeError):
                            pass
            
            # If not found by structure name, try by station interval
            if ekikan:
                matches = structure_df[
                    (structure_df['駅間'].astype(str).str.strip() == ekikan) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    length_val = matches.iloc[0]['長さ(m)']
                    if pd.notna(length_val) and str(length_val).strip() not in ['', 'nan']:
                        try:
                            return float(length_val)
                        except (ValueError, TypeError):
                            pass
            
            return None
            
        except Exception:
            return None

    def get_structure_weights_with_operator(self, structure_df, operator_df, row):
        """Get structure weights using 演算子‐2 formulas with A1, B1, C1 mapping"""
        try:
            rosen_name = str(row.get('路線名', '')).strip() if pd.notna(row.get('路線名', '')) else ''
            kozo_name = str(row.get('構造物名称', '')).strip() if pd.notna(row.get('構造物名称', '')) else ''
            
            eki_start = str(row.get('駅（始）', '')).strip() if pd.notna(row.get('駅（始）', '')) else ''
            eki_end = str(row.get('駅（至）', '')).strip() if pd.notna(row.get('駅（至）', '')) else ''
            
            ekikan = f"{eki_start}→{eki_end}" if eki_start and eki_end else ''
            
            # Get values from structure data
            structure_values = {'A1': 1.0, 'B1': 1.0, 'C1': 1.0}
            
            # First try to match by structure name
            if kozo_name:
                matches = structure_df[
                    (structure_df['構造物名称'].astype(str).str.strip() == kozo_name) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    match_row = matches.iloc[0]
                    
                    if '構造形式_重み' in match_row and pd.notna(match_row['構造形式_重み']):
                        try:
                            structure_values['A1'] = float(match_row['構造形式_重み'])
                        except (ValueError, TypeError):
                            pass
                    
                    if '角度_重み' in match_row and pd.notna(match_row['角度_重み']):
                        try:
                            structure_values['B1'] = float(match_row['角度_重み'])
                        except (ValueError, TypeError):
                            pass
                    
                    if '供用年数_重み' in match_row and pd.notna(match_row['供用年数_重み']):
                        try:
                            structure_values['C1'] = float(match_row['供用年数_重み'])
                        except (ValueError, TypeError):
                            pass
            
            # If not found by structure name, try by station interval
            elif ekikan:
                matches = structure_df[
                    (structure_df['駅間'].astype(str).str.strip() == ekikan) & 
                    (structure_df['路線名'].astype(str).str.strip() == rosen_name)
                ]
                
                if not matches.empty:
                    match_row = matches.iloc[0]
                    
                    if '構造形式_重み' in match_row and pd.notna(match_row['構造形式_重み']):
                        try:
                            structure_values['A1'] = float(match_row['構造形式_重み'])
                        except (ValueError, TypeError):
                            pass
                    
                    if '角度_重み' in match_row and pd.notna(match_row['角度_重み']):
                        try:
                            structure_values['B1'] = float(match_row['角度_重み'])
                        except (ValueError, TypeError):
                            pass
                    
                    if '供用年数_重み' in match_row and pd.notna(match_row['供用年数_重み']):
                        try:
                            structure_values['C1'] = float(match_row['供用年数_重み'])
                        except (ValueError, TypeError):
                            pass
            
            # Apply 演算子‐2 formulas with dynamic evaluation
            total_weight = self.evaluate_operator_formulas(operator_df, structure_values)
            
            return {
                'A1': structure_values['A1'],
                'B1': structure_values['B1'], 
                'C1': structure_values['C1'],
                'total_weight': total_weight
            }
            
        except Exception:
            return {'A1': 1.0, 'B1': 1.0, 'C1': 1.0, 'total_weight': 1.0}

    def evaluate_operator_formulas(self, operator_df, structure_values):
        """Evaluate 演算子‐2 formulas with A1, B1, C1 mapping"""
        try:
            # Find the formula in 演算子‐2 sheet
            if len(operator_df) > 0:
                formula_row = operator_df.iloc[0]
                
                # Look for formula column
                formula = None
                for col in operator_df.columns:
                    cell_value = str(formula_row[col]).strip()
                    if any(var in cell_value for var in ['A1', 'B1', 'C1']) and any(op in cell_value for op in ['*', '+', '-', '/']):
                        formula = cell_value
                        break
                
                if formula:
                    # Replace A1, B1, C1 with actual values
                    formula = formula.replace('A1', str(structure_values['A1']))
                    formula = formula.replace('B1', str(structure_values['B1']))
                    formula = formula.replace('C1', str(structure_values['C1']))
                    
                    # Safely evaluate the formula
                    try:
                        result = eval(formula)
                        return float(result)
                    except:
                        return 1.0
            
            # Default calculation if no formula found
            return structure_values['A1'] * structure_values['B1'] * structure_values['C1']
            
        except Exception:
            return 1.0

    def save_all_results(self, sheet1, sheet2, sheet3, sheet4, sheet5, sheet6, sheet7, sheet8, sheet9):
        """Save all 9 enhanced sheets to Excel workbook"""
        try:
            with pd.ExcelWriter(self.workbook_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                # Write all 9 enhanced sheets
                sheet1.to_excel(writer, sheet_name='割算結果(補修無視)', index=False)
                sheet2.to_excel(writer, sheet_name='割算結果(補修考慮)', index=False)
                sheet3.to_excel(writer, sheet_name='新しい演算(補修無視)', index=False)
                sheet4.to_excel(writer, sheet_name='新しい演算(補修考慮)', index=False)
                sheet5.to_excel(writer, sheet_name='割算結果-新しい演算(補修無視)', index=False)
                sheet6.to_excel(writer, sheet_name='割算結果-新しい演算(補修考慮)', index=False)
                sheet7.to_excel(writer, sheet_name='経時変化（橋長考慮）', index=False)
                sheet8.to_excel(writer, sheet_name='経時変化（橋長&形式考慮）', index=False)
                sheet9.to_excel(writer, sheet_name='経時変化（橋長無視&形式考慮）', index=False)
                
        except Exception as e:
            raise Exception(f"Error saving all enhanced results: {str(e)}")

    def auto_exit(self):
        """Auto exit after successful completion"""
        self.root.quit()

    def run(self):
        """Run the unified application"""
        self.root.mainloop()


# Main execution
if __name__ == "__main__":
    app = UnifiedExcelProcessorApp()
    app.run()
                