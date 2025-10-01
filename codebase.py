import subprocess
import os
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import Workbook, load_workbook
import pandas as pd
import xlsxwriter
import tempfile
import re

class PostObserProcessorApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Post-Obser Files Processor")
        self.root.geometry("600x500")
        self.root.configure(bg="#f0f0f0")
        
        self.base_dir = None
        self.fortran_program_path = None
        self.year_ranges = None
        
        self.create_main_gui()
    
    def create_main_gui(self):
        """Create main GUI"""
        main_frame = tk.Frame(self.root, bg="#f0f0f0", padx=30, pady=20)
        main_frame.pack(fill="both", expand=True)
        
        # Title
        title_label = tk.Label(main_frame, text="Post-Obser Files Processor", 
                              font=("Arial", 16, "bold"), fg="#1565C0", bg="#f0f0f0")
        title_label.pack(pady=(0, 20))
        
        # Instructions
        instruction_text = ("This tool processes obser files through Fortran program\n"
                          "and creates enhanced Excel outputs with charts.\n\n"
                          "Required files in directory:\n"
                          "• obser1.txt to obser8.txt\n"
                          "• 劣化予測プログラム .exe (Fortran program)\n"
                          "• tamagawa-new method  - Copy .xlsx (for structure data)")
        
        instruction_label = tk.Label(main_frame, text=instruction_text, 
                                   font=("Arial", 10), justify="center", bg="#f0f0f0")
        instruction_label.pack(pady=(0, 20))
        
        # Status
        self.status_label = tk.Label(main_frame, text="Ready to select directory...", 
                                    font=("Arial", 10), fg="#666", bg="#f0f0f0")
        self.status_label.pack(pady=(0, 15))
        
        # Directory selection
        select_btn = tk.Button(main_frame, text="Select Processing Directory", 
                             command=self.select_directory, 
                             bg="#4CAF50", fg="white", 
                             width=25, height=2, font=("Arial", 11))
        select_btn.pack(pady=5)
        
        # Processing frame (initially hidden)
        self.processing_frame = tk.Frame(main_frame, bg="#f0f0f0")
        
        # Year range configuration
        config_frame = tk.LabelFrame(self.processing_frame, text="Chart Configuration", 
                                   font=("Arial", 11, "bold"), bg="#f0f0f0", padx=15, pady=10)
        config_frame.pack(fill="x", pady=(20, 10))
        
        tk.Label(config_frame, text="Year range for logdensity charts:", 
                font=("Arial", 10), bg="#f0f0f0").pack(anchor="w")
        
        year_frame = tk.Frame(config_frame, bg="#f0f0f0")
        year_frame.pack(fill="x", pady=5)
        
        tk.Label(year_frame, text="Start Year:", bg="#f0f0f0").pack(side="left")
        self.start_year_var = tk.StringVar(value="1")
        tk.Entry(year_frame, textvariable=self.start_year_var, width=5).pack(side="left", padx=5)
        
        tk.Label(year_frame, text="End Year:", bg="#f0f0f0").pack(side="left", padx=(20, 0))
        self.end_year_var = tk.StringVar(value="5")
        tk.Entry(year_frame, textvariable=self.end_year_var, width=5).pack(side="left", padx=5)
        
        # Action buttons
        button_frame = tk.Frame(self.processing_frame, bg="#f0f0f0")
        button_frame.pack(pady=20)
        
        tk.Button(button_frame, text="Start Processing", 
                 command=self.start_processing,
                 bg="#FF9800", fg="white", width=20, height=2, 
                 font=("Arial", 11)).pack(side="left", padx=5)
        
        tk.Button(button_frame, text="Clear Output Folder", 
                 command=self.clear_output_folder,
                 bg="#f44336", fg="white", width=20, height=2, 
                 font=("Arial", 11)).pack(side="left", padx=5)

    def select_directory(self):
        """Select and validate processing directory"""
        self.base_dir = filedialog.askdirectory(title="Select Directory with Obser Files")
        
        if not self.base_dir:
            return
        
        # Validate required files
        required_files = [
            f"obser{i}.txt" for i in range(1, 9)
        ] + ["劣化予測プログラム .exe", "tamagawa-new method  - Copy .xlsx"]
        
        missing_files = []
        for file in required_files:
            if not os.path.exists(os.path.join(self.base_dir, file)):
                missing_files.append(file)
        
        if missing_files:
            messagebox.showerror("Missing Files", 
                               f"Required files not found:\n" + 
                               "\n".join(missing_files))
            return
        
        self.fortran_program_path = os.path.join(self.base_dir, "劣化予測プログラム .exe")
        
        self.status_label.config(text=f"Directory selected: {os.path.basename(self.base_dir)}", 
                               fg="#4CAF50")
        self.processing_frame.pack(fill="x", pady=20)

    def clear_output_folder(self):
        """Clear the output folder"""
        if not self.base_dir:
            return
        
        output_dir = os.path.join(self.base_dir, "output")
        
        files_to_clear = [
            "出力1.xlsx", "出力2.xlsx", "出力3.xlsx", "出力4.xlsx",
            "出力5.xlsx", "出力6.xlsx", "出力7.xlsx", "出力8.xlsx",
            "作図付き出力1.xlsx", "作図付き出力2.xlsx", "作図付き出力3.xlsx",
            "作図付き出力4.xlsx", "作図付き出力5.xlsx", "作図付き出力6.xlsx",
            "作図付き出力7.xlsx", "作図付き出力8.xlsx"
        ]
        
        if os.path.exists(output_dir):
            cleared_count = 0
            for filename in files_to_clear:
                file_path = os.path.join(output_dir, filename)
                try:
                    if os.path.exists(file_path):
                        os.unlink(file_path)
                        cleared_count += 1
                except Exception as e:
                    print(f"Failed to delete {file_path}: {e}")
            
            messagebox.showinfo("Cleared", f"Cleared {cleared_count} files from output folder")
        else:
            os.makedirs(output_dir)
            messagebox.showinfo("Created", "Output folder created")

    def start_processing(self):
        """Start the main processing workflow"""
        if not self.base_dir:
            messagebox.showerror("Error", "Please select a directory first")
            return
        
        # Validate year ranges
        try:
            start_year = int(self.start_year_var.get())
            end_year = int(self.end_year_var.get())
            if start_year < 1 or end_year < start_year:
                raise ValueError
            self.year_ranges = (start_year, end_year)
        except ValueError:
            messagebox.showerror("Error", "Invalid year range")
            return
        
        # Show processing dialog
        self.show_processing_dialog()

    def show_processing_dialog(self):
        """Show processing progress dialog"""
        self.progress_window = tk.Toplevel(self.root)
        self.progress_window.title("Processing Obser Files")
        self.progress_window.geometry("500x300")
        self.progress_window.grab_set()
        self.progress_window.configure(bg="#f0f0f0")
        
        progress_frame = tk.Frame(self.progress_window, bg="#f0f0f0", padx=20, pady=20)
        progress_frame.pack(fill="both", expand=True)
        
        tk.Label(progress_frame, text="Processing Obser Files", 
                font=("Arial", 14, "bold"), bg="#f0f0f0").pack(pady=(0, 20))
        
        self.progress_text = tk.Text(progress_frame, height=12, width=60, 
                                   font=("Consolas", 9))
        scrollbar = ttk.Scrollbar(progress_frame, orient="vertical", 
                                command=self.progress_text.yview)
        self.progress_text.configure(yscrollcommand=scrollbar.set)
        
        self.progress_text.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        self.progress_bar = ttk.Progressbar(progress_frame, mode='determinate', maximum=8)
        self.progress_bar.pack(fill="x", pady=(10, 0))
        
        # Start processing
        self.root.after(100, self.execute_processing)

    def log_progress(self, message):
        """Log progress message"""
        self.progress_text.insert(tk.END, f"{message}\n")
        self.progress_text.see(tk.END)
        self.progress_window.update()

    def execute_processing(self):
        """Execute the main processing logic"""
        try:
            output_dir = os.path.join(self.base_dir, "output")
            temp_dir = os.path.join(self.base_dir, "temp_obser")
            
            # Clear and create directories
            self.log_progress("Setting up directories...")
            if os.path.exists(output_dir):
                shutil.rmtree(output_dir)
            os.makedirs(output_dir)
            
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
            os.makedirs(temp_dir)
            
            files_to_write = ["pml.txt", "logdensity.txt", "ex1000.txt"]
            
            # Process first obser file (obser1.txt)
            self.log_progress("Processing obser1.txt...")
            self.run_fortran_program()
            self.write_to_excel(self.base_dir, output_dir, files_to_write, "出力1.xlsx")
            
            # Store original obser1.txt
            original_obser1 = os.path.join(self.base_dir, "obser1.txt")
            temp_obser1 = os.path.join(temp_dir, "obser1_temp_1.txt")
            os.rename(original_obser1, temp_obser1)
            
            self.progress_bar['value'] = 1
            self.progress_window.update()

                        
            # Process remaining obser files (obser2.txt to obser8.txt)
            for i in range(2, 9):
                self.log_progress(f"Processing obser{i}.txt...")
                
                # Rename current obser file to obser1.txt
                current_obser = os.path.join(self.base_dir, f"obser{i}.txt")
                if not os.path.exists(current_obser):
                    self.log_progress(f"Warning: obser{i}.txt not found, skipping...")
                    continue
                
                os.rename(current_obser, original_obser1)
                
                # Run Fortran program
                self.run_fortran_program()
                
                # Create Excel output
                workbook_name = f"出力{i}.xlsx"
                self.write_to_excel(self.base_dir, output_dir, files_to_write, workbook_name)
                
                # Store processed file
                temp_obser_path = os.path.join(temp_dir, f"obser1_temp_{i}.txt")
                os.rename(original_obser1, temp_obser_path)
                
                self.progress_bar['value'] = i
                self.progress_window.update()
            
            # Restore original obser files
            self.log_progress("Restoring original obser files...")
            for i in range(1, 9):
                temp_obser_path = os.path.join(temp_dir, f"obser1_temp_{i}.txt")
                original_path = os.path.join(self.base_dir, f"obser{i}.txt")
                if os.path.exists(temp_obser_path):
                    os.rename(temp_obser_path, original_path)
            
            # Remove temp directory
            shutil.rmtree(temp_dir)
            
            self.log_progress("Starting chart generation process...")
            self.execute_chart_generation(output_dir)
            
            self.progress_window.destroy()
            
            # Show completion dialog
            self.show_completion_dialog()
            
        except Exception as e:
            self.progress_window.destroy()
            messagebox.showerror("Error", f"Processing failed:\n{str(e)}")

    def run_fortran_program(self):
        """Run the Fortran program without timeout"""
        try:
            result = subprocess.run([self.fortran_program_path], 
                                cwd=self.base_dir, 
                                check=True, 
                                capture_output=True, 
                                text=True)
            # No timeout parameter - let it run as long as needed
        except subprocess.CalledProcessError as e:
            raise Exception(f"Fortran program execution failed: {e}")
    
    def write_to_excel(self, source_dir, output_dir, files_to_write, workbook_name):
        """Write text files to Excel workbook - copy data exactly as-is"""
        try:
            workbook = Workbook()
            
            for file_name in files_to_write:
                sheet_name = file_name.split('.')[0]
                sheet = workbook.create_sheet(title=sheet_name)
                file_path = os.path.join(source_dir, file_name)
                
                if os.path.exists(file_path):
                    with open(file_path, 'r', encoding='utf-8') as file:
                        for row_idx, line in enumerate(file):
                            line = line.strip()
                            if line:  # Skip empty lines
                                values = line.split()
                                for col_idx, value in enumerate(values):
                                    # Just copy the value as text - no conversion
                                    sheet.cell(row=row_idx + 1, column=col_idx + 1, value=value)
            
            # Remove the default sheet
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
            
            excel_path = os.path.join(output_dir, workbook_name)
            workbook.save(excel_path)
            
        except Exception as e:
            raise Exception(f"Error writing to Excel {workbook_name}: {e}")

    def execute_chart_generation(self, output_folder):
        """Execute chart generation for all output files"""
        try:
            tamagawa_file = os.path.join(self.base_dir, 'tamagawa-new method  - Copy .xlsx')
            
            if not os.path.exists(tamagawa_file):
                self.log_progress("Warning: tamagawa-new method  - Copy .xlsx not found, skipping chart generation")
                return
            
            # Load structure data
            wb = load_workbook(tamagawa_file, data_only=True)
            values_dict = {}
            
            for sheet in wb.worksheets:
                values = {'構造物名称': [], '構造物番号': []}
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row and len(row) >= 4:
                        values['構造物名称'].append(row[2] if row[2] else '')
                        values['構造物番号'].append(row[3] if row[3] else '')
                values_dict[sheet.title] = values
            
            # Process output files
            output_files = [f for f in os.listdir(output_folder) 
                          if f.endswith('.xlsx') and f.startswith('出力') 
                          and not f.startswith('作図付き')]
            
            for file in output_files:
                self.log_progress(f"Creating charts for {file}...")
                file_path = os.path.join(output_folder, file)
                sheet_name = self.get_sheet_name(file)
                
                if sheet_name in values_dict:
                    self.process_file_with_charts(file_path, values_dict[sheet_name], 
                                                output_folder, self.year_ranges)
                else:
                    self.log_progress(f"Warning: No structure data found for {sheet_name}")
            
        except Exception as e:
            self.log_progress(f"Chart generation error: {e}")

    def get_sheet_name(self, output_filename):
        """Get sheet name mapping for output file"""
        sheet_mapping = {
            '出力1.xlsx': '割算結果(補修考慮)',              # ← Fixed: Added parentheses
            '出力2.xlsx': '割算結果(補修無視)',              # ← Fixed: Added parentheses  
            '出力3.xlsx': '補修無視',                       # ← Already correct
            '出力4.xlsx': '補修考慮',                       # ← Already correct
            '出力5.xlsx': '新しい演算(補修無視)',            # ← Fixed: Added parentheses
            '出力6.xlsx': '新しい演算(補修考慮)',            # ← Fixed: Added parentheses
            '出力7.xlsx': '割算結果-新しい演算(補修無視)',    # ← Fixed: Added parentheses
            '出力8.xlsx': '割算結果-新しい演算(補修考慮)'     # ← Fixed: Added parentheses
        }
        return sheet_mapping.get(output_filename, '割算結果(補修考慮)')  # ← Fixed default too

    def process_file_with_charts(self, file_path, values, output_folder_path, year_ranges):
        """Process file and add charts"""
        try:
            wb = load_workbook(file_path)
            
            # Get or create sheets
            ex_ws = wb['ex1000'] if 'ex1000' in wb.sheetnames else None
            log_ws = wb['logdensity'] if 'logdensity' in wb.sheetnames else wb.create_sheet('logdensity')
            pml_ws = wb['pml'] if 'pml' in wb.sheetnames else None

            # Format sheets
            if ex_ws:
                self.format_ex1000(ex_ws, values)
            self.format_logdensity(log_ws)
            if pml_ws:
                self.format_pml(pml_ws)

            # Create new filename
            base_name = os.path.basename(file_path)
            if '_' in base_name:
                modified_filename = f"作図付き出力{base_name.split('_')[1].split('.')[0]}.xlsx"
            else:
                modified_filename = f"作図付き出力{base_name.split('.')[0][-1]}.xlsx"

            new_file_path = os.path.join(output_folder_path, modified_filename)

            # Save to temporary file first
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                temp_path = tmp.name
            wb.save(temp_path)
            wb.close()

            # Read data and create charts with xlsxwriter
            self.create_charts_with_xlsxwriter(temp_path, new_file_path, year_ranges)
            
            # Clean up temp file
            os.remove(temp_path)
            
        except Exception as e:
            raise Exception(f"Error processing file {file_path}: {e}")

    def format_ex1000(self, ws, values):
        """Format ex1000 sheet"""
        if ws.cell(row=1, column=1).value != '順位':
            ws.insert_rows(1, amount=3)
            ws.cell(row=1, column=1, value='順位')
            
            for col_num in range(2, ws.max_column + 1):
                ws.cell(row=1, column=col_num, value=col_num - 1)
            
            ws.cell(row=2, column=1, value='構造物番号・駅間')
            for col_num, number in enumerate(values['構造物番号'], start=2):
                ws.cell(row=2, column=col_num, value=f'({number})')
            
            ws.cell(row=3, column=1, value='構造物名称')
            for col_num, name in enumerate(values['構造物名称'], start=2):
                ws.cell(row=3, column=col_num, value=name)
        
        # Format data
        for row in range(4, ws.max_row + 1):
            try:
                ws.cell(row=row, column=1).value = int(float(ws.cell(row=row, column=1).value))
            except:
                pass
            for col in range(2, ws.max_column + 1):
                try:
                    ws.cell(row=row, column=col).value = float(ws.cell(row=row, column=col).value)
                except:
                    ws.cell(row=row, column=col).value = None

    def format_logdensity(self, ws):
        """Format logdensity sheet"""
        ws.insert_rows(1)
        for col in range(2, ws.max_column + 1):
            ws.cell(row=1, column=col, value=f'経過{col - 1}年目')
        
        for row in range(2, ws.max_row + 1):
            try:
                ws.cell(row=row, column=1).value = int(float(ws.cell(row=row, column=1).value))
            except:
                pass
            for col in range(2, ws.max_column + 1):
                try:
                    ws.cell(row=row, column=col).value = float(ws.cell(row=row, column=col).value)
                except:
                    ws.cell(row=row, column=col).value = None

    def format_pml(self, ws):
        """Format pml sheet"""
        ws.insert_cols(1)
        ws.insert_rows(1)
        ws.cell(row=1, column=1, value='経過年数')
        ws.cell(row=1, column=2, value='年')
        ws.cell(row=1, column=3, value='NEL (0.5:0.5)')
        ws.cell(row=1, column=4, value='PML(0.9:0.1)')
        ws.cell(row=1, column=5, value='PML_0.95 (0.95:0.05)')
        
        for i in range(2, ws.max_row + 1):
            ws.cell(row=i, column=1, value=i - 1)
        
        for col in range(2, ws.max_column + 1):
            for row in range(2, ws.max_row + 1):
                try:
                    ws.cell(row=row, column=col).value = int(float(ws.cell(row=row, column=col).value))
                except:
                    pass

    def create_charts_with_xlsxwriter(self, temp_path, new_file_path, year_ranges):
        """Create charts using xlsxwriter"""
        try:
            # Read data from temp file
            df_ex = None
            df_log = None
            df_pml = None
            
            try:
                df_ex = pd.read_excel(temp_path, sheet_name='ex1000', header=None)
            except:
                pass
            
            try:
                df_log = pd.read_excel(temp_path, sheet_name='logdensity', header=None)
            except:
                pass
            
            try:
                df_pml = pd.read_excel(temp_path, sheet_name='pml', header=0)
            except:
                pass

            # Create new workbook with charts
            with pd.ExcelWriter(new_file_path, engine='xlsxwriter') as writer:
                workbook = writer.book

                # Process ex1000 sheet with chart
                if df_ex is not None:
                    df_ex = df_ex.where(pd.notnull(df_ex), None)
                    ws = workbook.add_worksheet('ex1000')
                    writer.sheets['ex1000'] = ws

                    # Formatting
                    fmt_bold = workbook.add_format({'bold': True})
                    fmt_int = workbook.add_format({'num_format': '0'})
                    fmt_float = workbook.add_format({'num_format': '0.00000'})

                    # Write data
                    for row_num in range(df_ex.shape[0]):
                        for col_num in range(df_ex.shape[1]):
                            val = df_ex.iat[row_num, col_num]
                            if pd.isna(val):
                                ws.write(row_num, col_num, None)
                            elif row_num < 3:
                                ws.write(row_num, col_num, val, fmt_bold)
                            elif col_num == 0:
                                ws.write(row_num, col_num, val, fmt_int)
                            else:
                                ws.write(row_num, col_num, val, fmt_float)

                    # Create chart for ex1000
                    last = df_ex.iloc[3:].dropna(how='all').index[-1] if not df_ex.iloc[3:].dropna(how='all').empty else 3
                    chart = workbook.add_chart({'type': 'line'})
                    
                    for i in range(1, df_ex.shape[1]):
                        chart.add_series({
                            'name': ['ex1000', 2, i],
                            'categories': ['ex1000', 3, 0, last, 0],
                            'values': ['ex1000', 3, i, last, i],
                        })
                    
                    chart.set_title({'name': '経過年 vs. しきい値の強度確率'})
                    chart.set_x_axis({'name': '経過年', 'position_axis': 'on_tick'})
                    chart.set_y_axis({'name': 'しきい値の強度確率', 'num_format': '0%'})
                    ws.insert_chart(f'A{last + 6}', chart)

                # Process logdensity sheet with chart
                if df_log is not None:
                    df_log = df_log.where(pd.notnull(df_log), None)
                    df_log.to_excel(writer, sheet_name='logdensity', index=False, header=False)
                    ws2 = writer.sheets['logdensity']

                    # Formatting
                    col_a_format = workbook.add_format({'num_format': '0'})
                    col_rest_format = workbook.add_format({'num_format': '0.00000'})
                    ws2.set_column('A:A', 8, col_a_format)
                    end_col_letter = chr(ord('A') + df_log.shape[1] - 1)
                    ws2.set_column(f'B:{end_col_letter}', 12, col_rest_format)

                    # Create chart for logdensity
                    start_year, end_year = year_ranges
                    
                    if (start_year is not None and end_year is not None and 
                        start_year >= 1 and end_year <= df_log.shape[1] - 1 and 
                        start_year <= end_year):
                        
                        last2 = df_log.iloc[1:].dropna(how='all').index[-1] if not df_log.iloc[1:].dropna(how='all').empty else 1
                        chart2 = workbook.add_chart({'type': 'line'})
                        
                        for i in range(start_year, end_year + 1):
                            chart2.add_series({
                                'name': ['logdensity', 0, i],
                                'categories': ['logdensity', 1, 0, last2, 0],
                                'values': ['logdensity', 1, i, last2, i],
                            })
                        
                        chart2.set_title({'name': '経過年 vs. しきい値の強度確率'})
                        chart2.set_x_axis({'name': '劣化点数', 'position_axis': 'on_tick'})
                        chart2.set_y_axis({'name': '確率密度関数', 'num_format': '0.00000'})

                        last_col = chr(ord('A') + df_log.shape[1] - 1)
                        insert_col = chr(ord(last_col) + 3)
                        ws2.insert_chart(f'{insert_col}2', chart2)

                # Process pml sheet with chart
                if df_pml is not None:
                    df_pml = df_pml.where(pd.notnull(df_pml), None)
                    df_pml.to_excel(writer, sheet_name='pml', index=False, header=True)
                    ws3 = writer.sheets['pml']

                    # Formatting
                    col_a_format = workbook.add_format({'num_format': '0'})
                    col_rest_format = workbook.add_format({'num_format': '0'})
                    ws3.set_column('A:A', 8, col_a_format)
                    end_col_letter = chr(ord('A') + df_pml.shape[1] - 1)
                    ws3.set_column(f'B:{end_col_letter}', 12, col_rest_format)

                    # Create chart for pml
                    last3 = df_pml.shape[0]
                    chart3 = workbook.add_chart({'type': 'line'})
                    
                    for i in range(2, df_pml.shape[1]):
                        chart3.add_series({
                            'name': ['pml', 0, i],
                            'categories': ['pml', 1, 0, last3, 0],
                            'values': ['pml', 1, i, last3, i],
                            'marker': {'type': 'circle'}
                        })
                    
                    chart3.set_title({'name': 'PML Data'})
                    chart3.set_x_axis({'name': '経過年数', 'position_axis': 'on_tick'})
                    chart3.set_y_axis({'name': '劣化点数'})

                    last_col_pml = chr(ord('A') + df_pml.shape[1] - 1)
                    insert_col_pml = chr(ord(last_col_pml) + 3)
                    ws3.insert_chart(f'{insert_col_pml}2', chart3)

        except Exception as e:
            raise Exception(f"Error creating charts: {e}")

    def show_completion_dialog(self):
        """Show completion dialog"""
        completion_window = tk.Toplevel(self.root)
        completion_window.title("Processing Complete")
        completion_window.geometry("500x400")
        completion_window.grab_set()
        completion_window.configure(bg="#f0f0f0")
        
        # Center window
        completion_window.update_idletasks()
        x = (completion_window.winfo_screenwidth() // 2) - (500 // 2)
        y = (completion_window.winfo_screenheight() // 2) - (400 // 2)
        completion_window.geometry(f"500x400+{x}+{y}")
        
        main_frame = tk.Frame(completion_window, bg="#f0f0f0", padx=25, pady=20)
        main_frame.pack(fill="both", expand=True)
        
        # Success title
        success_label = tk.Label(main_frame, text="✅ Processing Complete!", 
                                font=("Arial", 16, "bold"), fg="#4CAF50", bg="#f0f0f0")
        success_label.pack(pady=(0, 20))
        
        # Summary
        summary_text = ("Successfully processed all obser files!\n\n"
                       "Generated Files:\n"
                       "• 8 Basic output files (出力1.xlsx - 出力8.xlsx)\n"
                       "• 8 Chart-enhanced files (作図付き出力1.xlsx - 作図付き出力8.xlsx)\n\n"
                       "Each chart file contains:\n"
                       "• ex1000 sheet with line charts\n"
                       "• logdensity sheet with probability density charts\n"
                       "• pml sheet with PML data charts")
        
        summary_label = tk.Label(main_frame, text=summary_text, 
                                font=("Arial", 11), justify="left", bg="#f0f0f0")
        summary_label.pack(pady=(0, 20))
        
        # Output location
        output_dir = os.path.join(self.base_dir, "output")
        location_label = tk.Label(main_frame, text=f"📁 Output Location:\n{output_dir}", 
                                 font=("Arial", 10), fg="#666", bg="#f0f0f0")
        location_label.pack(pady=(0, 20))
        
        # Buttons
        button_frame = tk.Frame(main_frame, bg="#f0f0f0")
        button_frame.pack(pady=20)
        
        def open_output_folder():
            try:
                os.startfile(output_dir)
            except:
                messagebox.showinfo("Info", f"Please open folder manually:\n{output_dir}")
        
        def close_app():
            completion_window.destroy()
            self.root.quit()
        
        tk.Button(button_frame, text="Open Output Folder", 
                 command=open_output_folder,
                 bg="#4CAF50", fg="white", width=18, height=2, 
                 font=("Arial", 11)).pack(side="left", padx=10)
        
        tk.Button(button_frame, text="Process Another Directory", 
                 command=lambda: [completion_window.destroy(), self.reset_app()],
                 bg="#2196F3", fg="white", width=20, height=2, 
                 font=("Arial", 11)).pack(side="left", padx=10)
        
        tk.Button(button_frame, text="Exit Application", 
                 command=close_app,
                 bg="#f44336", fg="white", width=15, height=2, 
                 font=("Arial", 11)).pack(side="left", padx=10)

    def reset_app(self):
        """Reset application for processing another directory"""
        self.base_dir = None
        self.fortran_program_path = None
        self.year_ranges = None
        
        # Hide processing frame and reset status
        self.processing_frame.pack_forget()
        self.status_label.config(text="Ready to select directory...", fg="#666")
        
        # Reset year range inputs
        self.start_year_var.set("1")
        self.end_year_var.set("5")

    def run(self):
        """Run the application"""
        self.root.mainloop()


class EnhancedPostProcessorApp:
    """Enhanced version with additional features and better error handling"""
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Enhanced Post-Obser Files Processor")
        self.root.geometry("700x600")
        self.root.configure(bg="#f0f0f0")
        
        self.base_dir = None
        self.fortran_program_path = None
        self.year_ranges = None
        self.processing_settings = {
            'create_charts': True,
            'backup_original': True,
            'detailed_logging': True
        }
        
        self.create_enhanced_gui()
    
    def create_enhanced_gui(self):
        """Create enhanced GUI with additional options"""
        main_frame = tk.Frame(self.root, bg="#f0f0f0", padx=30, pady=20)
        main_frame.pack(fill="both", expand=True)
        
        # Title with subtitle
        title_label = tk.Label(main_frame, text="Enhanced Post-Obser Files Processor", 
                              font=("Arial", 18, "bold"), fg="#1565C0", bg="#f0f0f0")
        title_label.pack(pady=(0, 5))
        
        subtitle_label = tk.Label(main_frame, text="Process obser files with Fortran program and create enhanced Excel outputs", 
                                 font=("Arial", 10), fg="#666", bg="#f0f0f0")
        subtitle_label.pack(pady=(0, 20))
        
        # Requirements section
        req_frame = tk.LabelFrame(main_frame, text="Requirements", 
                                 font=("Arial", 11, "bold"), bg="#f0f0f0", padx=15, pady=10)
        req_frame.pack(fill="x", pady=(0, 20))
        
        req_text = ("📁 Directory must contain:\n"
                   "  • obser1.txt to obser8.txt (generated by obser file generator)\n"
                   "  • 劣化予測プログラム .exe (Fortran program)\n"
                   "  • tamagawa-new method  - Copy .xlsx (structure data for charts)\n\n"
                   "📊 Output will be created in 'output' subfolder:\n"
                   "  • Basic Excel files (出力1.xlsx - 出力8.xlsx)\n"
                   "  • Chart-enhanced files (作図付き出力1.xlsx - 作図付き出力8.xlsx)")
        
        tk.Label(req_frame, text=req_text, font=("Arial", 9), 
                justify="left", bg="#f0f0f0").pack(anchor="w")
        
        # Status
        self.status_label = tk.Label(main_frame, text="Ready to select directory...", 
                                    font=("Arial", 11), fg="#666", bg="#f0f0f0")
        self.status_label.pack(pady=(0, 15))
        
        # Directory selection
        select_btn = tk.Button(main_frame, text="📁 Select Processing Directory", 
                             command=self.select_directory, 
                             bg="#4CAF50", fg="white", 
                             width=30, height=2, font=("Arial", 12, "bold"))
        select_btn.pack(pady=10)
        
        # Enhanced processing frame (initially hidden)
        self.processing_frame = tk.Frame(main_frame, bg="#f0f0f0")
        
        # Settings section
        settings_frame = tk.LabelFrame(self.processing_frame, text="Processing Settings", 
                                     font=("Arial", 11, "bold"), bg="#f0f0f0", padx=15, pady=10)
        settings_frame.pack(fill="x", pady=(20, 10))
        
        # Chart configuration
        chart_frame = tk.Frame(settings_frame, bg="#f0f0f0")
        chart_frame.pack(fill="x", pady=5)
        
        tk.Label(chart_frame, text="Year range for logdensity charts:", 
                font=("Arial", 10), bg="#f0f0f0").pack(anchor="w")
        
        year_controls = tk.Frame(chart_frame, bg="#f0f0f0")
        year_controls.pack(fill="x", pady=5)
        
        tk.Label(year_controls, text="Start:", bg="#f0f0f0").pack(side="left")
        self.start_year_var = tk.StringVar(value="1")
        tk.Entry(year_controls, textvariable=self.start_year_var, width=5).pack(side="left", padx=5)
        
        tk.Label(year_controls, text="End:", bg="#f0f0f0").pack(side="left", padx=(20, 0))
        self.end_year_var = tk.StringVar(value="5")
        tk.Entry(year_controls, textvariable=self.end_year_var, width=5).pack(side="left", padx=5)
        
        tk.Label(year_controls, text="(1-based indexing)", 
                font=("Arial", 8), fg="#666", bg="#f0f0f0").pack(side="left", padx=(10, 0))
        
        # Processing options
        options_frame = tk.Frame(settings_frame, bg="#f0f0f0")
        options_frame.pack(fill="x", pady=(10, 5))
        
        self.create_charts_var = tk.BooleanVar(value=True)
        tk.Checkbutton(options_frame, text="Create charts in Excel files", 
                      variable=self.create_charts_var, bg="#f0f0f0").pack(anchor="w")
        
        self.backup_var = tk.BooleanVar(value=True)
        tk.Checkbutton(options_frame, text="Backup original obser files during processing", 
                      variable=self.backup_var, bg="#f0f0f0").pack(anchor="w")
        
        self.detailed_log_var = tk.BooleanVar(value=True)
        tk.Checkbutton(options_frame, text="Show detailed processing logs", 
                      variable=self.detailed_log_var, bg="#f0f0f0").pack(anchor="w")
        
        # Action buttons
        button_frame = tk.Frame(self.processing_frame, bg="#f0f0f0")
        button_frame.pack(pady=20)
        
        tk.Button(button_frame, text="🚀 Start Processing", 
                 command=self.start_enhanced_processing,
                 bg="#FF9800", fg="white", width=20, height=2, 
                 font=("Arial", 12, "bold")).pack(side="left", padx=10)
        
        tk.Button(button_frame, text="🗑️ Clear Output Folder", 
                 command=self.clear_output_folder,
                 bg="#f44336", fg="white", width=18, height=2, 
                 font=("Arial", 11)).pack(side="left", padx=10)
        
        tk.Button(button_frame, text="📊 Open Output Folder", 
                 command=self.open_output_folder,
                 bg="#2196F3", fg="white", width=18, height=2, 
                 font=("Arial", 11)).pack(side="left", padx=10)

    def select_directory(self):
        """Enhanced directory selection with better validation"""
        self.base_dir = filedialog.askdirectory(title="Select Directory with Obser Files")
        
        if not self.base_dir:
            return
        
        # Enhanced validation
        validation_results = self.validate_directory()
        
        if validation_results['valid']:
            self.fortran_program_path = os.path.join(self.base_dir, "劣化予測プログラム .exe")
            self.status_label.config(text=f"✅ Directory validated: {os.path.basename(self.base_dir)}", 
                                   fg="#4CAF50")
            self.processing_frame.pack(fill="x", pady=20)
        else:
            self.show_validation_error(validation_results)

    def validate_directory(self):

        print("Files in directory:")
        for file in os.listdir(self.base_dir):
            print(f"  '{file}'")
        
        """Enhanced directory validation"""
        results = {
            'valid': True,
            'missing_obser': [],
            'missing_programs': [],
            'warnings': []
        }
        
        # Check obser files
        for i in range(1, 9):
            obser_file = f"obser{i}.txt"
            if not os.path.exists(os.path.join(self.base_dir, obser_file)):
                results['missing_obser'].append(obser_file)
        
        # Check required programs
        required_programs = ["劣化予測プログラム .exe"]
        for program in required_programs:
            if not os.path.exists(os.path.join(self.base_dir, program)):
                results['missing_programs'].append(program)
        
        # Check optional files
        if not os.path.exists(os.path.join(self.base_dir, "tamagawa-new method  - Copy .xlsx")):
            results['warnings'].append("tamagawa-new method  - Copy .xlsx not found - charts will not be generated")
        
        # Determine if valid
        if results['missing_obser'] or results['missing_programs']:
            results['valid'] = False
        
        return results

    def show_validation_error(self, results):
        """Show detailed validation error"""
        error_window = tk.Toplevel(self.root)
        error_window.title("Directory Validation Failed")
        error_window.geometry("500x400")
        error_window.grab_set()
        error_window.configure(bg="#f0f0f0")
        
        main_frame = tk.Frame(error_window, bg="#f0f0f0", padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)
        
        tk.Label(main_frame, text="❌ Directory Validation Failed", 
                font=("Arial", 14, "bold"), fg="#f44336", bg="#f0f0f0").pack(pady=(0, 20))
        
        # Create scrollable text area
        text_frame = tk.Frame(main_frame, bg="#f0f0f0")
        text_frame.pack(fill="both", expand=True, pady=(0, 20))
        
        text_area = tk.Text(text_frame, wrap=tk.WORD, font=("Consolas", 10))
        scrollbar = ttk.Scrollbar(text_frame, orient="vertical", command=text_area.yview)
        text_area.configure(yscrollcommand=scrollbar.set)
        
        text_area.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Build error message
        error_msg = f"Selected directory: {self.base_dir}\n\n"
        
        if results['missing_obser']:
            error_msg += "❌ Missing Obser Files:\n"
            for file in results['missing_obser']:
                error_msg += f"  • {file}\n"
            error_msg += "\n"
        
        if results['missing_programs']:
            error_msg += "❌ Missing Required Programs:\n"
            for program in results['missing_programs']:
                error_msg += f"  • {program}\n"
            error_msg += "\n"
        
        if results['warnings']:
            error_msg += "⚠️ Warnings:\n"
            for warning in results['warnings']:
                error_msg += f"  • {warning}\n"
            error_msg += "\n"
        
        error_msg += ("📋 Required Files:\n"
                     "  • obser1.txt to obser8.txt\n"
                     "  • 劣化予測プログラム .exe\n"
                     "  • tamagawa-new method  - Copy .xlsx (optional, for charts)\n\n"
                     "Please ensure all required files are in the selected directory.")
        
        text_area.insert("1.0", error_msg)
        text_area.config(state="disabled")
        
        # Close button
        tk.Button(main_frame, text="Close", command=error_window.destroy,
                 bg="#f44336", fg="white", width=15, height=2).pack()

    def start_enhanced_processing(self):
        """Start enhanced processing with settings"""
        if not self.base_dir:
            messagebox.showerror("Error", "Please select a directory first")
            return
        
        # Validate settings
        try:
            start_year = int(self.start_year_var.get())
            end_year = int(self.end_year_var.get())
            if start_year < 1 or end_year < start_year:
                raise ValueError
            self.year_ranges = (start_year, end_year)
        except ValueError:
            messagebox.showerror("Error", "Invalid year range. Please enter valid positive integers.")
            return
        
        # Update processing settings
        self.processing_settings.update({
            'create_charts': self.create_charts_var.get(),
            'backup_original': self.backup_var.get(),
            'detailed_logging': self.detailed_log_var.get()
        })
        
        # Show enhanced processing dialog
        self.show_enhanced_processing_dialog()

    def show_enhanced_processing_dialog(self):
        """Show enhanced processing dialog with better progress tracking"""
        self.progress_window = tk.Toplevel(self.root)
        self.progress_window.title("Processing Obser Files")
        self.progress_window.geometry("600x450")
        self.progress_window.grab_set()
        self.progress_window.configure(bg="#f0f0f0")
        
        # Center window
        self.progress_window.update_idletasks()
        x = (self.progress_window.winfo_screenwidth() // 2) - (600 // 2)
        y = (self.progress_window.winfo_screenheight() // 2) - (450 // 2)
        self.progress_window.geometry(f"600x450+{x}+{y}")
        
        progress_frame = tk.Frame(self.progress_window, bg="#f0f0f0", padx=20, pady=20)
        progress_frame.pack(fill="both", expand=True)
        
        tk.Label(progress_frame, text="🔄 Processing Obser Files", 
                font=("Arial", 16, "bold"), bg="#f0f0f0").pack(pady=(0, 20))
        
        # Progress bars
        tk.Label(progress_frame, text="Overall Progress:", 
                font=("Arial", 11), bg="#f0f0f0").pack(anchor="w")
        self.overall_progress = ttk.Progressbar(progress_frame, mode='determinate', maximum=10)
        self.overall_progress.pack(fill="x", pady=(5, 15))
        
        tk.Label(progress_frame, text="Current Step:", 
                font=("Arial", 11), bg="#f0f0f0").pack(anchor="w")
        self.step_progress = ttk.Progressbar(progress_frame, mode='determinate', maximum=8)
        self.step_progress.pack(fill="x", pady=(5, 15))
        
        # Status and log
        self.current_status = tk.Label(progress_frame, text="Initializing...", 
                                      font=("Arial", 11, "bold"), fg="#1565C0", bg="#f0f0f0")
        self.current_status.pack(pady=(0, 10))
        
        # Log area
        log_frame = tk.Frame(progress_frame, bg="#f0f0f0")
        log_frame.pack(fill="both", expand=True, pady=(0, 15))
        
        self.progress_text = tk.Text(log_frame, height=12, width=70, 
                                   font=("Consolas", 9), bg="#f8f8f8")
        log_scrollbar = ttk.Scrollbar(log_frame, orient="vertical", 
                                    command=self.progress_text.yview)
        self.progress_text.configure(yscrollcommand=log_scrollbar.set)
        
        self.progress_text.pack(side="left", fill="both", expand=True)
        log_scrollbar.pack(side="right", fill="y")
        
        # Cancel button (initially enabled)
        self.cancel_btn = tk.Button(progress_frame, text="Cancel Processing", 
                                   command=self.cancel_processing,
                                   bg="#f44336", fg="white", width=20, height=2)
        self.cancel_btn.pack()
        
        # Start processing
        self.processing_cancelled = False
        self.root.after(100, self.execute_enhanced_processing)

    def cancel_processing(self):
        """Cancel the processing"""
        self.processing_cancelled = True
        self.cancel_btn.config(text="Cancelling...", state="disabled")
        self.log_enhanced_progress("❌ Processing cancelled by user")

    def log_enhanced_progress(self, message, level="INFO"):
        """Enhanced progress logging with levels"""
        timestamp = pd.Timestamp.now().strftime("%H:%M:%S")
        
        if level == "ERROR":
            prefix = "❌"
        elif level == "WARNING":
            prefix = "⚠️"
        elif level == "SUCCESS":
            prefix = "✅"
        else:
            prefix = "ℹ️"
        
        log_message = f"[{timestamp}] {prefix} {message}\n"
        
        self.progress_text.insert(tk.END, log_message)
        self.progress_text.see(tk.END)
        
        if self.processing_settings['detailed_logging']:
            print(log_message.strip())
        
        self.progress_window.update()

    def update_status(self, status, step_progress=None, overall_progress=None):
        """Update status and progress bars"""
        self.current_status.config(text=status)
        
        if step_progress is not None:
            self.step_progress['value'] = step_progress
        
        if overall_progress is not None:
            self.overall_progress['value'] = overall_progress
        
        self.progress_window.update()

    def execute_enhanced_processing(self):
        """Execute enhanced processing with better error handling"""
        try:
            self.update_status("🔧 Setting up directories...", 0, 1)
            
            output_dir = os.path.join(self.base_dir, "output")
            temp_dir = os.path.join(self.base_dir, "temp_obser")
            backup_dir = os.path.join(self.base_dir, "backup_obser") if self.processing_settings['backup_original'] else None
            
            # Setup directories
            self.setup_directories(output_dir, temp_dir, backup_dir)
            
            if self.processing_cancelled:
                return
            
            self.update_status("📝 Processing obser files...", 0, 2)
            files_to_write = ["pml.txt", "logdensity.txt", "ex1000.txt"]
            
            # Create backup if requested
            if backup_dir:
                self.create_backup(backup_dir)
            
            # Process first obser file
            self.update_status("🔄 Processing obser1.txt...", 1, 3)
            self.log_enhanced_progress("Processing obser1.txt...")
            
            if not self.processing_cancelled:
                self.run_fortran_program_safe()
                self.write_to_excel_safe(self.base_dir, output_dir, files_to_write, "出力1.xlsx")
                
                # Store original obser1.txt
                original_obser1 = os.path.join(self.base_dir, "obser1.txt")
                temp_obser1 = os.path.join(temp_dir, "obser1_temp_1.txt")
                os.rename(original_obser1, temp_obser1)
                
                self.step_progress['value'] = 1
            
            # Process remaining obser files
            for i in range(2, 9):
                if self.processing_cancelled:
                    break
                
                self.update_status(f"🔄 Processing obser{i}.txt...", i, 3)
                self.log_enhanced_progress(f"Processing obser{i}.txt...")
                
                current_obser = os.path.join(self.base_dir, f"obser{i}.txt")
                if not os.path.exists(current_obser):
                    self.log_enhanced_progress(f"Warning: obser{i}.txt not found, skipping...", "WARNING")
                    continue
                
                # Rename and process
                os.rename(current_obser, original_obser1)
                self.run_fortran_program_safe()
                
                workbook_name = f"出力{i}.xlsx"
                self.write_to_excel_safe(self.base_dir, output_dir, files_to_write, workbook_name)
                
                # Store processed file
                temp_obser_path = os.path.join(temp_dir, f"obser1_temp_{i}.txt")
                os.rename(original_obser1, temp_obser_path)
                
                self.step_progress['value'] = i
            
            if not self.processing_cancelled:
                # Restore original files
                self.update_status("🔄 Restoring original files...", 8, 4)
                self.restore_original_files(temp_dir)
                
                # Create charts if requested
                if self.processing_settings['create_charts']:
                    self.update_status("📊 Creating charts...", 0, 5)
                    self.execute_chart_generation_safe(output_dir)
                
                # Cleanup
                self.cleanup_processing(temp_dir, backup_dir)
                
                self.update_status("✅ Processing completed successfully!", 8, 10)
                self.log_enhanced_progress("All processing completed successfully!", "SUCCESS")
                
                # Update UI
                self.cancel_btn.config(text="Close", bg="#4CAF50", state="normal", 
                                      command=self.progress_window.destroy)
                
                # Auto-close after delay and show completion
                self.root.after(2000, lambda: [self.progress_window.destroy(), 
                                              self.show_enhanced_completion_dialog()])
            else:
                self.handle_cancellation(temp_dir, backup_dir)
                
        except Exception as e:
            self.log_enhanced_progress(f"Critical error: {str(e)}", "ERROR")
            self.update_status("❌ Processing failed", 0, 0)
            self.cancel_btn.config(text="Close", bg="#f44336", state="normal", 
                                  command=self.progress_window.destroy)
            messagebox.showerror("Processing Error", f"Processing failed:\n{str(e)}")

    def setup_directories(self, output_dir, temp_dir, backup_dir):
        """Setup required directories"""
        self.log_enhanced_progress("Setting up directories...")
        
        # Clear and create output directory
        if os.path.exists(output_dir):
            shutil.rmtree(output_dir)
        os.makedirs(output_dir)
        
        # Create temp directory
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        os.makedirs(temp_dir)
        
        # Create backup directory if needed
        if backup_dir:
            if os.path.exists(backup_dir):
                shutil.rmtree(backup_dir)
            os.makedirs(backup_dir)
        
        self.log_enhanced_progress("Directories setup completed", "SUCCESS")

    def create_backup(self, backup_dir):
        """Create backup of original obser files"""
        self.log_enhanced_progress("Creating backup of original obser files...")
        
        for i in range(1, 9):
            obser_file = f"obser{i}.txt"
            source_path = os.path.join(self.base_dir, obser_file)
            backup_path = os.path.join(backup_dir, obser_file)
            
            if os.path.exists(source_path):
                shutil.copy2(source_path, backup_path)
        
        self.log_enhanced_progress("Backup created successfully", "SUCCESS")

    def run_fortran_program_safe(self):
        """Run Fortran program with error handling but no timeout"""
        try:
            result = subprocess.run([self.fortran_program_path], 
                                cwd=self.base_dir, 
                                check=True, 
                                capture_output=True, 
                                text=True)
            # Removed timeout=60 - let it run as long as needed
            
            if self.processing_settings['detailed_logging'] and result.stdout:
                self.log_enhanced_progress(f"Fortran output: {result.stdout.strip()}")
                    
        except subprocess.CalledProcessError as e:
            error_msg = f"Fortran program failed with return code {e.returncode}"
            if e.stderr:
                error_msg += f"\nError output: {e.stderr}"
            raise Exception(error_msg)
        except FileNotFoundError:
            raise Exception(f"Fortran program not found: {self.fortran_program_path}")

    def write_to_excel_safe(self, source_dir, output_dir, files_to_write, workbook_name):
        """Write text files to Excel workbook - copy data exactly as-is"""
        try:
            workbook = Workbook()
            
            for file_name in files_to_write:
                sheet_name = file_name.split('.')[0]
                sheet = workbook.create_sheet(title=sheet_name)
                file_path = os.path.join(source_dir, file_name)
                
                if os.path.exists(file_path):
                    with open(file_path, 'r', encoding='utf-8') as file:
                        for row_idx, line in enumerate(file):
                            line = line.strip()
                            if line:  # Skip empty lines
                                values = line.split()
                                for col_idx, value in enumerate(values):
                                    # Just copy the value as text - no conversion
                                    sheet.cell(row=row_idx + 1, column=col_idx + 1, value=value)
                else:
                    self.log_enhanced_progress(f"Warning: {file_name} not found", "WARNING")
            
            # Remove the default sheet
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
            
            excel_path = os.path.join(output_dir, workbook_name)
            workbook.save(excel_path)
            self.log_enhanced_progress(f"Created {workbook_name}", "SUCCESS")
            
        except Exception as e:
            raise Exception(f"Error writing to Excel {workbook_name}: {e}")

    def restore_original_files(self, temp_dir):
        """Restore original obser files"""
        self.log_enhanced_progress("Restoring original obser files...")
        
        for i in range(1, 9):
            temp_obser_path = os.path.join(temp_dir, f"obser1_temp_{i}.txt")
            original_path = os.path.join(self.base_dir, f"obser{i}.txt")
            
            if os.path.exists(temp_obser_path):
                os.rename(temp_obser_path, original_path)
        
        self.log_enhanced_progress("Original files restored", "SUCCESS")

    def execute_chart_generation_safe(self, output_folder):
        """Execute chart generation with error handling"""
        try:
            tamagawa_file = os.path.join(self.base_dir, 'tamagawa-new method  - Copy .xlsx')
            
            if not os.path.exists(tamagawa_file):
                self.log_enhanced_progress("tamagawa-new method  - Copy .xlsx not found, skipping chart generation", "WARNING")
                return
            
            self.log_enhanced_progress("Loading structure data...")
            
            # Load structure data with error handling
            try:
                wb = load_workbook(tamagawa_file, data_only=True)
                values_dict = {}
                
                for sheet in wb.worksheets:
                    values = {'構造物名称': [], '構造物番号': []}
                    try:
                        for row in sheet.iter_rows(min_row=2, values_only=True):
                            if row and len(row) >= 4:
                                values['構造物名称'].append(row[2] if row[2] else '')
                                values['構造物番号'].append(row[3] if row[3] else '')
                        values_dict[sheet.title] = values
                    except Exception as e:
                        self.log_enhanced_progress(f"Error reading sheet {sheet.title}: {e}", "WARNING")
                
                wb.close()
                
            except Exception as e:
                self.log_enhanced_progress(f"Error loading structure data: {e}", "WARNING")
                return
            
            # Process output files
            output_files = [f for f in os.listdir(output_folder) 
                          if f.endswith('.xlsx') and f.startswith('出力') 
                          and not f.startswith('作図付き')]
            
            for i, file in enumerate(output_files):
                if self.processing_cancelled:
                    break
                
                self.log_enhanced_progress(f"Creating charts for {file}...")
                file_path = os.path.join(output_folder, file)
                sheet_name = self.get_sheet_name(file)
                
                if sheet_name in values_dict:
                    try:
                        self.process_file_with_charts(file_path, values_dict[sheet_name], 
                                                    output_folder, self.year_ranges)
                        self.log_enhanced_progress(f"Charts created for {file}", "SUCCESS")
                    except Exception as e:
                        self.log_enhanced_progress(f"Error creating charts for {file}: {e}", "ERROR")
                else:
                    self.log_enhanced_progress(f"No structure data found for {sheet_name}", "WARNING")
                
                # Update progress
                chart_progress = ((i + 1) / len(output_files)) * 8
                self.step_progress['value'] = chart_progress
                self.progress_window.update()
            
        except Exception as e:
            self.log_enhanced_progress(f"Chart generation error: {e}", "ERROR")

    def cleanup_processing(self, temp_dir, backup_dir):
        """Cleanup temporary files"""
        self.log_enhanced_progress("Cleaning up temporary files...")
        
        # Remove temp directory
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        
        # Keep backup directory if created
        if backup_dir and os.path.exists(backup_dir):
            self.log_enhanced_progress(f"Backup files preserved in: {backup_dir}")
        
        self.log_enhanced_progress("Cleanup completed", "SUCCESS")

    def handle_cancellation(self, temp_dir, backup_dir):
        """Handle processing cancellation"""
        self.log_enhanced_progress("Handling cancellation...", "WARNING")
        
        try:
            # Try to restore original files if possible
            if os.path.exists(temp_dir):
                self.restore_original_files(temp_dir)
                shutil.rmtree(temp_dir)
            
            # Remove backup if created
            if backup_dir and os.path.exists(backup_dir):
                shutil.rmtree(backup_dir)
                
            self.update_status("❌ Processing cancelled", 0, 0)
            self.cancel_btn.config(text="Close", bg="#f44336", state="normal", 
                                  command=self.progress_window.destroy)
            
        except Exception as e:
            self.log_enhanced_progress(f"Error during cancellation cleanup: {e}", "ERROR")

    def clear_output_folder(self):
        """Clear the output folder with confirmation"""
        if not self.base_dir:
            messagebox.showwarning("Warning", "Please select a directory first")
            return
        
        output_dir = os.path.join(self.base_dir, "output")
        
        if not os.path.exists(output_dir):
            messagebox.showinfo("Info", "Output folder does not exist")
            return
        
        # Count files to be deleted
        files_to_clear = [f for f in os.listdir(output_dir) if f.endswith('.xlsx')]
        
        if not files_to_clear:
            messagebox.showinfo("Info", "Output folder is already empty")
            return
        
        # Confirm deletion
        if messagebox.askyesno("Confirm", 
                              f"Delete {len(files_to_clear)} files from output folder?\n\n"
                              f"This action cannot be undone."):
            try:
                cleared_count = 0
                for filename in files_to_clear:
                    file_path = os.path.join(output_dir, filename)
                    os.unlink(file_path)
                    cleared_count += 1
                
                messagebox.showinfo("Success", f"Cleared {cleared_count} files from output folder")
                
            except Exception as e:
                messagebox.showerror("Error", f"Error clearing files: {e}")

    def open_output_folder(self):
        """Open output folder in file explorer"""
        if not self.base_dir:
            messagebox.showwarning("Warning", "Please select a directory first")
            return
        
        output_dir = os.path.join(self.base_dir, "output")
        
        if not os.path.exists(output_dir):
            messagebox.showinfo("Info", "Output folder does not exist yet")
            return
        
        try:
            os.startfile(output_dir)
        except:
            messagebox.showinfo("Info", f"Please open folder manually:\n{output_dir}")

    def show_enhanced_completion_dialog(self):
        """Show enhanced completion dialog"""
        completion_window = tk.Toplevel(self.root)
        completion_window.title("Processing Complete")
        completion_window.geometry("600x500")
        completion_window.grab_set()
        completion_window.configure(bg="#f0f0f0")
        
        # Center window
        completion_window.update_idletasks()
        x = (completion_window.winfo_screenwidth() // 2) - (600 // 2)
        y = (completion_window.winfo_screenheight() // 2) - (500 // 2)
        completion_window.geometry(f"600x500+{x}+{y}")

                
        main_frame = tk.Frame(completion_window, bg="#f0f0f0", padx=25, pady=20)
        main_frame.pack(fill="both", expand=True)
        
        # Success header
        header_frame = tk.Frame(main_frame, bg="#f0f0f0")
        header_frame.pack(fill="x", pady=(0, 25))
        
        tk.Label(header_frame, text="🎉", font=("Arial", 36), bg="#f0f0f0").pack(side="left")
        tk.Label(header_frame, text="Processing Complete!", 
                font=("Arial", 18, "bold"), fg="#4CAF50", bg="#f0f0f0").pack(side="left", padx=(10, 0))
        
        # Summary statistics
        summary_frame = tk.LabelFrame(main_frame, text="Processing Summary", 
                                    font=("Arial", 12, "bold"), bg="#f0f0f0", padx=15, pady=10)
        summary_frame.pack(fill="x", pady=(0, 20))
        
        output_dir = os.path.join(self.base_dir, "output")
        basic_files = len([f for f in os.listdir(output_dir) if f.startswith('出力') and not f.startswith('作図付き')])
        chart_files = len([f for f in os.listdir(output_dir) if f.startswith('作図付き出力')])
        
        summary_text = (f"✅ Successfully processed {basic_files} obser files\n"
                       f"📊 Created {basic_files} basic Excel outputs\n"
                       f"📈 Created {chart_files} chart-enhanced outputs\n"
                       f"📁 All files saved to: output folder")
        
        tk.Label(summary_frame, text=summary_text, font=("Arial", 11), 
                justify="left", bg="#f0f0f0").pack(anchor="w")
        
        # File details
        details_frame = tk.LabelFrame(main_frame, text="Generated Files", 
                                    font=("Arial", 12, "bold"), bg="#f0f0f0", padx=15, pady=10)
        details_frame.pack(fill="both", expand=True, pady=(0, 20))
        
        # Scrollable file list
        files_text = tk.Text(details_frame, height=8, font=("Consolas", 9), bg="#f8f8f8")
        files_scrollbar = ttk.Scrollbar(details_frame, orient="vertical", command=files_text.yview)
        files_text.configure(yscrollcommand=files_scrollbar.set)
        
        files_text.pack(side="left", fill="both", expand=True)
        files_scrollbar.pack(side="right", fill="y")
        
        # List all generated files
        files_text.insert("1.0", "Basic Output Files:\n")
        for i in range(1, 9):
            filename = f"出力{i}.xlsx"
            if os.path.exists(os.path.join(output_dir, filename)):
                files_text.insert(tk.END, f"  ✅ {filename}\n")
            else:
                files_text.insert(tk.END, f"  ❌ {filename} (missing)\n")
        
        files_text.insert(tk.END, "\nChart-Enhanced Files:\n")
        for i in range(1, 9):
            filename = f"作図付き出力{i}.xlsx"
            if os.path.exists(os.path.join(output_dir, filename)):
                files_text.insert(tk.END, f"  ✅ {filename}\n")
            else:
                files_text.insert(tk.END, f"  ❌ {filename} (missing)\n")
        
        files_text.config(state="disabled")
        
        # Action buttons
        button_frame = tk.Frame(main_frame, bg="#f0f0f0")
        button_frame.pack(fill="x", pady=20)
        
        def open_output():
            try:
                os.startfile(output_dir)
                completion_window.after(1000, completion_window.destroy)
            except:
                messagebox.showinfo("Info", f"Please open folder manually:\n{output_dir}")
        
        def process_another():
            completion_window.destroy()
            self.reset_enhanced_app()
        
        def exit_app():
            completion_window.destroy()
            self.root.quit()
        
        tk.Button(button_frame, text="📁 Open Output Folder", 
                 command=open_output,
                 bg="#4CAF50", fg="white", width=18, height=2, 
                 font=("Arial", 11)).pack(side="left", padx=10)
        
        tk.Button(button_frame, text="🔄 Process Another", 
                 command=process_another,
                 bg="#2196F3", fg="white", width=18, height=2, 
                 font=("Arial", 11)).pack(side="left", padx=10)
        
        tk.Button(button_frame, text="❌ Exit Application", 
                 command=exit_app,
                 bg="#f44336", fg="white", width=18, height=2, 
                 font=("Arial", 11)).pack(side="left", padx=10)

    def reset_enhanced_app(self):
        """Reset application for processing another directory"""
        self.base_dir = None
        self.fortran_program_path = None
        self.year_ranges = None
        
        # Hide processing frame and reset status
        self.processing_frame.pack_forget()
        self.status_label.config(text="Ready to select directory...", fg="#666")
        
        # Reset settings to defaults
        self.start_year_var.set("1")
        self.end_year_var.set("5")
        self.create_charts_var.set(True)
        self.backup_var.set(True)
        self.detailed_log_var.set(True)

    def get_sheet_name(self, output_filename):
        """Get sheet name mapping for output file"""
        sheet_mapping = {
            '出力1.xlsx': '割算結果(補修考慮)',              # ← Fixed: Added parentheses
            '出力2.xlsx': '割算結果(補修無視)',              # ← Fixed: Added parentheses  
            '出力3.xlsx': '補修無視',                       # ← Already correct
            '出力4.xlsx': '補修考慮',                       # ← Already correct
            '出力5.xlsx': '新しい演算(補修無視)',            # ← Fixed: Added parentheses
            '出力6.xlsx': '新しい演算(補修考慮)',            # ← Fixed: Added parentheses
            '出力7.xlsx': '割算結果-新しい演算(補修無視)',    # ← Fixed: Added parentheses
            '出力8.xlsx': '割算結果-新しい演算(補修考慮)'     # ← Fixed: Added parentheses
        }
        return sheet_mapping.get(output_filename, '割算結果(補修考慮)')  # ← Fixed default too

    def process_file_with_charts(self, file_path, values, output_folder_path, year_ranges):
        """Process file and add charts"""
        try:
            wb = load_workbook(file_path)
            
            # Get or create sheets
            ex_ws = wb['ex1000'] if 'ex1000' in wb.sheetnames else None
            log_ws = wb['logdensity'] if 'logdensity' in wb.sheetnames else wb.create_sheet('logdensity')
            pml_ws = wb['pml'] if 'pml' in wb.sheetnames else None

            # Format sheets
            if ex_ws:
                self.format_ex1000(ex_ws, values)
            self.format_logdensity(log_ws)
            if pml_ws:
                self.format_pml(pml_ws)

            # Create new filename
            base_name = os.path.basename(file_path)
            if '_' in base_name:
                modified_filename = f"作図付き出力{base_name.split('_')[1].split('.')[0]}.xlsx"
            else:
                modified_filename = f"作図付き出力{base_name.split('.')[0][-1]}.xlsx"

            new_file_path = os.path.join(output_folder_path, modified_filename)

            # Save to temporary file first
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                temp_path = tmp.name
            wb.save(temp_path)
            wb.close()

            # Read data and create charts with xlsxwriter
            self.create_charts_with_xlsxwriter(temp_path, new_file_path, year_ranges)
            
            # Clean up temp file
            os.remove(temp_path)
            
        except Exception as e:
            raise Exception(f"Error processing file {file_path}: {e}")

    # def format_ex1000(self, ws, values):
    #     """Format ex1000 sheet"""
    #     if ws.cell(row=1, column=1).value != '順位':
    #         ws.insert_rows(1, amount=3)
    #         ws.cell(row=1, column=1, value='順位')
            
    #         for col_num in range(2, ws.max_column + 1):
    #             ws.cell(row=1, column=col_num, value=col_num - 1)
            
    #         ws.cell(row=2, column=1, value='構造物番号')
    #         for col_num, number in enumerate(values['構造物番号'], start=2):
    #             ws.cell(row=2, column=col_num, value=f'({number})')
            
    #         ws.cell(row=3, column=1, value='構造物名称')
    #         for col_num, name in enumerate(values['構造物名称'], start=2):
    #             ws.cell(row=3, column=col_num, value=name)
        
    #     # Format data
    #     for row in range(4, ws.max_row + 1):
    #         try:
    #             ws.cell(row=row, column=1).value = int(float(ws.cell(row=row, column=1).value))
    #         except:
    #             pass
    #         for col in range(2, ws.max_column + 1):
    #             try:
    #                 ws.cell(row=row, column=col).value = float(ws.cell(row=row, column=col).value)
    #             except:
    #                 ws.cell(row=row, column=col).value = None


    def format_ex1000(self, ws, values):
        """Format ex1000 sheet"""
        if ws.cell(row=1, column=1).value != '順位':
            ws.insert_rows(1, amount=3)
            ws.cell(row=1, column=1, value='順位')
            
            for col_num in range(2, ws.max_column + 1):
                ws.cell(row=1, column=col_num, value=col_num - 1)
            
            # Get the actual data from the Excel sheet to find correct column values
            # Read the original Excel file to get the structure data
            try:
                # Load the Excel file that was processed
                tamagawa_file = os.path.join(self.base_dir, 'tamagawa-new method  - Copy .xlsx')
                if os.path.exists(tamagawa_file):
                    # Find the correct sheet name for this output file
                    sheet_name = self.get_sheet_name(os.path.basename(ws.parent.path) if hasattr(ws.parent, 'path') else 'default')
                    
                    # Load the sheet data
                    df = pd.read_excel(tamagawa_file, sheet_name=sheet_name)
                    
                    # Find column indices
                    kouzou_bangou_col = None
                    kouzou_meisho_col = None
                    eki_hajime_col = None
                    eki_shuuryou_col = None
                    
                    for i, col_name in enumerate(df.columns):
                        col_str = str(col_name).strip()
                        if '構造物番号' in col_str:
                            kouzou_bangou_col = i
                        elif '構造物名称' in col_str:
                            kouzou_meisho_col = i
                        elif '駅（始）' in col_str or '駅(始)' in col_str:
                            eki_hajime_col = i
                        elif '駅（至）' in col_str or '駅(至)' in col_str:
                            eki_shuuryou_col = i
                    
                    # Row 2: 構造物番号
                    ws.cell(row=2, column=1, value='構造物番号')
                    if kouzou_bangou_col is not None:
                        for col_num in range(2, min(ws.max_column + 1, len(df) + 2)):
                            row_idx = col_num - 2  # Convert to DataFrame index
                            if row_idx < len(df):
                                bangou_value = df.iloc[row_idx, kouzou_bangou_col]
                                if pd.notna(bangou_value):
                                    ws.cell(row=2, column=col_num, value=f'({bangou_value})')
                                else:
                                    ws.cell(row=2, column=col_num, value='(-)')
                    else:
                        # Fallback to using values from tamagawa data
                        for col_num, number in enumerate(values['構造物番号'], start=2):
                            ws.cell(row=2, column=col_num, value=f'({number})')
                    
                    # Row 3: 構造物名称 or 駅（始）→駅（至）
                    ws.cell(row=3, column=1, value='構造物名称')
                    
                    # First try to use 構造物名称
                    if kouzou_meisho_col is not None:
                        for col_num in range(2, min(ws.max_column + 1, len(df) + 2)):
                            row_idx = col_num - 2
                            if row_idx < len(df):
                                meisho_value = df.iloc[row_idx, kouzou_meisho_col]
                                if pd.notna(meisho_value) and str(meisho_value).strip():
                                    ws.cell(row=3, column=col_num, value=str(meisho_value))
                                else:
                                    # If 構造物名称 is empty, try 駅（始）→駅（至）
                                    if eki_hajime_col is not None and eki_shuuryou_col is not None:
                                        hajime = df.iloc[row_idx, eki_hajime_col]
                                        shuuryou = df.iloc[row_idx, eki_shuuryou_col]
                                        if pd.notna(hajime) and pd.notna(shuuryou):
                                            combined_name = f"{hajime}→{shuuryou}"
                                            ws.cell(row=3, column=col_num, value=combined_name)
                                        else:
                                            ws.cell(row=3, column=col_num, value='-')
                                    else:
                                        ws.cell(row=3, column=col_num, value='-')
                    
                    # If no 構造物名称 column, use 駅（始）→駅（至）
                    elif eki_hajime_col is not None and eki_shuuryou_col is not None:
                        for col_num in range(2, min(ws.max_column + 1, len(df) + 2)):
                            row_idx = col_num - 2
                            if row_idx < len(df):
                                hajime = df.iloc[row_idx, eki_hajime_col]
                                shuuryou = df.iloc[row_idx, eki_shuuryou_col]
                                if pd.notna(hajime) and pd.notna(shuuryou):
                                    combined_name = f"{hajime}→{shuuryou}"
                                    ws.cell(row=3, column=col_num, value=combined_name)
                                else:
                                    ws.cell(row=3, column=col_num, value='-')
                    else:
                        # Fallback to original values
                        for col_num, name in enumerate(values['構造物名称'], start=2):
                            ws.cell(row=3, column=col_num, value=name)
                            
            except Exception as e:
                print(f"Error reading Excel data for formatting: {e}")
                # Fallback to original method
                ws.cell(row=2, column=1, value='構造物番号')
                for col_num, number in enumerate(values['構造物番号'], start=2):
                    ws.cell(row=2, column=col_num, value=f'({number})')
                
                ws.cell(row=3, column=1, value='構造物名称')
                for col_num, name in enumerate(values['構造物名称'], start=2):
                    ws.cell(row=3, column=col_num, value=name)
        
        # Format data (existing code)
        for row in range(4, ws.max_row + 1):
            try:
                ws.cell(row=row, column=1).value = int(float(ws.cell(row=row, column=1).value))
            except:
                pass
            for col in range(2, ws.max_column + 1):
                try:
                    ws.cell(row=row, column=col).value = float(ws.cell(row=row, column=col).value)
                except:
                    ws.cell(row=row, column=col).value = None



    def format_logdensity(self, ws):
        """Format logdensity sheet"""
        ws.insert_rows(1)
        for col in range(2, ws.max_column + 1):
            ws.cell(row=1, column=col, value=f'経過{col - 1}年目')
        
        for row in range(2, ws.max_row + 1):
            try:
                ws.cell(row=row, column=1).value = int(float(ws.cell(row=row, column=1).value))
            except:
                pass
            for col in range(2, ws.max_column + 1):
                try:
                    ws.cell(row=row, column=col).value = float(ws.cell(row=row, column=col).value)
                except:
                    ws.cell(row=row, column=col).value = None

    def format_pml(self, ws):
        """Format pml sheet"""
        ws.insert_cols(1)
        ws.insert_rows(1)
        ws.cell(row=1, column=1, value='経過年数')
        ws.cell(row=1, column=2, value='年')
        ws.cell(row=1, column=3, value='NEL (0.5:0.5)')
        ws.cell(row=1, column=4, value='PML(0.9:0.1)')
        ws.cell(row=1, column=5, value='PML_0.95 (0.95:0.05)')
        
        for i in range(2, ws.max_row + 1):
            ws.cell(row=i, column=1, value=i - 1)
        
        for col in range(2, ws.max_column + 1):
            for row in range(2, ws.max_row + 1):
                try:
                    ws.cell(row=row, column=col).value = int(float(ws.cell(row=row, column=col).value))
                except:
                    pass

    def create_charts_with_xlsxwriter(self, temp_path, new_file_path, year_ranges):
        """Create charts using xlsxwriter"""
        try:
            # Read data from temp file
            df_ex = None
            df_log = None
            df_pml = None
            
            try:
                df_ex = pd.read_excel(temp_path, sheet_name='ex1000', header=None)
            except:
                pass
            
            try:
                df_log = pd.read_excel(temp_path, sheet_name='logdensity', header=None)
            except:
                pass
            
            try:
                df_pml = pd.read_excel(temp_path, sheet_name='pml', header=0)
            except:
                pass

            # Create new workbook with charts
            with pd.ExcelWriter(new_file_path, engine='xlsxwriter') as writer:
                workbook = writer.book

                # Process ex1000 sheet with chart
                if df_ex is not None:
                    df_ex = df_ex.where(pd.notnull(df_ex), None)
                    ws = workbook.add_worksheet('ex1000')
                    writer.sheets['ex1000'] = ws

                    # Formatting
                    fmt_bold = workbook.add_format({'bold': True})
                    fmt_int = workbook.add_format({'num_format': '0'})
                    fmt_float = workbook.add_format({'num_format': '0.00000'})

                    # Write data
                    for row_num in range(df_ex.shape[0]):
                        for col_num in range(df_ex.shape[1]):
                            val = df_ex.iat[row_num, col_num]
                            if pd.isna(val):
                                ws.write(row_num, col_num, None)
                            elif row_num < 3:
                                ws.write(row_num, col_num, val, fmt_bold)
                            elif col_num == 0:
                                ws.write(row_num, col_num, val, fmt_int)
                            else:
                                ws.write(row_num, col_num, val, fmt_float)

                    # Create chart for ex1000
                    last = df_ex.iloc[3:].dropna(how='all').index[-1] if not df_ex.iloc[3:].dropna(how='all').empty else 3
                    chart = workbook.add_chart({'type': 'line'})
                    
                    for i in range(1, df_ex.shape[1]):
                        chart.add_series({
                            'name': ['ex1000', 2, i],
                            'categories': ['ex1000', 3, 0, last, 0],
                            'values': ['ex1000', 3, i, last, i],
                        })
                    
                    chart.set_title({'name': '経過年 vs. しきい値の強度確率'})
                    chart.set_x_axis({'name': '経過年', 'position_axis': 'on_tick'})
                    chart.set_y_axis({'name': 'しきい値の強度確率', 'num_format': '0%'})
                    ws.insert_chart(f'A{last + 6}', chart)

                # Process logdensity sheet with chart
                if df_log is not None:
                    df_log = df_log.where(pd.notnull(df_log), None)
                    df_log.to_excel(writer, sheet_name='logdensity', index=False, header=False)
                    ws2 = writer.sheets['logdensity']

                    # Formatting
                    col_a_format = workbook.add_format({'num_format': '0'})
                    col_rest_format = workbook.add_format({'num_format': '0.00000'})
                    ws2.set_column('A:A', 8, col_a_format)
                    end_col_letter = chr(ord('A') + df_log.shape[1] - 1)
                    ws2.set_column(f'B:{end_col_letter}', 12, col_rest_format)

                    # Create chart for logdensity
                    start_year, end_year = year_ranges
                    
                    if (start_year is not None and end_year is not None and 
                        start_year >= 1 and end_year <= df_log.shape[1] - 1 and 
                        start_year <= end_year):
                        
                        last2 = df_log.iloc[1:].dropna(how='all').index[-1] if not df_log.iloc[1:].dropna(how='all').empty else 1
                        chart2 = workbook.add_chart({'type': 'line'})
                        
                        for i in range(start_year, end_year + 1):
                            chart2.add_series({
                                'name': ['logdensity', 0, i],
                                'categories': ['logdensity', 1, 0, last2, 0],
                                'values': ['logdensity', 1, i, last2, i],
                            })
                        
                        chart2.set_title({'name': '経過年 vs. しきい値の強度確率'})
                        chart2.set_x_axis({'name': '劣化点数', 'position_axis': 'on_tick'})
                        chart2.set_y_axis({'name': '確率密度関数', 'num_format': '0.00000'})

                        last_col = chr(ord('A') + df_log.shape[1] - 1)
                        insert_col = chr(ord(last_col) + 3)
                        ws2.insert_chart(f'{insert_col}2', chart2)

                # Process pml sheet with chart
                if df_pml is not None:
                    df_pml = df_pml.where(pd.notnull(df_pml), None)
                    df_pml.to_excel(writer, sheet_name='pml', index=False, header=True)
                    ws3 = writer.sheets['pml']

                    # Formatting
                    col_a_format = workbook.add_format({'num_format': '0'})
                    col_rest_format = workbook.add_format({'num_format': '0'})
                    ws3.set_column('A:A', 8, col_a_format)
                    end_col_letter = chr(ord('A') + df_pml.shape[1] - 1)
                    ws3.set_column(f'B:{end_col_letter}', 12, col_rest_format)

                    # Create chart for pml
                    last3 = df_pml.shape[0]
                    chart3 = workbook.add_chart({'type': 'line'})
                    
                    for i in range(2, df_pml.shape[1]):
                        chart3.add_series({
                            'name': ['pml', 0, i],
                            'categories': ['pml', 1, 0, last3, 0],
                            'values': ['pml', 1, i, last3, i],
                            'marker': {'type': 'circle'}
                        })
                    
                    chart3.set_title({'name': 'PML Data'})
                    chart3.set_x_axis({'name': '経過年数', 'position_axis': 'on_tick'})
                    chart3.set_y_axis({'name': '劣化点数'})

                    last_col_pml = chr(ord('A') + df_pml.shape[1] - 1)
                    insert_col_pml = chr(ord(last_col_pml) + 3)
                    ws3.insert_chart(f'{insert_col_pml}2', chart3)

        except Exception as e:
            raise Exception(f"Error creating charts: {e}")

    def run(self):
        """Run the enhanced application"""
        self.root.mainloop()


def main():
    """Main function to choose between basic and enhanced versions"""
    choice_root = tk.Tk()
    choice_root.withdraw()  # Hide the root window
    
    choice = messagebox.askyesnocancel(
        "Version Selection",
        "Choose Post-Obser Processor Version:\n\n"
        "YES = Enhanced Version (recommended)\n"
        "  • Better error handling\n"
        "  • Detailed progress tracking\n" 
        "  • Backup options\n"
        "  • Advanced settings\n\n"
        "NO = Basic Version\n"
        "  • Simple interface\n"
        "  • Basic functionality\n\n"
        "CANCEL = Exit"
    )
    
    choice_root.destroy()
    
    if choice is True:
        print("Starting Enhanced Post-Obser Files Processor...")
        print("=" * 60)
        print("🚀 Enhanced Features:")
        print("• Advanced error handling and recovery")
        print("• Detailed progress tracking with logs")
        print("• Automatic backup of original files")
        print("• Configurable chart generation")
        print("• Better user interface")
        print("• Processing cancellation support")
        print("=" * 60)
        app = EnhancedPostProcessorApp()
    elif choice is False:
        print("Starting Basic Post-Obser Files Processor...")
        print("=" * 60)
        print("📋 Basic Features:")
        print("• Process obser1.txt to obser8.txt")
        print("• Run Fortran program for each file")
        print("• Generate Excel outputs with charts")
        print("• Simple progress tracking")
        print("=" * 60)
        app = PostObserProcessorApp()
    else:
        print("Exiting...")
        return
    
    try:
        app.run()
    except KeyboardInterrupt:
        print("\nApplication interrupted by user")
    except Exception as e:
        print(f"Application error: {e}")
        messagebox.showerror("Application Error", f"An unexpected error occurred:\n{str(e)}")


if __name__ == "__main__":
    print("Post-Obser Files Processor")
    print("========================")
    print("This tool processes obser files generated by the Obser File Generator")
    print("and creates enhanced Excel outputs with charts using Fortran programs.")
    print()
    print("Requirements:")
    print("• obser1.txt to obser8.txt (from Obser File Generator)")
    print("• 劣化予測プログラム .exe (Fortran program)")
    print("• tamagawa-new method  - Copy .xlsx (structure data, optional for charts)")
    print()
    print("Output:")
    print("• 8 basic Excel files (出力1.xlsx - 出力8.xlsx)")
    print("• 8 chart-enhanced files (作図付き出力1.xlsx - 作図付き出力8.xlsx)")
    print("========================")
    print()
    
    main()